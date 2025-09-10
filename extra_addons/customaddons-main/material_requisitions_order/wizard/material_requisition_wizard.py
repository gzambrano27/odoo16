# -*- coding: utf-8 -*-
# -*- encoding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.
from odoo import api,fields, models,_
from xlrd import open_workbook
import base64
import tempfile
import xlrd
from odoo.exceptions import ValidationError,UserError
from ...calendar_days.tools import CalendarManager,DateManager
from ...message_dialog.tools import FileManager
dtObj=DateManager()
clObj=CalendarManager()
flObj=FileManager()
from openpyxl import load_workbook

class MaterialRequisitionWizard(models.TransientModel):
    _name="material.requisition.wizard"
    _description="Asistente de Material de Requisicion"
    
    @api.model
    def get_default_movement_id(self):
        return self._context.get("active_ids",False) and self._context["active_ids"][0] or False
    
    @api.model
    def get_default_company_id(self):
        if self._context.get("active_ids",False):
            for brw_each in self.env["material.requisition.order"].sudo().browse(self._context["active_ids"]):
                return brw_each.company_id.id
        return False


    movement_id=fields.Many2one("material.requisition.order", "Requisicion",required=False,default=get_default_movement_id)
    company_id=fields.Many2one("res.company",string="Compañia",required=False,default=get_default_company_id)
    currency_id = fields.Many2one("res.currency", "Moneda", related="company_id.currency_id", store=False,
                                  readonly=True)
    origin=fields.Selection([('file','Archivo'),],string="Origen",default='file')
    file=fields.Binary("Archivo",required=False,filters='*.xlsx')
    file_name=fields.Char("Nombre de Archivo",required=False,size=255)
    
    

    def process(self):
        for brw_each in self:
            if brw_each.origin=="file":
                brw_each.process_file()
        return True
    
    def process_file(self):
        OBJ_PRODUCT = self.env["product.product"].sudo()
        OBJ_ANALITICA = self.env["account.analytic.account"].sudo()
        OBJ_LOCATION = self.env["stock.location"].sudo()
        OBJ_LINE = self.env["material.requisition.order.line"].sudo()
        ACTIVE_LINE_ID, Producto, Descripcion, Unidad, Cantidad = 0, 1, 2, 3, 4, 5
        for brw_each in self:
            line_ids = []
            ext = flObj.get_ext(brw_each.file_name)
            fileName = flObj.create(ext)
            flObj.write(fileName, flObj.decode64(brw_each.file))
    
            # Cargar el archivo con openpyxl
            book = load_workbook(fileName, data_only=True)
            sheet = book.active  # Selecciona la hoja activa
        
            for row_index, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):  # Comienza desde la fila 4
                line_value = row[ACTIVE_LINE_ID]
                line_id = 0
                if line_value:
                    if isinstance(line_value, (int, float)):
                        line_id = int(line_value)
                    else:
                        if line_value != "":
                            raise ValidationError(_("El # ID debe ser numérico o vacío. Revisa la fila %s") % (row_index,))
                # Este código se mantiene igual que el original, adaptando cómo se acceden a los valores de las celdas.
                default_code = str(row[Producto])
                qty_value = row[Cantidad]
                descripcion = row[Descripcion]
                quantity = int(qty_value)
                if not isinstance(qty_value, (int, float)):
                    raise ValidationError(_("La cantidad debe ser numérica. Revisa la fila %s") % (row_index,))
                
                srch_product = OBJ_PRODUCT.search([('default_code', '=', default_code)])
                if not srch_product:
                    raise ValidationError(_("No hay producto con código %s. Revisa la fila %s") % (default_code, row_index))
                if len(srch_product) > 1:
                    raise ValidationError(_("Existe más de un producto con código %s en la base de datos. Revisa la fila %s") % (default_code, row_index))    
                srch_cta = ''
                values = {
                    "qty": quantity,
                    "product_id": srch_product[0].id,
                    "description": descripcion,
                    'uom': srch_product[0].uom_po_id.id,
                }
                line_ids.append((1 if line_id > 0 else 0, line_id, values))
            brw_each.movement_id.write({"requisition_line_ids": line_ids})
        return True