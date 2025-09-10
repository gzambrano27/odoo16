# coding: utf-8
from odoo import api, fields, models, exceptions, tools, _
from ...message_dialog.tools import FileManager
from ...calendar_days.tools import DateManager
from ...calendar_days.tools import CalendarManager
dtFile=FileManager()        
dateO=DateManager()
calendarO=CalendarManager()
import openpyxl
import logging
_logger = logging.getLogger(__name__)

from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font, PatternFill
import openpyxl.utils as utils
from openpyxl.styles import NamedStyle
import datetime

class report_vencidos_pagar_xlsx(models.AbstractModel):
    _name = "report.gps_informes.report_vencidos_pagar_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Vencidos por Pagar"
    
    def create_xlsx_report(self, docids, data):
        def add_header(RANGE_COLUMNS,last_row,ultima_columna,headers):
            primera_letra_sum, ultima_letra_sum = False, False
            for j in RANGE_COLUMNS:
                column_name = headers[j].lower()
                if column_name in BLACK_LIST:
                    continue
                EACH_CELL = ws.cell(row=last_row, column=ultima_columna)
                EACH_CELL.value = columns_dscr.get(column_name, column_name).capitalize()
                EACH_CELL.fill = PATTERN_FILL  # Yellow background
                EACH_CELL.font = FONT_HEADER
                EACH_CELL.alignment = ALIGNMENT
                ultima_columna += 1
                if column_name == primera_letra_ky:
                    primera_letra_sum = utils.get_column_letter(ultima_columna - 1)
                if column_name == ultima_letra_ky:
                    ultima_letra_sum = utils.get_column_letter(ultima_columna - 1)
            EACH_CELL = ws.cell(row=last_row, column=ultima_columna)
            EACH_CELL.value = "Total"
            EACH_CELL.fill = PATTERN_FILL  # Yellow background
            EACH_CELL.font = FONT_HEADER
            EACH_CELL.alignment = ALIGNMENT
            return primera_letra_sum, ultima_letra_sum
        def sum_totals(values,total_columns):
            suma_total = 0
            # Iterar sobre cada diccionario en la lista
            for columna in total_columns:
                    # Verificar si la clave existe en el diccionario y sumarla
                if columna in values:
                    valor=values.get(columna,0.00)
                    if type(valor) in (float,int):
                        suma_total += valor
            return suma_total
        # def col_letter_to_index(col_letter):
        #     index = 0
        #     for char in col_letter:
        #         index = index * 26 + (ord(char.upper()) - 64)
        #     return index
        EXT="xlsx"
        dir_path=dtFile.get_path_file(__file__)
        new_filename=dtFile.create(EXT)  
        filename=dtFile.join(dir_path,"reporte_vencidos_pagar.xlsx")
        dtFile.copyfile(filename,new_filename)
        wb=False
        try:
            brw_wizard=self.env["l10n_ec.account.report"].sudo().browse(docids[-1])
            range_days = brw_wizard.range_days
            periods = brw_wizard.periods
            companys = ', '.join(brw_wizard.mapped('company_ids').mapped('name'))
            company_ids = brw_wizard.company_ids.ids+[-1,-1]
            COLUMN_BASE_QUERY = f"""SELECT
	S.START_DAY +1::int  AS START_DAY,
	CASE
		WHEN (S.START_DAY + S.DIAS_RANGO) > S.DIAS_RANGO * S.PERIODOS THEN 99999999
		ELSE S.START_DAY + S.DIAS_RANGO
	END::int END_DAY 
FROM
	(   
		SELECT
			GENERATE_SERIES(0, {periods} * {range_days}, {range_days}) AS START_DAY,
			{range_days}::INT AS DIAS_RANGO,
			{periods}::INT PERIODOS
	) AS S
UNION
SELECT
	-99999999::int AS START_DAY,
	0::int AS END_DAY
ORDER BY 1 ASC   """
            QUERY_RANGE=(f"SELECT X.*,X.START_DAY::VARCHAR AS DSCR  FROM ({COLUMN_BASE_QUERY} ) X   ORDER BY X.START_DAY::int ASC")
            self._cr.execute(QUERY_RANGE )
            result_columns = self._cr.fetchall()
            columns_dscr = {}
            COLUMN_DATA_TYPE = ""
            primera_letra_ky,ultima_letra_ky,total_columns=False,False,[]
            if result_columns:
                for start_day,end_day,dscr in result_columns:
                    if not primera_letra_ky:
                        primera_letra_ky=dscr
                    COLUMN_DATA_TYPE += f',"{dscr}" float '
                    dscr_tiempo="De %s a %s" % (start_day,end_day)
                    if start_day<=0:
                        dscr_tiempo = "Vencido"
                    if end_day>=99999999:
                        dscr_tiempo = "Más Antiguo"
                    columns_dscr[dscr.lower()] = dscr_tiempo
                    total_columns.append(dscr.lower())
                    ultima_letra_ky = dscr
            columns_dscr["company_name"] = "Empresa"
            columns_dscr["tipo"] = "Tipo"
            columns_dscr["partner_name"] = "Proveedor"
            columns_dscr["partner_vat"] = "# Identificación"
            columns_dscr["journal_name"] = "Diario"
            columns_dscr["name"] = "# Documento"
            columns_dscr["invoice_date"] = "Fecha de Documento"
            columns_dscr["payment_term_dscr"] = "Terminos de Pago"
            columns_dscr["internal_id"] = "# ID"
            #####
            QUERY_COLUMNS = (f"SELECT X.START_DAY::VARCHAR AS DSCR "
                             f" FROM ({COLUMN_BASE_QUERY} ) X  "
                             f" ORDER BY X.START_DAY::INT ASC")

            QUERY_VALUES=f"""SELECT
                                'move,'||AM.ID::VARCHAR AS DOC_ID,
                                PERIODS.DSCR,
                                -1.00*SUM(AML.AMOUNT_RESIDUAL) AS AMOUNT_RESIDUAL
                            FROM
                                ACCOUNT_MOVE AM
                                INNER JOIN ACCOUNT_MOVE_LINE AML ON AML.MOVE_ID = AM.ID
                                INNER JOIN ACCOUNT_ACCOUNT AA ON AA.ID = AML.ACCOUNT_ID
                                AND AA.ACCOUNT_TYPE = 'liability_payable'
                                INNER JOIN ({QUERY_RANGE} ) PERIODS ON  (NOW()::DATE-  COALESCE(AML.DATE_MATURITY,AM.DATE,NOW()::DATE) )    BETWEEN   PERIODS.START_DAY    AND  PERIODS.END_DAY									
                            WHERE
                                AM.STATE = 'posted'
                                AND AM.COMPANY_ID = ANY (ARRAY{ company_ids} )
                            GROUP BY
                                AM.ID,PERIODS.DSCR 
                            HAVING
                                SUM(AML.AMOUNT_RESIDUAL) != 0.00
                            UNION                             
                            select 'purchase,'||x.order_id::VARCHAR AS DOC_ID,
                                PERIODS.DSCR,
                                sum(x.AMOUNT_RESIDUAL) as AMOUNT_RESIDUAL from
                            (
                                select po.id as order_id,
                                case when(ptm.index=1) then  po.date_advance_payment 
                                    else (po.date_approve + PTM.interval_time)::date end as date_maturity,
                                round(po.to_invoice *ptm.final_amount/100.00,2) AS AMOUNT_RESIDUAL
                                from (                                
                                    select po.id as id ,
											po.date_advance_payment ::DATE as date_advance_payment,
											po.date_approve::DATE as date_approve,
											po.payment_Term_id,
                                       sum( 
										(case when(pol.product_qty>0) then 
											pol.price_total /pol.product_qty else 0.00 end) * 
										(pol.product_qty-pol.qty_invoiced) 
										
										) as to_invoice 
                                    from purchase_order po	
                                    inner join purchase_order_line pol on po.id=pol.order_id
                                    
                                     where po.state in ('purchase','done')  and po.invoice_status!='invoiced' 
                                           AND po.COMPANY_ID = ANY (ARRAY{ company_ids})  
                                    group by 
                                    po.id  ,
											po.date_advance_payment,
											po.date_approve::DATE ,
											po.payment_Term_id,po.amount_total
                                    having sum( 
										(case when(pol.product_qty>0) then 
											pol.price_total /pol.product_qty else 0.00 end) * 
										(pol.product_qty-pol.qty_invoiced) 
										
										)>0.00 
                                ) po 	
                                inner join (
                                    SELECT 
                                    ROW_NUMBER() OVER (PARTITION BY APT.ID ORDER BY CASE 
                                        WHEN (COALESCE(APTL.DAYS, 0) > 0) THEN (COALESCE(APTL.DAYS, 0) || ' days') 
                                        WHEN (COALESCE(APTL.MONTHS, 0) > 0) THEN (COALESCE(APTL.MONTHS, 0) || ' months') 
                                        ELSE '0 days'
                                    END::INTERVAL ) as index  ,
                                    APT.ID,
                                    COALESCE(APTL.MONTHS, 0) AS MONTHS,
                                    COALESCE(APTL.DAYS, 0) AS DAYS,
                                    APTL.VALUE,
                                    case when (APTL.VALUE='balance') then 0 else APTL.VALUE_AMOUNT end as VALUE_AMOUNT ,
                                    CASE 
                                        WHEN (COALESCE(APTL.DAYS, 0) > 0) THEN (COALESCE(APTL.DAYS, 0) || ' days') 
                                        WHEN (COALESCE(APTL.MONTHS, 0) > 0) THEN (COALESCE(APTL.MONTHS, 0) || ' months') 
                                        ELSE '0 days'
                                    END::INTERVAL AS interval_time   ,
                                    round((case when (APTL.VALUE='balance') then 100.00- (
                                        SUM(case when (APTL.VALUE='balance') then 0 else APTL.VALUE_AMOUNT end ) OVER (PARTITION BY APT.ID ORDER BY CASE 
                                        WHEN (COALESCE(APTL.DAYS, 0) > 0) THEN (COALESCE(APTL.DAYS, 0) || ' days') 
                                        WHEN (COALESCE(APTL.MONTHS, 0) > 0) THEN (COALESCE(APTL.MONTHS, 0) || ' months') 
                                        ELSE '0 days'
                                        END::INTERVAL)-(case when (APTL.VALUE='balance') then 0 else APTL.VALUE_AMOUNT end)
                                    ) else APTL.VALUE_AMOUNT end),2)	 as final_amount
                                    FROM 
                                        ACCOUNT_PAYMENT_TERM APT
                                    INNER JOIN 
                                        ACCOUNT_PAYMENT_TERM_LINE APTL 
                                        ON APTL.PAYMENT_ID = APT.ID
                                        where coalesce(apt.required_advance_payment,false)=true 
                                ) ptm on ptm.id=po.payment_Term_id  
                                
                            ) x 
                            INNER JOIN ({QUERY_RANGE} ) PERIODS ON  (x.date_maturity-NOW()::DATE )    BETWEEN   PERIODS.START_DAY    AND  PERIODS.END_DAY									
                            group by x.order_id,PERIODS.DSCR
                            HAVING
                                SUM(x.AMOUNT_RESIDUAL) != 0.00
	                   union	                    
	                    select 'bank,'||x.id::VARCHAR AS DOC_ID,
                                PERIODS.DSCR,
                                sum(x.to_paid) as AMOUNT_RESIDUAL from
                        (
                                select dbk.id ,
                                        dbkl.date_process ::DATE as date_maturity,
                                        dbk.date_process::DATE as date_approve,
                                        null payment_Term_id,
                                         sum(  dbkl.total_pending ) as to_paid 
                                from document_financial dbk
                                inner join document_financial_line dbkl on dbkl.document_id=dbk.id 
                                where dbk.state='posted' AND dbk.COMPANY_ID = ANY (ARRAY{ company_ids})  
                                and  dbkl.total_pending!=0.00 and dbk.internal_type='out' 
                                group by dbk.id , dbkl.date_process, dbk.date_process
                        
                        ) x 
                        INNER JOIN ({QUERY_RANGE} ) PERIODS ON  (x.date_maturity-NOW()::DATE ) 
                           BETWEEN   PERIODS.START_DAY    AND  PERIODS.END_DAY									
                        group by x.id,PERIODS.DSCR
                        HAVING SUM(x.to_paid) != 0.00 
	                    """

            QUERY_SUM_VALUES=f""" union select 'total,'||(string_to_array(x.DOC_ID, ','))[1] AS DOC_ID,
                                x.DSCR,
                                sum(x.AMOUNT_RESIDUAL) as AMOUNT_RESIDUAL from
                        (
{QUERY_VALUES}
                        ) x group by x.DSCR,(string_to_array(x.DOC_ID, ','))[1]  union 
                select 'total,total' AS DOC_ID,
                                x.DSCR,
                                sum(x.AMOUNT_RESIDUAL) as AMOUNT_RESIDUAL from
                        (
{QUERY_VALUES}
                        ) x group by x.DSCR 

"""

            #####
            ACCOUNT_REPORT_QUERY = f""";WITH TOTAL_FOR_PAYMENT AS (
                        SELECT * FROM CROSSTAB(
                          $$  
	                  {QUERY_VALUES}  {QUERY_SUM_VALUES} order by 1 asc,2 asc 
                        $$,
                        $$ 
                            {QUERY_COLUMNS}
                        $$
                            ) AS CT(
                        DOC_ID VARCHAR {COLUMN_DATA_TYPE}   
                            ) 
                        )

            SELECT 0::INT AS SEQUENCE_FORMAT,            
            RC.NAME AS COMPANY_NAME,
            'DOCUMENTO CONTABLE'::VARCHAR AS TIPO,
            COALESCE(RP.NAME,'DESCONOCIDO') AS PARTNER_NAME,
			COALESCE(RP.VAT,'') AS PARTNER_VAT,
			AJ.NAME AS JOURNAL_NAME,
			AM.NAME,
			COALESCE(AM.INVOICE_DATE,AM.DATE) AS INVOICE_DATE, 
			APT.NAME::JSON->>'es_EC'::varchar  AS PAYMENT_TERM_DSCR,
			AM.ID AS INTERNAL_ID,
			I.*             
            FROM TOTAL_FOR_PAYMENT I 
            INNER JOIN ACCOUNT_MOVE AM ON 'move,'||AM.ID=I.DOC_ID  
            LEFT JOIN ACCOUNT_PAYMENT_TERM APT ON APT.ID=AM.INVOICE_PAYMENT_TERM_ID
            INNER JOIN RES_COMPANY RC ON RC.ID=AM.COMPANY_ID 
            INNER JOIN ACCOUNT_JOURNAL AJ ON AJ.ID=AM.JOURNAL_ID 
            LEFT JOIN RES_PARTNER RP ON RP.ID=AM.PARTNER_ID
            WHERE I.DOC_ID LIKE 'move%'
			UNION			
			SELECT 1::INT AS SEQUENCE_FORMAT,            
            RC.NAME AS COMPANY_NAME,
            'ORDENES DE COMPRA'::VARCHAR AS TIPO,
            COALESCE(RP.NAME,'DESCONOCIDO') AS PARTNER_NAME,
			COALESCE(RP.VAT,'') AS PARTNER_VAT,
			'' AS JOURNAL_NAME,
			PO.NAME,
			PO.DATE_APPROVE::DATE AS INVOICE_DATE,
			APT.NAME::JSON->>'es_EC'::varchar  AS PAYMENT_TERM_DSCR,
			PO.ID AS INTERNAL_ID,
			I.*             
            FROM TOTAL_FOR_PAYMENT I 
            INNER JOIN PURCHASE_ORDER PO ON 'purchase,'||PO.ID=I.DOC_ID  
            INNER JOIN ACCOUNT_PAYMENT_TERM APT ON APT.ID=PO.PAYMENT_TERM_ID
            INNER JOIN RES_COMPANY RC ON RC.ID=PO.COMPANY_ID  
            INNER JOIN RES_PARTNER RP ON RP.ID=PO.PARTNER_ID
            WHERE I.DOC_ID LIKE 'purchase%'
            
            UNION
			
			SELECT 2::INT AS SEQUENCE_FORMAT,            
            RC.NAME AS COMPANY_NAME,
            'OPERACION FINANCIERA'::VARCHAR AS TIPO,
            COALESCE(RP.NAME,'DESCONOCIDO') AS PARTNER_NAME,
			COALESCE(RP.VAT,'') AS PARTNER_VAT,
			case when(dbk.type='pagares') then 'PAGARES'
			 when(dbk.type='emision') then 'EMISION DE OBLIGACIONES'
			 when(dbk.type='prestamo') then 'OBLIGACIONES BANCARIAS' else '' end AS JOURNAL_NAME, 
			dbk.NAME,
			dbk.DATE_process::DATE AS INVOICE_DATE,
			(dbk.YEARS::VARCHAR)||' AÑOS'  AS PAYMENT_TERM_DSCR,
			dbk.ID AS INTERNAL_ID,
			I.*             
            FROM TOTAL_FOR_PAYMENT I 
            INNER JOIN document_financial dbk ON 'bank,'||dbk.ID=I.DOC_ID   
            INNER JOIN RES_COMPANY RC ON RC.ID=dbk.COMPANY_ID  
            INNER JOIN RES_PARTNER RP ON RP.ID=dbk.PARTNER_ID
            WHERE I.DOC_ID LIKE 'bank%'
            
            
            UNION
			
			SELECT 3::INT AS SEQUENCE_FORMAT,            
            '' AS COMPANY_NAME,
            ''::VARCHAR AS TIPO,
            '' AS PARTNER_NAME,
			'' AS PARTNER_VAT,
			'' AS JOURNAL_NAME,
			CASE WHEN(I.DOC_ID='total,move') THEN  'ASIENTO,FACTURAS...' 
			    WHEN(I.DOC_ID='total,purchase') THEN  'ORDENES DE COMPRAS' 
			    WHEN(I.DOC_ID='total,bank') THEN  'OPERACIONES FINANCIERAS'
			    WHEN(I.DOC_ID='total,total') THEN  'TOTAL' END as NAME,
			null AS INVOICE_DATE,
			''  AS PAYMENT_TERM_DSCR,
			null AS INTERNAL_ID,
			I.*             
            FROM TOTAL_FOR_PAYMENT I 
            WHERE I.DOC_ID LIKE 'total%'
            
            ORDER BY 1 ASC,4 asc,3 asc
"""
            self._cr.execute(ACCOUNT_REPORT_QUERY)
            headers = [d.name for d in self._cr.description]

            result_account_report = self._cr.dictfetchall()
            wb, ws, target = self.open_xlsx(new_filename, load_sheet=True)
            i, INDEX_ROW = 0, 4
            last_row = INDEX_ROW
            ultima_columna_registros = 1
            COUNT = len(result_account_report)
            SUM_GROUPS={}
            if result_account_report:
                RANGE_COLUMNS = range(1, len(headers))
                PATTERN_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9',
                                           fill_type='solid')
                FONT_HEADER = Font( size=12, bold=True, color='000000')
                ALIGNMENT = Alignment(horizontal='center', vertical='center')
                ALIGNMENT_RIGHT = Alignment(horizontal='right', vertical='center')
                ALIGNMENT_LEFT = Alignment(horizontal='left', vertical='center')

                DECIMAL_STYLE = NamedStyle(name="decimal_style", number_format="0.00")

                BLACK_LIST = [ "doc_id"]
                FONT_BOLD = Font(size=12, bold=True)
                FONT_RED = Font( size=12, color="FF0000")
                FONT_RED_BOLD= Font( size=12, color="FF0000", bold=True)
                FONT_BOLD_LIST = ["total"]
                ultima_columna = 1
                primera_letra_sum,ultima_letra_sum=add_header(RANGE_COLUMNS, last_row, ultima_columna,headers)
                last_row += 1
                sequence_format3=0
                for each_payslip in result_account_report:
                    if each_payslip["sequence_format"]==3:
                        if sequence_format3==0:
                            last_row += 1
                            last_row += 1
                            add_header(RANGE_COLUMNS, last_row, ultima_columna,headers)
                            last_row += 1
                        sequence_format3+=1
                    ultima_columna_registros = 1
                    for j in RANGE_COLUMNS:
                        column_name = headers[j].lower()
                        if column_name in BLACK_LIST:
                            continue
                        ROW_EACH_CELL = ws.cell(row=last_row, column=ultima_columna_registros)
                        value_write = each_payslip[column_name]
                        if type(value_write) == str:
                            value_write = each_payslip[column_name].replace('"', '').upper()
                        ROW_EACH_CELL.value = value_write
                        if type(value_write) == str:
                            ROW_EACH_CELL.alignment = ALIGNMENT_LEFT
                        if type(value_write) in (float, int):
                            ROW_EACH_CELL.alignment = ALIGNMENT_RIGHT
                            if type(value_write) == float:
                                ROW_EACH_CELL.style = DECIMAL_STYLE
                            if value_write<0:
                                ROW_EACH_CELL.font = FONT_RED
                        if column_name in FONT_BOLD_LIST or each_payslip["sequence_format"] == 3:#
                            ROW_EACH_CELL.font = FONT_BOLD
                        ultima_columna_registros += 1
                    if primera_letra_sum and ultima_letra_sum:
                        ROW_EACH_CELL_TOTAL = ws.cell(row=last_row, column=ultima_columna_registros)
                        total_linea=sum_totals(each_payslip,total_columns)
                        ROW_EACH_CELL_TOTAL.value= total_linea
                        ROW_EACH_CELL_TOTAL.alignment = ALIGNMENT_RIGHT
                        ROW_EACH_CELL_TOTAL.style = DECIMAL_STYLE
                        if total_linea<0:
                            ROW_EACH_CELL_TOTAL.font = FONT_RED_BOLD
                        else:
                            ROW_EACH_CELL_TOTAL.font = FONT_BOLD
                        #total_columns
                    last_row += 1


            ws['A1'] = companys
            ws['B2'] = range_days
            ws['E2'] = periods
            ws['H2'] = COUNT
            ultima_letra = utils.get_column_letter(ultima_columna_registros)
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                #self.set_border(ws, 'A' + str(INDEX_ROW) + ':' + ultima_letra + str(last_row - 1), border)

            wb=self.save_wb(wb,target)  
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb=self.close_wb(wb)
        filecontent=dtFile.get_binary(new_filename)        
        return filecontent, EXT

class report_pagos_anticipados_xlsx(models.AbstractModel):
    _name = "report.gps_informes.report_pagos_anticipados_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Pagos Anticipados"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)
        filename = dtFile.join(dir_path, "reporte_pagos_anticipados.xlsx")
        dtFile.copyfile(filename, new_filename)
        wb = False
        try:
            brw_wizard = self.env["l10n_ec.account.report"].sudo().browse(docids[-1])
            companys = ', '.join(brw_wizard.mapped('company_ids').mapped('name'))
            company_ids = brw_wizard.company_ids.ids + [-1, -1]
            date_from=brw_wizard.date_from
            date_to = brw_wizard.date_to
            type_view_balance=brw_wizard.type_view_balance
            ACCOUNT_REPORT_QUERY=""";WITH VARIABLES AS (
	SELECT %s::INT[] AS COMPANY_IDS,
		'ANTICIPOS A PROVEEDOR'::VARCHAR(50) AS DSCR_PARENT_ACCT,
		'CUENTAS POR PAGAR PROVEEDORES'::VARCHAR(50) AS DSCR_CXP_PARENT_ACCT,
		%s::DATE AS DATE_FROM,
		%s::DATE AS DATE_TO,
		%s::VARCHAR(50) AS type_view_balance
),
CUENTAS AS (
	SELECT AAC.ID,AAC.NAME FROM 
		VARIABLES 
		INNER JOIN ACCOUNT_ACCOUNT AA ON AA.NAME=VARIABLES.DSCR_PARENT_ACCT AND AA.COMPANY_ID=ANY(VARIABLES.COMPANY_IDS) 
		INNER JOIN ACCOUNT_ACCOUNT AAC ON AAC.PARENT_ID=AA.ID 
	UNION 
	SELECT AAC.ID,AAC.NAME FROM 
		VARIABLES 
		INNER JOIN ACCOUNT_ACCOUNT AA ON AA.NAME=VARIABLES.DSCR_CXP_PARENT_ACCT AND AA.COMPANY_ID=ANY(VARIABLES.COMPANY_IDS) 
		INNER JOIN ACCOUNT_ACCOUNT AAC ON AAC.PARENT_ID=AA.ID 
) ,
REGISTRO_ASIENTOS AS (

	SELECT 
		AML.ID AS LINE_ID,
		AM.ID AS MOVE_ID,
		AML.DEBIT AS DEBIT,
		AML.CREDIT AS CREDIT,
		AML.BALANCE AS BALANCE,
		AML.AMOUNT_RESIDUAL AS AMOUNT_RESIDUAL 
		FROM
		VARIABLES 
		INNER JOIN ACCOUNT_MOVE AM ON AM.COMPANY_ID=ANY(VARIABLES.COMPANY_IDS) AND AM.STATE='posted'
		INNER JOIN ACCOUNT_MOVE_LINE AML ON AML.MOVE_ID=AM.ID 
		INNER JOIN CUENTAS CTAS ON CTAS.ID=AML.ACCOUNT_ID 
		INNER JOIN RES_COMPANY RC ON RC.ID=AM.COMPANY_ID 
		INNER JOIN ACCOUNT_JOURNAL AJ ON AJ.ID=AM.JOURNAL_ID 
		INNER JOIN ACCOUNT_ACCOUNT AA ON AA.ID=AML.ACCOUNT_ID
		WHERE 
		(
		    ((VARIABLES.DATE_FROM IS NOT NULL AND VARIABLES.DATE_TO IS NOT NULL) 
		    AND (AM.DATE>=VARIABLES.DATE_FROM AND AM.DATE<=VARIABLES.DATE_TO) )
		    OR 
		    ((VARIABLES.DATE_FROM IS NOT NULL AND VARIABLES.DATE_TO IS NULL) 
		    AND (AM.DATE>=VARIABLES.DATE_FROM ) )
		    OR 
		    ((VARIABLES.DATE_TO IS NOT NULL AND VARIABLES.DATE_FROM IS NULL) 
		    AND (AM.DATE<=VARIABLES.DATE_TO ) )
		    or 
			(VARIABLES.DATE_FROM IS NULL AND VARIABLES.DATE_TO IS NULL) 
		 )  
),
REGISTRO_SALDOS_aSIENTOS AS (

	SELECT X.LINE_ID,SUM(X.SALDO ) AS SALDO FROM 
	(
		SELECT
		APR.DEBIT_MOVE_ID AS LINE_ID,ROUND(SUM(APR.AMOUNT),2) AS SALDO
		FROM ACCOUNT_PARTIAL_rECONCILE APR
		INNER JOIN ACCOUNT_MOVE_LINE AMLC ON AMLC.ID=APR.CREDIT_MOVE_ID 
		INNER JOIN ACCOUNT_MOVE AM ON AM.ID=AMLC.MOVE_ID
		INNER JOIN REGISTRO_ASIENTOS R ON R.LINE_ID=APR.DEBIT_MOVE_ID
		INNER JOIN VARIABLES ON AM.DATE<=VARIABLES.DATE_TO 
		GROUP BY APR.DEBIT_MOVE_ID
		UNION
		SELECT
		APR.CREDIT_MOVE_ID AS LINE_ID,ROUND(SUM(APR.AMOUNT),2) AS SALDO
		FROM ACCOUNT_PARTIAL_rECONCILE APR 
		INNER JOIN ACCOUNT_MOVE_LINE AMLD ON AMLD.ID=APR.DEBIT_MOVE_ID
		INNER JOIN ACCOUNT_MOVE AM ON AM.ID=AMLD.MOVE_ID
		INNER JOIN REGISTRO_ASIENTOS R ON R.LINE_ID=APR.CREDIT_MOVE_ID 
		INNER JOIN VARIABLES ON AM.DATE<=VARIABLES.DATE_TO 
		GROUP BY APR.CREDIT_MOVE_ID
	) X
	
	GROUP BY X.LINE_ID
),
SALDOS_FECHA AS (
	SELECT RA.LINE_ID,RA.MOVE_ID,
	RA.DEBIT,
	RA.CREDIT,
	RA.BALANCE,
	RAS.SALDO,
	CASE WHEN(RA.DEBIT>0) THEN (RA.DEBIT-COALESCE(RAS.SALDO,0.00))
		ELSE
		-(RA.CREDIT-COALESCE(RAS.SALDO,0.00))
	END AS RESIDUAL
	FROM REGISTRO_ASIENTOS RA
	LEFT JOIN REGISTRO_SALDOS_aSIENTOS  RAS ON RA.LINE_ID=RAS.LINE_ID
)
	
SELECT 
'ANTICIPO'::VARCHAR AS TIPO_TRANSACCION,
RC.NAME as COMPANY_NAME,
AJ.NAME AS JOURNAL_NAME,
AML.ID AS LINE_ID,
AM.ID AS MOVE_ID,
AM.DATE AS MOVE_DATE,
AM.NAME AS MOVE_NAME,
AM.REF AS MOVE_REF,
AML.DEBIT AS LINE_DEBIT,
AML.CREDIT AS LINE_CREDIT,
COALESCE(AML.AMOUNT_RESIDUAL,0.00) AS LINE_RESIDUAL,
COALESCE(SF.RESIDUAL,0.00) SALDO_FECHA_CORTE,
AML.BALANCE AS LINE_BALANCE,
AP.ID AS PAYMENT_ID,
COALESCE(AP.IS_PREPAYMENT,FALSE) AS IS_PREPAYMENT,
AP.AMOUNT AS PAYMENT_AMOUNT,
AP.AMOUNT_RESIDUAL AS PAYMENT_RESIDUAL,
RP.NAME AS PARTNER_NAME,
RP.VAT AS PARTNER_VAT,
AA.CODE AS ACCOUNT_CODE,
AA.NAME AS ACCOUNT_NAME,
COALESCE(AM.PREPAYMENT_ASSIGNMENT,FALSE) AS ASIGNACION_ANTICIPADO,
COALESCe(AP.REF2,'') AS REFERENCIA2
FROM
VARIABLES 
INNER JOIN ACCOUNT_MOVE AM ON AM.COMPANY_ID=ANY(VARIABLES.COMPANY_IDS) AND AM.STATE='posted'
INNER JOIN ACCOUNT_MOVE_LINE AML ON AML.MOVE_ID=AM.ID 
inner JOIN CUENTAS CTAS ON CTAS.ID=AML.ACCOUNT_ID 
INNER JOIN RES_COMPANY RC ON RC.ID=AM.COMPANY_ID 
INNER JOIN ACCOUNT_JOURNAL AJ ON AJ.ID=AM.JOURNAL_ID 
INNER JOIN ACCOUNT_ACCOUNT AA ON AA.ID=AML.ACCOUNT_ID
LEFT JOIN SALDOS_FECHA SF ON SF.LINE_ID=AML.ID 
LEFT JOIN ACCOUNT_PAYMENT AP ON AP.MOVE_ID=AM.ID
LEFT JOIN RES_PARTNER RP ON RP.ID=COALESCE(AP.PARTNER_ID,AML.PARTNER_ID,AM.PARTNER_ID) 
WHERE 
(
    (VARIABLES.type_view_balance='all' ) 
      or 
    (VARIABLES.type_view_balance='with_balance' and COALESCE(SF.RESIDUAL,0.00)!=0.00)
) AND 
 (
    ((VARIABLES.DATE_FROM IS NOT NULL AND VARIABLES.DATE_TO IS NOT NULL) 
    AND (AM.DATE>=VARIABLES.DATE_FROM AND AM.DATE<=VARIABLES.DATE_TO) )
    OR 
    ((VARIABLES.DATE_FROM IS NOT NULL AND VARIABLES.DATE_TO IS NULL) 
    AND (AM.DATE>=VARIABLES.DATE_FROM ) )
    OR 
    ((VARIABLES.DATE_TO IS NOT NULL AND VARIABLES.DATE_FROM IS NULL) 
    AND (AM.DATE<=VARIABLES.DATE_TO ) )
    or 
	(VARIABLES.DATE_FROM IS NULL AND VARIABLES.DATE_TO IS NULL) 
 ) 
 """
            self._cr.execute(ACCOUNT_REPORT_QUERY,(company_ids,date_from or None,date_to or None,type_view_balance))
            result_account_report = self._cr.dictfetchall()
            wb, ws, target = self.open_xlsx(new_filename, load_sheet=True)
            COUNT = len(result_account_report)
            i, INDEX_ROW = 0, 6
            last_row = INDEX_ROW
            if result_account_report:
                for each_result in result_account_report:
                    row = str(INDEX_ROW + i)
                    ws['A' + row] = each_result["company_name"]
                    ws['B' + row] = each_result["tipo_transaccion"]
                    ws['C' + row] = each_result["journal_name"]
                    ws['D' + row] = each_result["partner_vat"]
                    ws['E' + row] = each_result["partner_name"]
                    ws['F' + row] = each_result["move_id"]
                    ws['G' + row] = each_result["move_name"]
                    ws['H' + row] = each_result["move_ref"]

                    ws['I' + row] = each_result["move_date"]

                    ws['J' + row] = each_result["account_code"]
                    ws['K' + row] = each_result["account_name"]

                    ws['L' + row] = each_result["line_debit"]
                    ws['M' + row] = each_result["line_credit"]
                    ws['N' + row] = each_result["line_balance"]#balance
                    ws['O' + row] = each_result["line_residual"]#saldo
                    ws['P' + row] = each_result["saldo_fecha_corte"]  # saldo
                    ws['Q' + row] = each_result["payment_id"]
                    ws['R' + row] = each_result["payment_id"] and (each_result["is_prepayment"] and "PAGO ANTICIPADO" or "PAGO") or ""
                    ws['S' + row] = each_result["payment_residual"]
                    ws['T' + row] = each_result["asignacion_anticipado"] and "SI" or "NO"

                    ws['U' + row] = each_result["referencia2"]

                    i += 1
                    last_row = INDEX_ROW + i
                    #ultima_columna_registros+=1
            ws['A1'] = companys
            ws['B2'] = date_from or 'INICIO DEL USO DEL SISTEMA'
            ws['B3'] = date_to or ''
            ws['F2'] = type_view_balance=='all' and 'Todos' or 'Con Saldos'
            ws['F3'] = COUNT
            ultima_letra ="U"
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':' + ultima_letra + str(last_row - 1), border)
            #print(result_account_report)
            wb = self.save_wb(wb, target)
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

class report_flujo_caja_xlsx(models.AbstractModel):
    _name = "report.gps_informes.report_flujo_caja_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Flujo de Caja"

    PATTERN_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9',
                               fill_type='solid')

    RED_FILL=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Relleno rojo
    FONT_HEADER = Font(size=12, bold=True, color='000000')
    ALIGNMENT = Alignment(horizontal='center', vertical='center')
    ALIGNMENT_RIGHT = Alignment(horizontal='right', vertical='center')
    ALIGNMENT_LEFT = Alignment(horizontal='left', vertical='center')

    DECIMAL_STYLE = NamedStyle(name="decimal_style", number_format="0.00")

    FONT_BOLD = Font(size=12, bold=True)
    FONT_RED = Font(size=12, color="FF0000")
    FONT_RED_BOLD = Font(size=12, color="FF0000", bold=True)

    def add_header(self,RANGE_COLUMNS, ws,last_row, ultima_columna,headers,BLACK_LIST,columns_dscr,primera_letra_ky,ultima_letra_ky):
        proyectado_cols, real_cols , dif_cols = [], [],[]
        for j in RANGE_COLUMNS:
            column_name = headers[j].lower()
            if column_name in BLACK_LIST:
                continue
            EACH_CELL = ws.cell(row=last_row, column=ultima_columna)
            is_sum_values=(column_name.startswith("sem_") and (column_name.endswith("proyectado") \
                                                               or column_name.endswith("real") \
                                                               or column_name.endswith("total_dif")))
            if not is_sum_values:
                EACH_CELL.value = columns_dscr.get(column_name, column_name).capitalize()
            else:
                dscr_sum_values=""
                header_value=columns_dscr.get(column_name, column_name)
                if column_name.startswith("sem_") and column_name.endswith("proyectado"):
                    dscr_sum_values="Proyectado"
                    header_value=header_value.replace('sem_','').replace('proyectado','')
                    proyectado_cols.append(column_name)
                if column_name.startswith("sem_") and column_name.endswith("real"):
                    dscr_sum_values="Real"
                    header_value = header_value.replace('sem_','').replace('real','')
                    real_cols.append(column_name)
                if column_name.startswith("sem_") and column_name.endswith("total_dif"):
                    dscr_sum_values="Diferencia"
                    header_value = header_value.replace('sem_','').replace('total_dif','')
                    dif_cols.append(column_name)
                EACH_CELL.value = dscr_sum_values
                ########################################
                if dscr_sum_values=="Proyectado":
                    ws.merge_cells(start_row=last_row - 1, start_column=ultima_columna, end_row=last_row - 1,
                                   end_column=ultima_columna + 2)
                    HEADER_CELL = ws.cell(row=last_row-1, column=ultima_columna)
                    HEADER_CELL.fill = self.PATTERN_FILL  # Yellow background
                    HEADER_CELL.font = self.FONT_HEADER
                    HEADER_CELL.alignment = self.ALIGNMENT
                    HEADER_CELL.value= header_value.capitalize()
                    thin = Side(border_style="thin", color="000000")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    primera_letra_col=utils.get_column_letter(ultima_columna )
                    ultima_letra_col = utils.get_column_letter(ultima_columna +2)
                    self.set_border(ws, primera_letra_col + str(last_row - 1) + ':' + ultima_letra_col + str(last_row - 1), border)
                ########################################
            EACH_CELL.fill = self.PATTERN_FILL  # Yellow background
            EACH_CELL.font = self.FONT_HEADER
            EACH_CELL.alignment = self.ALIGNMENT
            ultima_columna += 1
        ########################
        ws.merge_cells(start_row=last_row - 1, start_column=ultima_columna, end_row=last_row - 1,
                       end_column=ultima_columna + 2)
        HEADER_CELL = ws.cell(row=last_row - 1, column=ultima_columna)
        HEADER_CELL.fill = self.PATTERN_FILL  # Yellow background
        HEADER_CELL.font = self.FONT_HEADER
        HEADER_CELL.alignment = self.ALIGNMENT
        HEADER_CELL.value = "Total"
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        primera_letra_col = utils.get_column_letter(ultima_columna)
        ultima_letra_col = utils.get_column_letter(ultima_columna + 2)
        self.set_border(ws, primera_letra_col + str(last_row - 1) + ':' + ultima_letra_col + str(last_row - 1), border)
        ########################
        EACH_CELL = ws.cell(row=last_row, column=ultima_columna)
        EACH_CELL.value="Proyectado"
        EACH_CELL.fill = self.PATTERN_FILL  # Yellow background
        EACH_CELL.font = self.FONT_HEADER
        EACH_CELL.alignment = self.ALIGNMENT
        ultima_columna += 1
        #########
        EACH_CELL = ws.cell(row=last_row, column=ultima_columna)
        EACH_CELL.value = "Real"
        EACH_CELL.fill = self.PATTERN_FILL  # Yellow background
        EACH_CELL.font = self.FONT_HEADER
        EACH_CELL.alignment = self.ALIGNMENT
        ultima_columna += 1
        EACH_CELL = ws.cell(row=last_row, column=ultima_columna)
        EACH_CELL.value = "Diferencia"
        EACH_CELL.fill = self.PATTERN_FILL  # Yellow background
        EACH_CELL.font = self.FONT_HEADER
        EACH_CELL.alignment = self.ALIGNMENT
        ultima_columna += 1
        return proyectado_cols, real_cols , dif_cols

    def sum_totals(self,values, total_columns):
        suma_total = 0
        # Iterar sobre cada diccionario en la lista
        for columna in total_columns:
            # Verificar si la clave existe en el diccionario y sumarla
            if columna in values:
                valor = values.get(columna, 0.00)
                if type(valor) in (float, int):
                    suma_total += valor
        return suma_total

    def col_letter_to_index(self,col_letter):
        index = 0
        for char in col_letter:
            index = index * 26 + (ord(char.upper()) - 64)
        return index

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)
        filename = dtFile.join(dir_path, "reporte_flujo_caja.xlsx")
        dtFile.copyfile(filename, new_filename)

        wb = False
        try:
            brw_wizard = self.env["l10n_ec.account.report"].sudo().browse(docids[-1])
            range_days = brw_wizard.range_days
            COLUMN_BASE_QUERY = f"""SELECT 
    'sem_' || ROW_NUMBER() OVER (PARTITION BY tipo ORDER BY inicio_semana ASC,TIPO ASC)||'_'||tipo AS semana_indice,
    inicio_semana AS semana_inicio,
    (inicio_semana + INTERVAL '{range_days-1} days')::DATE AS semana_fin,
    tipo 
FROM (
    SELECT generate_series('{brw_wizard.date_from}'::date, '{brw_wizard.date_to}'::date, '{range_days} day'::interval)::date AS inicio_semana
) semanas
CROSS JOIN (VALUES ('proyectado'), ('real'), ('total_dif')) AS tipos(tipo)
"""
            QUERY_RANGE = (
                f"SELECT X.* FROM ({COLUMN_BASE_QUERY} ) X  ORDER BY X.semana_inicio ASC,X.TIPO ASC")

            self._cr.execute(QUERY_RANGE)
            result_columns = self._cr.fetchall()
            #print(result_columns)
            columns_dscr = {}
            COLUMN_DATA_TYPE = ""
            primera_letra_ky, ultima_letra_ky, total_columns = False, False, []

            if result_columns:
                for semana_indice, semana_inicio, semana_fin,tipo in result_columns:
                    # if not primera_letra_ky:
                    #     primera_letra_ky = dscr
                    COLUMN_DATA_TYPE += f',"{semana_indice}" float '
                    dscr_tiempo = "%s del %s a %s" % (tipo,semana_inicio, semana_fin)
                    columns_dscr[semana_indice] = dscr_tiempo
                    total_columns.append(semana_indice)
                    #ultima_letra_ky = dscr

            QUERY_COLUMNS = (f"SELECT X.semana_indice::VARCHAR AS DSCR "
                             f" FROM ({COLUMN_BASE_QUERY} ) X  "
                             f" ORDER BY X.semana_inicio ASC,X.TIPO ASC")


            #print(result_account_report)
            wb, ws, target = self.open_xlsx(new_filename, load_sheet=False)
            ws=wb["ANEXO CXP"]
            self.crear_cxp( ws, brw_wizard, COLUMN_BASE_QUERY,
                            QUERY_COLUMNS,
                            COLUMN_DATA_TYPE,columns_dscr,total_columns,primera_letra_ky, ultima_letra_ky)
            ws = wb["ANEXO CXC"]
            self.crear_cxc(ws, brw_wizard, COLUMN_BASE_QUERY,
                           QUERY_COLUMNS,
                           COLUMN_DATA_TYPE, columns_dscr, total_columns, primera_letra_ky, ultima_letra_ky)
            wb = self.save_wb(wb, target)
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

    def crear_cxp(self,ws,brw_wizard,COLUMN_BASE_QUERY, \
                  QUERY_COLUMNS, \
                  COLUMN_DATA_TYPE,columns_dscr,total_columns,primera_letra_ky, ultima_letra_ky):
        DEC=2
        columns_dscr=columns_dscr.copy()
        companys = ', '.join(brw_wizard.mapped('company_ids').mapped('name'))
        company_ids = brw_wizard.company_ids.ids + [-1, -1]

        QUERY_VALUES = f""" SELECT  'move,'||AM.ID::VARCHAR AS DOC_ID,
                 PERIODS.semana_indice,
                 SUM(AML.credit) AS valor
         FROM
            ACCOUNT_MOVE AM
            INNER JOIN ACCOUNT_MOVE_LINE AML ON AML.MOVE_ID = AM.ID
            INNER JOIN ACCOUNT_ACCOUNT AA ON AA.ID = AML.ACCOUNT_ID
               AND AA.ACCOUNT_TYPE = 'liability_payable'
            INNER JOIN ( {COLUMN_BASE_QUERY} ) PERIODS ON 
        		COALESCE(AML.DATE_MATURITY,AM.DATE,NOW()::DATE)   
        		BETWEEN   PERIODS.semana_inicio    AND  PERIODS.semana_fin									
        WHERE AM.STATE = 'posted' AND PERIODS.TIPO='proyectado'
           AND AM.COMPANY_ID = ANY (ARRAY{company_ids}  ) and aml.credit>0.00
         GROUP BY
           AM.ID,PERIODS.semana_indice   
           having SUM(AML.credit)!=0.00
           union all 
            SELECT  'move,'||AM.ID::VARCHAR AS DOC_ID,
                 PERIODS.semana_indice,
                 SUM(AML.credit+AML.AMOUNT_RESIDUAL) AS valor
         FROM
            ACCOUNT_MOVE AM
            INNER JOIN ACCOUNT_MOVE_LINE AML ON AML.MOVE_ID = AM.ID
            INNER JOIN ACCOUNT_ACCOUNT AA ON AA.ID = AML.ACCOUNT_ID
               AND AA.ACCOUNT_TYPE = 'liability_payable'
            INNER JOIN ( {COLUMN_BASE_QUERY} ) PERIODS ON 
        		COALESCE(AML.DATE_MATURITY,AM.DATE,NOW()::DATE)   
        		BETWEEN   PERIODS.semana_inicio    AND  PERIODS.semana_fin									
        WHERE AM.STATE = 'posted' AND PERIODS.TIPO='real'
           AND AM.COMPANY_ID = ANY (ARRAY{company_ids} ) and aml.credit>0.00 
         GROUP BY
           AM.ID,PERIODS.semana_indice 
            having round(SUM(AML.credit+AML.AMOUNT_RESIDUAL),{DEC})!=0.00 
        
        union all 
            SELECT  'move,'||AM.ID::VARCHAR AS DOC_ID,
                 PERIODS.semana_indice,
                 -1*SUM(AML.AMOUNT_RESIDUAL) AS valor
         FROM
            ACCOUNT_MOVE AM
            INNER JOIN ACCOUNT_MOVE_LINE AML ON AML.MOVE_ID = AM.ID
            INNER JOIN ACCOUNT_ACCOUNT AA ON AA.ID = AML.ACCOUNT_ID
               AND AA.ACCOUNT_TYPE = 'liability_payable'
            INNER JOIN ( {COLUMN_BASE_QUERY} ) PERIODS ON 
        		COALESCE(AML.DATE_MATURITY,AM.DATE,NOW()::DATE)   
        		BETWEEN   PERIODS.semana_inicio    AND  PERIODS.semana_fin									
        WHERE AM.STATE = 'posted' AND PERIODS.TIPO='total_dif'
           AND AM.COMPANY_ID = ANY (ARRAY{company_ids} ) and aml.credit>0.00 
         GROUP BY
           AM.ID,PERIODS.semana_indice 
            having round(SUM(AML.AMOUNT_RESIDUAL),{DEC})!=0.00

union all

SELECT  'purchase,'||po.ID::VARCHAR AS DOC_ID,
                 PERIODS.semana_indice,
                 SUM(APR.amount) AS valor
	from			 
PURCHASE_ORDER PO
INNER JOIN account_payment_request APR ON APR.order_id=PO.ID
INNER JOIN (  {COLUMN_BASE_QUERY}
 ) PERIODS ON APR.date_maturity   
        		BETWEEN   PERIODS.semana_inicio    AND  PERIODS.semana_fin	
WHERE
                                PO.STATE in ('purchase','done') 
								and po.invoice_status!='invoiced' 
                                AND PO.COMPANY_ID = ANY (ARRAY{company_ids} ) 
								 AND PERIODS.TIPO='proyectado'
                            GROUP BY
                                po.ID,PERIODS.semana_indice
                            HAVING
                                 SUM(APR.amount)!= 0.00
union all 
  SELECT  'purchase,'||po.ID::VARCHAR AS DOC_ID,
                 PERIODS.semana_indice,
                 SUM(APR.paid) AS valor
	from			 
PURCHASE_ORDER PO
INNER JOIN account_payment_request APR ON APR.order_id=PO.ID
INNER JOIN (  {COLUMN_BASE_QUERY}
 ) PERIODS ON APR.date_maturity   
        		BETWEEN   PERIODS.semana_inicio    AND  PERIODS.semana_fin	
WHERE
                                PO.STATE in ('purchase','done') 
								and po.invoice_status!='invoiced' 
                                AND PO.COMPANY_ID = ANY (ARRAY{company_ids} ) 
								 AND PERIODS.TIPO='real'
                            GROUP BY
                                po.ID,PERIODS.semana_indice
                            HAVING
                                 SUM(APR.paid)!= 0.00
union all 
  SELECT  'purchase,'||po.ID::VARCHAR AS DOC_ID,
                 PERIODS.semana_indice,
                 SUM(APR.pending) AS valor
	from			 
PURCHASE_ORDER PO
INNER JOIN account_payment_request APR ON APR.order_id=PO.ID
INNER JOIN (  {COLUMN_BASE_QUERY}
 ) PERIODS ON APR.date_maturity   
        		BETWEEN   PERIODS.semana_inicio    AND  PERIODS.semana_fin	
WHERE
                                PO.STATE in ('purchase','done') 
								and po.invoice_status!='invoiced' 
                                AND PO.COMPANY_ID = ANY (ARRAY{company_ids} ) 
								 AND PERIODS.TIPO='total_dif'
                            GROUP BY
                                po.ID,PERIODS.semana_indice
                            HAVING
                                 SUM(APR.pending)!= 0.00
                                 
union all 

SELECT  'bank,'||AM.ID::VARCHAR AS DOC_ID,
                 PERIODS.semana_indice,
                 SUM(AML.total_to_paid) AS valor
         FROM
            document_financial AM
            INNER JOIN document_financial_LINE AML ON AML.document_id = AM.ID and am.internal_type='out'  
            INNER JOIN ( {COLUMN_BASE_QUERY} ) PERIODS ON 
        	(aml.date_process::DATE)  
        		BETWEEN   PERIODS.semana_inicio    AND  PERIODS.semana_fin									
        WHERE AM.STATE in ('posted','paid') AND PERIODS.TIPO='proyectado'
           AND AM.COMPANY_ID = ANY (ARRAY{company_ids}  ) 
         GROUP BY
           AM.ID,PERIODS.semana_indice   
           having SUM(AML.total_to_paid)!=0.00
           union all 
            SELECT  'bank,'||AM.ID::VARCHAR AS DOC_ID,
                 PERIODS.semana_indice,
                 SUM(AML.total_paid) AS valor
         FROM
             document_financial AM
            INNER JOIN document_financial_LINE AML ON AML.document_id = AM.ID and am.internal_type='out'  
            INNER JOIN ( {COLUMN_BASE_QUERY} ) PERIODS ON 
        		(aml.date_process::DATE)   
        		BETWEEN   PERIODS.semana_inicio    AND  PERIODS.semana_fin									
         WHERE AM.STATE in ('posted','paid') AND PERIODS.TIPO='real'
           AND AM.COMPANY_ID = ANY (ARRAY{company_ids} )  
         GROUP BY
           AM.ID,PERIODS.semana_indice 
            having round(SUM(AML.total_paid),{DEC})!=0.00 
        
        union all 
            SELECT  'bank,'||AM.ID::VARCHAR AS DOC_ID,
                 PERIODS.semana_indice,
                 SUM(AML.total_pending) AS valor
         FROM
          document_financial AM
            INNER JOIN document_financial_LINE AML ON AML.document_id = AM.ID and am.internal_type='out' 
            INNER JOIN ( {COLUMN_BASE_QUERY} ) PERIODS ON 
        		(aml.date_process::DATE)   
        		BETWEEN   PERIODS.semana_inicio    AND  PERIODS.semana_fin									
        WHERE AM.STATE in ('posted','paid') AND PERIODS.TIPO='total_dif'
           AND AM.COMPANY_ID = ANY (ARRAY{company_ids} ) 
         GROUP BY
           AM.ID,PERIODS.semana_indice 
            having round(SUM(AML.total_pending),{DEC})!=0.00


"""

        #####
        ACCOUNT_REPORT_QUERY = f""";WITH ACCOUNT_REPORT_QUERY AS (
                                SELECT * FROM CROSSTAB(
                                  $$  
        	                  {QUERY_VALUES}  order by 1 asc,2 asc 
                                $$,
                                $$ 
                                    {QUERY_COLUMNS}
                                $$
                                    ) AS CT(
                                DOC_ID VARCHAR {COLUMN_DATA_TYPE}   
                                    ) 
                                ),CLASIFICACION_CTAS AS (
                                
                                select agrl.name,aat.code,AAT.company_id,aat.id
from account_Group_report agr
inner join account_group_report_line agrl on agrl.report_id=agr.id
inner join report_template_account_acc_rel rtpl on rtpl.report_line_id=agrl.id 
inner join account_account  aat on aat.id=rtpl.account_id
where agr.CODE='FLUJO' and aat.company_id = ANY (ARRAY{company_ids} ) 
                                
                                
                                
                                ) ,

                                CTAS_GASTOS as (
        	SELECT ARQ.DOC_ID,  AA.NAME AS CUENTA_NAME,AA.CODE AS CUENTA_CODE,C.name AS CLASIFICACION ,
SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.DOC_ID) AS TOTAL,
		case when(SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.DOC_ID)>0) 
		then SUM(AML.PRICE_TOTAL)/SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.DOC_ID)
			else 1.00 end  as percentage 
			
        	FROM ACCOUNT_REPORT_QUERY ARQ
        	INNER JOIN ACCOUNT_MOVE_LINE AML ON 'move,'||AML.MOVE_ID=ARQ.DOC_ID 
        	INNER JOIN ACCOUNT_ACCOUNT AA ON AA.ID=AML.ACCOUNT_ID 	AND 
        		AA.ACCOUNT_TYPE != 'liability_payable' AND AML.PRODUCT_ID IS NOT NULL 
        		AND AML.DEBIT>0 
        	LEFT JOIN CLASIFICACION_CTAS C ON C.id=AA.ID 
        	GROUP BY ARQ.DOC_ID,AA.NAME,AA.CODE,C.name

        )	,CTAS_GASTOS_COMPRAS as (
        
        SELECT ARQ.DOC_ID,COALESCE(AA.NAME,AAC.NAME) AS CUENTA_NAME,
                    COALESCE(AA.CODE,AAC.CODE) AS CUENTA_CODE,C.name AS CLASIFICACION ,
                    
                    SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.DOC_ID) AS TOTAL,
		case when(SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.DOC_ID)>0) 
		then SUM(AML.PRICE_TOTAL)/SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.DOC_ID)
			else 1.00 end  as percentage 

        	FROM ACCOUNT_REPORT_QUERY ARQ
        	INNER JOIN PURCHASE_ORDER_LINE AML ON 'purchase,'||AML.ORDER_ID=ARQ.DOC_ID AND AML.PRODUCT_ID IS NOT NULL 
        	INNER JOIN PURCHASE_ORDER PO ON PO.ID=AML.ORDER_ID 
			INNER JOIN PRODUCT_PRODUCT PP ON PP.ID=AML.PRODUCT_ID 
			INNER JOIN PRODUCT_TEMPLATE PT ON PT.ID=PP.PRODUCT_TMPL_ID 
			
			LEFT JOIN IR_PROPERTY IPPT ON IPPT.RES_ID=('product.template,'||PT.id )::VARCHAR
				AND IPPT.TYPE='many2one' AND IPPT.name='property_account_expense_id'
				AND IPPT.COMPANY_ID=PO.COMPANY_ID 
			
			LEFT JOIN ACCOUNT_ACCOUNT AA ON ('account.account,'||AA.ID)::VARCHAR=IPPT.VALUE_REFERENCE::VARCHAR AND IPPT.ID IS NOT NULL AND AA.COMPANY_ID=PO.COMPANY_ID 

	

			LEFT JOIN IR_PROPERTY IPPTC ON IPPTC.RES_ID=('product.category,'||PT.CATEG_ID )::VARCHAR
				AND IPPTC.TYPE='many2one' AND IPPTC.name='property_account_expense_categ_id'
				AND IPPTc.COMPANY_ID=PO.COMPANY_ID 
		LEFT JOIN ACCOUNT_ACCOUNT AAC ON ('account.account,'||AAC.ID)::VARCHAR=IPPTC.VALUE_REFERENCE::VARCHAR AND IPPTC.ID IS NOT NULL AND AAC.COMPANY_ID=PO.COMPANY_ID 

			LEFT JOIN CLASIFICACION_CTAS C ON C.id=COALESCE(AA.ID ,AAC.ID) 
			
				group by ARQ.DOC_ID,COALESCE(AA.NAME,AAC.NAME),COALESCE(AA.CODE,AAC.CODE),C.name
        	
        
        )
        
        SELECT 0::INT AS SEQUENCE_FORMAT,            
                    RC.NAME AS COMPANY_NAME,
                    'DOCUMENTO CONTABLE'::VARCHAR AS TIPO,
                    COALESCE(RP.NAME,'DESCONOCIDO') AS PARTNER_NAME,
        			COALESCE(RP.VAT,'') AS PARTNER_VAT,
        			AJ.NAME AS JOURNAL_NAME,
        			AM.NAME,
        			COALESCE(AM.INVOICE_DATE,AM.DATE) AS INVOICE_DATE, 
        			APT.NAME::JSON->>'es_EC'::varchar  AS PAYMENT_TERM_DSCR,
        			AM.ID AS INTERNAL_ID,
        			CTAS.CLASIFICACION,
        			CTAS.CUENTA_CODE,
        	        CTAS.CUENTA_NAME,    
        	        CTAS.percentage,    	        
        			I.*             
                    FROM ACCOUNT_REPORT_QUERY I 
        			INNER JOIN CTAS_GASTOS CTAS ON CTAS.DOC_ID=I.DOC_ID
                    INNER JOIN ACCOUNT_MOVE AM ON 'move,'||AM.ID=I.DOC_ID  
                    LEFT JOIN ACCOUNT_PAYMENT_TERM APT ON APT.ID=AM.INVOICE_PAYMENT_TERM_ID
                    INNER JOIN RES_COMPANY RC ON RC.ID=AM.COMPANY_ID 
                    INNER JOIN ACCOUNT_JOURNAL AJ ON AJ.ID=AM.JOURNAL_ID 
                    LEFT JOIN RES_PARTNER RP ON RP.ID=AM.PARTNER_ID
                    WHERE I.DOC_ID LIKE 'move%'  
        UNION 
        
        SELECT 1::INT AS SEQUENCE_FORMAT,            
            RC.NAME AS COMPANY_NAME,
            'ORDENES DE COMPRA'::VARCHAR AS TIPO,
            COALESCE(RP.NAME,'DESCONOCIDO') AS PARTNER_NAME,
			COALESCE(RP.VAT,'') AS PARTNER_VAT,
			'' AS JOURNAL_NAME,
			PO.NAME,
			PO.DATE_APPROVE::DATE AS INVOICE_DATE,
			APT.NAME::JSON->>'es_EC'::varchar  AS PAYMENT_TERM_DSCR,
			PO.ID AS INTERNAL_ID,
			CTAS.CLASIFICACION,
			CTAS.CUENTA_CODE,
        	CTAS.CUENTA_NAME,    
        	CTAS.percentage,  
			I.*             
            FROM ACCOUNT_REPORT_QUERY I 
            INNER JOIN CTAS_GASTOS_COMPRAS CTAS ON CTAS.DOC_ID=I.DOC_ID
            INNER JOIN PURCHASE_ORDER PO ON 'purchase,'||PO.ID=I.DOC_ID  
            INNER JOIN ACCOUNT_PAYMENT_TERM APT ON APT.ID=PO.PAYMENT_TERM_ID
            INNER JOIN RES_COMPANY RC ON RC.ID=PO.COMPANY_ID  
            INNER JOIN RES_PARTNER RP ON RP.ID=PO.PARTNER_ID
            WHERE I.DOC_ID LIKE 'purchase%'
     
     UNION 
        
        SELECT 2::INT AS SEQUENCE_FORMAT,            
            RC.NAME AS COMPANY_NAME,
           'OPERACION FINANCIERA'::VARCHAR AS TIPO,
            COALESCE(RP.NAME,'DESCONOCIDO') AS PARTNER_NAME,
			COALESCE(RP.VAT,'') AS PARTNER_VAT,
			CASE po.type
    WHEN 'pagares' THEN 'PAGARES'
    WHEN 'emision' THEN 'EMISION DE OBLIGACIONES'
    WHEN 'prestamo' THEN 'OBLIGACIONES BANCARIAS'
    WHEN 'contrato' THEN 'CONTRATO'
    ELSE 'DESCONOCIDO'
END AS JOURNAL_NAME,
			PO.NAME,
			PO.date_process::DATE AS INVOICE_DATE,
			''  AS PAYMENT_TERM_DSCR,
			PO.ID AS INTERNAL_ID,
			'' as CLASIFICACION,
			'' as CUENTA_CODE,
        	'' as CUENTA_NAME,    
        	1.00 as percentage,  
			I.*             
            FROM ACCOUNT_REPORT_QUERY I 
            INNER JOIN document_financial PO ON 'bank,'||PO.ID=I.DOC_ID   and po.internal_type='out'  
            INNER JOIN RES_COMPANY RC ON RC.ID=PO.COMPANY_ID  
            INNER JOIN RES_PARTNER RP ON RP.ID=PO.PARTNER_ID
            WHERE I.DOC_ID LIKE 'bank%'
                    
"""
        print(ACCOUNT_REPORT_QUERY)
        self._cr.execute(ACCOUNT_REPORT_QUERY)
        headers = [d.name for d in self._cr.description]

        result_account_report = self._cr.dictfetchall()

        i, INDEX_ROW = 0, 5
        last_row = INDEX_ROW
        ultima_columna_registros = 1
        COUNT = len(result_account_report)
        SUM_GROUPS = {}

        columns_dscr["company_name"] = "Empresa"
        columns_dscr["category_name"] = "Categoría"
        columns_dscr["clasificacion"] = "Clasificación"
        columns_dscr["partner_vat"] = "# Identificación"
        columns_dscr["partner_name"] = "Proveedor"
        columns_dscr["journal_name"] = "Diario"
        columns_dscr["name"] = "# Documento"
        columns_dscr["invoice_date"] = "Fecha de Documento"
        columns_dscr["payment_term_dscr"] = "Termino de Pago"
        columns_dscr["internal_id"] = "# ID"
        columns_dscr["cuenta_code"] = "Cod. Cuenta"
        columns_dscr["cuenta_name"] = "Nombre Cuenta"
        #####
        if result_account_report:
            RANGE_COLUMNS = range(1, len(headers))


            BLACK_LIST = ["doc_id"]

            FONT_BOLD_LIST = ["total"]
            ultima_columna = 1
            proyectado_cols, real_cols , dif_cols = self.add_header(RANGE_COLUMNS, ws,last_row, ultima_columna,headers,BLACK_LIST,columns_dscr,primera_letra_ky,ultima_letra_ky)
            compute_fields={"proyectado":proyectado_cols,
                            "real":real_cols,
                            "dif":dif_cols
                            }
            last_row += 1
            for each_linea in result_account_report:
                ultima_columna_registros = 1
                for j in RANGE_COLUMNS:
                    column_name = headers[j].lower()
                    if column_name in BLACK_LIST:
                        continue
                    ROW_EACH_CELL = ws.cell(row=last_row, column=ultima_columna_registros)
                    value_write = each_linea[column_name]
                    if type(value_write) == str:
                        value_write = each_linea[column_name].replace('"', '').upper()
                    ROW_EACH_CELL.value = value_write
                    if type(value_write) == str:
                        ROW_EACH_CELL.alignment = self.ALIGNMENT_LEFT
                    if type(value_write) in (float, int):
                        ROW_EACH_CELL.alignment = self.ALIGNMENT_RIGHT
                        if type(value_write) == float:
                            ROW_EACH_CELL.style = self.DECIMAL_STYLE
                        if column_name.startswith("sem_"):##columnas de totales por proyeccion
                            factor=each_linea["percentage"]
                            ROW_EACH_CELL.value = value_write*factor
                            if column_name.endswith("total_dif"):
                                ROW_EACH_CELL.fill = self.RED_FILL
                    if column_name in FONT_BOLD_LIST or each_linea["sequence_format"] == 3:  #
                        ROW_EACH_CELL.font = self.FONT_BOLD
                    ultima_columna_registros += 1
                for compute_field_ky in compute_fields:
                    ROW_EACH_CELL = ws.cell(row=last_row, column=ultima_columna_registros)
                    ROW_EACH_CELL.alignment = self.ALIGNMENT_RIGHT
                    ROW_EACH_CELL.style = self.DECIMAL_STYLE
                    sum_total=sum((each_linea.get(key, 0.00) or 0.00)for key in compute_fields[compute_field_ky])
                    sum_total=sum_total*(each_linea.get("percentage", 1.00) or 1.00 )
                    ROW_EACH_CELL.value = sum_total
                    if sum_total != 0.00 and compute_field_ky=="dif":
                        ROW_EACH_CELL.fill = self.RED_FILL
                    ultima_columna_registros+=1
                last_row += 1
        ws['A1'] = companys
        ws['B2'] = brw_wizard.date_from
        ws['B3'] = brw_wizard.date_to
        ws['E3'] = brw_wizard.range_days
        ws['E2'] = COUNT
        ultima_letra = utils.get_column_letter(ultima_columna_registros-1)
        if last_row >= INDEX_ROW:
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            self.set_border(ws, 'A' + str(INDEX_ROW) + ':' + ultima_letra + str(last_row - 1), border)

    def crear_cxc(self, ws, brw_wizard, COLUMN_BASE_QUERY, \
                  QUERY_COLUMNS, \
                  COLUMN_DATA_TYPE, columns_dscr, total_columns, primera_letra_ky, ultima_letra_ky):
        DEC = 2
        columns_dscr = columns_dscr.copy()
        companys = ', '.join(brw_wizard.mapped('company_ids').mapped('name'))
        company_ids = brw_wizard.company_ids.ids + [-1, -1]

        QUERY_VALUES = f""" SELECT  'move,'||AM.ID::VARCHAR AS DOC_ID,
                     PERIODS.semana_indice,
                     SUM(AML.debit) AS valor
             FROM
                ACCOUNT_MOVE AM
                INNER JOIN ACCOUNT_MOVE_LINE AML ON AML.MOVE_ID = AM.ID
                INNER JOIN ACCOUNT_ACCOUNT AA ON AA.ID = AML.ACCOUNT_ID
                   AND AA.ACCOUNT_TYPE = 'asset_receivable'
                INNER JOIN ( {COLUMN_BASE_QUERY} ) PERIODS ON 
            		COALESCE(AML.DATE_MATURITY,AM.DATE,NOW()::DATE)   
            		BETWEEN   PERIODS.semana_inicio    AND  PERIODS.semana_fin									
            WHERE AM.STATE = 'posted' AND PERIODS.TIPO='proyectado'
               AND AM.COMPANY_ID = ANY (ARRAY{company_ids}  ) and aml.debit>0.00
             GROUP BY
               AM.ID,PERIODS.semana_indice   
               having SUM(AML.debit)!=0.00
               union all 
                SELECT  'move,'||AM.ID::VARCHAR AS DOC_ID,
                     PERIODS.semana_indice,
                     SUM(AML.debit-AML.AMOUNT_RESIDUAL) AS valor
             FROM
                ACCOUNT_MOVE AM
                INNER JOIN ACCOUNT_MOVE_LINE AML ON AML.MOVE_ID = AM.ID
                INNER JOIN ACCOUNT_ACCOUNT AA ON AA.ID = AML.ACCOUNT_ID
                   AND AA.ACCOUNT_TYPE = 'asset_receivable'
                INNER JOIN ( {COLUMN_BASE_QUERY} ) PERIODS ON 
            		COALESCE(AML.DATE_MATURITY,AM.DATE,NOW()::DATE)   
            		BETWEEN   PERIODS.semana_inicio    AND  PERIODS.semana_fin									
            WHERE AM.STATE = 'posted' AND PERIODS.TIPO='real'
               AND AM.COMPANY_ID = ANY (ARRAY{company_ids} ) and aml.debit>0.00 
             GROUP BY
               AM.ID,PERIODS.semana_indice 
                having round(SUM(AML.debit-AML.AMOUNT_RESIDUAL),{DEC})!=0.00 

            union all 
                SELECT  'move,'||AM.ID::VARCHAR AS DOC_ID,
                     PERIODS.semana_indice,
                     SUM(AML.AMOUNT_RESIDUAL) AS valor
             FROM
                ACCOUNT_MOVE AM
                INNER JOIN ACCOUNT_MOVE_LINE AML ON AML.MOVE_ID = AM.ID
                INNER JOIN ACCOUNT_ACCOUNT AA ON AA.ID = AML.ACCOUNT_ID
                   AND AA.ACCOUNT_TYPE = 'asset_receivable'
                INNER JOIN ( {COLUMN_BASE_QUERY} ) PERIODS ON 
            		COALESCE(AML.DATE_MATURITY,AM.DATE,NOW()::DATE)   
            		BETWEEN   PERIODS.semana_inicio    AND  PERIODS.semana_fin									
            WHERE AM.STATE = 'posted' AND PERIODS.TIPO='total_dif'
               AND AM.COMPANY_ID = ANY (ARRAY{company_ids} ) and aml.debit>0.00 
             GROUP BY
               AM.ID,PERIODS.semana_indice 
                having round(SUM(AML.AMOUNT_RESIDUAL),{DEC})!=0.00

    


    """

        #####
        ACCOUNT_REPORT_QUERY = f""";WITH ACCOUNT_REPORT_QUERY AS (
                                    SELECT * FROM CROSSTAB(
                                      $$  
            	                  {QUERY_VALUES}  order by 1 asc,2 asc 
                                    $$,
                                    $$ 
                                        {QUERY_COLUMNS}
                                    $$
                                        ) AS CT(
                                    DOC_ID VARCHAR {COLUMN_DATA_TYPE}   
                                        ) 
                                    ),CLASIFICACION_CLIENTES AS (

                             select coalesce(agr.name,'SIN CLASIFICAR') as clasificacion_cte,
		rp.id AS PARTNER_ID 
    from res_partner rp   
    LEFT join account_partner_group_category agr on rp.group_category_id=agr.id
       

                                    ) 

            SELECT 0::INT AS SEQUENCE_FORMAT,            
                        RC.NAME AS COMPANY_NAME,
                        'DOCUMENTO CONTABLE'::VARCHAR AS TIPO,
            			CCL.clasificacion_cte,
                        COALESCE(RP.NAME,'DESCONOCIDO') AS PARTNER_NAME,
            			COALESCE(RP.VAT,'') AS PARTNER_VAT,
            			AJ.NAME AS JOURNAL_NAME,
            			AM.NAME,
            			COALESCE(AM.INVOICE_DATE,AM.DATE) AS INVOICE_DATE, 
            			APT.NAME::JSON->>'es_EC'::varchar  AS PAYMENT_TERM_DSCR,
            			AM.ID AS INTERNAL_ID  , 	        
            			I.*             
                        FROM ACCOUNT_REPORT_QUERY I  
                        INNER JOIN ACCOUNT_MOVE AM ON 'move,'||AM.ID=I.DOC_ID  
                        LEFT JOIN ACCOUNT_PAYMENT_TERM APT ON APT.ID=AM.INVOICE_PAYMENT_TERM_ID
                        INNER JOIN RES_COMPANY RC ON RC.ID=AM.COMPANY_ID 
                        INNER JOIN ACCOUNT_JOURNAL AJ ON AJ.ID=AM.JOURNAL_ID 
                        LEFT JOIN RES_PARTNER RP ON RP.ID=AM.PARTNER_ID
                        LEFT JOIN CLASIFICACION_CLIENTES CCL ON CCL.PARTNER_ID=AM.PARTNER_ID
                        WHERE I.DOC_ID LIKE 'move%'  
            

    """
        self._cr.execute(ACCOUNT_REPORT_QUERY)
        headers = [d.name for d in self._cr.description]

        result_account_report = self._cr.dictfetchall()

        i, INDEX_ROW = 0, 5
        last_row = INDEX_ROW
        ultima_columna_registros = 1
        COUNT = len(result_account_report)
        SUM_GROUPS = {}

        columns_dscr["company_name"] = "Empresa"
        columns_dscr["clasificacion_cte"] = "Clasificación de Cliente"
        columns_dscr["partner_vat"] = "# Identificación"
        columns_dscr["partner_name"] = "Cliente"
        columns_dscr["partner_vat"] = "# Identificación"
        columns_dscr["journal_name"] = "Diario"
        columns_dscr["name"] = "# Documento"
        columns_dscr["percentage"] = "% Dist."
        columns_dscr["invoice_date"] = "Fecha de Documento"
        columns_dscr["payment_term_dscr"] = "Termino de Cobro"
        columns_dscr["internal_id"] = "# ID"
        #####
        if result_account_report:
            RANGE_COLUMNS = range(1, len(headers))

            BLACK_LIST = ["doc_id"]

            FONT_BOLD_LIST = ["total"]
            ultima_columna = 1
            proyectado_cols, real_cols , dif_cols = self.add_header(RANGE_COLUMNS, ws, last_row, ultima_columna,
                                                                  headers, BLACK_LIST, columns_dscr,
                                                                  primera_letra_ky, ultima_letra_ky)

            compute_fields = {"proyectado": proyectado_cols,
                              "real": real_cols,
                              "dif": dif_cols
                              }
            last_row += 1
            for each_linea in result_account_report:
                ultima_columna_registros = 1
                for j in RANGE_COLUMNS:
                    column_name = headers[j].lower()
                    if column_name in BLACK_LIST:
                        continue
                    ROW_EACH_CELL = ws.cell(row=last_row, column=ultima_columna_registros)
                    value_write = each_linea[column_name]
                    if type(value_write) == str:
                        value_write = each_linea[column_name].replace('"', '').upper()
                    ROW_EACH_CELL.value = value_write
                    if type(value_write) == str:
                        ROW_EACH_CELL.alignment = self.ALIGNMENT_LEFT
                    if type(value_write) in (float, int):
                        ROW_EACH_CELL.alignment = self.ALIGNMENT_RIGHT
                        if type(value_write) == float:
                            ROW_EACH_CELL.style = self.DECIMAL_STYLE
                        if column_name.startswith("sem_") and (column_name.endswith("total_dif")):
                            ROW_EACH_CELL.fill = self.RED_FILL
                    if column_name in FONT_BOLD_LIST or each_linea["sequence_format"] == 3:  #
                        ROW_EACH_CELL.font = self.FONT_BOLD
                    ultima_columna_registros += 1
                for compute_field_ky in compute_fields:
                    ROW_EACH_CELL = ws.cell(row=last_row, column=ultima_columna_registros)
                    ROW_EACH_CELL.alignment = self.ALIGNMENT_RIGHT
                    ROW_EACH_CELL.style = self.DECIMAL_STYLE
                    sum_total = sum(each_linea.get(key, 0.00) or 0.00 for key in compute_fields[compute_field_ky])
                    ROW_EACH_CELL.value = sum_total
                    if sum_total != 0.00 and compute_field_ky=="dif":
                        ROW_EACH_CELL.fill = self.RED_FILL
                    ultima_columna_registros += 1
                last_row += 1

        ws['A1'] = companys
        ws['B2'] = brw_wizard.date_from
        ws['B3'] = brw_wizard.date_to
        ws['E3'] = brw_wizard.range_days
        ws['E2'] = COUNT
        ultima_letra = utils.get_column_letter(ultima_columna_registros - 1)
        if last_row >= INDEX_ROW:
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            self.set_border(ws, 'A' + str(INDEX_ROW) + ':' + ultima_letra + str(last_row - 1), border)

class report_cuentas_pagar_report_xlsx(models.AbstractModel):
    _name = "report.gps_informes.report_cuentas_pagar_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Cuentas por Pagar"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)
        filename = dtFile.join(dir_path, "reporte_cuentas_pagar.xlsx")
        dtFile.copyfile(filename, new_filename)
        wb = False
        try:
            brw_wizard = self.env["l10n_ec.account.report"].sudo().browse(docids[-1])
            wb, ws, target = self.open_xlsx(new_filename, load_sheet=False)
            ws = wb["REPORTE  DE CUENTAS POR PAGAR"]
            self.create_report_cxp(brw_wizard, ws)
            ws = wb["NC,SALDOS Y ANTICIPOS"]
            self.create_report_cxp_nc_anticipos(brw_wizard, ws)
            wb = self.save_wb(wb, target)
            #########################################
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

    def create_report_cxp(self,brw_wizard, ws):
        try:
            companys = ', '.join(brw_wizard.mapped('company_ids').mapped('name'))
            company_ids = brw_wizard.company_ids.ids + [-1, -1]
            date_from = brw_wizard.date_from
            date_to = brw_wizard.date_to
            #type_view_balance = brw_wizard.type_view_balance
            ACCOUNT_REPORT_QUERY = f""";with base_por_pagar as (
	SELECT  'move,'||AM.ID::VARCHAR AS DOC_ID,
			AML.id,
			COALESCE(AML.DATE_MATURITY,AM.DATE,NOW()::DATE) as DATE_MATURITY,
                 -1*SUM(AML.AMOUNT_RESIDUAL) AS saldo,
                 SUM(AML.credit) AS valor
         FROM
            ACCOUNT_MOVE AM
            INNER JOIN ACCOUNT_MOVE_LINE AML ON AML.MOVE_ID = AM.ID
            INNER JOIN ACCOUNT_ACCOUNT AA ON AA.ID = AML.ACCOUNT_ID
               AND AA.ACCOUNT_TYPE = 'liability_payable'
        WHERE AM.STATE = 'posted' 
           AND AM.COMPANY_ID = ANY (array{company_ids}) and aml.credit>0.00 
         GROUP BY
           AM.ID,AML.id,
			COALESCE(AML.DATE_MATURITY,AM.DATE,NOW()::DATE) 
            having round(SUM(AML.AMOUNT_RESIDUAL),2)!=0.00
union all
  SELECT  'purchase,'||po.ID::VARCHAR AS DOC_ID,
  apr.id,
  APR.date_maturity ,
                 SUM(APR.pending) AS saldo ,
                 SUM(APR.amount) AS valor
	from			 
PURCHASE_ORDER PO
INNER JOIN account_payment_request APR ON APR.order_id=PO.ID
WHERE
                                PO.STATE in ('purchase','done') 
								and po.invoice_status!='invoiced' 
                                AND PO.COMPANY_ID = ANY (array{company_ids}) 
						    GROUP BY
                                po.ID,apr.id, APR.date_maturity--,PERIODS.semana_indice
                            HAVING
                                 SUM(APR.pending)!= 0.00
                                 
union all
  SELECT  'bank,'||df.ID::VARCHAR AS DOC_ID,
  dfl.id,
  dfl.date_process as date_maturity ,
                 SUM(dfl.total_pending) AS saldo ,
                 SUM(dfl.total_to_paid) AS valor
	from			 
document_financial df
INNER JOIN document_financial_line dfl ON dfl.document_id=df.ID
WHERE df.internal_Type='out' and 
df.state in ('posted','paid') 
                                AND df.COMPANY_ID = ANY (array{company_ids}) 
						    GROUP BY
                                df.ID,dfl.id, dfl.date_process 
                            HAVING
                                 SUM(dfl.total_pending)!= 0.00
                                 
),
ACCOUNT_REPORT_QUERY as (
	select doc_id from base_por_pagar
	group by doc_id
),
CLASIFICACION_CTAS AS (                               
                                select agrl.name,aat.code,AAT.company_id,aat.id
from account_Group_report agr
inner join account_group_report_line agrl on agrl.report_id=agr.id
inner join report_template_account_acc_rel rtpl on rtpl.report_line_id=agrl.id 
inner join account_account  aat on aat.id=rtpl.account_id
where agr.CODE='FLUJO' and aat.company_id = ANY (array{company_ids})      
                                ) ,
                                CTAS_GASTOS as (
        	SELECT ARQ.DOC_ID,  AA.NAME AS CUENTA_NAME,AA.CODE AS CUENTA_CODE,C.name AS CLASIFICACION ,
SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.DOC_ID) AS TOTAL,
		case when(SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.DOC_ID)>0) 
		then SUM(AML.PRICE_TOTAL)/SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.DOC_ID)
			else 1.00 end  as percentage 
			
        	FROM ACCOUNT_REPORT_QUERY ARQ
        	INNER JOIN ACCOUNT_MOVE_LINE AML ON 'move,'||AML.MOVE_ID=ARQ.DOC_ID 
        	INNER JOIN ACCOUNT_ACCOUNT AA ON AA.ID=AML.ACCOUNT_ID 	AND 
        		AA.ACCOUNT_TYPE != 'liability_payable' AND AML.PRODUCT_ID IS NOT NULL 
        		AND AML.DEBIT>0 
        	LEFT JOIN CLASIFICACION_CTAS C ON C.id=AA.ID 
        	GROUP BY ARQ.DOC_ID,AA.NAME,AA.CODE,C.name

        )	,CTAS_GASTOS_COMPRAS as (
        
        SELECT ARQ.DOC_ID,COALESCE(AA.NAME,AAC.NAME) AS CUENTA_NAME,
                    COALESCE(AA.CODE,AAC.CODE) AS CUENTA_CODE,C.name AS CLASIFICACION ,
                    
                    SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.DOC_ID) AS TOTAL,
		case when(SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.DOC_ID)>0) 
		then SUM(AML.PRICE_TOTAL)/SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.DOC_ID)
			else 1.00 end  as percentage 

        	FROM ACCOUNT_REPORT_QUERY ARQ
        	INNER JOIN PURCHASE_ORDER_LINE AML ON 'purchase,'||AML.ORDER_ID=ARQ.DOC_ID AND AML.PRODUCT_ID IS NOT NULL 
        	INNER JOIN PURCHASE_ORDER PO ON PO.ID=AML.ORDER_ID 
			INNER JOIN PRODUCT_PRODUCT PP ON PP.ID=AML.PRODUCT_ID 
			INNER JOIN PRODUCT_TEMPLATE PT ON PT.ID=PP.PRODUCT_TMPL_ID 
			
			LEFT JOIN IR_PROPERTY IPPT ON IPPT.RES_ID=('product.template,'||PT.id )::VARCHAR
				AND IPPT.TYPE='many2one' AND IPPT.name='property_account_expense_id'
				AND IPPT.COMPANY_ID=PO.COMPANY_ID 
			
			LEFT JOIN ACCOUNT_ACCOUNT AA ON ('account.account,'||AA.ID)::VARCHAR=IPPT.VALUE_REFERENCE::VARCHAR AND IPPT.ID IS NOT NULL AND AA.COMPANY_ID=PO.COMPANY_ID 

	

			LEFT JOIN IR_PROPERTY IPPTC ON IPPTC.RES_ID=('product.category,'||PT.CATEG_ID )::VARCHAR
				AND IPPTC.TYPE='many2one' AND IPPTC.name='property_account_expense_categ_id'
				AND IPPTc.COMPANY_ID=PO.COMPANY_ID 
		LEFT JOIN ACCOUNT_ACCOUNT AAC ON ('account.account,'||AAC.ID)::VARCHAR=IPPTC.VALUE_REFERENCE::VARCHAR AND IPPTC.ID IS NOT NULL AND AAC.COMPANY_ID=PO.COMPANY_ID 

			LEFT JOIN CLASIFICACION_CTAS C ON C.id=COALESCE(AA.ID ,AAC.ID) 
			
				group by ARQ.DOC_ID,COALESCE(AA.NAME,AAC.NAME),COALESCE(AA.CODE,AAC.CODE),C.name
        	
        
        )
        
        SELECT            
                    RC.NAME AS COMPANY_NAME,
                    'DOCUMENTO CONTABLE'::VARCHAR AS TIPO,
                    COALESCE(RP.NAME,'DESCONOCIDO') AS PARTNER_NAME,
        			COALESCE(RP.VAT,'') AS PARTNER_VAT,
        			AJ.NAME AS JOURNAL_NAME,
        			AM.NAME,
        			COALESCE(AM.INVOICE_DATE,AM.DATE) AS INVOICE_DATE, 
					am.amount_total,
        			APT.NAME::JSON->>'es_EC'::varchar  AS PAYMENT_TERM_DSCR,
        			AM.ID AS INTERNAL_ID,
        			CTAS.CLASIFICACION,
        			CTAS.CUENTA_CODE,
        	        CTAS.CUENTA_NAME,    
        	coalesce(CTAS.percentage,  1.00) as percentage,   	        
        			I.*             
                    FROM base_por_pagar I 
        			left JOIN CTAS_GASTOS CTAS ON CTAS.DOC_ID=I.DOC_ID
                    INNER JOIN ACCOUNT_MOVE AM ON 'move,'||AM.ID=I.DOC_ID  
                    LEFT JOIN ACCOUNT_PAYMENT_TERM APT ON APT.ID=AM.INVOICE_PAYMENT_TERM_ID
                    INNER JOIN RES_COMPANY RC ON RC.ID=AM.COMPANY_ID 
                    INNER JOIN ACCOUNT_JOURNAL AJ ON AJ.ID=AM.JOURNAL_ID 
                    LEFT JOIN RES_PARTNER RP ON RP.ID=AM.PARTNER_ID
                    WHERE I.DOC_ID LIKE 'move%'  
        UNION 
        
        SELECT  RC.NAME AS COMPANY_NAME,
            'ORDENES DE COMPRA'::VARCHAR AS TIPO,
            COALESCE(RP.NAME,'DESCONOCIDO') AS PARTNER_NAME,
			COALESCE(RP.VAT,'') AS PARTNER_VAT,
			'' AS JOURNAL_NAME,
			PO.NAME,
			PO.DATE_APPROVE::DATE AS INVOICE_DATE,
			po.amount_total,
			APT.NAME::JSON->>'es_EC'::varchar  AS PAYMENT_TERM_DSCR,
			PO.ID AS INTERNAL_ID,
			CTAS.CLASIFICACION,
			CTAS.CUENTA_CODE,
        	CTAS.CUENTA_NAME,    
        	coalesce(CTAS.percentage,  1.00) as percentage,
			I.*             
            FROM base_por_pagar I 
            INNER JOIN CTAS_GASTOS_COMPRAS CTAS ON CTAS.DOC_ID=I.DOC_ID
            INNER JOIN PURCHASE_ORDER PO ON 'purchase,'||PO.ID=I.DOC_ID  
            INNER JOIN ACCOUNT_PAYMENT_TERM APT ON APT.ID=PO.PAYMENT_TERM_ID
            INNER JOIN RES_COMPANY RC ON RC.ID=PO.COMPANY_ID  
            INNER JOIN RES_PARTNER RP ON RP.ID=PO.PARTNER_ID
            WHERE I.DOC_ID LIKE 'purchase%' 

UNION 
        
        SELECT  RC.NAME AS COMPANY_NAME,
            'OPERACION FINANCIERA'::VARCHAR AS TIPO,
            COALESCE(RP.NAME,'DESCONOCIDO') AS PARTNER_NAME,
			COALESCE(RP.VAT,'') AS PARTNER_VAT,
			'' AS JOURNAL_NAME,
			PO.NAME,
			PO.DATE_PROCESS::DATE AS INVOICE_DATE,
			PO.TOTAL AS amount_total,
			''  AS PAYMENT_TERM_DSCR,
			PO.ID AS INTERNAL_ID,
			'' AS cLASIFICACION,
			'' AS CUENTA_CODE,
        	'' AS CUENTA_NAME,    
        	1.00 as percentage,
			I.*             
            FROM base_por_pagar I 
            INNER JOIN document_financial PO ON 'bank,'||PO.ID=I.DOC_ID   
            INNER JOIN RES_COMPANY RC ON RC.ID=PO.COMPANY_ID  
            INNER JOIN RES_PARTNER RP ON RP.ID=PO.PARTNER_ID
            WHERE I.DOC_ID LIKE 'bank%'
            
"""
            self._cr.execute(ACCOUNT_REPORT_QUERY)
            result_account_report = self._cr.dictfetchall()
            COUNT = len(result_account_report)
            i, INDEX_ROW = 0, 6
            last_row = INDEX_ROW
            if result_account_report:
                for each_result in result_account_report:
                    row = str(INDEX_ROW + i)
                    ws['A' + row] = each_result["company_name"]
                    ws['B' + row] = each_result["tipo"]
                    ws['C' + row] = each_result["journal_name"]
                    ws['D' + row] = each_result["partner_vat"]
                    ws['E' + row] = each_result["partner_name"]
                    ws['F' + row] = each_result["internal_id"]
                    ws['G' + row] = each_result["name"]
                    ws['H' + row] = each_result["invoice_date"]

                    ws['I' + row] = each_result["amount_total"]

                    ws['J' + row] = each_result["clasificacion"]
                    ws['K' + row] = each_result["cuenta_code"]

                    ws['L' + row] = each_result["cuenta_name"]
                    ws['M' + row] = each_result["id"]
                    ws['N' + row] = each_result["date_maturity"]  # balance
                    ws['O' + row] = each_result["valor"]
                    ws['P' + row] = each_result["saldo"] # saldo
                    ws['Q' + row] = each_result["percentage"]
                    ws['R' + row] =  each_result["percentage"]* each_result["saldo"]

                    i += 1
                    last_row = INDEX_ROW + i
                    # ultima_columna_registros+=1
            ws['A1'] = companys
            ws['B2'] = COUNT
            ultima_letra = "R"
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':' + ultima_letra + str(last_row - 1), border)
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))

    def create_report_cxp_nc_anticipos(self, brw_wizard, ws):
        try:
            companys = ', '.join(brw_wizard.mapped('company_ids').mapped('name'))
            company_ids = brw_wizard.company_ids.ids + [-1, -1]
            ACCOUNT_REPORT_QUERY = f""";with variables as (
	SELECT ARRAY[3] AS COMPany_ids 
),
base_documentos as ( 
	SELECT  am.company_id,AM.ID AS DOC_ID,AM.move_type,am.name,coalesce(am.date,am.invoice_date) as doc_date,
	ap.id as payment_id,
			AML.id,am.journal_id,aml.partner_id,
			COALESCE(AML.DATE_MATURITY,AM.DATE,NOW()::DATE) as DATE_MATURITY,
                 SUM(AML.AMOUNT_RESIDUAL) AS saldo,
                 SUM(AML.debit) AS valor,
				 coalesce(ap.is_prepayment,false) as anticipo,
				 ap.purchase_id as purchase_id
         FROM
            ACCOUNT_MOVE AM
            INNER JOIN ACCOUNT_MOVE_LINE AML ON AML.MOVE_ID = AM.ID
            INNER JOIN ACCOUNT_ACCOUNT AA ON AA.ID = AML.ACCOUNT_ID
               and ( AA.ACCOUNT_TYPE = 'liability_payable'
			   	or (aa.ACCOUNT_TYPE='asset_prepayments' and aa.prepayment_account)
				)
			 left join account_payment ap on ap.move_id=am.id 
        WHERE AM.STATE = 'posted' and am.move_type='entry'
           AND AM.COMPANY_ID = ANY (array{company_ids}) and aml.DEBIT>0.00 
         GROUP BY am.company_id,am.journal_id,aml.partner_id,
           AM.ID,AML.id,AM.move_type,ap.id,
			COALESCE(AML.DATE_MATURITY,AM.DATE,NOW()::DATE) , coalesce(ap.is_prepayment,false)
            having round(SUM(AML.AMOUNT_RESIDUAL),2)!=0.00
	union 
	SELECT  am.company_id,AM.ID AS DOC_ID,AM.move_type,am.name,coalesce(am.date,am.invoice_date) as doc_date,
	null as payment_id,
			AML.id,am.journal_id,aml.partner_id,
			COALESCE(AML.DATE_MATURITY,AM.DATE,NOW()::DATE) as DATE_MATURITY,
                 SUM(AML.AMOUNT_RESIDUAL) AS saldo,
                 SUM(AML.debit) AS valor,
				 false as anticipo,
				 null as purchase_id
         FROM
            ACCOUNT_MOVE AM
            INNER JOIN ACCOUNT_MOVE_LINE AML ON AML.MOVE_ID = AM.ID
            INNER JOIN ACCOUNT_ACCOUNT AA ON AA.ID = AML.ACCOUNT_ID
               AND AA.ACCOUNT_TYPE = 'liability_payable' 
        WHERE AM.STATE = 'posted' and am.move_type='in_refund'
           AND AM.COMPANY_ID = ANY (array{company_ids}) and aml.DEBIT>0.00 
         GROUP BY am.company_id,aml.partner_id,
           AM.ID,AML.id,AM.move_type, am.name,coalesce(am.date,am.invoice_date),am.journal_id,
			COALESCE(AML.DATE_MATURITY,AM.DATE,NOW()::DATE) 
            having round(SUM(AML.AMOUNT_RESIDUAL),2)!=0.00
),
solo_pagos_anticipados as (
	select payment_id,DOC_ID,purchase_id from base_documentos where anticipo=true
	group by payment_id,DOC_ID,purchase_id
),
CLASIFICACION_CTAS AS (                              
 	select agrl.name,aat.code,AAT.company_id,aat.id
		from account_Group_report agr
		inner join variables on 1=1
		inner join account_group_report_line agrl on agrl.report_id=agr.id
		inner join report_template_account_acc_rel rtpl on rtpl.report_line_id=agrl.id 
		inner join account_account  aat on aat.id=rtpl.account_id
		where agr.CODE='FLUJO' and aat.company_id = ANY (variables.company_ids )                                                                   
) ,
distribucion as (
   
        SELECT ARQ.DOC_ID,COALESCE(AA.NAME,AAC.NAME) AS CUENTA_NAME,
                    COALESCE(AA.CODE,AAC.CODE) AS CUENTA_CODE,C.name AS CLASIFICACION ,
                    
                    SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.DOC_ID) AS TOTAL,
		case when(SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.DOC_ID)>0) 
		then SUM(AML.PRICE_TOTAL)/SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.DOC_ID)
			else 1.00 end  as percentage ,
			PO.ID AS ORDER_ID,
			PO.NAME AS ORDER_NAME 
        	FROM solo_pagos_anticipados ARQ
        	INNER JOIN PURCHASE_ORDER_LINE AML ON AML.ORDER_ID=ARQ.purchase_id AND AML.PRODUCT_ID IS NOT NULL 
        	INNER JOIN PURCHASE_ORDER PO ON PO.ID=AML.ORDER_ID 
			INNER JOIN PRODUCT_PRODUCT PP ON PP.ID=AML.PRODUCT_ID 
			INNER JOIN PRODUCT_TEMPLATE PT ON PT.ID=PP.PRODUCT_TMPL_ID 
			
			LEFT JOIN IR_PROPERTY IPPT ON IPPT.RES_ID=('product.template,'||PT.id )::VARCHAR
				AND IPPT.TYPE='many2one' AND IPPT.name='property_account_expense_id'
				AND IPPT.COMPANY_ID=PO.COMPANY_ID 
			
			LEFT JOIN ACCOUNT_ACCOUNT AA ON ('account.account,'||AA.ID)::VARCHAR=IPPT.VALUE_REFERENCE::VARCHAR AND IPPT.ID IS NOT NULL AND AA.COMPANY_ID=PO.COMPANY_ID 

	

			LEFT JOIN IR_PROPERTY IPPTC ON IPPTC.RES_ID=('product.category,'||PT.CATEG_ID )::VARCHAR
				AND IPPTC.TYPE='many2one' AND IPPTC.name='property_account_expense_categ_id'
				AND IPPTc.COMPANY_ID=PO.COMPANY_ID 
		LEFT JOIN ACCOUNT_ACCOUNT AAC ON ('account.account,'||AAC.ID)::VARCHAR=IPPTC.VALUE_REFERENCE::VARCHAR AND IPPTC.ID IS NOT NULL AND AAC.COMPANY_ID=PO.COMPANY_ID 

			LEFT JOIN CLASIFICACION_CTAS C ON C.id=COALESCE(AA.ID ,AAC.ID) 
			
				group by ARQ.DOC_ID,COALESCE(AA.NAME,AAC.NAME),COALESCE(AA.CODE,AAC.CODE),C.name,PO.ID ,PO.name  
)


select rc.name as company_name,
case when(bd.move_type='in_refund') then 'NOTA DE CREDITO'
	when(bd.move_type='entry' and bd.payment_id is null) then 'ASIENTO CONTABLE' 
	when(bd.move_type='entry' and bd.payment_id is not null and not bd.anticipo) then 'PAGO'
	when(bd.move_type='entry' and bd.payment_id is not null and bd.anticipo) then 'ANTICIPO'
end as tipo,
aj.name as journal_name,
coalesce(rp.vat,'') AS partner_VAT,
coalesce(rp.name,'DESCONOCIDO') AS partner_name,
bd.DOC_ID as report_Doc_id,
bd.*,
d.order_id,
d.order_name,
d.CUENTA_NAME,
d.CUENTA_CODE,
d.CLASIFICACION,
d.total,
coalesce(d.percentage,1.00) as percentage,			
coalesce(bd.saldo,0.00)*coalesce(d.percentage,1.00) as saldo_aplicado 
from 
base_documentos bd
inner join res_company rc on rc.id=bd.company_id
inner join account_journal aj on aj.id=bd.journal_id 
left join res_partner rp on rp.id=bd.partner_id 
left join distribucion d on d.DOC_ID=bd.DOC_ID
"""
            self._cr.execute(ACCOUNT_REPORT_QUERY)
            result_account_report = self._cr.dictfetchall()
            COUNT = len(result_account_report)
            i, INDEX_ROW = 0, 6
            last_row = INDEX_ROW
            if result_account_report:
                for each_result in result_account_report:
                    row = str(INDEX_ROW + i)
                    ws['A' + row] = each_result["company_name"]
                    ws['B' + row] = each_result["tipo"]
                    ws['C' + row] = each_result["journal_name"]
                    ws['D' + row] = each_result["partner_vat"]
                    ws['E' + row] = each_result["partner_name"]
                    ws['F' + row] = each_result["report_doc_id"]
                    ws['G' + row] = each_result["name"]
                    ws['H' + row] = each_result["doc_date"]
                    ws['I' + row] = each_result["valor"]
                    ws['J' + row] = each_result["saldo"]

                    ws['K' + row] = each_result["order_id"]
                    ws['L' + row] = each_result["order_name"]

                    ws['M' + row] = each_result["clasificacion"]
                    ws['N' + row] = each_result["cuenta_code"]
                    ws['N' + row] = each_result["cuenta_name"]
                    ws['O' + row] = each_result["total"]
                    ws['P' + row] = each_result["saldo"]
                    ws['Q' + row] = each_result["percentage"]
                    ws['R' + row] = each_result["saldo_aplicado"]

                    i += 1
                    last_row = INDEX_ROW + i
            ws['A1'] = companys
            ws['B2'] = COUNT
            ultima_letra = "R"
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':' + ultima_letra + str(last_row - 1), border)
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))

class report_cobros_anticipados_xlsx(models.AbstractModel):
    _name = "report.gps_informes.report_cobros_anticipados_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Cobros Anticipados"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)
        filename = dtFile.join(dir_path, "reporte_cobros_anticipados.xlsx")
        dtFile.copyfile(filename, new_filename)
        wb = False
        try:
            brw_wizard = self.env["l10n_ec.account.report"].sudo().browse(docids[-1])
            wb, ws, target = self.open_xlsx(new_filename, load_sheet=False)
            ws = wb["REPORTE  DE COBROS ANTICIPADOS"]
            self.create_report_cxc_anticipo(brw_wizard, ws)
            #ws = wb["DETALLE DE COBROS"]
            #self.create_report_cxp_nc_anticipos(brw_wizard, ws)
            wb = self.save_wb(wb, target)
            #########################################
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

    def create_report_cxc_anticipo(self, brw_wizard, ws):
        try:
            companys = ', '.join(brw_wizard.mapped('company_ids').mapped('name'))
            company_ids = brw_wizard.company_ids.ids + [-1, -1]
            date_from = brw_wizard.date_from or datetime.date(2020,1,1)
            date_to = brw_wizard.date_to or datetime.date(2090,1,1)
            ACCOUNT_REPORT_QUERY=f"""WITH variables AS (
    SELECT ARRAY{company_ids} AS company_ids,
           '{date_from}'::date AS date_from,
           '{date_to}'::date AS date_to
)

 SELECT
rc.name AS COMPany_name,
rp.vat as partner_Vat,
rp.name AS PARTNER_name , 
sum(ap.amount) as total,
sum(ap.amount_Residual) as pendiente 	
FROM variables
INNER JOIN account_payment  ap   ON 1=1
inner join account_move am on am.id=ap.move_id and 
	am.company_id = ANY(variables.company_ids)
    AND am.date::date>=variables.date_from and am.date::date<=variables.date_to 
INNER JOIN res_company rc ON rc.id = am.company_id
INNER JOIN res_partner rp ON rp.id = am.partner_id
inner join account_account aa on aa.id=ap.prepayment_account_id
WHERE am.state IN ('draft', 'posted' ) and ap.is_prepayment
	and aa.name ilike 'ANTICIPOS DE CLIENTE%'
	GROUP BY 
	rc.name ,rp.vat,rp.name  """
            self._cr.execute(ACCOUNT_REPORT_QUERY)
            result_account_report = self._cr.dictfetchall()
            COUNT = len(result_account_report)
            i, INDEX_ROW = 0, 6
            last_row = INDEX_ROW
            if result_account_report:
                for each_result in result_account_report:
                    row = str(INDEX_ROW + i)
                    ws['A' + row] = each_result["company_name"]
                    ws['B' + row] = each_result["partner_vat"]
                    ws['C' + row] = each_result["partner_name"]
                    ws['D' + row] = each_result["total"]
                    ws['E' + row] = each_result["pendiente"]
                    i += 1
                    last_row = INDEX_ROW + i
                    # ultima_columna_registros+=1
            ws['A1'] = companys
            ws['D2'] = COUNT
            ws['B2'] = brw_wizard.date_from or ''
            ws['B3'] = brw_wizard.date_to or ''
            ultima_letra = "E"
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':' + ultima_letra + str(last_row - 1), border)
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
