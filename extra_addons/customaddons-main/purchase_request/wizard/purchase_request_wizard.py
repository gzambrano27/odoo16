# -*- coding: utf-8 -*-
# -*- encoding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.
from odoo import api,fields, models,_
from xlrd import open_workbook
import base64
import tempfile
import xlrd
from odoo.exceptions import ValidationError,UserError
from ...calendar_days.tools import CalendarManager,DateManager
from ...message_dialog.tools import FileManager
dtObj=DateManager()
clObj=CalendarManager()
flObj=FileManager()
from openpyxl import load_workbook

class PurchaseRequestWizard(models.TransientModel):
    _name="purchase.request.wizard"
    _description="Asistente de Requisicion de Compra"
    
    @api.model
    def get_default_movement_id(self):
        return self._context.get("active_ids",False) and self._context["active_ids"][0] or False
    
    @api.model
    def get_default_company_id(self):
        if self._context.get("active_ids",False):
            for brw_each in self.env["purchase.request"].sudo().browse(self._context["active_ids"]):
                return brw_each.company_id.id
        return False


    movement_id=fields.Many2one("purchase.request", "Requisicion",required=False,default=get_default_movement_id)
    company_id=fields.Many2one("res.company",string="Compañia",required=False,default=get_default_company_id)
    currency_id = fields.Many2one("res.currency", "Moneda", related="company_id.currency_id", store=False,
                                  readonly=True)
    origin=fields.Selection([('file','Archivo'),],string="Origen",default='file')
    file=fields.Binary("Archivo",required=False,filters='*.xlsx')
    file_name=fields.Char("Nombre de Archivo",required=False,size=255)
    
    

    def process(self):
        for brw_each in self:
            if brw_each.origin=="file":
                brw_each.process_file()
        return True
    
    def process_file(self):
        OBJ_PRODUCT   = self.env["product.product"].sudo()
        OBJ_ANALITICA = self.env["account.analytic.account"].sudo()
        OBJ_LINE      = self.env["purchase.request.line"].sudo()

        # Índices de columnas
        COL_ID, COL_PROD, COL_DESC, COL_CTA, COL_UND, COL_QTY, COL_EMPS, COL_SOLO = 0, 1, 2, 3, 4, 5, 6, 7

        # --- helpers ---
        def to_number(value):
            if value is None:
                raise ValueError("vacío")
            if isinstance(value, (int, float)):
                return float(value)
            if isinstance(value, str):
                s = value.strip().replace(' ', '')
                if not s:
                    raise ValueError("vacío")
                # Normaliza 1.234,56 y 1,234.56
                if ',' in s and '.' in s:
                    if s.rfind(',') > s.rfind('.'):
                        s = s.replace('.', '').replace(',', '.')
                    else:
                        s = s.replace(',', '')
                else:
                    s = s.replace(',', '.')
                return float(s)
            raise ValueError("tipo no soportado")

        def to_bool(value):
            return str(value or '').strip().lower() in ('si', 'sí', 'true', '1', 'yes', 'x')

        HEADER_ROW = 3        # fila donde están los títulos (#ID, Producto, ...)
        FIRST_DATA_ROW = HEADER_ROW + 1

        for wiz in self:
            line_ids = []
            ext = flObj.get_ext(wiz.file_name)
            fileName = flObj.create(ext)
            flObj.write(fileName, flObj.decode64(wiz.file))

            book = load_workbook(fileName, data_only=True)
            sheet = book.active

            for row_index, row in enumerate(
                sheet.iter_rows(min_row=FIRST_DATA_ROW, values_only=True),
                start=FIRST_DATA_ROW
            ):
                # Salta filas completamente vacías
                if not row or not any(c not in (None, '') for c in row):
                    continue

                default_code = (row[COL_PROD] or '').strip()

                # Salta si no hay código o si por alguna razón se coló la fila de títulos
                if not default_code or default_code.lower() == 'producto':
                    continue

                # Lee/castea cantidad
                try:
                    quantity = to_number(row[COL_QTY])
                except Exception:
                    raise ValidationError(_("La cantidad debe ser numérica. Revisa la fila %s") % row_index)

                # Empleados
                employee_raw = row[COL_EMPS] or ''
                employee_codes = [c.strip() for c in str(employee_raw).split(',') if c.strip()]
                employee_codes = [('0' + c) if len(c) == 9 else c for c in employee_codes]
                # elimina duplicados conservando orden
                employee_codes = list(dict.fromkeys(employee_codes))
                employee_ids = self.env['hr.employee'].sudo().search([('identification_id', 'in', employee_codes)])
                if len(employee_codes) != len(employee_ids):
                    raise UserError(_("Algunos empleados no encontrados en: %s") % employee_codes)

                # Producto
                srch_product = OBJ_PRODUCT.search([('default_code', '=', default_code)], limit=2)
                if not srch_product:
                    raise ValidationError(_("No hay producto con código %s. Revisa la fila %s") % (default_code, row_index))
                if len(srch_product) > 1:
                    raise ValidationError(_("Existe más de un producto con código %s en la base de datos. Revisa la fila %s") % (default_code, row_index))
                product = srch_product[0]

                # Cuenta analítica (opcional)
                analytic_distribution = {}
                ctaanalitica = (row[COL_CTA] or '').strip()
                if ctaanalitica:
                    acc = OBJ_ANALITICA.search([('name', '=', ctaanalitica), ('company_id', '=', wiz.company_id.id)], limit=1)
                    if not acc:
                        raise ValidationError(_("No hay cuenta analítica con nombre %s. Revisa la fila %s") % (ctaanalitica, row_index))
                    analytic_distribution[str(acc.id)] = 100

                # ID de línea (para actualizar si viene)
                line_value = row[COL_ID]
                line_id = int(line_value) if isinstance(line_value, (int, float)) else 0

                values = {
                    "product_qty": quantity,
                    "product_id": product.id,
                    "name": row[COL_DESC] or '',
                    'product_uom_id': product.uom_po_id.id,
                    'employees_ids': [(6, 0, employee_ids.ids)],
                    'un_solo_custodio': to_bool(row[COL_SOLO]),
                }
                if analytic_distribution:
                    values['analytic_distribution'] = analytic_distribution

                line_ids.append((1, line_id, values) if line_id else (0, 0, values))

            wiz.movement_id.write({"line_ids": line_ids})
        return True