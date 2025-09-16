# coding: utf-8
from odoo import api, fields, models, exceptions, tools, _
from ...message_dialog.tools import FileManager
from ...calendar_days.tools import DateManager
from ...calendar_days.tools import CalendarManager
dtFile=FileManager()
dateO=DateManager()
calendarO=CalendarManager()

import logging
_logger = logging.getLogger(__name__)
import re
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill

# Relleno amarillo

class ReportFormulario103Xlsx(models.AbstractModel):
    _name = 'report.gps_anexos.report_formulario_103_xlsx'
    _inherit = 'report.report_xlsx.abstract'

    def create_xlsx_report2(self, docids, data):
        EXT = "xlsx"
        new_filename = dtFile.create(EXT)
        def get_ws_objects(template_file_name):
            dir_path = dtFile.get_path_file(__file__)
            filename = dtFile.join(dir_path, template_file_name)
            dtFile.copyfile(filename, new_filename)
            wb, ws, target = self.open_xlsx(new_filename, load_sheet=True)
            return wb, ws, target
        wb = False
        try:
            wb, ws, target = get_ws_objects("Formulario_103.xlsx")
            fill_amarillo = PatternFill(start_color="FFFF00",  # Color en formato HEX ARGB
                                        end_color="FFFF00",
                                        fill_type="solid")
            for item in data.get('result', []):
                cell = item.get('cell_row')  # Ejemplo: "S13"
                value = item.get('value')
                if cell and value is not None:
                    ws[cell] = value  # Escribe el valor en la celda indicada
                if not cell and value is not None:
                    ky= item.get('ky_id')
                    if ky=="year":
                        ws["R5"] =value
                        ws.merge_cells('R5:U6')
                    if ky=="mes":
                        celdas_pintar={1:"C5",2:"D5",3:"E5",4:"F5",5:"G5",6:"H5",
                                       7:"I5",8:"J5",9:"K5",10:"L5",11:"M5",12:"N5"}
                        ws[celdas_pintar[int(value)]].fill = fill_amarillo
                    if ky=="vat":
                        start_col = "B"  # <-- aquí va tu "?"
                        row = 10
                        col = column_index_from_string(start_col)
                        for each_val in value:
                            ws.cell(row=row, column=col,
                                    value=each_val)  # equivalente a ws[f"{get_column_letter(col)}{row}"] = each_val
                            col += 1
            #ws = wb["SOLICITUDES"]
            wb = self.save_wb(wb, target)
            print("Datos escritos correctamente")
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        new_filename = dtFile.create(EXT)
        def get_ws_objects(template_file_name):
            dir_path = dtFile.get_path_file(__file__)
            filename = dtFile.join(dir_path, template_file_name)
            dtFile.copyfile(filename, new_filename)
            wb, ws, target = self.open_xlsx(new_filename, load_sheet=True)
            return wb, ws, target
        wb = False
        try:
            wb, ws, target = get_ws_objects("Formulario_103.xlsx")
            # fill_amarillo = PatternFill(start_color="FFFF00",  # Color en formato HEX ARGB
            #                             end_color="FFFF00",
            #                             fill_type="solid")
            for item in data.get('result', []):
                print(item)
                cell = item.get('cell_row')  # Ejemplo: "S13"
                value = item.get('value')
                if cell and value is not None:
                    ws[cell] = value  # Escribe el valor en la celda indicada
                if not cell and value is not None:
                    ky= item.get('ky_id')
                    # if ky=="year":
                    #     ws["R5"] =value
                    #     ws.merge_cells('R5:U6')
                    # if ky=="mes":
                    #     celdas_pintar={1:"C5",2:"D5",3:"E5",4:"F5",5:"G5",6:"H5",
                    #                    7:"I5",8:"J5",9:"K5",10:"L5",11:"M5",12:"N5"}
                    #     ws[celdas_pintar[int(value)]].fill = fill_amarillo
                    # if ky=="vat":
                    #     start_col = "B"  # <-- aquí va tu "?"
                    #     row = 10
                    #     col = column_index_from_string(start_col)
                    #     for each_val in value:
                    #         ws.cell(row=row, column=col,
                    #                 value=each_val)  # equivalente a ws[f"{get_column_letter(col)}{row}"] = each_val
                    #         col += 1
            #ws = wb["SOLICITUDES"]
            wb = self.save_wb(wb, target)
            print("Datos escritos correctamente")
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT