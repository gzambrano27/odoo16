# coding: utf-8
from odoo import api, fields, models, exceptions, tools, _
from ...message_dialog.tools import FileManager
from ...calendar_days.tools import DateManager
from ...calendar_days.tools import CalendarManager

dtFile = FileManager()
dateO = DateManager()
calendarO = CalendarManager()
import openpyxl
import logging

_logger = logging.getLogger(__name__)

from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font, PatternFill
import openpyxl.utils as utils
from openpyxl.styles import NamedStyle


class menu_group_report_xlsx(models.AbstractModel):
    _name = "report.gps_reports.menu_group_report_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Permisos"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)
        filename = dtFile.join(dir_path, "grupo_menus.xlsx")
        dtFile.copyfile(filename, new_filename)
        wb = False
        try:
            brw_wizard = self.env["security.group.wizard"].sudo().browse(docids[-1])
            wb, ws, target = self.open_xlsx(new_filename, load_sheet=False)
            ws = wb["MENUS POR GRUPO"]
            self.create_report_menus(brw_wizard, ws)
            wb = self.save_wb(wb, target)
            #########################################
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

    def create_report_menus(self, brw_wizard, ws):
        menu_ids=brw_wizard.menu_ids.ids
        menu_ids+=[-1,-1]
        menu_ids = tuple(menu_ids)
        self._cr.execute(f""";WITH RECURSIVE group_hierarchy AS (
    -- Obtener los grupos asignados directamente al usuario
    SELECT 
        rg.id AS group_id,
        coalesce(rg.name->>'es_EC',rg.name->>'en_US') AS group_name,
	    coalesce(imc.name->>'es_EC',imc.name->>'en_US') as category_name,
		rg.id as pk_id
    FROM res_groups rg
	LEFT JOIN IR_module_category imc on imc.id=rg.category_id
    UNION ALL
    -- Obtener los grupos heredados a través de la relación implied_ids
    SELECT 
        rg2.id AS group_id,
        gh.group_name,
		gh.category_name ,
		coalesce(rgi.hid,gh.pk_id) as pk_id
    FROM res_groups rg2
    inner JOIN res_groups_implied_rel rgi ON rg2.id = rgi.hid
    inner JOIN group_hierarchy gh ON rgi.gid = gh.group_id
	
),
menu_group as (
	select gh.*,img.menu_id
	from
	ir_ui_menu_group_rel img
	inner join group_hierarchy gh on gh.pk_id=img.gid
	
),
group_usuarios_menu as (
	select rgi.uid,mg.* 
	from menu_group mg
	inner join res_groups_users_rel rgi on rgi.gid=mg.pk_id
)


select gu.category_name,gu.group_name,
uim.id as menu_id,
coalesce(uim.name->>'es_EC',uim.name->>'en_US') as menu_name,gu.uid,
ru.login as login_name,
rpu.name as nombre_usuario 
from ir_ui_menu uim 
inner join group_usuarios_menu gu on gu.menu_id=uim.id
inner join res_users ru on ru.id=gu.uid and ru.active 
inner join res_partner rpu on rpu.id=ru.partner_id  and ru.active  
left join user_menu_rel hide on hide.user_id=ru.id and hide.menu_id=uim.id
where uim.id in %s
and uim.active=true
and hide.user_id  is null 
        """,(menu_ids,))
        result = self._cr.dictfetchall()
        if result:
            i, INDEX_ROW = 0, 4
            for each_result in result:
                row = str(INDEX_ROW + i)
                ws['A' + row] = each_result["category_name"]
                ws['B' + row] = each_result["group_name"]
                ws['C' + row] = each_result["menu_id"]
                ws['D' + row] = each_result["menu_name"]
                ws['E' + row] = each_result["uid"]
                ws['F' + row] = each_result["login_name"]
                ws['G' + row] = each_result["nombre_usuario"]
                i += 1
            last_row = INDEX_ROW + i
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':G' + str(last_row - 1), border)
            ws['B2'] = len(result)
