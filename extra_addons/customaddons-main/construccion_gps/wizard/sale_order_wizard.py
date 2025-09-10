import re
import unicodedata
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
from openpyxl import load_workbook


def _normalize_header(text):
    """Convierte a minúsculas, elimina espacios y tildes."""
    if not text:
        return ''
    text = str(text).strip().lower()
    # quitar tildes
    text = ''.join(
        c for c in unicodedata.normalize('NFKD', text)
        if not unicodedata.combining(c)
    )
    return text


class SaleOrderImportWizard(models.TransientModel):
    _name = 'sale.order.import.wizard'
    _description = 'Importar Líneas al Sale Order'

    sale_order_id = fields.Many2one(
        'sale.order', string='Pedido de Venta', required=True, readonly=True)
    file = fields.Binary('Archivo Excel', required=True)
    filename = fields.Char('Nombre del Archivo')

    def action_import_lines(self):
        if not self.file:
            raise UserError(_("Por favor, suba un archivo Excel."))

        file_data = base64.b64decode(self.file)
        try:
            workbook = load_workbook(filename=io.BytesIO(file_data), data_only=True)
        except Exception as e:
            raise UserError(_("Error leyendo el archivo Excel: %s") % e)

        sheet = workbook.active
        # 🔹 añadimos columna APU
        expected_headers = [
            'id', 'apu', 'producto', 'descripcion',
            'cuentas analiticas por plan (%)', 'cantidad', 'precio unitario'
        ]
        header_row = None

        # Busca fila de encabezados
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=20), start=1):
            values = [_normalize_header(cell.value) for cell in row[:len(expected_headers)]]
            if values == expected_headers:
                header_row = i
                break
        if not header_row:
            raise UserError(_("La estructura del Excel no es válida. Se esperan las columnas: %s") %
                            ', '.join(expected_headers))

        # Procesa cada fila
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            # 🔹 ahora hay 7 columnas
            line_id, apu_val, prod_name, name, analytic_cell, qty, price_unit = row
            if not prod_name:
                continue

            # Extrae default_code
            default_code = False
            if isinstance(prod_name, str) and prod_name.startswith('[') and ']' in prod_name:
                default_code = prod_name.split(']')[0].strip('[]').strip()

            # Buscar product.template
            search_term = prod_name.strip()
            code_to_search = default_code or search_term
            domain = ['|', ('default_code', '=', code_to_search), ('name', '=', search_term)]
            tpl = self.env['product.template'].search(domain, limit=1)
            if not tpl:
                raise UserError(_("No se encontró el Producto Template '%s'.") % prod_name)

            vals = {
                'order_id': self.sale_order_id.id,
                'product_id': tpl.product_variant_id.id,
                'name': name or tpl.display_name,
                'product_uom_qty': qty or 0.0,
            }
            if price_unit not in (None, 0):
                vals['price_unit'] = price_unit

            # 🔹 Buscar y asignar APU
            if apu_val:
                apu = False
                if isinstance(apu_val, (int, float)):
                    apu = self.env['apu.apu'].browse(int(apu_val))
                if not apu:
                    apu = self.env['apu.apu'].search([('code', '=', str(apu_val))], limit=1)
                if not apu:
                    raise UserError(_("No se encontró el APU con referencia '%s'.") % apu_val)
                vals['apu_id'] = apu.id

            # Procesa las cuentas analíticas agrupadas por plan
            if analytic_cell:
                lines = str(analytic_cell).splitlines()
                distribution = {}
                for ln in lines:
                    parts = ln.split(':', 1)
                    if len(parts) != 2:
                        continue
                    accounts_str = parts[1]
                    for match in re.finditer(r'([^,]+?)\s*\(\s*([\d\.,]+)%\s*\)', accounts_str):
                        acct_name = match.group(1).strip()
                        pct_str = match.group(2).replace(',', '.')
                        try:
                            pct = float(pct_str)
                        except ValueError:
                            continue
                        aa = self.env['account.analytic.account'].search([('name', '=', acct_name)], limit=1)
                        if not aa:
                            raise UserError(_("No se encontró la Cuenta Analítica '%s'.") % acct_name)
                        distribution[aa.id] = pct
                if distribution:
                    vals['analytic_distribution'] = distribution

            # Actualiza o crea la línea
            if line_id:
                existing = self.sale_order_id.order_line.filtered(lambda l: l.id == int(line_id))
                if existing:
                    existing.write(vals)
                else:
                    self.env['sale.order.line'].create(vals)
            else:
                self.env['sale.order.line'].create(vals)

        return {'type': 'ir.actions.act_window_close'}

    def _create_sale_line(self, **vals):
        """Crea una línea nueva en el pedido de venta activo."""
        vals['order_id'] = self.sale_order_id.id
        self.sale_order_id.order_line.create(vals)
