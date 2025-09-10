# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.
from odoo import api, fields, models, _, Command
from datetime import datetime, timedelta
from odoo.exceptions import UserError, ValidationError
from odoo.osv.expression import AND, OR
from odoo.tools import float_round
from odoo.addons import decimal_precision as dp
from collections import defaultdict
import logging
import io, base64, xlsxwriter
import json
import math

_logger = logging.getLogger(__name__)


def _format_value_for_tracking(value):
	"""Formatea valores para mostrarlos en el chatter de manera segura."""
	# Vacíos, pero preserva 0
	if not value and value != 0:
		return "—"

	# ¿Es un recordset de Odoo? (evitando tocar display_name)
	# Nota: 'mapped' y '_name' existen en recordsets; no disparan ensure_one.
	if hasattr(value, '_name') and hasattr(value, 'mapped'):
		# Recordset vacío
		if not value:
			return "—"
		# Para 1 o N registros usamos name_get() (no requiere ensure_one)
		try:
			names = [name for _id, name in value.name_get()]
			return ", ".join(names)
		except Exception:
			# Fallback por si el modelo no implementa name_get de forma estándar
			try:
				return ", ".join(value.mapped('name')) or ", ".join(map(str, value.ids))
			except Exception:
				return str(value)

	# Tipos primitivos / otros
	return str(value)


try:
	from openpyxl import load_workbook
except ImportError:
	raise UserError(_("El módulo openpyxl es requerido para importar archivos Excel."))


class APUCategoria(models.Model):
	_name = 'apu.categoria'
	_description = 'Categoría de APU'
	_inherit = ['mail.thread']
	_order = 'codigo asc'
	_rec_name = 'codigo'

	name = fields.Char(string='Nombre', required=True)
	codigo = fields.Char(string='Código', required=True)
	subcategoria_ids = fields.One2many('apu.subcategoria', 'categoria_id', string='Subcategorías')

	def name_get(self):
		result = []
		for record in self:
			# Combina el código y el nombre para la representación
			rec_name = "%s - %s" % (record.codigo or '', record.name or '')
			result.append((record.id, rec_name))
		return result


class APUSubcategoria(models.Model):
	_name = 'apu.subcategoria'
	_description = 'Subcategoría de APU'
	_inherit = ['mail.thread']
	_order = 'codigo asc'
	_rec_name = 'codigo'

	name = fields.Char(string='Nombre', required=True)
	codigo = fields.Char(string='Código', required=True)
	categoria_id = fields.Many2one('apu.categoria', string='Categoría', required=True)

	def name_get(self):
		result = []
		for record in self:
			# Combina el código y el nombre para la representación
			rec_name = "%s - %s" % (record.codigo or '', record.name or '')
			result.append((record.id, rec_name))
		return result


class ProductoTemplate(models.Model):
	_inherit = 'product.template'

	subcategoria_id = fields.Many2one('apu.subcategoria', string="Subcategoría",
	                                  help="Selecciona la subcategoría asociada a este APU")
	notas_descripcion = fields.Text(string="Nota", tracking=True)

	apu_custom_ids = fields.One2many(
		'product.template.apu',
		'product_tmpl_id',
		string='APU Personalizadas',
	)
	actividad_custom_ids = fields.One2many(
		'product.template.actividad',
		'product_tmpl_id',
		string='Actividades Personalizadas',
	)


class ProductTemplateAPU(models.Model):
	_name = 'product.template.apu'
	_description = 'APU personalizado para Product Template'

	product_tmpl_id = fields.Many2one('product.template', string='Plantilla de Producto', ondelete='cascade',
	                                  required=True, index=True, )
	apu_line_id = fields.Many2one('apu.apu.line', string='Línea APU', ondelete='cascade',
	                              help='Enlace con apu.apu.line', domain=[('bom_id', '!=', False)], )
	fecha_modificacion = fields.Datetime(string='Fecha de Modificación', default=fields.Datetime.now, )
	cantidad = fields.Float(string='Cantidad', digits=(16, 4), default=0.0, )
	precio_unitario = fields.Float(string='Precio Unitario', digits=(16, 4), default=0.0, )
	costo_unitario = fields.Float(string='Costo Unitario', digits=(16, 4), default=0.0, )
	precio_total = fields.Float(string='Precio Total', digits=(16, 4), default=0.0, )
	costo_total = fields.Float(string='Costo Total', digits=(16, 4), default=0.0, )


class ProductTemplateActividad(models.Model):
	_name = 'product.template.actividad'
	_description = 'Actividad personalizada para Product Template'

	product_tmpl_id = fields.Many2one('product.template', string='Plantilla de Producto', ondelete='cascade',
	                                  required=True, index=True, )
	apu_line_tareas_id = fields.Many2one('apu.apu.line.tareas', string='Línea de Tareas APU', ondelete='cascade',
	                                     help='Enlace con apu.apu.line.tareas', )
	fecha_modificacion = fields.Datetime(string='Fecha de Modificación', default=fields.Datetime.now, )
	cantidad = fields.Float(string='Cantidad', digits=(16, 4), default=0.0, )
	precio_unitario = fields.Float(string='Costo/U', digits=(16, 4), default=0.0, )
	costo_hora_actividad = fields.Float(string='Costo Hora', digits=(16, 4), default=0.0, )
	rendimiento = fields.Float('Hora unidad', default=1.0, digits=(16, 4), required=True)
	costo_final_actividad = fields.Float('Costo Final', digits=(16, 4))


class ApuApu(models.Model):
	""" Defines bills of material for a product or a product template """
	_name = 'apu.apu'
	_description = 'Bill of Material'
	_inherit = ['mail.thread']
	_rec_name = 'product_tmpl_id'
	_rec_names_search = ['product_tmpl_id', 'code']
	_order = "sequence, id"
	_check_company_auto = True

	def _get_default_product_uom_id(self):
		return self.env['uom.uom'].search([], limit=1, order='id').id

	code = fields.Char('Reference', copy=False)
	active = fields.Boolean('Active', default=True)
	type = fields.Selection([
		# ('normal', 'Manufacture this product'),
		('phantom', 'Kit')], 'Tipo',
		default='phantom', required=True)
	product_tmpl_id = fields.Many2one(
		'product.template', 'Product',
		check_company=True, index=True,
		domain="[('type', 'in', ['product', 'consu']), '|', ('company_id', '=', False), ('company_id', '=', company_id)]")
	product_id = fields.Many2one(
		'product.product', 'Product Variant',
		check_company=True, index=True,
		domain="['&', ('product_tmpl_id', '=', product_tmpl_id), ('type', 'in', ['product', 'consu']),  '|', ('company_id', '=', False), ('company_id', '=', company_id)]",
		help="If a product variant is defined the BOM is available only for this product.")
	subcategoria_id = fields.Many2one(
		'apu.subcategoria',
		string="Subcategoría",
		help="Selecciona la subcategoría asociada a este APU",
		copy=False
	)
	line_ids = fields.One2many('apu.apu.line', 'bom_id', 'BoM Lines', copy=True)
	byproduct_ids = fields.One2many('apu.apu.byproduct', 'bom_id', 'By-products', copy=True)
	product_qty = fields.Float(
		'Quantity', default=1.0,
		digits=(16, 4), required=True,
		help="This should be the smallest quantity that this product can be produced in. If the BOM contains operations, make sure the work center capacity is accurate.")
	product_uom_id = fields.Many2one(
		'uom.uom', 'Unit of Measure',
		default=_get_default_product_uom_id, required=True,
		help="Unit of Measure (Unit of Measure) is the unit of measurement for the inventory control",
		domain="[('category_id', '=', product_uom_category_id)]")
	product_uom_category_id = fields.Many2one(related='product_tmpl_id.uom_id.category_id')
	sequence = fields.Integer('Sequence')
	operation_ids = fields.One2many('mrp.routing.workcenter', 'bom_id', 'Operations', copy=True)
	ready_to_produce = fields.Selection([
		('all_available', ' When all components are available'),
		('asap', 'When components for 1st operation are available')], string='Manufacturing Readiness',
		default='all_available', required=True)
	picking_type_id = fields.Many2one(
		'stock.picking.type', 'Operation Type',
		domain="[('code', '=', 'mrp_operation'), ('company_id', '=', company_id)]",
		check_company=True,
		help=u"When a procurement has a ‘produce’ route with a operation type set, it will try to create "
		     "a Manufacturing Order for that product using a BoM of the same operation type. That allows "
		     "to define stock rules which trigger different manufacturing orders with different BoMs.")
	company_id = fields.Many2one(
		'res.company', 'Company', index=True,
		default=lambda self: self.env.company)
	cliente_id = fields.Many2one(
		'res.partner', string='Cliente',
		domain="[('type','!=','private'), ('company_id','in',(False, company_id))]",
		help="Cliente asociado al APU", copy=False)
	consumption = fields.Selection([
		('flexible', 'Allowed'),
		('warning', 'Allowed with warning'),
		('strict', 'Blocked')],
		help="Defines if you can consume more or less components than the quantity defined on the BoM:\n"
		     "  * Allowed: allowed for all manufacturing users.\n"
		     "  * Allowed with warning: allowed for all manufacturing users with summary of consumption differences when closing the manufacturing order.\n"
		     "  Note that in the case of component Manual Consumption, where consumption is registered manually exclusively, consumption warnings will still be issued when appropriate also.\n"
		     "  * Blocked: only a manager can close a manufacturing order when the BoM consumption is not respected.",
		default='warning',
		string='Flexible Consumption',
		required=True
	)
	possible_product_template_attribute_value_ids = fields.Many2many(
		'product.template.attribute.value',
		compute='_compute_possible_product_template_attribute_value_ids')
	allow_operation_dependencies = fields.Boolean('Operation Dependencies',
	                                              help="Create operation level dependencies that will influence both planning and the status of work orders upon MO confirmation. If this feature is ticked, and nothing is specified, Odoo will assume that all operations can be started simultaneously."
	                                              )

	total_manoobra = fields.Float(
		string="Total Mano de Obra",
		compute="_compute_totals",
		store=True
	)
	total_material = fields.Float(
		string="Total Material",
		compute="_compute_totals",
		store=True
	)
	total_equipo = fields.Float(
		string="Total Equipos",
		compute="_compute_totals",
		store=True
	)
	total_transporte = fields.Float(
		string="Total Transporte",
		compute="_compute_totals",
		store=True
	)
	total_general = fields.Float(
		string="Total General",
		compute="_compute_totals",
		store=True
	)

	apply_tax = fields.Boolean(string="¿Aplicar impuesto?", default=False)
	tax_percentage = fields.Float(string="Porcentaje de impuesto (%)", default=0.0)
	total_manoobra_with_tax = fields.Float(
		string="Total Mano de Obra",
		compute="_compute_totals_with_tax",
		store=True,
		readonly=True
	)
	total_material_with_tax = fields.Float(
		string="Total Material",
		compute="_compute_totals_with_tax",
		store=True,
		readonly=True
	)
	total_equipo_with_tax = fields.Float(
		string="Total Equipos",
		compute="_compute_totals_with_tax",
		store=True,
		readonly=True
	)
	total_transporte_with_tax = fields.Float(
		string="Total Transporte",
		compute="_compute_totals_with_tax",
		store=True,
		readonly=True
	)
	total_with_tax = fields.Float(string="Total con impuesto", compute="_compute_totals_with_tax")
	total_apu = fields.Float(
		string="Total APU",
		compute="_compute_total_apu",
		store=True
	)

	# Totales por componente
	total_equipos_precio = fields.Float(compute="_compute_totales")
	total_equipos_costo_directo = fields.Float(compute="_compute_totales")
	total_equipos_margen = fields.Float(compute="_compute_totales")
	total_equipos_margen_porc = fields.Float(compute="_compute_totales")

	total_mano_obra_precio = fields.Float(compute="_compute_totales")
	total_mano_obra_costo_directo = fields.Float(compute="_compute_totales")
	total_mano_obra_margen = fields.Float(compute="_compute_totales")
	total_mano_obra_margen_porc = fields.Float(compute="_compute_totales")

	total_materiales_precio = fields.Float(compute="_compute_totales")
	total_materiales_costo_directo = fields.Float(compute="_compute_totales")
	total_materiales_margen = fields.Float(compute="_compute_totales")
	total_materiales_margen_porc = fields.Float(compute="_compute_totales")

	total_transporte_precio = fields.Float(compute="_compute_totales")
	total_transporte_costo_directo = fields.Float(compute="_compute_totales")
	total_transporte_margen = fields.Float(compute="_compute_totales")
	total_transporte_margen_porc = fields.Float(compute="_compute_totales")

	# Totales generales
	total_precio = fields.Float(compute="_compute_totales")
	total_costo_directo = fields.Float(compute="_compute_totales")
	total_margen_bruto = fields.Float(compute="_compute_totales")
	total_porc_margen_bruto = fields.Float(compute="_compute_totales")
	total_porc_margen_brutoc = fields.Float(compute="_compute_totales")

	# Indirectos
	indirectos_porcentaje = fields.Float(string="Porcentaje Indirectos (%)", default=0.0)
	indirectos_costo = fields.Float(string="Costo Indirectos", compute="_compute_totals", store=True)
	indirectos_precio = fields.Float(string="Precio Indirectos", compute="_compute_totals", store=True)
	indirectos_margen = fields.Float(string="Margen Indirectos ($)", compute="_compute_totals", store=True)
	indirectos_margen_porc = fields.Float(string="Margen Indirectos (%)", compute="_compute_totals", store=True)

	# Utilidad
	utilidad_porcentaje = fields.Float(string="Porcentaje Utilidad (%)", default=0.0)
	utilidad_costo = fields.Float(string="Costo Utilidad ($)", compute="_compute_totals", store=True)
	utilidad_precio = fields.Float(string="Precio Utilidad", default=0.0)
	utilidad_margen = fields.Float(string="Margen Utilidad ($)", compute="_compute_totals", store=True)
	utilidad_margen_porc = fields.Float(string="Margen Utilidad (%)", compute="_compute_totals", store=True)

	# Totales Generales
	total_apu_costo = fields.Float(string="Costo Total APU", compute="_compute_totals", store=True)
	total_apu_precio = fields.Float(string="Precio Total APU", compute="_compute_totals", store=True)
	total_apu_margen = fields.Float(string="Margen Total APU ($)", compute="_compute_totals", store=True)
	total_apu_margen_porc = fields.Float(string="Margen Total APU (%)", compute="_compute_totals", store=True)
	total_apu_margen_porc_100 = fields.Float(string="Margen Total APU (%)", compute="_compute_totals", store=True)
	semaforo_margen = fields.Selection(
		[('verde', 'Verde'), ('naranja', 'Naranja'), ('rojo', 'Rojo')],
		string="Semáforo Margen", compute="_compute_semaforo_margen", store=True
	)
	tag_ids = fields.Many2many(
		'apu.apu.tags',
		'apu_apu_tag_rel',
		'apu_id',
		'tag_id',
		string='Etiquetas', copy=False
	)

	_sql_constraints = [
		('qty_positive', 'check (product_qty > 0)', 'The quantity to produce must be positive!'),
	]

	@api.constrains('product_tmpl_id', 'code', 'company_id', 'cliente_id', 'subcategoria_id', 'tag_ids')
	def _check_unique_combination(self):
		for record in self:
			# Buscar candidatos con todos los campos iguales excepto tag_ids
			domain = [
				('id', '!=', record.id),
				('code', '=', record.code),
				('product_tmpl_id', '=', record.product_tmpl_id.id),
				('company_id', '=', record.company_id.id),
				('cliente_id', '=', record.cliente_id.id),
				('subcategoria_id', '=', record.subcategoria_id.id),
			]
			candidates = self.search(domain)
			record_tags = set(record.tag_ids.ids)

			for dup in candidates:
				if set(dup.tag_ids.ids) == record_tags:
					raise ValidationError(_(
						"La combinación de Código: %s, Producto: %s, Compañía: %s, Cliente: %s, "
						"Subcategoría: %s y Etiquetas (%s) ya existe."
					) % (
						                      record.code,
						                      record.product_tmpl_id.name,
						                      record.company_id.name,
						                      record.cliente_id.name,
						                      record.subcategoria_id.name,
						                      ", ".join(record.tag_ids.mapped('name'))
					                      ))

	@api.constrains('code', 'tag_ids')
	def _check_unique_code_tags(self):
		for record in self:
			# Buscar candidatos con mismo code
			same_code_records = self.search([
				('id', '!=', record.id),
				('code', '=', record.code),
			])
			record_tags = set(record.tag_ids.ids)

			for duplicate in same_code_records:
				if set(duplicate.tag_ids.ids) == record_tags:
					raise ValidationError(
						"Ya existe un APU con el mismo Code '%s' y las mismas Etiquetas (%s)." %
						(record.code, ", ".join(record.tag_ids.mapped('name')))
					)

	def _update_product_total_apu_precio(self):
		self.ensure_one()
		Pricelist = self.env['product.pricelist']
		PricelistItem = self.env['product.pricelist.item']

		# Paso 1: Trabajamos solo con el template del producto
		product_tmpl = self.product_tmpl_id

		# Paso 2: Buscar o crear la lista de precios con el nombre del cliente
		if self.cliente_id:
			pricelist = Pricelist.search([('name', '=', self.cliente_id.name)], limit=1)
			if not pricelist:
				pricelist = Pricelist.create({
					'name': self.cliente_id.name,
					'company_id': False,  # Aseguramos que el campo company_id esté vacío
				})

			# Paso 3: Buscar TODOS los ítems en la lista de precios para este template de producto
			domain = [
				('pricelist_id', '=', pricelist.id),
				('applied_on', '=', '1_product'),  # Aplica sobre el template del producto
				('product_tmpl_id', '=', product_tmpl.id)
			]
			existing_items = PricelistItem.search(domain)

			if existing_items:
				# Si ya existe al menos un ítem, actualizamos el primero (si el precio es diferente)
				first_item = existing_items[0]
				if first_item.fixed_price != self.total_apu_precio:
					first_item.write({
						'fixed_price': self.total_apu_precio,
					})
				# Eliminar cualquier duplicado que pueda existir
				duplicates = existing_items - first_item
				if duplicates:
					duplicates.unlink()
			else:
				# Paso 4: Si no existe ningún ítem y el precio es mayor que cero, se crea uno nuevo
				if self.total_apu_precio > 0:
					PricelistItem.create({
						'pricelist_id': pricelist.id,
						'applied_on': '1_product',  # Aplicamos sobre el template del producto
						'product_tmpl_id': product_tmpl.id,
						'fixed_price': self.total_apu_precio,
						'min_quantity': 1,  # Cantidad mínima es 1 (ajusta según sea necesario)
						'company_id': False,  # Aseguramos que el campo company_id esté vacío
					})
		else:
			for record in self:
				if record.product_tmpl_id:
					record.product_tmpl_id.write({'list_price': record.total_apu_precio})

	@api.model
	def create(self, vals):
		record = super(ApuApu, self).create(vals)
		record._update_product_total_apu_precio()
		return record

	@api.depends('total_apu_margen_porc')
	def _compute_semaforo_margen(self):
		for line in self:
			margin_type = 'margen-apu-cab'  # Tipo de margen a evaluar
			color = self.env['bim.rangos.margen'].evaluate_margin_in_range(line.total_apu_margen_porc * 100,
			                                                               margin_type)

			# Mapear valores de color a opciones válidas
			mapping = {
				'success': 'verde',
				'warning': 'naranja',
				'danger': 'rojo'
			}
			line.semaforo_margen = mapping.get(color, 'rojo')  # Valor predeterminado es 'rojo'

	@api.depends('total_precio', 'total_costo_directo', 'total_margen_bruto', 'total_porc_margen_bruto',
	             'indirectos_porcentaje', 'utilidad_porcentaje')
	def _compute_totals(self):
		for record in self:
			# Totales Directos
			total_directo_costo = record.total_costo_directo
			total_directo_precio = record.total_precio
			total_directo_margen = record.total_margen_bruto
			total_directo_margen_porc = record.total_porc_margen_bruto

			# Indirectos
			indirectos_costo = 0  # (total_directo_costo * record.indirectos_porcentaje) / 100
			indirectos_precio = (total_directo_precio * record.indirectos_porcentaje) / 100
			indirectos_margen = indirectos_precio - indirectos_costo
			indirectos_margen_porc = indirectos_margen / indirectos_precio if indirectos_precio else 0.0

			# Utilidad
			utilidad_costo = 0  # (total_directo_costo * record.utilidad_porcentaje) / 100
			utilidad_precio = (total_directo_precio * record.utilidad_porcentaje) / 100
			utilidad_margen = utilidad_precio - utilidad_costo
			utilidad_margen_porc = utilidad_margen / utilidad_precio if utilidad_precio else 0.0

			# Totales Generales
			total_apu_costo = total_directo_costo  # + indirectos_costo + utilidad_costo
			total_apu_precio = total_directo_precio + indirectos_precio + utilidad_precio
			total_apu_margen = total_apu_precio - total_apu_costo
			total_apu_margen_porc = (total_apu_margen / total_apu_precio) if total_apu_precio else 0.0
			total_apu_margen_porc_100 = ((total_apu_margen / total_apu_precio) if total_apu_precio else 0.0) * 100

			record.indirectos_costo = indirectos_costo
			record.indirectos_precio = indirectos_precio
			record.indirectos_margen = indirectos_margen
			record.indirectos_margen_porc = indirectos_margen_porc

			record.utilidad_costo = utilidad_costo
			record.utilidad_precio = utilidad_precio
			record.utilidad_margen = utilidad_margen
			record.utilidad_margen_porc = utilidad_margen_porc

			record.total_apu_costo = total_apu_costo
			record.total_apu_precio = total_apu_precio
			record.total_apu_margen = total_apu_margen
			record.total_apu_margen_porc = total_apu_margen_porc
			record.total_apu_margen_porc_100 = total_apu_margen_porc_100

	@api.depends('line_ids')
	def _compute_totales(self):
		for record in self:
			# Inicializar acumuladores
			total_precio = total_costo = total_margen = 0.0
			equipos_precio = equipos_costo = equipos_margen = 0.0
			mano_obra_precio = mano_obra_costo = mano_obra_margen = 0.0
			materiales_precio = materiales_costo = materiales_margen = 0.0
			transporte_precio = transporte_costo = transporte_margen = 0.0

			# Acumular valores por tipo
			for line in record.line_ids:
				if line.tipo_componente == 'equipo':
					equipos_precio += line.precio_general
					equipos_costo += line.subtotal
					equipos_margen += line.margen_bruto
				elif line.tipo_componente == 'manoobra':
					mano_obra_precio += line.precio_general
					mano_obra_costo += line.subtotal
					mano_obra_margen += line.margen_bruto
				elif line.tipo_componente == 'material':
					materiales_precio += line.precio_general
					materiales_costo += line.subtotal
					materiales_margen += line.margen_bruto
				elif line.tipo_componente == 'transporte':
					transporte_precio += line.precio_general
					transporte_costo += line.subtotal
					transporte_margen += line.margen_bruto

			# Asignar totales al registro
			record.total_equipos_precio = equipos_precio
			record.total_equipos_costo_directo = equipos_costo
			record.total_equipos_margen = equipos_margen
			record.total_equipos_margen_porc = (equipos_margen / equipos_precio) if equipos_precio else 0

			record.total_mano_obra_precio = mano_obra_precio
			record.total_mano_obra_costo_directo = mano_obra_costo
			record.total_mano_obra_margen = mano_obra_margen
			record.total_mano_obra_margen_porc = (mano_obra_margen / mano_obra_precio) if mano_obra_precio else 0

			record.total_materiales_precio = materiales_precio
			record.total_materiales_costo_directo = materiales_costo
			record.total_materiales_margen = materiales_margen
			record.total_materiales_margen_porc = (materiales_margen / materiales_precio) if materiales_precio else 0

			record.total_transporte_precio = transporte_precio
			record.total_transporte_costo_directo = transporte_costo
			record.total_transporte_margen = transporte_margen
			record.total_transporte_margen_porc = (transporte_margen / transporte_precio) if transporte_precio else 0

			# Totales generales
			record.total_precio = equipos_precio + mano_obra_precio + materiales_precio + transporte_precio
			record.total_costo_directo = equipos_costo + mano_obra_costo + materiales_costo + transporte_costo
			record.total_margen_bruto = equipos_margen + mano_obra_margen + materiales_margen + transporte_margen
			record.total_porc_margen_bruto = (
					record.total_margen_bruto / record.total_precio) if record.total_precio else 0
			record.total_porc_margen_brutoc = ((
					                                   record.total_margen_bruto / record.total_precio) if record.total_precio else 0) * 100

	@api.depends('total_general', 'total_with_tax')
	def _compute_total_apu(self):
		for record in self:
			record.total_apu = record.total_general + record.total_with_tax

	def action_recalculate_totals(self):
		"""Recalcula los totales y los totales con impuestos."""
		for record in self:
			record._compute_totals()  # Recalcular totales base
			record._compute_totals_with_tax()  # Recalcular totales con impuestos
			record._compute_total_apu()  # Recalcular total apu

	@api.depends(
		'product_tmpl_id.attribute_line_ids.value_ids',
		'product_tmpl_id.attribute_line_ids.attribute_id.create_variant',
		'product_tmpl_id.attribute_line_ids.product_template_value_ids.ptav_active',
	)
	def _compute_possible_product_template_attribute_value_ids(self):
		for bom in self:
			bom.possible_product_template_attribute_value_ids = bom.product_tmpl_id.valid_product_template_attribute_line_ids._without_no_variant_attributes().product_template_value_ids._only_active()

	@api.depends('apply_tax', 'tax_percentage', 'total_manoobra', 'total_material', 'total_equipo', 'total_transporte')
	def _compute_totals_with_tax(self):
		for record in self:
			if record.apply_tax and record.tax_percentage > 0:
				# Calcular los totales con impuesto
				record.total_manoobra_with_tax = record.total_manoobra + (
						record.total_manoobra * (record.tax_percentage / 100))
				record.total_material_with_tax = record.total_material + (
						record.total_material * (record.tax_percentage / 100))
				record.total_equipo_with_tax = record.total_equipo + (
						record.total_equipo * (record.tax_percentage / 100))
				record.total_transporte_with_tax = record.total_transporte + (
						record.total_transporte * (record.tax_percentage / 100))
			else:
				# Si no se aplica impuesto, los totales con impuesto son iguales a los originales
				record.total_manoobra_with_tax = record.total_manoobra
				record.total_material_with_tax = record.total_material
				record.total_equipo_with_tax = record.total_equipo
				record.total_transporte_with_tax = record.total_transporte

			# Calcular el total general con impuesto
			record.total_with_tax = (
					record.total_manoobra_with_tax +
					record.total_material_with_tax +
					record.total_equipo_with_tax +
					record.total_transporte_with_tax
			)

	@api.onchange('product_id')
	def _onchange_product_id(self):
		if self.product_id:
			self.line_ids.bom_product_template_attribute_value_ids = False
			self.operation_ids.bom_product_template_attribute_value_ids = False
			self.byproduct_ids.bom_product_template_attribute_value_ids = False

	@api.constrains('active', 'product_id', 'product_tmpl_id', 'line_ids')
	def _check_bom_cycle(self):
		subcomponents_dict = dict()

		def _check_cycle(components, finished_products):
			"""
			Check whether the components are part of the finished products (-> cycle). Then, if
			these components have a BoM, repeat the operation with the subcomponents (recursion).
			The method will return the list of product variants that creates the cycle
			"""
			products_to_find = self.env['product.product']

			for component in components:
				if component in finished_products:
					names = finished_products.mapped('display_name')
					raise ValidationError(_("The current configuration is incorrect because it would create a cycle "
					                        "between these products: %s.") % ', '.join(names))
				if component not in subcomponents_dict:
					products_to_find |= component

			bom_find_result = self._bom_find(products_to_find)
			for component in components:
				if component not in subcomponents_dict:
					bom = bom_find_result[component]
					subcomponents = bom.line_ids.filtered(lambda l: not l._skip_bom_line(component)).product_id
					subcomponents_dict[component] = subcomponents
				subcomponents = subcomponents_dict[component]
				if subcomponents:
					_check_cycle(subcomponents, finished_products | component)

		boms_to_check = self
		domain = []
		for product in self.line_ids.product_id:
			domain = OR([domain, self._bom_find_domain(product)])
		if domain:
			boms_to_check |= self.env['apu.apu'].search(domain)

		for bom in boms_to_check:
			if not bom.active:
				continue
			finished_products = bom.product_id or bom.product_tmpl_id.product_variant_ids
			if bom.line_ids.bom_product_template_attribute_value_ids:
				grouped_by_components = defaultdict(lambda: self.env['product.product'])
				for finished in finished_products:
					components = bom.line_ids.filtered(lambda l: not l._skip_bom_line(finished)).product_id
					grouped_by_components[components] |= finished
				for components, finished in grouped_by_components.items():
					_check_cycle(components, finished)
			else:
				_check_cycle(bom.line_ids.product_id, finished_products)

	def write(self, vals):
		res = super().write(vals)
		if 'sequence' in vals and self and self[-1].id == self._prefetch_ids[-1]:
			self.browse(self._prefetch_ids)._check_bom_cycle()
		self._update_product_total_apu_precio()
		return res

	@api.constrains('product_id', 'product_tmpl_id', 'line_ids', 'byproduct_ids', 'operation_ids')
	def _check_bom_lines(self):
		for bom in self:
			apply_variants = bom.line_ids.bom_product_template_attribute_value_ids | bom.operation_ids.bom_product_template_attribute_value_ids | bom.byproduct_ids.bom_product_template_attribute_value_ids
			if bom.product_id and apply_variants:
				raise ValidationError(
					_("You cannot use the 'Apply on Variant' functionality and simultaneously create a BoM for a specific variant."))
			for ptav in apply_variants:
				if ptav.product_tmpl_id != bom.product_tmpl_id:
					raise ValidationError(_(
						"The attribute value %(attribute)s set on product %(product)s does not match the BoM product %(bom_product)s.",
						attribute=ptav.display_name,
						product=ptav.product_tmpl_id.display_name,
						bom_product=bom.product_tmpl_id.display_name
					))
			for byproduct in bom.byproduct_ids:
				if bom.product_id:
					same_product = bom.product_id == byproduct.product_id
				else:
					same_product = bom.product_tmpl_id == byproduct.product_id.product_tmpl_id
				if same_product:
					raise ValidationError(_("By-product %s should not be the same as BoM product.") % bom.display_name)
				if byproduct.cost_share < 0:
					raise ValidationError(_("By-products cost shares must be positive."))
			if sum(bom.byproduct_ids.mapped('cost_share')) > 100:
				raise ValidationError(_("The total cost share for a BoM's by-products cannot exceed 100."))

	@api.onchange('line_ids', 'product_qty')
	def onchange_bom_structure(self):
		if self.type == 'phantom' and self._origin and self.env['stock.move'].search(
				[('bom_line_id', 'in', self._origin.line_ids.ids)], limit=1):
			return {
				'warning': {
					'title': _('Warning'),
					'message': _(
						'The product has already been used at least once, editing its structure may lead to undesirable behaviours. '
						'You should rather archive the product and create a new one with a new bill of materials.'),
				}
			}

	@api.onchange('product_uom_id')
	def onchange_product_uom_id(self):
		res = {}
		if not self.product_uom_id or not self.product_tmpl_id:
			return
		if self.product_uom_id.category_id.id != self.product_tmpl_id.uom_id.category_id.id:
			self.product_uom_id = self.product_tmpl_id.uom_id.id
			res['warning'] = {'title': _('Warning'), 'message': _(
				'The Product Unit of Measure you chose has a different category than in the product form.')}
		return res

	@api.onchange('product_tmpl_id')
	def onchange_product_tmpl_id(self):
		if self.product_tmpl_id:
			self.product_uom_id = self.product_tmpl_id.uom_id.id
			if self.product_id.product_tmpl_id != self.product_tmpl_id:
				self.product_id = False
			self.line_ids.bom_product_template_attribute_value_ids = False
			self.operation_ids.bom_product_template_attribute_value_ids = False
			self.byproduct_ids.bom_product_template_attribute_value_ids = False

			domain = [('product_tmpl_id', '=', self.product_tmpl_id.id)]
			if self.id.origin:
				domain.append(('id', '!=', self.id.origin))
			number_of_bom_of_this_product = self.env['apu.apu'].search_count(domain)
			if number_of_bom_of_this_product:  # add a reference to the bom if there is already a bom for this product
				self.code = _("%s (new) %s", self.product_tmpl_id.name, number_of_bom_of_this_product)
			else:
				self.code = False

	def copy(self, default=None):
		res = super().copy(default)
		if self.operation_ids:
			operations_mapping = {}
			for original, copied in zip(self.operation_ids, res.operation_ids.sorted()):
				operations_mapping[original] = copied
			for bom_line in res.line_ids:
				if bom_line.operation_id:
					bom_line.operation_id = operations_mapping[bom_line.operation_id]
			for operation in self.operation_ids:
				if operation.blocked_by_operation_ids:
					copied_operation = operations_mapping[operation]
					dependencies = []
					for dependency in operation.blocked_by_operation_ids:
						dependencies.append(Command.link(operations_mapping[dependency].id))
					copied_operation.blocked_by_operation_ids = dependencies

		return res

	@api.model
	def name_create(self, name):
		# prevent to use string as product_tmpl_id
		if isinstance(name, str):
			raise UserError(_("You cannot create a new Bill of Material from here."))
		return super(ApuApu, self).name_create(name)

	def toggle_active(self):
		self.with_context({'active_test': False}).operation_ids.toggle_active()
		return super().toggle_active()

	def name_get(self):
		result = []
		for bom in self:
			code_part = bom.code or ''
			tags_part = ", ".join(bom.tag_ids.mapped('name')) if bom.tag_ids else ''
			display_name = code_part
			if tags_part:
				display_name = f"{code_part} - {tags_part}"
			result.append((bom.id, display_name))
		return result

	@api.constrains('product_tmpl_id', 'product_id', 'type')
	def check_kit_has_not_orderpoint(self):
		product_ids = [pid for bom in self.filtered(lambda bom: bom.type == "phantom")
		               for pid in (bom.product_id.ids or bom.product_tmpl_id.product_variant_ids.ids)]
		if self.env['stock.warehouse.orderpoint'].search([('product_id', 'in', product_ids)], count=True):
			raise ValidationError(
				_("You can not create a kit-type bill of materials for products that have at least one reordering rule."))

	@api.ondelete(at_uninstall=False)
	def _unlink_except_running_mo(self):
		if self.env['mrp.production'].search([('bom_id', 'in', self.ids), ('state', 'not in', ['done', 'cancel'])],
		                                     limit=1):
			raise UserError(
				_('You can not delete a Bill of Material with running manufacturing orders.\nPlease close or cancel it first.'))

	@api.model
	def _bom_find_domain(self, products, picking_type=None, company_id=False, bom_type=False):
		domain = ['&', '|', ('product_id', 'in', products.ids), '&', ('product_id', '=', False),
		          ('product_tmpl_id', 'in', products.product_tmpl_id.ids), ('active', '=', True)]
		if company_id or self.env.context.get('company_id'):
			domain = AND([domain, ['|', ('company_id', '=', False),
			                       ('company_id', '=', company_id or self.env.context.get('company_id'))]])
		if picking_type:
			domain = AND([domain, ['|', ('picking_type_id', '=', picking_type.id), ('picking_type_id', '=', False)]])
		if bom_type:
			domain = AND([domain, [('type', '=', bom_type)]])
		return domain

	@api.model
	def _bom_find(self, products, picking_type=None, company_id=False, bom_type=False):
		""" Find the first BoM for each products

		:param products: `product.product` recordset
		:return: One bom (or empty recordset `apu.apu` if none find) by product (`product.product` record)
		:rtype: defaultdict(`lambda: self.env['apu.apu']`)
		"""
		bom_by_product = defaultdict(lambda: self.env['apu.apu'])
		products = products.filtered(lambda p: p.type != 'service')
		if not products:
			return bom_by_product
		domain = self._bom_find_domain(products, picking_type=picking_type, company_id=company_id, bom_type=bom_type)

		# Performance optimization, allow usage of limit and avoid the for loop `bom.product_tmpl_id.product_variant_ids`
		if len(products) == 1:
			bom = self.search(domain, order='sequence, product_id, id', limit=1)
			if bom:
				bom_by_product[products] = bom
			return bom_by_product

		boms = self.search(domain, order='sequence, product_id, id')

		products_ids = set(products.ids)
		for bom in boms:
			products_implies = bom.product_id or bom.product_tmpl_id.product_variant_ids
			for product in products_implies:
				if product.id in products_ids and product not in bom_by_product:
					bom_by_product[product] = bom

		return bom_by_product

	def explode(self, product, quantity, picking_type=False):
		"""
			Explodes the BoM and creates two lists with all the information you need: bom_done and line_done
			Quantity describes the number of times you need the BoM: so the quantity divided by the number created by the BoM
			and converted into its UoM
		"""
		from collections import defaultdict

		graph = defaultdict(list)
		V = set()

		def check_cycle(v, visited, recStack, graph):
			visited[v] = True
			recStack[v] = True
			for neighbour in graph[v]:
				if visited[neighbour] == False:
					if check_cycle(neighbour, visited, recStack, graph) == True:
						return True
				elif recStack[neighbour] == True:
					return True
			recStack[v] = False
			return False

		product_ids = set()
		product_boms = {}

		def update_product_boms():
			products = self.env['product.product'].browse(product_ids)
			product_boms.update(self._bom_find(products, picking_type=picking_type or self.picking_type_id,
			                                   company_id=self.company_id.id, bom_type='phantom'))
			# Set missing keys to default value
			for product in products:
				product_boms.setdefault(product, self.env['apu.apu'])

		boms_done = [(self, {'qty': quantity, 'product': product, 'original_qty': quantity, 'parent_line': False})]
		lines_done = []
		V |= set([product.product_tmpl_id.id])

		bom_lines = []
		for bom_line in self.line_ids:
			product_id = bom_line.product_id
			V |= set([product_id.product_tmpl_id.id])
			graph[product.product_tmpl_id.id].append(product_id.product_tmpl_id.id)
			bom_lines.append((bom_line, product, quantity, False))
			product_ids.add(product_id.id)
		update_product_boms()
		product_ids.clear()
		while bom_lines:
			current_line, current_product, current_qty, parent_line = bom_lines[0]
			bom_lines = bom_lines[1:]

			if current_line._skip_bom_line(current_product):
				continue

			line_quantity = current_qty * current_line.product_qty
			if not current_line.product_id in product_boms:
				update_product_boms()
				product_ids.clear()
			bom = product_boms.get(current_line.product_id)
			if bom:
				converted_line_quantity = current_line.product_uom_id._compute_quantity(line_quantity / bom.product_qty,
				                                                                        bom.product_uom_id)
				bom_lines += [(line, current_line.product_id, converted_line_quantity, current_line) for line in
				              bom.line_ids]
				for bom_line in bom.line_ids:
					graph[current_line.product_id.product_tmpl_id.id].append(bom_line.product_id.product_tmpl_id.id)
					if bom_line.product_id.product_tmpl_id.id in V and check_cycle(
							bom_line.product_id.product_tmpl_id.id, {key: False for key in V},
							{key: False for key in V}, graph):
						raise UserError(
							_('Recursion error!  A product with a Bill of Material should not have itself in its BoM or child BoMs!'))
					V |= set([bom_line.product_id.product_tmpl_id.id])
					if not bom_line.product_id in product_boms:
						product_ids.add(bom_line.product_id.id)
				boms_done.append((bom,
				                  {'qty': converted_line_quantity, 'product': current_product, 'original_qty': quantity,
				                   'parent_line': current_line}))
			else:
				# We round up here because the user expects that if he has to consume a little more, the whole UOM unit
				# should be consumed.
				rounding = current_line.product_uom_id.rounding
				line_quantity = float_round(line_quantity, precision_rounding=rounding, rounding_method='UP')
				lines_done.append((current_line,
				                   {'qty': line_quantity, 'product': current_product, 'original_qty': quantity,
				                    'parent_line': parent_line}))

		return boms_done, lines_done

	@api.model
	def get_import_templates(self):
		return [{
			'label': _('Import Template for Bills of Materials'),
			'template': '/mrp/static/xls/apu.apu.xls'
		}]

	def export_excel(self):
		"""
		Genera un Excel con tres hojas:
		- Rubro: Contiene la cabecera del APU con los campos:
		  product_tmpl_id, product_id, product_qty, product_uom_id, code, company_id, cliente_id, indirectos_porcentaje, utilidad_porcentaje, subcategoria.
		- Apu: Contiene las líneas del APU con los campos:
		  Rubro ID (el code del APU), codigo_componente, product_id, tipo_componente, product_qty, costo, precio_unit.
		- Actividades: Contiene las actividades asociadas a las líneas del APU con los campos:
		  APU ID (el codigo_componente de la línea), tarea_id, tipo_actividad, unidad, cantidad, jornada, tarifa, rendimiento.
		"""
		self.ensure_one()
		import io, base64, xlsxwriter
		output = io.BytesIO()
		workbook = xlsxwriter.Workbook(output, {'in_memory': True})

		# Definir formatos
		header_format = workbook.add_format({
			'bold': True,
			'bg_color': '#F7F7F7',
			'border': 1,
			'align': 'center'
		})
		numeric_format = workbook.add_format({
			'num_format': '#,##0.00',
			'border': 1,
			'align': 'right'
		})
		text_format = workbook.add_format({
			'border': 1,
			'align': 'left'
		})

		# --- Hoja Rubro ---
		worksheet_rubro = workbook.add_worksheet("Rubro")
		rubro_headers = [
			"Code", "Product Template", "Product", "Quantity", "UoM", "Company", "Cliente", "Indirectos (%)",
			"Utilidad (%)", "Subcategoria"
		]
		for col, header in enumerate(rubro_headers):
			worksheet_rubro.write(0, col, header, header_format)
		worksheet_rubro.write(1, 0, self.code or '', text_format)
		worksheet_rubro.write(1, 1, self.product_tmpl_id.default_code or '', text_format)
		worksheet_rubro.write(1, 2, self.product_id.default_code or '', text_format)
		worksheet_rubro.write_number(1, 3, float(self.product_qty), numeric_format)
		worksheet_rubro.write(1, 4, self.product_uom_id.name or '', text_format)
		worksheet_rubro.write(1, 5, self.company_id.name or '', text_format)
		worksheet_rubro.write(1, 6, self.cliente_id.name or '', text_format)
		worksheet_rubro.write_number(1, 7, float(self.indirectos_porcentaje), numeric_format)
		worksheet_rubro.write_number(1, 8, float(self.utilidad_porcentaje), numeric_format)
		worksheet_rubro.write(1, 9, self.subcategoria_id.name or '', text_format)

		# --- Hoja Apu ---
		worksheet_apu = workbook.add_worksheet("Apu")
		apu_headers = [
			"Rubro ID", "Codigo Componente", "Product",
			"Tipo Componente", "Unidad", "Cantidad", "Costo", "Precio Unit"
		]
		for col, header in enumerate(apu_headers):
			worksheet_apu.write(0, col, header, header_format)

		row = 1
		for line in self.line_ids:
			worksheet_apu.write(row, 0, self.code or '', text_format)
			worksheet_apu.write(row, 1, line.codigo_componente or '', text_format)
			worksheet_apu.write(row, 2, line.product_id.default_code or '', text_format)
			worksheet_apu.write(row, 3, line.tipo_componente or '', text_format)
			worksheet_apu.write(row, 4, line.product_uom_id.name or '', text_format)
			worksheet_apu.write_number(row, 5, float(line.product_qty), numeric_format)
			worksheet_apu.write_number(row, 6, float(line.subtotal), numeric_format)  # Costo total
			worksheet_apu.write_number(row, 7, float(line.precio_unit), numeric_format)  # Precio Unitario
			row += 1

		# --- Hoja Actividades ---
		worksheet_actividades = workbook.add_worksheet("Actividades")
		act_headers = [
			"APU ID", "Tarea", "Tipo Actividad",
			"Unidad", "Cantidad", "Costo/H", "Rendimiento"
		]
		for col, header in enumerate(act_headers):
			worksheet_actividades.write(0, col, header, header_format)

		row = 1
		for line in self.line_ids:
			for act in line.line_tarea_ids:
				worksheet_actividades.write(row, 0, line.codigo_componente or '', text_format)
				worksheet_actividades.write(row, 1, act.tarea_id.default_code or '', text_format)
				worksheet_actividades.write(row, 2, act.tipo_actividad or '', text_format)
				worksheet_actividades.write(row, 3, act.unidad.name if act.unidad else '', text_format)
				worksheet_actividades.write_number(row, 4, float(act.cantidad), numeric_format)
				worksheet_actividades.write_number(row, 5, float(act.costo), numeric_format)
				worksheet_actividades.write_number(row, 6, float(act.rendimiento), numeric_format)
				row += 1

		workbook.close()
		output.seek(0)
		file_data = output.read()

		# Crear el attachment y retornar la acción para descargar el fichero
		attachment = self.env['ir.attachment'].create({
			'name': 'APU_Template.xlsx',
			'datas': base64.b64encode(file_data),
			'type': 'binary',
			'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
		})
		url = "/web/content/%s?download=true" % (attachment.id)
		return {
			'type': 'ir.actions.act_url',
			'url': url,
			'target': 'new',
		}

	# Campos para importar archivos Excel en el mismo registro
	import_file = fields.Binary(string="Archivo Excel")
	import_filename = fields.Char(string="Nombre del Archivo")

	def import_excel_apu(self):
		"""
		Importa datos de un archivo Excel de carga masiva con tres hojas:

		- Hoja **Rubro**:
		  Encabezados en la fila 1:
			 "Code", "Product Template", "Product", "Quantity", "UoM", "Company", "Cliente", "Indirectos (%)", "Utilidad (%)", "Subcategoria"

		- Hoja **Apu**:
		  Encabezados en la fila 1:
			 "Rubro ID", "Codigo Componente", "Product", "Tipo Componente", "Unidad", "Quantity", "Costo", "Precio Unit"

		- Hoja **Actividades**:
		  Encabezados en la fila 1:
			 "APU ID", "Tarea", "Tipo Actividad", "Unidad", "Cantidad", "Jornada", "Tarifa", "Rendimiento"
		"""
		if not self.import_file:
			raise UserError(_("Por favor, cargue un archivo Excel para importar."))

		import io, base64
		file_data = base64.b64decode(self.import_file)
		try:
			workbook = load_workbook(filename=io.BytesIO(file_data), data_only=True)
		except Exception as e:
			raise UserError(_("No se pudo leer el archivo Excel. Error: %s") % e)

		# --- Hoja Rubro ---
		if "Rubro" not in workbook.sheetnames:
			raise UserError(_("La hoja 'Rubro' no se encontró en el archivo Excel."))
		ws_rubro = workbook["Rubro"]
		apu_headers = {}
		for row in range(2, ws_rubro.max_row + 1):
			code_val = ws_rubro.cell(row=row, column=1).value
			if not code_val:
				continue
			apu_headers[code_val] = {
				'code': code_val,
				'product_template_name': ws_rubro.cell(row=row, column=2).value or '',
				'product_name': ws_rubro.cell(row=row, column=3).value or '',
				'quantity': ws_rubro.cell(row=row, column=4).value or False,
				'uom_name': ws_rubro.cell(row=row, column=5).value or '',
				'company_name': ws_rubro.cell(row=row, column=6).value or '',
				'cliente_id': ws_rubro.cell(row=row, column=7).value or '',
				'indirectos_porcentaje': ws_rubro.cell(row=row, column=8).value or False,
				'utilidad_porcentaje': ws_rubro.cell(row=row, column=9).value or False,
				'subcategoria_id': ws_rubro.cell(row=row, column=10).value or '',
			}

		# --- Hoja Apu ---
		if "Apu" not in workbook.sheetnames:
			raise UserError(_("La hoja 'Apu' no se encontró en el archivo Excel."))
		ws_apu = workbook["Apu"]
		apu_lines_by_code = {}
		for row in range(2, ws_apu.max_row + 1):
			rubro_id_val = ws_apu.cell(row=row, column=1).value or ''
			codigo_componente = ws_apu.cell(row=row, column=2).value
			if not codigo_componente:
				continue
			product_line_name = ws_apu.cell(row=row, column=3).value or ''
			tipo_componente = ws_apu.cell(row=row, column=4).value or ''
			unidad_line_name = ws_apu.cell(row=row, column=5).value or ''
			product_qty = ws_apu.cell(row=row, column=6).value or False
			costo_importado = ws_apu.cell(row=row, column=7).value
			precio_unit_importado = ws_apu.cell(row=row, column=8).value

			line_data = {
				'codigo_componente': codigo_componente,
				'product_line_name': product_line_name,
				'tipo_componente': tipo_componente,
				'unidad_line_name': unidad_line_name,
				'product_qty': product_qty,
				'costo': costo_importado,
				'precio_unit': precio_unit_importado,
			}
			apu_lines_by_code.setdefault(rubro_id_val, []).append(line_data)

		# --- Hoja Actividades ---
		if "Actividades" not in workbook.sheetnames:
			raise UserError(_("La hoja 'Actividades' no se encontró en el archivo Excel."))
		ws_act = workbook["Actividades"]
		activities_by_line = {}
		for row in range(2, ws_act.max_row + 1):
			apu_id_val = ws_act.cell(row=row, column=1).value
			if not apu_id_val:
				continue
			act_data = {
				'tarea_name': ws_act.cell(row=row, column=2).value or '',
				'tipo_actividad': ws_act.cell(row=row, column=3).value or '',
				'unidad_act_name': ws_act.cell(row=row, column=4).value or '',
				'cantidad': ws_act.cell(row=row, column=5).value or False,
				'costo': ws_act.cell(row=row, column=6).value or False,
				'rendimiento': ws_act.cell(row=row, column=7).value or False,
			}
			activities_by_line.setdefault(apu_id_val, []).append(act_data)

		# --- Crear registros APU ---
		for code_key, header in apu_headers.items():
			product_template_rec = self.env['product.template'].search([
				'|', ('default_code', '=', header['product_template_name']),
				('name', '=', header['product_template_name'])
			], limit=1)
			if not product_template_rec:
				raise UserError(_("No se encontró el Product Template '%s'.") % header['product_template_name'])

			product_variant_rec = self.env['product.product'].search([
				'|', ('default_code', '=', header['product_name']),
				('name', '=', header['product_name'])
			], limit=1)
			if not product_variant_rec:
				raise UserError(_("No se encontró el Product '%s'.") % header['product_name'])

			uom_rec = self.env['uom.uom'].search([('name', '=', header['uom_name'])], limit=1)
			if not uom_rec:
				raise UserError(_("No se encontró la Unidad de Medida '%s'.") % header['uom_name'])

			company_rec = self.env['res.company'].search([('name', '=', header['company_name'])], limit=1)
			if not company_rec:
				raise UserError(_("No se encontró la Compañía '%s'.") % header['company_name'])

			cliente_rec = False
			if header['cliente_id']:
				cliente_rec = self.env['res.partner'].search([('name', '=', header['cliente_id'])], limit=1)
				if not cliente_rec:
					raise UserError(_("No se encontró el Cliente '%s'.") % header['cliente_id'])

			subcategoria_rec = False
			if header['subcategoria_id']:
				subcategoria_rec = self.env['apu.subcategoria'].search([('codigo', '=', header['subcategoria_id'])],
				                                                       limit=1)
				if not subcategoria_rec:
					raise UserError(_("No se encontró la Subcategoria '%s'.") % header['subcategoria_id'])

			vals_header = {
				'code': header['code'],
				'product_qty': header['quantity'],
				'product_tmpl_id': product_template_rec.id,
				'product_id': product_variant_rec.id,
				'product_uom_id': uom_rec.id,
				'company_id': company_rec.id,
				'cliente_id': cliente_rec.id if cliente_rec else False,
				'indirectos_porcentaje': header['indirectos_porcentaje'],
				'utilidad_porcentaje': header['utilidad_porcentaje'],
				'subcategoria_id': subcategoria_rec.id if subcategoria_rec else False,
			}

			vals_lines = []
			for line in apu_lines_by_code.get(code_key, []):
				product_line_rec = self.env['product.product'].search([
					'|', ('default_code', '=', line['product_line_name']),
					('name', '=', line['product_line_name'])
				], limit=1)
				if not product_line_rec:
					raise UserError(_("No se encontró el producto '%s' en la hoja Apu.") % line['product_line_name'])

				uom_line_rec = self.env['uom.uom'].search([('name', '=', line['unidad_line_name'])], limit=1)
				if not uom_line_rec:
					raise UserError(_("No se encontró la unidad '%s' en la hoja Apu.") % line['unidad_line_name'])

				acts = activities_by_line.get(line['codigo_componente'], [])
				costo_imp = line.get('costo')
				precio_unit_imp = line.get('precio_unit')

				if costo_imp not in (None, False, 0):
					cost_value = costo_imp
				elif acts:
					cost_value = sum(a['cantidad'] * a['costo'] for a in acts)
				else:
					cost_value = product_line_rec.standard_price

				line_vals = {
					'codigo_componente': line['codigo_componente'],
					'product_id': product_line_rec.id,
					'tipo_componente': line['tipo_componente'],
					'product_qty': line['product_qty'],
					'product_uom_id': uom_line_rec.id,
					'cost': cost_value,
					'precio_unit': precio_unit_imp or 0.0,  # ← aquí se guarda el precio unitario importado
				}

				if acts:
					tarea_vals = []
					for act in acts:
						tarea_rec = self.env['product.product'].search([
							'|', ('default_code', '=', act['tarea_name']),
							('name', '=', act['tarea_name'])
						], limit=1)
						if not tarea_rec:
							raise UserError(_("No se encontró la Tarea '%s'.") % act['tarea_name'])
						unidad_act_rec = self.env['uom.uom'].search([('name', '=', act['unidad_act_name'])], limit=1)
						if not unidad_act_rec:
							raise UserError(_("No se encontró la Unidad '%s'.") % act['unidad_act_name'])
						act_vals = {
							'tarea_id': tarea_rec.id,
							'tipo_actividad': act['tipo_actividad'],
							'unidad': unidad_act_rec.id,
							'cantidad': act['cantidad'],
							'costo': act['costo'],
							'rendimiento': act['rendimiento'],
						}
						tarea_vals.append((0, 0, act_vals))
					line_vals['line_tarea_ids'] = tarea_vals

				vals_lines.append((0, 0, line_vals))

			vals_header['line_ids'] = vals_lines
			self.create(vals_header)

		return {'type': 'ir.actions.client', 'tag': 'reload'}

	def export_excel_apu(self):
		import io, base64, xlsxwriter
		from datetime import datetime
		output = io.BytesIO()
		workbook = xlsxwriter.Workbook(output, {'in_memory': True})
		worksheet = workbook.add_worksheet("Informe Detallado Mano de Obra")

		# --- Configuración de Fuentes y Formatos ---
		common_font = {'font_name': 'Verdana', 'font_size': 10, 'valign': 'vcenter'}
		title_format = workbook.add_format(
			{**{'bold': True, 'font_size': 16, 'align': 'center', 'valign': 'vcenter'}, **common_font})
		header_format = workbook.add_format({**{'bold': True, 'bg_color': '#EAECEE', 'border': 1, 'align': 'center',
		                                        'valign': 'vcenter', 'text_wrap': True}, **common_font})
		cat_header_format = workbook.add_format({**{'bold': True, 'bg_color': '#DCE6F1', 'border': 1, 'align': 'center',
		                                            'valign': 'vcenter'}, **common_font})
		subcat_header_format = workbook.add_format(
			{**{'bold': True, 'bg_color': '#EBF5FB', 'border': 1, 'align': 'center',
			    'valign': 'vcenter'}, **common_font})
		center_text_format = workbook.add_format(
			{**{'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True},
			 **common_font})
		monetary_format = workbook.add_format({**{'num_format': '$#,##0.00', 'border': 1, 'align': 'right',
		                                          'valign': 'vcenter'}, **common_font})
		numeric_format = workbook.add_format({**{'num_format': '#,##0.00', 'border': 1, 'align': 'right',
		                                         'valign': 'vcenter'}, **common_font})
		text_format = workbook.add_format({**{'border': 1, 'align': 'left', 'valign': 'vcenter', 'text_wrap': True},
		                                   **common_font})
		date_format = workbook.add_format({**{'num_format': 'yyyy-mm-dd', 'border': 1, 'align': 'left',
		                                      'valign': 'vcenter'}, **common_font})
		apu_text_format = workbook.add_format(
			{**{'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True},
			 **common_font})
		apu_numeric_format = workbook.add_format({**{'num_format': '$#,##0.00', 'border': 1, 'align': 'right',
		                                             'valign': 'vcenter'}, **common_font})
		apu_numeric_plain = workbook.add_format({**{'num_format': '#,##0.00', 'border': 1, 'align': 'right',
		                                            'valign': 'vcenter'}, **common_font})

		# --- Ajustar ancho de columnas ---
		worksheet.set_column('A:A', 15)
		worksheet.set_column('B:B', 40)
		worksheet.set_column('C:C', 12)
		worksheet.set_column('D:D', 15)
		worksheet.set_column('E:E', 12)
		worksheet.set_column('F:F', 15)

		# --- Título ---
		worksheet.merge_range('A1:F1', "Informe Detallado Mano de Obra", title_format)

		# Se ha eliminado la sección de Datos de Cabecera

		# --- Precalcular totales por Categoría y Subcategoría ---
		# Se obtiene el conjunto de todas las líneas de los registros seleccionados
		# Y se filtra para trabajar únicamente con las líneas cuyo tipo de componente es 'manoobra'
		lines = self.mapped('line_ids').filtered(lambda l: l.tipo_componente == 'manoobra')
		cat_totals = {}
		subcat_totals = {}
		for line in lines:
			# Usamos el APU padre (bom_id) para obtener la subcategoría y la categoría
			parent_apu = line.bom_id
			if parent_apu and parent_apu.subcategoria_id:
				cat = parent_apu.subcategoria_id.categoria_id.name or "Sin Categoría"
				subcat = parent_apu.subcategoria_id.name or "Sin Subcategoría"
			else:
				cat = "Sin Categoría"
				subcat = "Sin Subcategoría"
			cat_totals[cat] = cat_totals.get(cat, 0.0) + ((line.precio_unit or 0.0) * (line.product_qty or 0.0))
			subcat_totals[(cat, subcat)] = subcat_totals.get((cat, subcat), 0.0) + (
					(line.precio_unit or 0.0) * (line.product_qty or 0.0))

		# --- Tabla Principal: Líneas de APU con Agrupación ---
		table_start = 2
		table_headers = ["ITEM", "DESCRIPCIÓN", "UNIDAD", "PRECIO", "CANTIDAD", "SUBTOTAL"]
		for col, h in enumerate(table_headers):
			worksheet.write(table_start, col, h, header_format)
		current_row = table_start + 1

		prev_cat = None
		prev_subcat = None

		for line in lines:
			parent_apu = line.bom_id
			if parent_apu and parent_apu.subcategoria_id:
				current_cat = parent_apu.subcategoria_id.categoria_id.name or "Sin Categoría"
				current_subcat = parent_apu.subcategoria_id.name or "Sin Subcategoría"
			else:
				current_cat = "Sin Categoría"
				current_subcat = "Sin Subcategoría"

			if current_cat != prev_cat:
				cat_code = parent_apu.subcategoria_id.categoria_id.codigo if (
						parent_apu and parent_apu.subcategoria_id and parent_apu.subcategoria_id.categoria_id.codigo) else ""
				worksheet.write(current_row, 0, cat_code, cat_header_format)
				worksheet.write(current_row, 1, current_cat, cat_header_format)
				worksheet.merge_range(current_row, 2, current_row, 4, "", cat_header_format)
				worksheet.write_number(current_row, 5, cat_totals.get(current_cat, 0.0), monetary_format)
				current_row += 1
				prev_cat = current_cat
				prev_subcat = None

			if current_subcat != prev_subcat:
				subcat_code = parent_apu.subcategoria_id.codigo if (
						parent_apu and parent_apu.subcategoria_id and parent_apu.subcategoria_id.codigo) else ""
				worksheet.write(current_row, 0, subcat_code, subcat_header_format)
				worksheet.write(current_row, 1, current_subcat, subcat_header_format)
				worksheet.merge_range(current_row, 2, current_row, 4, "", subcat_header_format)
				worksheet.write_number(current_row, 5, subcat_totals.get((current_cat, current_subcat), 0.0),
				                       monetary_format)
				current_row += 1
				prev_subcat = current_subcat

			# Fila principal de la línea de APU: se escribe el código del APU padre, la descripción (nombre del producto),
			# la unidad y se calcula el subtotal como precio_unit * product_qty.
			item_code = parent_apu.code if (parent_apu and parent_apu.code) else ""
			worksheet.write(current_row, 0, item_code, center_text_format)
			worksheet.write(current_row, 1, line.product_id.name or "", center_text_format)
			worksheet.write(current_row, 2, line.product_uom_id.name or "", center_text_format)
			worksheet.write_number(current_row, 3, line.precio_unit or 0.0, monetary_format)
			worksheet.write_number(current_row, 4, line.product_qty or 0.0, numeric_format)
			worksheet.write_number(current_row, 5, (line.precio_unit or 0.0) * (line.product_qty or 0.0),
			                       monetary_format)
			current_row += 1

		worksheet.write(current_row + 1, 0, "Total General", header_format)
		total_general = sum((line.precio_unit or 0.0) * (line.product_qty or 0.0) for line in lines)
		worksheet.write_number(current_row + 1, 5, total_general, monetary_format)

		workbook.close()
		output.seek(0)
		file_data = output.read()

		attachment = self.env['ir.attachment'].create({
			'name': 'SaleOrder_Export.xlsx',
			'datas': base64.b64encode(file_data),
			'type': 'binary',
			'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
		})
		url = "/web/content/%s?download=true" % (attachment.id)
		return {
			'type': 'ir.actions.act_url',
			'url': url,
			'target': 'new',
		}

	def _log_changes_in_chatter(self, changes):
		"""Centraliza la publicación de cambios en el chatter del APU."""
		for change in changes:
			if change.get("accion") == "delete":
				body = f"""
	            <div style="border:1px solid #e74c3c; border-radius:6px; padding:10px; margin:5px 0; background:#fdecea;">
	              <b>🗑️ {change['tipo']} Eliminada:</b> 
	              <span style="color:#c0392b;">{change['record']}</span>
	              <ul>{change['fields_html']}</ul>
	            </div>
	            """
			else:
				body = f"""
	            <div style="border:1px solid #ccc; border-radius:6px; padding:10px; margin:5px 0; background:#f9f9f9;">
	              <b>🔄 Modificación en {change['tipo']}:</b> 
	              <span style="color:#2E86C1;">{change['record']}</span>
	              <ul>{change['details']}</ul>
	            </div>
	            """
			self.message_post(body=body, subtype_xmlid="mail.mt_note")


class ApuApuLine(models.Model):
	_name = 'apu.apu.line'
	_description = 'Bill of Material Line'
	_parent_name = 'parent_id'
	_parent_store = True
	_parent_order = 'sequence'
	_order = 'parent_path, sequence, id'
	_check_company_auto = True

	def _get_default_product_uom_id(self):
		return self.env['uom.uom'].search([], limit=1, order='id').id

	codigo_componente = fields.Char('Codigo')
	product_id = fields.Many2one('product.product', 'Component', required=True, check_company=True)

	def name_get(self):
		result = []
		for record in self:
			codigo = record.codigo_componente or ''
			producto = record.product_id.name or ''
			name = f"{codigo} - {producto}"
			result.append((record.id, name))
		return result

	product_tmpl_id = fields.Many2one('product.template', 'Product Template', related='product_id.product_tmpl_id',
	                                  store=True, index=True)
	company_id = fields.Many2one(
		related='bom_id.company_id', store=True, index=True, readonly=True)
	product_qty = fields.Float(
		'Quantity', default=1.0,
		digits=(16, 4), required=True)
	product_uom_id = fields.Many2one(
		'uom.uom', 'Product Unit of Measure',
		default=_get_default_product_uom_id,
		required=True,
		help="Unit of Measure (Unit of Measure) is the unit of measurement for the inventory control",
		domain="[('category_id', '=', product_uom_category_id)]")
	product_uom_category_id = fields.Many2one(related='product_id.uom_id.category_id')
	sequence = fields.Integer(
		'Sequence', default=1,
		help="Gives the sequence order when displaying.")
	bom_id = fields.Many2one(
		'apu.apu', 'Parent BoM',
		index=True, ondelete='cascade', required=True)
	parent_product_tmpl_id = fields.Many2one('product.template', 'Parent Product Template',
	                                         related='bom_id.product_tmpl_id')
	possible_bom_product_template_attribute_value_ids = fields.Many2many(
		related='bom_id.possible_product_template_attribute_value_ids')
	bom_product_template_attribute_value_ids = fields.Many2many(
		'product.template.attribute.value', string="Apply on Variants", ondelete='restrict',
		domain="[('id', 'in', possible_bom_product_template_attribute_value_ids)]",
		help="BOM Product Variants needed to apply this line.")
	allowed_operation_ids = fields.One2many('mrp.routing.workcenter', related='bom_id.operation_ids')
	operation_id = fields.Many2one(
		'mrp.routing.workcenter', 'Consumed in Operation', check_company=True,
		domain="[('id', 'in', allowed_operation_ids)]",
		help="The operation where the components are consumed, or the finished products created.")
	child_bom_id = fields.Many2one(
		'apu.apu', 'Sub BoM', compute='_compute_child_bom_id')
	child_line_ids = fields.One2many(
		'apu.apu.line', string="BOM lines of the referred bom",
		compute='_compute_child_line_ids')
	attachments_count = fields.Integer('Attachments Count', compute='_compute_attachments_count')
	tracking = fields.Selection(related='product_id.tracking')
	manual_consumption = fields.Boolean(
		'Manual Consumption', default=False, compute='_compute_manual_consumption',
		readonly=False, store=True, copy=True,
		help="When activated, then the registration of consumption for that component is recorded manually exclusively.\n"
		     "If not activated, and any of the components consumption is edited manually on the manufacturing order, Odoo assumes manual consumption also.")
	manual_consumption_readonly = fields.Boolean(
		'Manual Consumption Readonly', compute='_compute_manual_consumption_readonly')
	tipo_componente = fields.Selection([
		('manoobra', 'Mano de Obra'),
		('material', 'Materiales'),
		('equipo', 'Equipos'),
		('transporte', 'Transporte')], 'Tipo Componente',
		default='material', required=True)
	# cost = fields.Float("Costo Unit", compute="_calc_costo", digits=dp.get_precision('Product Unit of Measure'))
	cost = fields.Float("Costo Unit", digits=(16, 4))

	cost_compute = fields.Float("Costo Unit C.", compute="_compute_costo_unit", readonly=False,
	                            digits=(16, 4))

	subtotal = fields.Float("Costo Tot", compute="_calc_subtotal", digits=(16, 4))
	# Nuevo campo precio_unit calculado como la suma de los costo_final_actividad de las actividades asociadas
	precio_unit = fields.Float("Precio Unit", readonly=False, store=True, digits=(16, 4))
	precio_general = fields.Float("Precio Gen", compute="_compute_precio_general", store=True, digits=(16, 4))
	margen_bruto = fields.Float("Margen Bruto", compute="_compute_margen_bruto", store=True)
	porc_margen_bruto = fields.Float("%Margen Bruto", compute="_compute_margen_bruto", store=True)
	semaforo_margen = fields.Selection(
		[('verde', 'Verde'), ('naranja', 'Naranja'), ('rojo', 'Rojo')],
		string="Semáforo Margen", compute="_compute_semaforo_margen", store=True
	)
	parent_id = fields.Many2one('apu.apu.line', string='Línea Padre')
	child_ids = fields.One2many('apu.apu.line', 'parent_id', string='Subcomponentes')
	parent_path = fields.Char(index=True)

	_sql_constraints = [
		('bom_qty_zero', 'CHECK (product_qty>=0)',
		 'All product quantities must be greater or equal to 0.\nLines with 0 quantities can be used as optional lines. \nYou should install the mrp_byproduct module if you want to manage extra products on BoMs !'),
		('unique_codigo_componente', 'unique(bom_id, codigo_componente)',
		 'El código del componente no puede repetirse dentro del mismo APU.')
	]

	# Nuevo campo para relacionar las tareas
	line_tarea_ids = fields.One2many(
		'apu.apu.line.tareas',  # Modelo relacionado
		'apu_line_id',  # Campo Many2one en el modelo relacionado
		string='Tareas', copy=True
	)

	@api.depends('line_tarea_ids.costo_final_actividad', 'product_id.standard_price')
	def _compute_costo_unit(self):
		"""Calcula el costo unitario:
		- Si existen tareas relacionadas, se suma el 'costo_final_actividad' de cada una.
		- Si no existen tareas, se asigna el standard_price del producto.
		"""
		for line in self:
			if line.line_tarea_ids:
				line.cost_compute = sum(line.line_tarea_ids.mapped('costo_final_actividad'))
			else:
				line.cost_compute = line.product_id.standard_price

	@api.onchange('cost_compute')
	def _onchange_cost_compute(self):
		self.cost = self.cost_compute

	@api.depends('precio_general', 'subtotal')
	def _compute_margen_bruto(self):
		for line in self:
			# Cálculo del margen bruto en $
			line.margen_bruto = line.precio_general - line.subtotal
			# Cálculo del margen bruto en %
			line.porc_margen_bruto = (line.margen_bruto / line.precio_general) * 100 if line.precio_general else 0.0

	@api.depends('porc_margen_bruto')
	def _compute_semaforo_margen(self):
		for line in self:
			margin_type = 'margen-apu'  # Tipo de margen a evaluar
			color = self.env['bim.rangos.margen'].evaluate_margin_in_range(line.porc_margen_bruto, margin_type)

			# Mapear valores de color a opciones válidas
			mapping = {
				'success': 'verde',
				'warning': 'naranja',
				'danger': 'rojo'
			}
			line.semaforo_margen = mapping.get(color, 'rojo')  # Valor predeterminado es 'rojo'

	@api.depends('product_qty', 'precio_unit')
	def _compute_precio_general(self):
		"""
		Calcula el precio general basado en la cantidad y el precio unitario.
		"""
		for line in self:
			line.precio_general = line.product_qty * line.precio_unit

	@api.depends('product_qty', 'product_id')
	def _calc_costo(self):
		for move in self:
			move.cost = self.env['product.product'].browse(move.product_id.id).standard_price

	@api.depends('product_qty', 'product_id', 'cost')
	def _calc_subtotal(self):
		for move in self:
			move.subtotal = move.product_qty * move.cost

	@api.depends('product_id', 'tracking', 'operation_id')
	def _compute_manual_consumption(self):
		for line in self:
			line.manual_consumption = (line.tracking != 'none' or line.operation_id)

	@api.depends('tracking', 'operation_id')
	def _compute_manual_consumption_readonly(self):
		for line in self:
			line.manual_consumption_readonly = (line.tracking != 'none' or line.operation_id)

	@api.depends('product_id', 'bom_id')
	def _compute_child_bom_id(self):
		products = self.product_id
		bom_by_product = self.env['apu.apu']._bom_find(products)
		for line in self:
			if not line.product_id:
				line.child_bom_id = False
			else:
				line.child_bom_id = bom_by_product.get(line.product_id, False)

	@api.depends('product_id')
	def _compute_attachments_count(self):
		for line in self:
			nbr_attach = self.env['mrp.document'].search_count([
				'|',
				'&', ('res_model', '=', 'product.product'), ('res_id', '=', line.product_id.id),
				'&', ('res_model', '=', 'product.template'), ('res_id', '=', line.product_id.product_tmpl_id.id)])
			line.attachments_count = nbr_attach

	@api.depends('child_bom_id')
	def _compute_child_line_ids(self):
		""" If the BOM line refers to a BOM, return the ids of the child BOM lines """
		for line in self:
			line.child_line_ids = line.child_bom_id.line_ids.ids or False

	@api.onchange('product_uom_id')
	def onchange_product_uom_id(self):
		res = {}
		if not self.product_uom_id or not self.product_id:
			return res
		if self.product_uom_id.category_id != self.product_id.uom_id.category_id:
			self.product_uom_id = self.product_id.uom_id.id
			res['warning'] = {'title': _('Warning'), 'message': _(
				'The Product Unit of Measure you chose has a different category than in the product form.')}
		return res

	@api.onchange('product_id')
	def onchange_product_id(self):
		if self.product_id:
			self.product_uom_id = self.product_id.uom_id.id
			self.cost = self.env['product.product'].browse(self.product_id.id).standard_price

	@api.model_create_multi
	def create(self, vals_list):
		for values in vals_list:
			if 'product_id' in values and 'product_uom_id' not in values:
				values['product_uom_id'] = self.env['product.product'].browse(values['product_id']).uom_id.id
		record = super(ApuApuLine, self).create(vals_list)
		record._compute_precio_general()
		record._compute_margen_bruto()

	def _skip_bom_line(self, product):
		""" Control if a BoM line should be produced, can be inherited to add
		custom control.
		"""
		self.ensure_one()
		if product._name == 'product.template':
			return False
		return not product._match_all_variant_values(self.bom_product_template_attribute_value_ids)

	def action_see_attachments(self):
		domain = [
			'|',
			'&', ('res_model', '=', 'product.product'), ('res_id', '=', self.product_id.id),
			'&', ('res_model', '=', 'product.template'), ('res_id', '=', self.product_id.product_tmpl_id.id)]
		attachment_view = self.env.ref('mrp.view_document_file_kanban_mrp')
		return {
			'name': _('Attachments'),
			'domain': domain,
			'res_model': 'mrp.document',
			'type': 'ir.actions.act_window',
			'view_id': attachment_view.id,
			'views': [(attachment_view.id, 'kanban'), (False, 'form')],
			'view_mode': 'kanban,tree,form',
			'help': _('''<p class="o_view_nocontent_smiling_face">
                        Upload files to your product
                    </p><p>
                        Use this feature to store any files, like drawings or specifications.
                    </p>'''),
			'limit': 80,
			'context': "{'default_res_model': '%s','default_res_id': %d, 'default_company_id': %s}" % (
				'product.product', self.product_id.id, self.company_id.id)
		}

	@api.model
	def create(self, vals):
		record = super(ApuApuLine, self).create(vals)
		record._propagar_cambio_a_product_template_apu()
		return record

	def write(self, vals):
		# 1) Snapshot ANTES (solo campos realmente escritos)
		tracked = [f for f in vals.keys() if f in self._fields]
		old_values = {
			rec.id: {f: getattr(rec, f) for f in tracked}
			for rec in self  # no uses self.exists() aquí, aún no se ha escrito nada
		}
		# Mapa de APU padre por línea (lo usaremos después para agrupar mensajes)
		parent_bom_by_line = {rec.id: (rec.bom_id.id if rec.bom_id else False) for rec in self}

		# 2) Escribir
		res = super(ApuApuLine, self).write(vals)

		# 3) Snapshot DESPUÉS y diff seguro (solo sobre registros existentes)
		grouped_by_bom = {}
		for rec in self.exists():
			if rec.id not in old_values:
				continue
			new_values = {f: getattr(rec, f) for f in old_values[rec.id].keys()}

			line_changes = []
			for field, old in old_values[rec.id].items():
				new = new_values.get(field)
				if old != new:
					field_str = rec._fields[field].string
					line_changes.append(
						f"<li><b>{field_str}</b>: "
						f"<span style='color:red;'>{_format_value_for_tracking(old)}</span> → "
						f"<span style='color:green;'>{_format_value_for_tracking(new)}</span></li>"
					)

			if line_changes:
				bom_id = parent_bom_by_line.get(rec.id)
				if bom_id:
					grouped_by_bom.setdefault(bom_id, []).append({
						"record": rec.display_name,
						"details": "".join(line_changes),
						"tipo": "Línea",
						"accion": "update",
					})

		# 4) Publicar un mensaje por APU
		for bom_id, changes in grouped_by_bom.items():
			bom = self.env['apu.apu'].browse(bom_id)
			if bom.exists():
				bom._log_changes_in_chatter(changes)

		# 5) Mantener tu propagación a logs custom si cambian cantidad/costo/precio
		if any(f in vals for f in ('product_qty', 'cost', 'precio_unit')):
			self._propagar_cambio_a_product_template_apu()

		return res

	def unlink(self):
		# 1) Preparar cambios por APU (NO usar self[0] luego del unlink)
		changes_by_bom = {}
		for line in self:
			bom = line.bom_id
			if not bom:
				continue
			fields_html = "".join(
				f"<li><b>{k}</b>: {v}</li>"
				for k, v in {
					"Producto": _format_value_for_tracking(line.product_id),
					"Cantidad": line.product_qty,
					"Costo": line.cost,
					"Precio Unitario": line.precio_unit,
					"Unidad": _format_value_for_tracking(line.product_uom_id),
				}.items()
			)
			changes_by_bom.setdefault(bom.id, []).append({
				"tipo": "Línea",
				"accion": "delete",
				"record": line.display_name,
				"fields_html": fields_html,
			})

		# 2) Limpiar tus logs custom relacionados
		self.env['product.template.apu'].search([('apu_line_id', 'in', self.ids)]).unlink()

		# 3) Borrar líneas
		res = super(ApuApuLine, self).unlink()

		# 4) Publicar en cada APU padre (solo IDs, nada del record borrado)
		for bom_id, payload in changes_by_bom.items():
			bom = self.env['apu.apu'].browse(bom_id)
			if bom.exists() and payload:
				bom._log_changes_in_chatter(payload)

		return res

	def _propagar_cambio_a_product_template_apu(self):
		PTA = self.env['product.template.apu']
		for line in self:
			datos = {
				'fecha_modificacion': fields.Datetime.now(),
				'cantidad': line.product_qty,
				'precio_unitario': line.precio_unit,
				'costo_unitario': line.cost,
				'precio_total': line.precio_general,
				'costo_total': line.subtotal,
			}
			# Siempre crear un nuevo registro en lugar de actualizar
			PTA.create({
				'product_tmpl_id': line.product_tmpl_id.id,
				'apu_line_id': line.id,
				**datos,
			})


class ApuApuLineTareas(models.Model):
	""" Define las tareas correspondientes a las APUS """
	_name = 'apu.apu.line.tareas'
	_description = 'Detalle de Apus con Tareas'
	_inherit = ['portal.mixin', 'mail.thread', 'mail.activity.mixin']

	# Nuevo campo para conectar con ApuApuLine
	apu_line_id = fields.Many2one('apu.apu.line', string='Apu Line', ondelete='cascade')

	# Campo relacionado para que el dominio de tarea_id pueda utilizar company_id
	company_id = fields.Many2one(related='apu_line_id.company_id', store=True, index=True, readonly=True)

	sequence = fields.Integer("Sequence")
	tarea_id = fields.Many2one('product.product', 'Actividad', required=True, check_company=True)

	def name_get(self):
		result = []
		for record in self:
			codigo = record.apu_line_id.codigo_componente or ''
			producto = record.tarea_id.name or ''
			name = f"Actividad: {codigo} - {producto}"
			result.append((record.id, name))
		return result

	tipo_actividad = fields.Selection([
		('manoobra', 'Mano de Obra'),
		('material', 'Materiales'),
		('equipo', 'Equipos'),
		('transporte', 'Transporte')], 'Tipo Actividad',
		default='material', required=True)
	unidad = fields.Many2one('uom.uom', 'Unidad')
	cantidad = fields.Float('Cantidad', default=1.0, digits=(16, 4), required=True)
	costo = fields.Float('Costo/H', default=0.0, digits=(16, 4), required=True)
	costo_hora_actividad = fields.Float('Costo Hora', compute='_compute_costos', digits=(16, 4))
	rendimiento = fields.Float('Hora unidad', default=1.0, digits=(16, 4), required=True)
	costo_final_actividad = fields.Float('Costo Final', compute='_compute_costos', digits=(16, 4))

	@api.depends('cantidad', 'costo', 'rendimiento')
	def _compute_costos(self):
		for rec in self:
			rec.costo_hora_actividad = rec.cantidad * rec.costo
			rec.costo_final_actividad = rec.costo_hora_actividad * rec.rendimiento

	@api.model_create_multi
	def create(self, vals_list):
		records = super().create(vals_list)
		records._compute_costos()
		records._propagar_cambio_a_product_template_actividad()
		return records

	def write(self, vals):
		# 1) Snapshot ANTES
		tracked = [f for f in vals.keys() if f in self._fields]
		old_values = {
			rec.id: {f: getattr(rec, f) for f in tracked}
			for rec in self
		}
		parent_bom_by_task = {
			rec.id: (rec.apu_line_id.bom_id.id if rec.apu_line_id and rec.apu_line_id.bom_id else False)
			for rec in self
		}

		# 2) Escribir
		res = super(ApuApuLineTareas, self).write(vals)

		# 3) Recalcular costos y propagar a tu modelo de trazas si aplica
		if any(f in vals for f in ('cantidad', 'costo', 'rendimiento')):
			self._compute_costos()
			self._propagar_cambio_a_product_template_actividad()

		# 4) Snapshot DESPUÉS + diff seguro
		grouped_by_bom = {}
		for rec in self.exists():
			if rec.id not in old_values:
				continue
			new_values = {f: getattr(rec, f) for f in old_values[rec.id].keys()}

			tarea_changes = []
			for field, old in old_values[rec.id].items():
				new = new_values.get(field)
				if old != new:
					field_str = rec._fields[field].string
					tarea_changes.append(
						f"<li><b>{field_str}</b>: "
						f"<span style='color:red;'>{_format_value_for_tracking(old)}</span> → "
						f"<span style='color:green;'>{_format_value_for_tracking(new)}</span></li>"
					)

			if tarea_changes:
				bom_id = parent_bom_by_task.get(rec.id)
				if bom_id:
					grouped_by_bom.setdefault(bom_id, []).append({
						"record": rec.display_name,
						"details": "".join(tarea_changes),
						"tipo": "Tarea",
						"accion": "update",
					})

		# 5) Publicar un mensaje por APU
		for bom_id, changes in grouped_by_bom.items():
			bom = self.env['apu.apu'].browse(bom_id)
			if bom.exists():
				bom._log_changes_in_chatter(changes)

		return res

	def unlink(self):
		# 1) Armar payload por APU padre
		changes_by_bom = {}
		for tarea in self:
			bom = tarea.apu_line_id.bom_id if tarea.apu_line_id else False
			if not bom:
				continue

			fields_html = "".join(
				f"<li><b>{k}</b>: {v}</li>"
				for k, v in {
					"Actividad": _format_value_for_tracking(tarea.tarea_id),
					"Tipo": tarea.tipo_actividad,
					"Cantidad": tarea.cantidad,
					"Costo": tarea.costo,
					"Rendimiento": tarea.rendimiento,
					"Unidad": _format_value_for_tracking(tarea.unidad),
				}.items()
			)
			changes_by_bom.setdefault(bom.id, []).append({
				"tipo": "Tarea",
				"accion": "delete",
				"record": tarea.display_name,
				"fields_html": fields_html,
			})

			# 1.1) (Opcional) Borrar followers del registro hijo antes del unlink real
			self.env['mail.followers'].sudo().search([
				('res_model', '=', self._name),
				('res_id', '=', tarea.id),
			]).unlink()

		# 2) Borrar tus logs custom ligados a la tarea
		self.env['product.template.actividad'].search([('apu_line_tareas_id', 'in', self.ids)]).unlink()

		# 3) Borrar tareas
		res = super(ApuApuLineTareas, self).unlink()

		# 4) Publicar en cada APU (no usar self[0] aquí)
		for bom_id, payload in changes_by_bom.items():
			bom = self.env['apu.apu'].browse(bom_id)
			if bom.exists() and payload:
				bom._log_changes_in_chatter(payload)

		return res

	def _propagar_cambio_a_product_template_actividad(self):
		PTA = self.env['product.template.actividad']
		for tarea in self:
			tarea._compute_costos()
			datos = {
				'fecha_modificacion': fields.Datetime.now(),
				'cantidad': tarea.cantidad,
				'precio_unitario': tarea.costo,
				'costo_hora_actividad': tarea.costo_hora_actividad,
				'rendimiento': tarea.rendimiento,
				'costo_final_actividad': tarea.costo_final_actividad,
			}
			# Crear un nuevo log para esta tarea en su producto correspondiente
			PTA.create({
				'product_tmpl_id': tarea.apu_line_id.product_tmpl_id.id,
				'apu_line_tareas_id': tarea.id,
				**datos,
			})


class MrpByProduct(models.Model):
	_name = 'apu.apu.byproduct'
	_description = 'Byproduct'
	_rec_name = "product_id"
	_check_company_auto = True
	_order = 'sequence, id'

	product_id = fields.Many2one('product.product', 'By-product', required=True, check_company=True)
	company_id = fields.Many2one(related='bom_id.company_id', store=True, index=True, readonly=True)
	product_qty = fields.Float(
		'Quantity',
		default=1.0, digits=(16, 4), required=True)
	product_uom_category_id = fields.Many2one(related='product_id.uom_id.category_id')
	product_uom_id = fields.Many2one('uom.uom', 'Unit of Measure', required=True,
	                                 compute="_compute_product_uom_id", store=True, readonly=False, precompute=True,
	                                 domain="[('category_id', '=', product_uom_category_id)]")
	bom_id = fields.Many2one('apu.apu', 'BoM', ondelete='cascade', index=True)
	allowed_operation_ids = fields.One2many('mrp.routing.workcenter', related='bom_id.operation_ids')
	operation_id = fields.Many2one(
		'mrp.routing.workcenter', 'Produced in Operation', check_company=True,
		domain="[('id', 'in', allowed_operation_ids)]")
	possible_bom_product_template_attribute_value_ids = fields.Many2many(
		related='bom_id.possible_product_template_attribute_value_ids')
	bom_product_template_attribute_value_ids = fields.Many2many(
		'product.template.attribute.value', string="Apply on Variants", ondelete='restrict',
		domain="[('id', 'in', possible_bom_product_template_attribute_value_ids)]",
		help="BOM Product Variants needed to apply this line.")
	sequence = fields.Integer("Sequence")
	cost_share = fields.Float(
		"Cost Share (%)", digits=(16, 4),  # decimal = 2 is important for rounding calculations!!
		help="The percentage of the final production cost for this by-product line (divided between the quantity produced)."
		     "The total of all by-products' cost share must be less than or equal to 100.")

	@api.depends('product_id')
	def _compute_product_uom_id(self):
		""" Changes UoM if product_id changes. """
		for record in self:
			record.product_uom_id = record.product_id.uom_id.id

	def _skip_byproduct_line(self, product):
		""" Control if a byproduct line should be produced, can be inherited to add
		custom control.
		"""
		self.ensure_one()
		if product._name == 'product.template':
			return False
		return not product._match_all_variant_values(self.bom_product_template_attribute_value_ids)


class SaleOrder(models.Model):
	_inherit = 'sale.order'

	state = fields.Selection(selection_add=[('presupuesto_revisado', 'Presupuesto revisado'),
	                                        ('generar_requisiciones', 'Requisiciones Generadas'),
	                                        ('generar_proyecto', 'Proyecto Generado'), ])
	project_name = fields.Char(string="Nombre del Proyecto", compute="_compute_project_name", store=True, copy=False,
	                           help="Nombre del proyecto asociado a este pedido.")
	show_generate_project = fields.Boolean(string="Mostrar botón de generación de proyecto",
	                                       compute="_compute_show_generate_project")
	proyecto = fields.Char(string="Proyecto", copy=False, )
	no_proyecto = fields.Char(string="No. Proyecto", copy=False, readonly=True)
	plantilla_tarea_id = fields.Many2one('plantilla.tareas', 'Plantilla de tareas', index=True)

	# Relacionar las requisiciones de compra con el pedido de venta
	material_purchase_requisition_ids = fields.One2many('material.purchase.requisition', 'sale_order_id',
	                                                    string="Material Purchase Requisitions", copy=False)
	purchase_requisition_ids = fields.One2many('purchase.request', 'sale_order_id',
	                                           string="Material Purchase Requisitions", copy=False)
	purchase_request_ids = fields.One2many('purchase.request', 'sale_order_id', string="Purchase Requests", copy=False)
	macro_purchase_request_ids = fields.One2many('macro.purchase.request', 'sale_order_id', string="Purchase Requests",
	                                             copy=False)
	project_id = fields.Many2one('project.project', 'Project', copy=False, index=True)
	plantilla_proyecto_xml = fields.Binary(string="Plantilla XML de Proyecto", copy=False)
	plantilla_proyecto_xml_filename = fields.Char(string="Nombre del archivo XML", copy=False, )
	tipo_pedido = fields.Selection([('apu', 'APU'), ('normal', 'Normal')], string="Tipo Pedido", default='normal',
	                               copy=False, )
	es_alcance = fields.Boolean('Es alcance?')
	motivo_alcance = fields.Text('Motivo alcance')
	macro_unificado = fields.Boolean('Macro Unificado?')
	pmo_asignado = fields.Many2one('res.users', 'PMO Asignado')
	tipo_presupuesto = fields.Selection([('civil', 'Civil'), ('mecanico', 'Mecanico')], string="Tipo Psp",
	                                    default='civil', copy=False)
	superintendente = fields.Many2one('res.users', 'Superintendente')
	supervisor = fields.Many2one('res.users', 'Supervisor')

	@api.model
	def fields_get(self, allfields=None, attributes=None):
		res = super(SaleOrder, self).fields_get(allfields, attributes)
		if 'state' in res and 'selection' in res['state']:
			# Reorganizar los estados en el orden deseado
			res['state']['selection'] = [
				('draft', 'Presupuesto'),
				('presupuesto_revisado', 'Presupuesto Revisado'),
				('sent', 'Presupuesto Enviado'),
				('sale', 'Pedido de Venta'),
				('generar_requisiciones', 'Requisiciones'),
				('generar_proyecto', 'Proyecto'),
				('done', 'Bloqueada'),
				('cancel', 'Cancelado'),
			]
		return res

	def action_set_budget_reviewed(self):
		"""Pasar a Presupuesto revisado desde Borrador."""
		for order in self:
			if order.tipo_pedido == 'apu':
				# Solo permitir si el usuario está en el grupo gerente
				if not self.env.user.has_group('construccion_gps.group_apus_usuario'):
					raise UserError(_('No tienes permisos para enviar a pedido revisado.'))
			if order.tipo_pedido != 'apu':
				raise UserError(_('Esta acción aplica solo a pedidos de tipo APU.'))
			if order.state != 'draft':
				raise UserError(_('Solo puedes marcar como revisado un presupuesto en borrador.'))
			order.state = 'presupuesto_revisado'
		return True

	def action_send_informe_detallado(self):
		self.ensure_one()

		template = self.env.ref('construccion_gps.email_template_informe_detallado', raise_if_not_found=False)

		# Generar el informe detallado
		pdf_content, _ = self.env['ir.actions.report']._render_qweb_pdf(
			'construccion_gps.informe_detallado', [self.id]
		)

		attachment = self.env['ir.attachment'].create({
			'name': f'{self.name}_informe_detallado.pdf',
			'type': 'binary',
			'datas': base64.b64encode(pdf_content),
			'res_model': 'sale.order',
			'res_id': self.id,
			'mimetype': 'application/pdf',
		})

		lang = self.env.context.get('lang') or self.partner_id.lang

		ctx = {
			'default_model': 'sale.order',
			'default_res_id': self.id,
			'default_use_template': True,
			'default_template_id': template.id if template else None,
			'default_composition_mode': 'comment',
			'default_attachment_ids': [(6, 0, [attachment.id])],
			'mark_so_as_sent': True,
			'force_email': True,
			'model_description': self.with_context(lang=lang).type_name,
		}

		return {
			'type': 'ir.actions.act_window',
			'view_mode': 'form',
			'res_model': 'mail.compose.message',
			'views': [(False, 'form')],
			'target': 'new',
			'context': ctx,
		}

	@api.depends('proyecto', 'no_proyecto', 'referencia_analitica')
	def _compute_project_name(self):
		for record in self:
			if record.referencia_analitica:
				partes = record.referencia_analitica.split('-')
				if len(partes) >= 5:
					segundo_segmento = partes[1]  # Ej: '7672025'
					penultimo_segmento = partes[-2]  # Ej: '001'
					num_parte = segundo_segmento[:3]  # Ej: '767'
					ultimo_digito = penultimo_segmento[-1]  # Ej: '1'
					record.no_proyecto = f"{num_parte}.{ultimo_digito}"
				else:
					record.no_proyecto = record.proyecto
					record.project_name = record.proyecto
			else:
				record.no_proyecto = record.proyecto
				record.project_name = record.proyecto

	def get_apu_lines_or_product(self, product_template_id, product_uom_qty=1.0):
		ApuApu = self.env['apu.apu']
		apu_record = ApuApu.search([('product_tmpl_id', '=', product_template_id)], limit=1)
		if apu_record:
			# Devuelve el recordset de apu.apu.line asociado al APU encontrado
			return apu_record.line_ids
		# Si no existe un APU para ese template, retorna el product.template (o un recordset vacío)
		return self.env['product.template'].browse(product_template_id)

	@api.depends('state')
	def _compute_show_generate_project(self):
		""" Muestra el botón solo cuando el pedido de venta ya no es un presupuesto """
		for order in self:
			order.show_generate_project = order.state not in ('draft', 'sent', 'presupuesto_revisado'
			                                                                   'generar_requisiciones')  # Oculta en borrador y presupuesto enviado

	def action_generate_project(self):
		self.ensure_one()
		project = self.env['project.project'].create({
			'name': f"Proyecto - {self.name}",
			'partner_id': self.partner_id.id,
			'allow_timesheets': True,
			'allow_analytic_account': True,
			'analytic_account_id': self.analytic_account_id.id,
			'company_id': self.company_id.id
		})
		# 1) Tomar el proyecto ya asociado al sale.order
		if not project:
			raise UserError(_("No hay ningún proyecto asociado. Primero debes Generar Requisiciones."))
		self.project_id = project.id
		analytic_account = self.analytic_account_id
		if not analytic_account:
			raise UserError(_("El proyecto %s no tiene cuenta analítica asociada.") % project.name)

		if self.plantilla_proyecto_xml:
			import xml.etree.ElementTree as ET
			import base64
			from io import BytesIO
			import pytz
			from datetime import datetime, time

			def _parse_iso_date(value):
				try:
					return datetime.strptime(value[:10], '%Y-%m-%d').date()
				except Exception:
					return False

			def _convert_local_date_to_utc_str(date_obj, is_start):
				if not date_obj:
					return False
				tz_local = pytz.timezone('America/Guayaquil')
				hour = time(3, 0, 0) if is_start else time(11, 30, 0)
				dt_local = datetime.combine(date_obj, hour)
				dt_utc = tz_local.localize(dt_local).astimezone(pytz.utc)
				return dt_utc.strftime('%Y-%m-%d %H:%M:%S')

			try:
				tree = ET.parse(BytesIO(base64.b64decode(self.plantilla_proyecto_xml)))
				root = tree.getroot()
				ns = {'ns': 'http://schemas.microsoft.com/project'}
			except Exception as e:
				raise UserError(f"Error leyendo XML de plantilla: {e}")

			task_data = {}
			for task in root.findall("ns:Tasks/ns:Task", ns):
				uid = task.findtext("ns:UID", "", ns).strip()
				name = task.findtext("ns:Name", "", ns).strip()
				start = task.findtext("ns:Start", "", ns)
				finish = task.findtext("ns:Finish", "", ns)
				duration = task.findtext("ns:Duration", "PT0H0M0S", ns)
				outline = task.findtext("ns:OutlineNumber", "", ns).strip()
				outline_level = int(task.findtext("ns:OutlineLevel", "0", ns))
				wbs = task.findtext("ns:WBS", "", ns).strip()

				preds = []
				for pl in task.findall("ns:PredecessorLink", ns):
					pu = pl.findtext("ns:PredecessorUID", namespaces=ns)
					if pu and pu != "0":
						preds.append(pu)

				if not (uid and uid != "0" and name and start and finish and outline):
					continue

				start_date = _parse_iso_date(start)
				finish_date = _parse_iso_date(finish)
				try:
					part = duration[2:]
					h = m = s = 0
					if 'H' in part:
						h, part = part.split('H', 1);
						h = int(h)
					if 'M' in part:
						m, part = part.split('M', 1);
						m = int(m)
					if 'S' in part:
						s = int(part.replace('S', ''))
					duration_days = round((h + m / 60 + s / 3600) / 8, 2)
				except:
					duration_days = 0.0

				task_data[outline] = {
					'uid': uid,
					'name': name,
					'start_date': start_date,
					'finish_date': finish_date,
					'duration': duration_days,
					'outline': outline,
					'level': outline_level,
					'wbs': wbs,
					'record': None,
					'predecessors': preds,
				}

			def _key_outline(t):
				return list(map(int, t['outline'].split('.')))

			sorted_tasks = sorted(task_data.values(), key=_key_outline)

			for task in sorted_tasks:
				parent_outline = '.'.join(task['outline'].split('.')[:-1])
				parent_rec = task_data.get(parent_outline, {}).get('record') if parent_outline else False
				start_utc = _convert_local_date_to_utc_str(task['start_date'], True)
				end_utc = _convert_local_date_to_utc_str(task['finish_date'], False)

				rec = self.env['project.task'].create({
					'name': task['name'],
					'project_id': project.id,
					'planned_date_begin': start_utc,
					'planned_date_end': end_utc,
					'duration': task['duration'],
					'parent_id': parent_rec.id if parent_rec else False,  # Si no hay jerarquía, va al root
					'sequence': int(task['uid']),
					'wbs_value': task['wbs'],
				})
				task['record'] = rec

			for task in sorted_tasks:
				for puid in task.get('predecessors', []):
					pred = next((x for x in sorted_tasks if x['uid'] == puid), None)
					if pred and pred['record']:
						self.env['project.task.linked'].create({
							'from_id': pred['record'].id,
							'to_id': task['record'].id,
							'lag': 0,
							'lag_unit': 'd',
							'type': 2,
							'dep_active': True,
						})

		# OBRA EN CAMPO al final
		max_uid = max([
			int(task['uid']) for task in task_data.values() if task.get('uid') and task['uid'].isdigit()
		], default=0)

		# Paso 2: calcular la fecha más lejana
		last_end_date = max([
			task['finish_date'] for task in task_data.values() if task['finish_date']
		], default=fields.Date.today())

		obra_start = last_end_date + timedelta(days=1)
		obra_end = obra_start

		# Paso 3: crear "OBRA EN CAMPO" como raíz y con secuencia final
		obra_en_campo = self.env['project.task'].create({
			'name': "OBRA EN CAMPO",
			'project_id': project.id,
			'parent_id': False,
			'sequence': max_uid + 10000,
			'planned_date_begin': obra_start,
			'planned_date_end': obra_end,
			'wbs_value': '99',
		})

		# Crear tareas APU como hijas de OBRA EN CAMPO
		categorias = {}
		for line in self.order_line:
			apu = self.env['apu.apu'].search([
				('product_tmpl_id', '=', line.product_id.product_tmpl_id.id)
			], limit=1)
			if not apu or not apu.subcategoria_id or not apu.subcategoria_id.categoria_id:
				continue

			categoria = apu.subcategoria_id.categoria_id
			subcategoria = apu.subcategoria_id
			categorias.setdefault(categoria, {}).setdefault(subcategoria, []).append((line, apu))

		sequence_counter = 1000
		for cat in sorted(categorias, key=lambda c: c.codigo or ''):
			tarea_cat = self.env['project.task'].create({
				'name': cat.name,
				'project_id': project.id,
				'parent_id': obra_en_campo.id,
				'sequence': sequence_counter,
			})
			sequence_counter += 10

			for subcat in sorted(categorias[cat], key=lambda s: s.codigo or ''):
				tarea_subcat = self.env['project.task'].create({
					'name': subcat.name,
					'project_id': project.id,
					'parent_id': tarea_cat.id,
					'sequence': sequence_counter,
				})
				sequence_counter += 10

				for line, apu in categorias[cat][subcat]:
					self.env['project.task'].create({
						'name': f"[{apu.code}] {apu.product_tmpl_id.name}" if apu.code else apu.product_tmpl_id.name,
						'project_id': project.id,
						'parent_id': tarea_subcat.id,
						'sequence': sequence_counter,
						'planned_hours': line.product_uom_qty,
						'description': f"Precio unitario: {line.price_unit}",
						'cantidad': line.product_uom_qty,
						'costo': line.price_unit,
					})
					sequence_counter += 10
		self.state = 'generar_proyecto'
		return {
			'type': 'ir.actions.act_window',
			'res_model': 'project.project',
			'res_id': project.id,
			'view_mode': 'form',
			'target': 'current',
		}

	def get_apu_lines_by_product_template(self, product_template_id, product_uom_qty=1.0):
		"""
		Busca el registro de APU asociado al product_template_id y, para cada línea APU,
		recalcula la cantidad y el precio final usando product_uom_qty (la cantidad del sale order).
		Devuelve una lista de diccionarios con la información necesaria.
		"""
		ApuApu = self.env['apu.apu']
		apu_record = ApuApu.search([('product_tmpl_id', '=', product_template_id)], limit=1)
		result = []
		if apu_record:
			for line in apu_record.line_ids:
				# Nueva cantidad: cantidad del sale order * cantidad definida en la línea APU
				new_qty = product_uom_qty * line.product_qty
				# Precio final: nueva cantidad * precio unitario de la línea APU
				final_price = new_qty * line.precio_unit
				result.append({
					'record': line,
					'new_qty': new_qty,
					'final_price': final_price,
					'tipo_componente': line.tipo_componente or 'Otro',
				})
		return result

	def get_component_totals(self):
		"""
		Recorre las líneas del sale order y, para cada línea, obtiene las APU Lines (recalculadas)
		y acumula la sumatoria del precio final agrupado por tipo de componente.
		Retorna un diccionario con la forma: {tipo_componente: sumatoria_total, ...}
		"""
		totals = {}
		for order in self:
			for line in order.order_line:
				apu_lines = order.get_apu_lines_by_product_template(line.product_id.product_tmpl_id.id,
				                                                    line.product_uom_qty)
				for data in apu_lines:
					tipo = data.get('tipo_componente')
					totals[tipo] = totals.get(tipo, 0) + data.get('final_price', 0.0)
		return totals

	def _get_all_components_grouped(self, apu, parent_qty, processed_kits=set()):
		""" Obtiene los componentes agrupados por tipo de componente """
		if apu.id in processed_kits:
			raise UserError(_("¡Error de recursión detectado! El kit %s está generando un ciclo." % apu.display_name))

		processed_kits.add(apu.id)
		grouped_components = defaultdict(lambda: defaultdict(lambda: {'cantidad_total': 0.0, 'uom_id': None}))

		for apu_line in apu.line_ids:
			component_qty = apu_line.product_qty * parent_qty / apu.product_qty

			sub_kit = self.env['apu.apu'].search([
				('product_tmpl_id', '=', apu_line.product_id.product_tmpl_id.id),
				('type', '=', 'phantom')
			], limit=1)

			if sub_kit:
				# Si el componente es otro kit, obtener sus componentes de forma recursiva
				sub_components = self._get_all_components_grouped(sub_kit, component_qty, processed_kits)
				for tipo, componentes in sub_components.items():
					for product_id, data in componentes.items():
						grouped_components[tipo][product_id]['cantidad_total'] += data['cantidad_total']
						grouped_components[tipo][product_id]['uom_id'] = data['uom_id']
			else:
				# Si es un componente normal, agruparlo por tipo de componente
				tipo = apu_line.tipo_componente or 'Otro'
				grouped_components[tipo][apu_line.product_id.id]['cantidad_total'] += component_qty
				grouped_components[tipo][apu_line.product_id.id]['uom_id'] = apu_line.product_uom_id.id

		return grouped_components

	def action_confirm(self):
		res = super(SaleOrder, self).action_confirm()
		picking = self.env['stock.picking']
		for order in self:
			if order.tipo_pedido == 'apu':
				# Solo permitir si el usuario está en el grupo gerente
				if not self.env.user.has_group('construccion_gps.group_apus_gerente'):
					raise UserError(_('No tienes permisos para confirmar pedidos de tipo APU.'))
			for line in order.order_line:
				# Buscar si el producto tiene un APU (kit)
				apu = self.env['apu.apu'].search([
					('product_tmpl_id', '=', line.product_id.product_tmpl_id.id),
					('type', '=', 'phantom')  # Solo procesar kits
				], limit=1)

				if apu:
					# Crear o buscar el picking existente
					picking = order.picking_ids.filtered(lambda p: p.state not in ['done', 'cancel'])
					if not picking:
						picking = self.env['stock.picking'].create({
							'partner_id': order.partner_shipping_id.id,
							'picking_type_id': order.warehouse_id.out_type_id.id,
							'location_id': order.warehouse_id.lot_stock_id.id,
							'location_dest_id': order.partner_shipping_id.property_stock_customer.id,
							'origin': order.name,
						})

					# Generar los movimientos de stock para los componentes de los kits anidados
					def get_all_components(apu, parent_qty):
						components = []
						for apu_line in apu.line_ids:
							component_qty = apu_line.product_qty * parent_qty / apu.product_qty

							# Verificar si el componente es otro kit
							sub_kit = self.env['apu.apu'].search([
								('product_tmpl_id', '=', apu_line.product_id.product_tmpl_id.id),
								('type', '=', 'phantom')
							], limit=1)

							if sub_kit:
								# Si el componente es otro kit, obtener sus componentes de forma recursiva
								components += get_all_components(sub_kit, component_qty)
							else:
								# Si es un componente normal, agregarlo a la lista
								components.append({
									'picking_id': picking.id,
									'product_id': apu_line.product_id.id,
									'product_uom_qty': component_qty,
									'product_uom': apu_line.product_id.uom_id.id,
									'location_id': picking.location_id.id,
									'location_dest_id': picking.location_dest_id.id,
									'name': f"Component of {line.product_id.display_name}",
									'sale_line_id': line.id,
								})
						return components

					# Obtener todos los componentes del kit y sus subkits
					all_components = get_all_components(apu, line.product_uom_qty)

					# Crear los movimientos de stock para los componentes finales
					self.env['stock.move'].create(all_components)

				# Eliminar el movimiento del producto principal (evita que aparezca en el picking)
				move_field = 'move_ids_without_package' if hasattr(picking,
				                                                   'move_ids_without_package') else 'move_lines'
				rule_moves = getattr(picking, move_field).filtered(lambda move: move.rule_id)
				for move in rule_moves:
					move._action_cancel()
					move.unlink()

		return res

	def export_excel(self):
		self.ensure_one()
		import io, base64, xlsxwriter
		output = io.BytesIO()
		workbook = xlsxwriter.Workbook(output, {'in_memory': True})
		worksheet = workbook.add_worksheet("Informe Detallado Mano de Obra")

		# --- Configuración de Fuentes y Formatos ---
		common_font = {'font_name': 'Verdana', 'font_size': 10, 'valign': 'vcenter'}
		title_format = workbook.add_format(
			{**{'bold': True, 'font_size': 16, 'align': 'center', 'valign': 'vcenter'}, **common_font})
		header_format = workbook.add_format(
			{**{'bold': True, 'bg_color': '#EAECEE', 'border': 1, 'align': 'center', 'valign': 'vcenter',
			    'text_wrap': True}, **common_font})
		cat_header_format = workbook.add_format(
			{**{'bold': True, 'bg_color': '#DCE6F1', 'border': 1, 'align': 'center', 'valign': 'vcenter'},
			 **common_font})
		subcat_header_format = workbook.add_format(
			{**{'bold': True, 'bg_color': '#EBF5FB', 'border': 1, 'align': 'center', 'valign': 'vcenter'},
			 **common_font})
		# Formato centrado para celdas de texto (ITEM, DESCRIPCIÓN, UNIDAD) con ajuste de texto
		center_text_format = workbook.add_format(
			{**{'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True}, **common_font})
		# Formato numérico para valores tipo moneda (alineado a la derecha)
		monetary_format = workbook.add_format(
			{**{'num_format': '$#,##0.00', 'border': 1, 'align': 'right', 'valign': 'vcenter'}, **common_font})
		numeric_format = workbook.add_format(
			{**{'num_format': '#,##0.00', 'border': 1, 'align': 'right', 'valign': 'vcenter'}, **common_font})
		text_format = workbook.add_format(
			{**{'border': 1, 'align': 'left', 'valign': 'vcenter', 'text_wrap': True}, **common_font})
		date_format = workbook.add_format(
			{**{'num_format': 'yyyy-mm-dd', 'border': 1, 'align': 'left', 'valign': 'vcenter'}, **common_font})

		# Formatos para la sección de APU (se centran todos los datos)
		apu_text_format = workbook.add_format(
			{**{'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True}, **common_font})
		apu_numeric_format = workbook.add_format(
			{**{'num_format': '$#,##0.00', 'border': 1, 'align': 'right', 'valign': 'vcenter'}, **common_font})
		apu_numeric_plain = workbook.add_format(
			{**{'num_format': '#,##0.00', 'border': 1, 'align': 'right', 'valign': 'vcenter'}, **common_font})

		# --- Ajustar ancho de columnas ---
		worksheet.set_column('A:A', 15)
		worksheet.set_column('B:B', 40)
		worksheet.set_column('C:C', 12)
		worksheet.set_column('D:D', 15)
		worksheet.set_column('E:E', 12)
		worksheet.set_column('F:F', 15)

		# --- Título ---
		worksheet.merge_range('A1:F1', "Informe Detallado Mano de Obra", title_format)

		# --- Datos de Cabecera del Pedido ---
		worksheet.write('A3', "Orden #", header_format)
		worksheet.write('B3', self.name or '', text_format)
		worksheet.write('A4', "Compañía", header_format)
		worksheet.write('B4', self.company_id.name or '', text_format)
		worksheet.write('A5', "Responsable", header_format)
		worksheet.write('B5', self.user_id.name or '', text_format)
		worksheet.write('A6', "Fecha", header_format)
		try:
			dt = datetime.strptime(self.date_order, "%Y-%m-%d %H:%M:%S")
		except Exception:
			dt = self.date_order
		if isinstance(dt, datetime):
			worksheet.write_datetime('B6', dt, date_format)
		else:
			worksheet.write('B6', dt or '', text_format)
		worksheet.write('A7', "Cliente", header_format)
		worksheet.write('B7', self.partner_id.name or '', text_format)

		# --- Precalcular totales por Categoría y Subcategoría ---
		cat_totals = {}
		subcat_totals = {}
		for line in self.order_line:
			apu_rec = self.env['apu.apu'].search([
				('product_tmpl_id', '=', line.product_id.product_tmpl_id.id)
			], limit=1)
			if apu_rec and apu_rec.subcategoria_id:
				cat = apu_rec.subcategoria_id.categoria_id.name or "Sin Categoría"
				subcat = apu_rec.subcategoria_id.name or "Sin Subcategoría"
			else:
				cat = "Sin Categoría"
				subcat = "Sin Subcategoría"
			cat_totals[cat] = cat_totals.get(cat, 0.0) + (line.price_subtotal or 0.0)
			subcat_totals[(cat, subcat)] = subcat_totals.get((cat, subcat), 0.0) + (line.price_subtotal or 0.0)

		# --- Tabla Principal: Líneas de Pedido con Agrupación ---
		table_start = 9
		table_headers = ["ITEM", "DESCRIPCIÓN", "UNIDAD", "PRECIO", "CANTIDAD", "SUBTOTAL"]
		for col, h in enumerate(table_headers):
			worksheet.write(table_start, col, h, header_format)
		current_row = table_start + 1

		prev_cat = None
		prev_subcat = None

		for line in self.order_line:
			apu_rec = self.env['apu.apu'].search([
				('product_tmpl_id', '=', line.product_id.product_tmpl_id.id)
			], limit=1)
			if apu_rec and apu_rec.subcategoria_id:
				current_cat = apu_rec.subcategoria_id.categoria_id.name or "Sin Categoría"
				current_subcat = apu_rec.subcategoria_id.name or "Sin Subcategoría"
			else:
				current_cat = "Sin Categoría"
				current_subcat = "Sin Subcategoría"

			# Si la categoría cambia, insertar encabezado de categoría con total
			if current_cat != prev_cat:
				cat_code = (apu_rec.subcategoria_id.categoria_id.codigo
				            if (apu_rec and apu_rec.subcategoria_id and apu_rec.subcategoria_id.categoria_id.codigo)
				            else "")
				worksheet.write(current_row, 0, cat_code, cat_header_format)
				worksheet.write(current_row, 1, current_cat, cat_header_format)
				worksheet.merge_range(current_row, 2, current_row, 4, "", cat_header_format)
				worksheet.write_number(current_row, 5, cat_totals.get(current_cat, 0.0), monetary_format)
				current_row += 1
				prev_cat = current_cat
				prev_subcat = None

			# Si la subcategoría cambia, insertar encabezado de subcategoría con total
			if current_subcat != prev_subcat:
				subcat_code = (apu_rec.subcategoria_id.codigo
				               if (apu_rec and apu_rec.subcategoria_id and apu_rec.subcategoria_id.codigo)
				               else "")
				worksheet.write(current_row, 0, subcat_code, subcat_header_format)
				worksheet.write(current_row, 1, current_subcat, subcat_header_format)
				worksheet.merge_range(current_row, 2, current_row, 4, "", subcat_header_format)
				worksheet.write_number(current_row, 5, subcat_totals.get((current_cat, current_subcat), 0.0),
				                       monetary_format)
				current_row += 1
				prev_subcat = current_subcat

			# Fila principal de la línea del pedido: usar formato centrado para ITEM, DESCRIPCIÓN y UNIDAD
			item_code = apu_rec.code if (apu_rec and apu_rec.code) else ""
			worksheet.write(current_row, 0, item_code, center_text_format)
			# Para DESCRIPCIÓN se usa wrap para mostrar salto de línea si es necesario
			worksheet.write(current_row, 1, line.name or "", center_text_format)
			worksheet.write(current_row, 2, apu_rec.product_uom_id.name or "", center_text_format)
			worksheet.write_number(current_row, 3, line.price_unit or 0.0, monetary_format)
			worksheet.write_number(current_row, 4, line.product_uom_qty or 0.0, numeric_format)
			worksheet.write_number(current_row, 5, line.price_subtotal or 0.0, monetary_format)
			current_row += 1

			# Bloque adicional: APU Lines de Mano de Obra para la línea
			apu_lines = self.get_apu_lines_or_product(line.product_id.product_tmpl_id.id, line.product_uom_qty)
			manoobra_lines = apu_lines.filtered(lambda r: r.tipo_componente == 'manoobra')
			for apu_line in manoobra_lines:
				new_qty = line.product_uom_qty * apu_line.product_qty
				final_price = new_qty * (apu_line.precio_unit or 0.0)
				worksheet.write(current_row, 0, apu_line.codigo_componente or "", apu_text_format)
				comp_info = ("[" + (apu_line.product_id.default_code or "") + "] " + (apu_line.product_id.name or "")) \
					if apu_line.product_id.default_code else (apu_line.product_id.name or "")
				worksheet.write(current_row, 1, comp_info, apu_text_format)
				worksheet.write(current_row, 2, apu_line.product_uom_id.name or "", apu_text_format)
				worksheet.write_number(current_row, 3, apu_line.precio_unit or 0.0, apu_numeric_format)
				worksheet.write_number(current_row, 4, new_qty, apu_numeric_plain)
				worksheet.write_number(current_row, 5, final_price, apu_numeric_format)
				current_row += 1

		# --- Total General ---
		worksheet.write(current_row + 1, 0, "Total General", header_format)
		total_general = sum(line.price_subtotal or 0.0 for line in self.order_line)
		worksheet.write_number(current_row + 1, 5, total_general, monetary_format)

		workbook.close()
		output.seek(0)
		file_data = output.read()

		attachment = self.env['ir.attachment'].create({
			'name': 'SaleOrder_Export.xlsx',
			'datas': base64.b64encode(file_data),
			'type': 'binary',
			'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
		})
		url = "/web/content/%s?download=true" % (attachment.id)
		return {
			'type': 'ir.actions.act_url',
			'url': url,
			'target': 'new',
		}

	porcentaje_inicial_requisicion = fields.Float(string='Porcentaje Inicial de Requisición (%)', digits=(12, 2),
	                                              default=100.0,
	                                              help="Porcentaje de la cantidad del pedido que se requisitará inicialmente (1-100).")
	tiempo_entrega = fields.Char(string="Tiempo de Entrega", copy=False, tracking=True)
	condiciones_comerciales = fields.Text(string="Condiciones Comerciales", copy=False, tracking=True)
	garantias = fields.Text(string="Garantías", copy=False, tracking=True)
	analytic_account_id = fields.Many2one('account.analytic.account', string="Cuenta analítica", store=True)

	@api.constrains('porcentaje_inicial_requisicion')
	def _check_porcentaje_inicial_requisicion(self):
		for order in self:
			pct = order.porcentaje_inicial_requisicion
			if pct < 1.0 or pct > 100.0:
				raise ValidationError(
					_("El porcentaje inicial de requisición debe estar entre 1 y 100%%.")
				)

	def _round_custom(self, value):
		"""Redondea: si la parte decimal < 0.5 hacia abajo, si ≥ 0.5 hacia arriba."""
		entero = math.floor(value)
		return entero + 1 if (value - entero) >= 0.5 else entero

	@api.depends('no_proyecto')
	def _create_analytic_account(self):
		for record in self:
			if not record.no_proyecto:
				return False

			cuenta = self.env['account.analytic.account'].search([
				('code', '=', record.no_proyecto),
				('company_id', '=', record.company_id.id)
			], limit=1)

			# Si es alcance, solo usamos la cuenta existente (no creamos nueva)
			if record.es_alcance:
				return cuenta if cuenta else False

			# Si no es alcance y no existe cuenta, la creamos
			if not cuenta:
				cuenta = self.env['account.analytic.account'].create({
					'name': record.no_proyecto + '-' + (record.proyecto or ''),
					'company_id': record.env.company.id,
					'active': True,
					'plan_id': 8,  # GE PROYECTOS (ajusta según corresponda)
					'code': record.no_proyecto,
					'partner_id': record.partner_id.id
				})

			record.analytic_account_id = cuenta.id
			return cuenta

	def generate_purchase_requisitionXYZ(self):
		self.ensure_one()
		pur_req_serv = self.env['macro.purchase.request']
		pur_req = self.env['macro.purchase.request']
		factor = (self.porcentaje_inicial_requisicion or 100.0) / 100.0

		# 1) Crear o reutilizar proyecto + cuenta analítica
		# debo enviar a crear la cuenta analitica
		cuenta = self._create_analytic_account()
		if not cuenta:
			raise UserError(_("Debes definir primero un nombre de proyecto."))

		analytic_account = cuenta  # project.analytic_account_id
		if not analytic_account:
			raise UserError(
				_("No se encontró la cuenta analítica para el proyecto %s.") % self.project_name
			)
		aa_id = analytic_account.id

		# Asignar al sale.order y sus líneas
		self.write({'analytic_account_id': aa_id})  # , 'project_id': project.id})
		for line in self.order_line:
			line.analytic_distribution = {aa_id: 100.0}

		if not self.macro_unificado:
			# 2) Requisición de servicios
			pick_type = self.env['stock.picking.type'].search([
				('warehouse_id.company_id', '=', self.company_id.id),
				('code', '=', 'incoming')
			], limit=1)
			pr_vals = {
				'sale_order_id': self.id,
				'state': 'draft',
				'company_id': self.company_id.id,
				'requested_by': self.env.user.id,
				'date_start': fields.Date.context_today(self),
				'date_planned': (datetime.today() + timedelta(days=15)).date(),
				'picking_type_id': pick_type.id,
				'permite_aprobar': True,
				'request_type': 'service',
				'analytic_distribution': {aa_id: 100.0},
			}
			pur_req_serv = self.env['macro.purchase.request'].create(pr_vals)

			for sol in self.order_line.sorted('id', reverse=True):
				apu = sol.apu_id or self.env['apu.apu'].search([
					('product_tmpl_id', '=', sol.product_template_id.id)
				], limit=1)
				if not apu:
					continue

				for apu_line in apu.line_ids.sorted('sequence', reverse=True):
					if apu_line.product_id.detailed_type in ('service', 'consu'):
						raw = sol.product_uom_qty * apu_line.product_qty * factor
						qty = raw  # self._round_custom(raw)
						if qty <= 0:
							continue
						tipo_componente_display = dict(
							self.env['apu.apu.line']._fields['tipo_componente'].selection
						).get(apu_line.tipo_componente, '').strip()

						tipo_costo_rec = self.env['tipo.costo'].search([
							('name', 'ilike', tipo_componente_display)
						], limit=1)
						self.env['macro.purchase.request.line'].create({
							'request_id': pur_req_serv.id,
							'sale_order_line_id': sol.id,
							'product_id': apu_line.product_id.id,
							'name': apu_line.product_id.name,
							'product_qty': qty,
							'date_required': fields.Date.context_today(self),
							'analytic_distribution': {aa_id: 100.0},
							'product_brand_id': apu_line.product_id.product_brand_id.id,
							'product_uom_id': apu_line.product_id.uom_id.id,
							'product_categ_id': apu_line.product_id.categ_id.id,
							'rubro': apu_line.codigo_componente,
							'costo_unit_apu': apu_line.cost,
							'tipo_costo': tipo_costo_rec.name if tipo_costo_rec else tipo_componente_display,
							# apu_line.tipo_componente
						})

			# 3) Requisición de productos
			pick_type = self.env['stock.picking.type'].search([
				('warehouse_id.company_id', '=', self.company_id.id),
				('code', '=', 'incoming')
			], limit=1)
			pr_vals = {
				'sale_order_id': self.id,
				'state': 'draft',
				'company_id': self.company_id.id,
				'requested_by': self.env.user.id,
				'date_start': fields.Date.context_today(self),
				'date_planned': (datetime.today() + timedelta(days=15)).date(),
				'picking_type_id': pick_type.id,
				'permite_aprobar': True,
				'request_type': 'product',
				'analytic_distribution': {aa_id: 100.0},
			}
			pur_req = self.env['macro.purchase.request'].create(pr_vals)

			for sol in self.order_line.sorted('id', reverse=True):
				apu = sol.apu_id or self.env['apu.apu'].search([
					('product_tmpl_id', '=', sol.product_template_id.id)
				], limit=1)
				if not apu:
					continue

				for apu_line in apu.line_ids.sorted('sequence', reverse=True):
					if apu_line.product_id.detailed_type in ('product'):
						raw = sol.product_uom_qty * apu_line.product_qty * factor
						qty = raw  # self._round_custom(raw)
						if qty <= 0:
							continue
						tipo_componente_display = dict(
							self.env['apu.apu.line']._fields['tipo_componente'].selection
						).get(apu_line.tipo_componente, '').strip()

						tipo_costo_rec = self.env['tipo.costo'].search([
							('name', 'ilike', tipo_componente_display)
						], limit=1)
						self.env['macro.purchase.request.line'].create({
							'request_id': pur_req.id,
							'sale_order_line_id': sol.id,
							'product_id': apu_line.product_id.id,
							'name': apu_line.product_id.name,
							'product_qty': qty,
							'date_required': fields.Date.context_today(self),
							'analytic_distribution': {aa_id: 100.0},
							'product_brand_id': apu_line.product_id.product_brand_id.id,
							'product_categ_id': apu_line.product_id.categ_id.id,
							'product_uom_id': apu_line.product_id.uom_id.id,
							'rubro': apu_line.codigo_componente,
							'costo_unit_apu': apu_line.cost,
							'tipo_costo': tipo_costo_rec.name if tipo_costo_rec else tipo_componente_display,
							# apu_line.tipo_componente
						})
		else:  # para las que vn directo
			# Clasificación de líneas por tipo
			serv_items = []
			prod_items = []

			for sol in self.order_line.sorted('id', reverse=True):
				raw_qty = sol.product_uom_qty * factor
				if raw_qty <= 0:
					continue
				# tendria que ir a buscar el apu para obtener el costo unitario
				apu = self.env['apu.apu'].search([
					('product_tmpl_id', '=', sol.product_id.product_tmpl_id.id)
				], limit=1)

				item_data = {
					'sale_order_line_id': sol.id,
					'product_id': sol.product_id.id,
					'name': sol.product_id.name,
					'product_qty': raw_qty,
					'date_required': fields.Date.context_today(self),
					'analytic_distribution': {aa_id: 100.0},
					'product_brand_id': sol.product_id.product_brand_id.id,
					'product_uom_id': sol.product_id.uom_id.id,
					'product_categ_id': sol.product_id.categ_id.id,
					'rubro': '',
					'tipo_costo': '',
					'costo_unit_apu': apu.total_apu_costo
				}
				if sol.product_id.detailed_type in ('service', 'consu'):
					serv_items.append(item_data)
				elif sol.product_id.detailed_type == 'product':
					prod_items.append(item_data)

			pick_type = self.env['stock.picking.type'].search([
				('warehouse_id.company_id', '=', self.company_id.id),
				('code', '=', 'incoming')
			], limit=1)

			# Crear requisición de servicios si hay elementos
			if serv_items:
				pur_req_serv = self.env['macro.purchase.request'].create({
					'sale_order_id': self.id,
					'state': 'draft',
					'company_id': self.company_id.id,
					'requested_by': self.env.user.id,
					'date_start': fields.Date.context_today(self),
					'date_planned': (datetime.today() + timedelta(days=15)).date(),
					'picking_type_id': pick_type.id,
					'permite_aprobar': True,
					'request_type': 'service',
					'analytic_distribution': {aa_id: 100.0},
				})
				for line in serv_items:
					line['request_id'] = pur_req_serv.id
					self.env['macro.purchase.request.line'].create(line)

			# Crear requisición de productos si hay elementos
			if prod_items:
				pur_req = self.env['macro.purchase.request'].create({
					'sale_order_id': self.id,
					'state': 'draft',
					'company_id': self.company_id.id,
					'requested_by': self.env.user.id,
					'date_start': fields.Date.context_today(self),
					'date_planned': (datetime.today() + timedelta(days=15)).date(),
					'picking_type_id': pick_type.id,
					'permite_aprobar': True,
					'request_type': 'product',
					'analytic_distribution': {aa_id: 100.0},
				})
				for line in prod_items:
					line['request_id'] = pur_req.id
					self.env['macro.purchase.request.line'].create(line)
		# 4) Estado y retorno
		self.write({'state': 'generar_requisiciones'})
		return {
			'material_requisition': pur_req_serv,
			'purchase_request': pur_req,
		}

	def generate_purchase_requisition(self):
		self.ensure_one()
		pur_req_serv = self.env['macro.purchase.request']
		pur_req = self.env['macro.purchase.request']
		factor = (self.porcentaje_inicial_requisicion or 100.0) / 100.0

		# 1) Crear o reutilizar proyecto + cuenta analítica
		# debo enviar a crear la cuenta analitica
		cuenta = self._create_analytic_account()
		if not cuenta:
			raise UserError(_("Debes definir primero un nombre de proyecto."))

		analytic_account = cuenta  # project.analytic_account_id
		if not analytic_account:
			raise UserError(
				_("No se encontró la cuenta analítica para el proyecto %s.") % self.project_name
			)
		aa_id = analytic_account.id

		# Asignar al sale.order y sus líneas
		self.write({'analytic_account_id': aa_id})  # , 'project_id': project.id})
		for line in self.order_line:
			line.analytic_distribution = {aa_id: 100.0}

		if not self.macro_unificado:
			# 2) Requisición de servicios
			pick_type = self.env['stock.picking.type'].search([
				('warehouse_id.company_id', '=', self.company_id.id),
				('code', '=', 'incoming')
			], limit=1)
			pr_vals = {
				'sale_order_id': self.id,
				'origin': self.referencia_analitica,
				'state': 'draft',
				'company_id': self.company_id.id,
				'requested_by': self.env.user.id,
				'date_start': fields.Date.context_today(self),
				'date_planned': (datetime.today() + timedelta(days=15)).date(),
				'picking_type_id': pick_type.id,
				'permite_aprobar': True,
				'request_type': 'service',
				'analytic_distribution': {self.analytic_account_id.id: 100.0},
				'superintendente': self.superintendente.id,
				'supervisor': self.supervisor.id
			}
			pur_req_serv = self.env['macro.purchase.request'].create(pr_vals)

			# --- AGRUPAR SERVICIOS ---
			agg_serv = {}  # key: (product_id, uom_id)
			for sol in self.order_line.sorted('id', reverse=True):
				apu = sol.apu_id or self.env['apu.apu'].search([
					('product_tmpl_id', '=', sol.product_template_id.id)
				], limit=1)
				if not apu:
					continue

				for apu_line in apu.line_ids.sorted('sequence', reverse=True):
					if apu_line.product_id.detailed_type not in ('service', 'consu'):
						continue
					qty = sol.product_uom_qty * apu_line.product_qty * factor
					if qty <= 0:
						continue

					key = (apu_line.product_id.id, apu_line.product_id.uom_id.id)
					tipo_costo_display = dict(self.env['apu.apu.line']._fields['tipo_componente'].selection) \
						.get(apu_line.tipo_componente, '').strip()

					if key not in agg_serv:
						agg_serv[key] = {
							'request_id': pur_req_serv.id,
							'sale_order_line_id': False,  # agregado: línea consolidada
							'product_id': apu_line.product_id.id,
							'name': apu_line.product_id.name,
							'product_qty': qty,
							'date_required': fields.Date.context_today(self),
							'analytic_distribution': {self.analytic_account_id.id: 100.0},
							'product_brand_id': apu_line.product_id.product_brand_id.id,
							'product_uom_id': apu_line.product_id.uom_id.id,
							'product_categ_id': apu_line.product_id.categ_id.id,
							'costo_unit_apu': apu_line.cost,
							'tipo_costo': tipo_costo_display,
							'_rubros': set([apu_line.codigo_componente or '']),
						}
					else:
						agg_serv[key]['product_qty'] += qty
						agg_serv[key]['_rubros'].add(apu_line.codigo_componente or '')

			for vals in agg_serv.values():
				vals['rubro'] = ", ".join(sorted(r for r in vals.pop('_rubros') if r))[:200]
				self.env['macro.purchase.request.line'].create(vals)

			# 3) Requisición de productos
			pick_type = self.env['stock.picking.type'].search([
				('warehouse_id.company_id', '=', self.company_id.id),
				('code', '=', 'incoming')
			], limit=1)
			pr_vals = {
				'sale_order_id': self.id,
				'origin': self.referencia_analitica,
				'state': 'draft',
				'company_id': self.company_id.id,
				'requested_by': self.env.user.id,
				'date_start': fields.Date.context_today(self),
				'date_planned': (datetime.today() + timedelta(days=15)).date(),
				'picking_type_id': pick_type.id,
				'permite_aprobar': True,
				'request_type': 'product',
				'analytic_distribution': {self.analytic_account_id.id: 100.0},
				'superintendente': self.superintendente.id,
				'supervisor': self.supervisor.id
			}
			pur_req = self.env['macro.purchase.request'].create(pr_vals)

			# --- AGRUPAR PRODUCTOS (stockables) ---
			agg_prod = {}  # key: (product_id, uom_id)
			for sol in self.order_line.sorted('id', reverse=True):
				apu = sol.apu_id or self.env['apu.apu'].search([
					('product_tmpl_id', '=', sol.product_template_id.id)
				], limit=1)
				if not apu:
					continue

				for apu_line in apu.line_ids.sorted('sequence', reverse=True):
					if apu_line.product_id.detailed_type != 'product':
						continue
					qty = sol.product_uom_qty * apu_line.product_qty * factor
					if qty <= 0:
						continue

					key = (apu_line.product_id.id, apu_line.product_id.uom_id.id)
					tipo_costo_display = dict(self.env['apu.apu.line']._fields['tipo_componente'].selection) \
						.get(apu_line.tipo_componente, '').strip()

					if key not in agg_prod:
						agg_prod[key] = {
							'request_id': pur_req.id,
							'sale_order_line_id': False,  # consolidada
							'product_id': apu_line.product_id.id,
							'name': apu_line.product_id.name,
							'product_qty': qty,
							'date_required': fields.Date.context_today(self),
							'analytic_distribution': {self.analytic_account_id.id: 100.0},
							'product_brand_id': apu_line.product_id.product_brand_id.id,
							'product_uom_id': apu_line.product_id.uom_id.id,
							'product_categ_id': apu_line.product_id.categ_id.id,
							'costo_unit_apu': apu_line.cost,
							'tipo_costo': tipo_costo_display,
							'_rubros': set([apu_line.codigo_componente or '']),
						}
					else:
						agg_prod[key]['product_qty'] += qty
						agg_prod[key]['_rubros'].add(apu_line.codigo_componente or '')

			for vals in agg_prod.values():
				vals['rubro'] = ", ".join(sorted(r for r in vals.pop('_rubros') if r))[:200]
				self.env['macro.purchase.request.line'].create(vals)
		else:  # para las que vn directo
			# --- macro_unificado = True: primero juntamos, luego creamos en bloque ---
			serv_items = []
			prod_items = []
			for sol in self.order_line.sorted('id', reverse=True):
				raw_qty = sol.product_uom_qty * factor
				if raw_qty <= 0:
					continue
				apu = self.env['apu.apu'].search([('product_tmpl_id', '=', sol.product_id.product_tmpl_id.id)], limit=1)
				base = {
					'sale_order_line_id': False,  # consolidada
					'product_id': sol.product_id.id,
					'name': sol.product_id.name,
					'product_qty': raw_qty,
					'date_required': fields.Date.context_today(self),
					'analytic_distribution': {self.analytic_account_id.id: 100.0},
					'product_brand_id': sol.product_id.product_brand_id.id,
					'product_uom_id': sol.product_id.uom_id.id,
					'product_categ_id': sol.product_id.categ_id.id,
					'rubro': '',
					'tipo_costo': '',
					'costo_unit_apu': apu.total_apu_costo if apu else 0.0,
				}
				if sol.product_id.detailed_type in ('service', 'consu'):
					serv_items.append(dict(base))
				elif sol.product_id.detailed_type == 'product':
					prod_items.append(dict(base))

			def _group_items(items):
				grouped = {}
				for it in items:
					key = (it['product_id'], it['product_uom_id'])
					if key not in grouped:
						grouped[key] = it
					else:
						grouped[key]['product_qty'] += it['product_qty']
				return list(grouped.values())

			pick_type = self.env['stock.picking.type'].search([
				('warehouse_id.company_id', '=', self.company_id.id),
				('code', '=', 'incoming')
			], limit=1)

			if serv_items:
				pur_req_serv = self.env['macro.purchase.request'].create({
					'sale_order_id': self.id,
					'origin': self.referencia_analitica,
					'state': 'draft',
					'company_id': self.company_id.id,
					'requested_by': self.env.user.id,
					'date_start': fields.Date.context_today(self),
					'date_planned': (datetime.today() + timedelta(days=15)).date(),
					'picking_type_id': pick_type.id,
					'permite_aprobar': True,
					'request_type': 'service',
					'analytic_distribution': {self.analytic_account_id.id: 100.0},
					'superintendente': self.superintendente.id,
					'supervisor': self.supervisor.id
				})
				for vals in _group_items(serv_items):
					vals['request_id'] = pur_req_serv.id
					self.env['macro.purchase.request.line'].create(vals)

			if prod_items:
				pur_req = self.env['macro.purchase.request'].create({
					'sale_order_id': self.id,
					'origin': self.referencia_analitica,
					'state': 'draft',
					'company_id': self.company_id.id,
					'requested_by': self.env.user.id,
					'date_start': fields.Date.context_today(self),
					'date_planned': (datetime.today() + timedelta(days=15)).date(),
					'picking_type_id': pick_type.id,
					'permite_aprobar': True,
					'request_type': 'product',
					'analytic_distribution': {self.analytic_account_id.id: 100.0},
					'superintendente': self.superintendente.id,
					'supervisor': self.supervisor.id
				})
				for vals in _group_items(prod_items):
					vals['request_id'] = pur_req.id
					self.env['macro.purchase.request.line'].create(vals)

		# 4) Estado y retorno
		self.write({'state': 'generar_requisiciones'})
		# 5) Notificar a Supervisor y Superintendente
		reqs_creadas = []
		if pur_req_serv and pur_req_serv.ids:
			reqs_creadas.append(pur_req_serv)
		if pur_req and pur_req.ids:
			reqs_creadas.append(pur_req)
		self._send_macro_requisition_notification(reqs_creadas)
		return {
			'material_requisition': pur_req_serv,
			'purchase_request': pur_req,
		}

	def _send_macro_requisition_notification(self, reqs):
		"""Enviar correo directo SOLO a supervisor y superintendente; sin reply-to; y log silencioso."""
		self.ensure_one()
		# Correos destino: supervisor/superintendente
		emails = []
		for u in (self.supervisor, self.superintendente):
			if u and u.partner_id and u.partner_id.email:
				emails.append(u.partner_id.email)
		emails = list(dict.fromkeys(emails))
		if not emails:
			return

		# Links a las requisiciones creadas
		base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url') or ''
		items = []
		for r in reqs:
			if r:
				url = f"{base_url}/web#id={r.id}&model=macro.purchase.request&view_type=form"
				label = r.display_name or r.name or f"Requisición {r.id}"
				tipo = getattr(r, 'request_type', '') or ''
				items.append(f"<li><a href='{url}'>{label}</a> ({tipo})</li>")
		if not items:
			return

		subject = f"Macro requisiciones creadas - {self.referencia_analitica or self.name}"
		body = f"""
			<p>Se han creado macro requisiciones para el proyecto
			<b>{(self.referencia_analitica or self.name) or ''}</b> del presupuesto <b>{self.name}</b>.</p>
			<ul>{''.join(items)}</ul>
		"""

		# Enviar correo directo (NO partner_ids, NO followers, SIN reply_to)
		email_from = (self.env.user.email_formatted
		              or self.company_id.partner_id.email_formatted
		              or False)
		self.env['mail.mail'].sudo().create({
			'subject': subject,
			'body_html': body,
			'email_from': email_from,
			'email_to': ",".join(emails),
			'reply_to': False,  # <-- sin "Responder a"
			'auto_delete': True,
		}).send()

		# Log en chatter sin notificar a nadie
		self.message_post(body=body, subject=subject, subtype_xmlid='mail.mt_note')

	def action_view_material_requisitions(self):
		action = self.env.ref('construccion_gps.action_view_material_requisitions').read()[0]
		action['domain'] = [('sale_order_id', '=', self.id), ('request_type', '=', 'service')]
		return action

	def action_view_purchase_requests(self):
		action = self.env.ref('construccion_gps.action_view_purchase_requests').read()[0]
		action['domain'] = [('sale_order_id', '=', self.id), ('request_type', '=', 'product')]
		return action

	def export_sale_order_excel(self):
		"""
		Exporta el Sale Order a un Excel.
		Líneas: id, apu_id, product_template_id, name, analytic account name(s) agrupadas por plan con porcentaje,
				product_uom_qty, price_unit.
		"""
		self.ensure_one()
		output = io.BytesIO()
		workbook = xlsxwriter.Workbook(output, {'in_memory': True})
		worksheet = workbook.add_worksheet("Pedido de Venta")

		# Definir formatos
		header_format = workbook.add_format({
			'bold': True, 'font_name': 'Verdana', 'font_size': 11,
			'align': 'center', 'valign': 'vcenter', 'bg_color': '#F7F7F7', 'border': 1
		})
		text_format = workbook.add_format({
			'font_name': 'Verdana', 'font_size': 10, 'border': 1,
			'align': 'left', 'valign': 'vcenter'
		})
		number_format = workbook.add_format({
			'font_name': 'Verdana', 'font_size': 10, 'border': 1,
			'align': 'right', 'valign': 'vcenter', 'num_format': '#,##0.00'
		})

		# Ajustar columnas A–G (agregamos una más por la nueva columna APU)
		worksheet.set_column('A:A', 10)  # ID
		worksheet.set_column('B:B', 10)  # APU (ID)
		worksheet.set_column('C:C', 30)  # Producto
		worksheet.set_column('D:D', 40)  # Descripción
		worksheet.set_column('E:E', 30)  # Cuentas Analíticas por Plan (%)
		worksheet.set_column('F:F', 10)  # Cantidad
		worksheet.set_column('G:G', 15)  # Precio Unitario

		# Escribir Encabezado: Nombre del pedido
		worksheet.merge_range('A1:G1', f"Pedido: {self.name or ''}", header_format)

		# Escribir títulos de columnas
		titles = ['ID', 'APU', 'Producto', 'Descripción',
		          'Cuentas Analíticas por Plan (%)', 'Cantidad', 'Precio Unitario']
		for col, title in enumerate(titles):
			worksheet.write(2, col, title, header_format)

		# Escribir datos de las líneas
		row = 3
		for line in self.order_line:
			# A: ID
			worksheet.write(row, 0, line.id, text_format)
			# B: APU (ID numérico)
			worksheet.write(row, 1, line.apu_id.id if line.apu_id else '', text_format)
			# C: Producto (template code)
			worksheet.write(row, 2,
			                line.product_template_id.default_code or '', text_format)
			# D: Descripción
			worksheet.write(row, 3, line.name or '', text_format)

			# E: Analytic Distribution agrupado por Plan
			analytic_display = ''
			dist = line.analytic_distribution
			if dist:
				plan_map = {}
				# Si es dict {account_id: pct, ...}
				items = dist.items() if isinstance(dist, dict) else (
					[(d.account_id.id, d.distribution) for d in dist]
					if hasattr(dist, '__iter__') else []
				)
				for acct_id, pct in items:
					acct = self.env['account.analytic.account'].browse(int(acct_id))
					plan = acct.plan_id.complete_name or 'Sin Plan'
					plan_map.setdefault(plan, []).append((acct.name, pct))
				analytic_lines = []
				for plan_name, entries in sorted(plan_map.items()):
					chunks = [f"{name} ({pct:.2f}%)" for name, pct in entries]
					analytic_lines.append(f"{plan_name}: " + ", ".join(chunks))
				analytic_display = "\n".join(analytic_lines)
			worksheet.write(row, 4, analytic_display, text_format)

			# F: Cantidad
			worksheet.write_number(row, 5, line.product_uom_qty or 0.0, number_format)
			# G: Precio Unitario
			worksheet.write_number(row, 6, line.price_unit or 0.0, number_format)

			row += 1

		# Finalizar y crear attachment
		workbook.close()
		output.seek(0)
		file_data = output.read()
		attachment = self.env['ir.attachment'].create({
			'name': f'{self.name}_SaleOrder.xlsx',
			'datas': base64.b64encode(file_data),
			'type': 'binary',
			'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
		})
		return {
			'type': 'ir.actions.act_url',
			'url': f"/web/content/{attachment.id}?download=1",
			'target': 'self',  # Fuerza descarga directa
		}

	@api.model_create_multi
	def create(self, vals_list):
		if self.env.context.get('from_reviewed_menu') or self.env.context.get('from_presupuesto_menu'):
			raise UserError(_("No se pueden crear pedidos desde 'Presupuestos Revisados' o 'Presupuestos'."))
		return super().create(vals_list)


class SaleOrderLine(models.Model):
	_inherit = 'sale.order.line'

	apu_id = fields.Many2one('apu.apu', string='APU', help='Selecciona un APU para cargar el producto correspondiente')
	show_apu_lines = fields.Boolean(string="Mostrar receta", default=False)

	apu_line_preview_ids = fields.One2many(
		comodel_name='apu.apu.line',
		compute='_compute_apu_line_preview',
		string="Líneas del APU",
		store=False
	)
	order_referencia_analitica = fields.Char(
		related='order_id.referencia_analitica',
		string="Proyecto",
		store=True, index=True, precompute=True)

	def action_ver_apu_popup(self):
		self.ensure_one()
		if not self.apu_id:
			raise UserError(_("No hay un APU vinculado a esta línea."))

		return {
			'name': _('Detalle del APU: %s') % (self.apu_id.display_name,),
			'type': 'ir.actions.act_window',
			'res_model': 'apu.apu.line',
			'view_mode': 'tree',
			'views': [(self.env.ref('construccion_gps.view_apu_line_tree_hierarchy').id, 'tree')],
			'domain': [('bom_id', '=', self.apu_id.id)],
			'target': 'new',
			'context': dict(self.env.context),
		}

	@api.depends('apu_id', 'show_apu_lines')
	def _compute_apu_line_preview(self):
		for line in self:
			if line.show_apu_lines and line.apu_id:
				line.apu_line_preview_ids = line.apu_id.line_ids
			else:
				line.apu_line_preview_ids = False

	def action_toggle_apu_lines(self):
		for rec in self:
			rec.show_apu_lines = not rec.show_apu_lines

	@api.onchange('apu_id')
	def _onchange_apu_id(self):
		"""Cuando se selecciona un APU, carga la plantilla y busca la variante
		por default_code; si no existe, avisa."""
		if not self.apu_id:
			# limpiar si quita el APU
			self.product_template_id = False
			self.product_id = False
			return

		template = self.apu_id.product_tmpl_id
		if not template:
			# seguridad: si no viene plantilla en el APU
			raise UserError(_('El APU seleccionado no tiene plantilla asociada.'))

		# asigno siempre la plantilla
		self.product_template_id = template.id

		# intento buscar la variante usando default_code de la plantilla
		variant = False
		if template.default_code:
			variant = self.env['product.product'].search([
				('default_code', '=', template.default_code),
				('product_tmpl_id', '=', template.id)
			], limit=1)
		self.name = f"{self.apu_id.subcategoria_id.codigo or ''} {self.apu_id.subcategoria_id.name or ''}".strip()
		# si no existe o no había default_code, tomo la primera variante
		if not variant:
			variant = template.product_variant_ids and template.product_variant_ids[0] or False

		if variant:
			self.product_id = variant.id
		else:
			# si aún así no hay variante, limpio y muestro advertencia
			self.product_id = False
			return {
				'warning': {
					'title': _('APU sin variante'),
					'message': _(
						'No se encontró ninguna variante para la plantilla "%s". '
						'Comprueba el código interno (default_code) definido en el producto.'
					) % template.name
				}
			}

	@api.onchange('product_id')
	def _onchange_price_from_apu(self):
		if not self.order_id or not self.product_id:
			return

		if self.order_id.tipo_pedido == 'apu':
			# Buscar el APU relacionado al template
			apu = self.env['apu.apu'].search([
				('product_tmpl_id', '=', self.product_id.product_tmpl_id.id),
				('id', '=', self.apu_id.id)
			], limit=1)

			if apu and apu.total_apu_precio > 0:
				self.price_unit = apu.total_apu_precio
				self.name = f"{self.apu_id.subcategoria_id.codigo or ''} {self.apu_id.subcategoria_id.name or ''}".strip()

	@api.onchange('product_uom_qty')
	def _onchange_qty_keep_apu_price(self):
		if self.order_id.tipo_pedido == 'apu' and self.apu_id:
			apu = self.env['apu.apu'].search([
				('id', '=', self.apu_id.id),
				('product_tmpl_id', '=', self.product_id.product_tmpl_id.id)
			], limit=1)
			if apu and apu.total_apu_precio > 0:
				self.price_unit = apu.total_apu_precio


class MaterialPurchaseRequisition(models.Model):
	_inherit = 'material.purchase.requisition'

	# Relaciona el requisición de compra con el pedido de venta
	sale_order_id = fields.Many2one('sale.order', string='Presupuesto',
	                                help="Asocia esta requisición de compra con un pedido de venta", )
	permite_aprobar = fields.Boolean('Permite Aprobar', default=False)

	def requisition_confirm(self):
		for rec in self:
			if not rec.sale_order_id and not rec.permite_aprobar:
				raise UserError(
					_("No puedes enviar a aprobación una requisición que no está vinculada a un presupuesto."))
			return super(MaterialPurchaseRequisition, self).requisition_confirm()


class MaterialPurchaseRequisitionLine(models.Model):
	_inherit = 'material.purchase.requisition.line'

	sale_order_line_id = fields.Many2one('sale.order.line', string='Línea de Pedido', index=True,
	                                     help="Referencia a la línea del pedido de venta que originó esta línea de requisición", )

	@api.constrains('qty', 'sale_order_line_id', 'product_id', 'analytic_account_id')
	def _check_qty_not_exceed_apux(self):
		for line in self:
			# ------------------------------------------------------
			# 1) Requisición desde presupuesto
			# ------------------------------------------------------
			if line.sale_order_line_id:
				so_line = line.sale_order_line_id
				apu = self.env['apu.apu'].search([
					('product_tmpl_id', '=', so_line.product_id.product_tmpl_id.id)
				], limit=1)
				apl = apu.line_ids.filtered(lambda l: l.product_id == line.product_id) if apu else self.env[
					'apu.apu.line']
				if not apu or not apl:
					print('deberia poner una bandera')
				# raise UserError(_(
				#     "El producto %s no está contemplado en el APU del presupuesto %s."
				# ) % (line.product_id.display_name, so_line.order_id.name))

				total_presupuestado = apl.product_qty * so_line.product_uom_qty
				antes = sum(
					rec.qty for rec in self.search([
						('sale_order_line_id', '=', so_line.id),
						('product_id', '=', line.product_id.id),
					]) if rec.id != line.id
				)
				restante = max(total_presupuestado - antes, 0.0)

				if line.qty > restante:
					raise UserError(_(
						"Has solicitado %s unidades de %s, pero la cantidad máxima presupuestada X es %s."
					) % (
						                float_round(line.qty, 2),
						                line.product_id.display_name,
						                float_round(total_presupuestado, 2),
					                ))
				continue

			# ------------------------------------------------------
			# 2) Requisición “manual”
			# ------------------------------------------------------
			# 2.1) Extraer analytic_distribution (JSON o dict)
			# dist = {}
			# if isinstance(line.analytic_account_id.id)#(line.analytic_distribution, str):
			#     try:
			#         dist = json.loads(line.analytic_distribution)
			#     except ValueError:
			#         dist = {}
			# elif isinstance(line.analytic_distribution, dict):
			#     dist = line.analytic_distribution
			# acct_ids = [int(a) for a in dist.keys()]
			# if not acct_ids:
			#     continue
			# analytic = self.env['account.analytic.account'].browse(acct_ids[0])
			analytic = self.env['account.analytic.account'].browse(line.analytic_account_id.id)
			so = self.env['sale.order'].search([
				('state', 'in', ('generar_proyecto', 'generar_requisiciones')),
				('analytic_account_id', '=', analytic.id),
			], limit=1)
			if not so:
				continue  # si no hay presupuesto, permitimos crear

			# 2.3) Total presupuestado consolidando todas las APU de ese SO
			total_presupuestado = 0.0
			band_existe = 0
			for sol in so.order_line.filtered(lambda l: l.product_id == line.product_id):
				apu = self.env['apu.apu'].search([
					('product_tmpl_id', '=', sol.product_id.product_tmpl_id.id)
				], limit=1)
				for apl in apu.line_ids.filtered(lambda l: l.product_id == line.product_id):
					total_presupuestado += apl.product_qty * sol.product_uom_qty
					band_existe = 1

			if total_presupuestado <= 0 and band_existe != 0:
				# print('comentado por pruebas')
				raise UserError(_(
					"El producto %s no está contemplado en las APU del presupuesto asociado a la cuenta '%s'."
				) % (line.product_id.display_name, analytic.display_name))

			antes = sum(
				rec.qty for rec in self.search([
					('product_id', '=', line.product_id.id),
					('analytic_account_id', '=', analytic.id),
				]) if rec.id != line.id
			)
			restante = max(total_presupuestado - antes, 0.0)

			if line.qty > restante and band_existe != 0:
				raise UserError(_(
					"Para el presupuesto con cuenta analítica '%s':\n"
					"  • Total presupuestado: %s unidades\n"
					"  • Ya solicitado: %s unidades\n"
					"  • Te quedan disponibles: %s unidades"
				) % (
					                analytic.display_name,
					                float_round(total_presupuestado, 2),
					                float_round(antes, 2),
					                float_round(restante, 2),
				                ))


class PurchaseRequest(models.Model):
	_inherit = 'purchase.request'

	# Relaciona el requisición de compra con el pedido de venta
	sale_order_id = fields.Many2one('sale.order', string='Presupuesto',
	                                help="Asocia esta requisición de compra con un pedido de venta", )
	permite_aprobar = fields.Boolean('Permite Aprobar', default=False)

	def button_to_approve(self):
		for rec in self:
			for line in rec.line_ids:
				# debo aumentar la logica de que si la cuenta analitica es antes del 15/jul/2025 lo detenga
				if line.analytic_distribution:
					for analytic_id_str in line.analytic_distribution.keys():
						try:
							analytic_id = int(analytic_id_str)
							# voy a sacar el code y hacerlo numerico para compararlo
							codigo = self.env['account.analytic.account'].browse(analytic_id).code
							try:
								codigo_num = float(codigo)
								if codigo_num > 823 and not rec.permite_aprobar:
									raise ValidationError(
										_("La línea con producto '%s' tiene asignada una cuenta analítica no permitida (Código: %s), debe ser creada la requisicion desde el presupueto.") %
										(line.product_id.display_name, codigo)
									)
							except ValueError:
								pass
						except ValueError:
							continue  # ignorar claves mal formateadas
		# if not rec.sale_order_id and not rec.permite_aprobar :
		# raise UserError(_("No puedes enviar a aprobación una requisición que no está vinculada a un presupuesto."))
		for request in self:
			for line in request.line_ids:
				if not line.sale_order_line_id or not line.product_id:
					continue  # sin línea de venta o sin producto, omitir

				apu_id = line.sale_order_line_id.apu_id.id
				product_id = line.product_id.id

				# Buscar si el producto está en las líneas del APU correspondiente
				apu_line_exists = self.env['apu.apu.line'].search_count([
					('bom_id', '=', apu_id),
					('product_id', '=', product_id)
				]) > 0

				if not apu_line_exists:
					raise ValidationError(
						f"El producto '{line.product_id.display_name}' no está incluido en el APU del pedido de venta asociado. "
						"No se puede aprobar la requisición."
					)
		return super(PurchaseRequest, self).button_to_approve()


class PurchaseRequestLine(models.Model):
	_inherit = 'purchase.request.line'

	sale_order_line_id = fields.Many2one('sale.order.line', string='Línea de Pedido', index=True,
	                                     help="Referencia a la línea del pedido de venta que originó esta línea de requisición", )

	@api.constrains('product_qty', 'sale_order_line_id', 'product_id', 'analytic_distribution')
	def _check_qty_not_exceed_apu(self):
		producto_origen = None
		for line in self:
			# ------------------------------------------------------
			# 1) Requisición desde presupuesto
			# ------------------------------------------------------
			producto_origen = None
			if line.sale_order_line_id:
				so_line = line.sale_order_line_id
				apu = self.env['apu.apu'].search([
					('product_tmpl_id', '=', so_line.product_id.product_tmpl_id.id)
				], limit=1)
				apl = apu.line_ids.filtered(lambda l: l.product_id == line.product_id) if apu else self.env[
					'apu.apu.line']
				producto_origen = None
				if not apu or not apl:
					# no esta contemplado en el apu del presupuesto pero el producto es un opcional
					product_tmpl = line.product_id.product_tmpl_id  # producto que estoy evaluando
					self.env['apu.apu.line'].search([('bom_id', '=', line.sale_order_line_id.apu_id.id)])
					# aqui debe ir a buscar
					producto_origen = None
					tmpl_ids_apu = line.sale_order_line_id.apu_id.line_ids.mapped('product_id.product_tmpl_id')
					for tmpl in tmpl_ids_apu:
						producto_origen = None
						if line.product_id.product_tmpl_id.id in tmpl.optional_product_ids.ids:
							# si lo tengo encontrado obtengo el producto origen
							producto_origen = tmpl
							break  # permitido
					else:
						raise UserError(_(
							"El producto %s no está contemplado en el APU del presupuesto %s ni como opcional"
						) % (line.product_id.display_name, so_line.order_id.name))
				if producto_origen:
					# Buscar línea APU original del producto origen
					apl_origen = apu.line_ids.filtered(lambda l: l.product_id.product_tmpl_id == producto_origen)
					if not apl_origen:
						raise UserError(
							_("No se encontró la línea APU del producto origen: %s") % producto_origen.display_name)
					total_presupuestado = apl_origen.product_qty * so_line.product_uom_qty

					# Sumar requisiciones ya hechas usando el producto opcional (line.product_id),
					# pero que se originaron de la misma línea del pedido
					antes = sum(
						rec.product_qty for rec in self.search([
							('sale_order_line_id', '=', so_line.id),
							('product_id', '=', line.product_id.id),
						]) if rec.id != line.id
					)
				else:
					# Validación directa (producto sí está en APU como línea)
					total_presupuestado = apl.product_qty * so_line.product_uom_qty
					antes = sum(
						rec.product_qty for rec in self.search([
							('sale_order_line_id', '=', so_line.id),
							('product_id', '=', line.product_id.id),
						]) if rec.id != line.id
					)

				restante = max(total_presupuestado - antes, 0.0)

				if line.product_qty > restante:
					# voy a verificar si el producto es un opcional
					raise UserError(_(
						"Has solicitado %s unidades de %s, pero la cantidad máxima presupuestada es %s."
					) % (
						                float_round(line.product_qty, 2),
						                line.product_id.display_name,
						                float_round(total_presupuestado, 2),
					                ))
				continue

			# ------------------------------------------------------
			# 2) Requisición “manual”
			# ------------------------------------------------------
			# 2.1) Extraer analytic_distribution (JSON o dict)
			dist = {}
			if isinstance(line.analytic_distribution, str):
				try:
					dist = json.loads(line.analytic_distribution)
				except ValueError:
					dist = {}
			elif isinstance(line.analytic_distribution, dict):
				dist = line.analytic_distribution
			acct_ids = [int(a) for a in dist.keys()]
			if not acct_ids:
				continue

			analytic = self.env['account.analytic.account'].browse(acct_ids[0])
			so = self.env['sale.order'].search([
				('state', 'in', ('generar_proyecto', 'generar_requisiciones')),
				('analytic_account_id', '=', analytic.id),
			], limit=1)
			if not so:
				continue

			# 2.3) Total presupuestado consolidando todas las APU de ese SO
			total_presupuestado = 0.0
			band_existe = 0
			for sol in so.order_line.filtered(lambda l: l.product_id == line.product_id):
				apu = self.env['apu.apu'].search([
					('product_tmpl_id', '=', sol.product_id.product_tmpl_id.id)
				], limit=1)
				for apl in apu.line_ids.filtered(lambda l: l.product_id == line.product_id):
					band_existe = 1
					total_presupuestado += apl.product_qty * sol.product_uom_qty

			if total_presupuestado <= 0 and band_existe != 0:
				# print('comentado por pruebas')
				raise UserError(_(
					"El producto %s no está contemplado en las APU del presupuesto asociado a la cuenta '%s'."
				) % (line.product_id.display_name, analytic.display_name))
			antes = 0
			if line.display_type != 'line_section':
				self.env.cr.execute("""
                                    SELECT SUM(product_qty)
                                    FROM purchase_request_line
                                    WHERE product_id = %s
                                      AND id != %s
                                      AND analytic_distribution::jsonb ? %s
				                    """, (line.product_id.id, line.id, str(analytic.id)))
				antes = self.env.cr.fetchone()[0] or 0.0
			# antes = sum(
			#     rec.product_qty for rec in self.search([
			#         ('product_id', '=', line.product_id.id),
			#         ('analytic_distribution', '=', analytic.id),
			#     ]) if rec.id != line.id
			# )
			restante = max(total_presupuestado - antes, 0.0)

			if line.product_qty > restante and band_existe != 0:
				raise UserError(_(
					"Para el presupuesto con cuenta analítica '%s':\n"
					"  • Total presupuestado: %s unidades\n"
					"  • Ya solicitado: %s unidades\n"
					"  • Te quedan disponibles: %s unidades"
				) % (
					                analytic.display_name,
					                float_round(total_presupuestado, 2),
					                float_round(antes, 2),
					                float_round(restante, 2),
				                ))