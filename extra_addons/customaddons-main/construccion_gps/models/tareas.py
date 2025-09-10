# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

from odoo import api, fields, models, _, Command
from odoo.exceptions import UserError, ValidationError
from odoo.osv.expression import AND, OR
from odoo.tools import float_round
from odoo.addons import decimal_precision as dp
from collections import defaultdict
import logging  
_logger = logging.getLogger(__name__)

class Tareas(models.Model):
    """ Define las tareas generales """
    _name = 'tareas'
    _description = 'Tareas'
    _inherit = ['portal.mixin', 'mail.thread', 'mail.activity.mixin']


    code = fields.Char('Reference')
    name = fields.Char('Descripcion')
    state = fields.Selection([
        ('activo', 'Activo'),
        ('inactivo', 'Inactivo')], 'Estado',
        default='activo', required=True)
    company_id=fields.Many2one("res.company",string="Compañia",default=lambda self: self.env.company)

    _sql_constraints = [
        ('unique_tarea', 'unique(name)', 'La tarea debe ser única.')
    ]

    @api.model
    def create(self, vals):
        if not vals.get('code'):
            # Generar el código de referencia utilizando la secuencia
            vals['code'] = self.env['ir.sequence'].next_by_code('TAREA') or '/'
            return super(Tareas, self).create(vals)

class TareasApu(models.Model):
    """ Define las tareas correspondientes a las APUS """
    _name = 'apu.tareas'
    _description = 'Apus con Tareas'
    _inherit = ['portal.mixin', 'mail.thread', 'mail.activity.mixin']
    _rec_name = 'product_apu_id'


    product_apu_id = fields.Many2one('apu.apu.line', 'Apu', index=True)
    descripcion = fields.Char('Descripcion')
    company_id = fields.Many2one('res.company', 'Company', index=True, default=lambda self: self.env.company)
    line_ids = fields.One2many('apu.tareas.line', 'apu_tarea_id', 'Tareas Lines', copy=True)

    def export_excel(self):
        """Genera un Excel con:
        - Encabezado: product_apu_id, descripcion, company_id.
        - Tabla: tarea_id, rendimiento, company.
        Los valores numéricos se muestran con dos decimales.
        Se crea un attachment y se retorna una acción URL para descargar el fichero.
        """
        self.ensure_one()
        import io, base64
        import xlsxwriter

        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet("Tareas APU Export")

        # Formatos
        title_format = workbook.add_format({
            'bold': True, 'font_size': 16,
            'align': 'center', 'valign': 'vcenter'
        })
        header_format = workbook.add_format({
            'bold': True, 'bg_color': '#F7F7F7',
            'border': 1, 'align': 'center'
        })
        text_format = workbook.add_format({
            'border': 1, 'align': 'left'
        })
        numeric_format = workbook.add_format({
            'num_format': '#,##0.00', 'border': 1, 'align': 'right'
        })

        # --- Título ---
        # Merge de las celdas A1 a C1 para el título
        worksheet.merge_range(0, 0, 0, 2, "Exportación Tareas APU", title_format)

        # --- Encabezado ---
        # Usamos la fila 2 para etiquetas y la fila 3 para valores
        header_row = 2
        worksheet.write(header_row, 0, "Product APU", header_format)
        worksheet.write(header_row, 1, "Descripción", header_format)
        worksheet.write(header_row, 2, "Compañía", header_format)

        data_row = header_row + 1
        # Utilizamos el nombre del Product Template relacionado
        product_apu_name = self.product_apu_id.product_tmpl_id.name if (
                    self.product_apu_id and self.product_apu_id.product_tmpl_id) else ""
        descripcion = self.descripcion or ""
        company_name = self.company_id.name or ""
        worksheet.write(data_row, 0, product_apu_name, text_format)
        worksheet.write(data_row, 1, descripcion, text_format)
        worksheet.write(data_row, 2, company_name, text_format)

        # --- Tabla de Líneas ---
        # La tabla comienza dos filas después de la cabecera (dejamos una fila vacía)
        table_start_row = data_row + 2
        table_headers = ["Tarea", "Rendimiento", "Compañía"]
        for col, col_header in enumerate(table_headers):
            worksheet.write(table_start_row, col, col_header, header_format)

        row = table_start_row + 1
        for line in self.line_ids:
            tarea_name = line.tarea_id.name if line.tarea_id else ""
            rendimiento_val = line.rendimiento or 0.0
            line_company = line.company_id.name if line.company_id else ""
            worksheet.write(row, 0, tarea_name, text_format)
            worksheet.write_number(row, 1, float(rendimiento_val), numeric_format)
            worksheet.write(row, 2, line_company, text_format)
            row += 1

        # Establecer un ancho uniforme para todas las columnas de la tabla
        uniform_width = 25  # Puedes modificar este valor según tus necesidades
        for col in range(len(table_headers)):
            worksheet.set_column(col, col, uniform_width)

        workbook.close()
        output.seek(0)
        file_data = output.read()

        # Crear el attachment con el fichero Excel
        attachment = self.env['ir.attachment'].create({
            'name': 'TareasAPU.xlsx',
            'datas': base64.b64encode(file_data),
            'type': 'binary',
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        url = "/web/content/%s?download=true" % (attachment.id)
        return {
            'type': 'ir.actions.act_url',
            'url': url,
            'target': 'new',
        }

    # Campos para importar archivos Excel en el mismo registro
    import_file = fields.Binary(string="Archivo Excel")
    import_filename = fields.Char(string="Nombre del Archivo")

    def import_excel(self):
        """Importa datos desde un archivo Excel con el siguiente formato:

        - Encabezado (fila 4, 1-based):
          • Columna A: Product APU (se buscará en 'apu.apu' utilizando product_tmpl_id.name)
          • Columna B: Descripción
          • Columna C: Compañía (se buscará en 'res.company' por nombre)

        - Tabla de líneas (encabezados en la fila 6 y datos desde la fila 7):
          • Columna A: Tarea (se buscará en 'tareas' por nombre)
          • Columna B: Rendimiento
          • Columna C: Compañía (se buscará en 'res.company' por nombre)

        Se actualiza el registro actual con estos datos, reemplazando las líneas existentes.
        """
        self.ensure_one()
        if not self.import_file:
            raise UserError(_("Por favor, cargue un archivo Excel para importar."))

        import io, base64
        try:
            file_data = base64.b64decode(self.import_file)
        except Exception as e:
            raise UserError(_("Error al decodificar el archivo: %s") % e)
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(_("El módulo openpyxl es requerido para importar archivos Excel."))
        try:
            workbook = load_workbook(io.BytesIO(file_data), data_only=True)
        except Exception as e:
            raise UserError(_("No se pudo leer el archivo Excel. Error: %s") % e)
        ws = workbook.active

        # Leer cabecera: usar la fila 4 (ya que en exportación se escribieron los datos en la fila 4)
        product_apu_name = ws.cell(row=4, column=1).value  # Columna A, fila 4
        descripcion = ws.cell(row=4, column=2).value  # Columna B, fila 4
        company_name = ws.cell(row=4, column=3).value  # Columna C, fila 4

        # Buscar el registro de APU usando el nombre del Product Template (como se exportó)
        product_apu_rec = self.env['apu.apu'].search([('product_tmpl_id.name', '=', product_apu_name)], limit=1)
        if not product_apu_rec:
            raise UserError(_("No se encontró el registro de APU con Product Template '%s'.") % product_apu_name)
        # Buscar la compañía en 'res.company'
        company_rec = self.env['res.company'].search([('name', '=', company_name)], limit=1)
        if not company_rec:
            raise UserError(_("No se encontró la compañía con nombre '%s'.") % company_name)

        vals_update = {
            'product_apu_id': product_apu_rec.id,
            'descripcion': descripcion or "",
            'company_id': company_rec.id,
        }

        # Procesar la tabla de líneas: encabezados en fila 6, datos desde la fila 7
        table_header_row = 6
        line_vals = []
        row_idx = table_header_row + 1  # Comienza en fila 7
        while True:
            tarea_name = ws.cell(row=row_idx, column=1).value
            if not tarea_name:
                break  # Fin de datos si la primera columna está vacía
            rendimiento = ws.cell(row=row_idx, column=2).value
            line_company_name = ws.cell(row=row_idx, column=3).value

            # Buscar la tarea en 'tareas' por nombre
            tarea_rec = self.env['tareas'].search([('name', '=', tarea_name)], limit=1)
            if not tarea_rec:
                raise UserError(_("No se encontró la Tarea '%s'.") % tarea_name)
            # Buscar la compañía para la línea (si se indica)
            if line_company_name:
                line_company_rec = self.env['res.company'].search([('name', '=', line_company_name)], limit=1)
                if not line_company_rec:
                    raise UserError(_("No se encontró la compañía '%s' para la línea.") % line_company_name)
            else:
                line_company_rec = False

            vals_line = {
                'tarea_id': tarea_rec.id,
                'rendimiento': rendimiento or 0.0,
                'company_id': line_company_rec.id if line_company_rec else False,
            }
            line_vals.append((0, 0, vals_line))
            row_idx += 1

        vals_update['line_ids'] = line_vals
        self.write(vals_update)
        return {'type': 'ir.actions.client', 'tag': 'reload'}

class TareasApuLine(models.Model):
    """ Define las tareas correspondientes a las APUS """
    _name = 'apu.tareas.line'
    _description = 'Detalle de Apus con Tareas'
    _inherit = ['portal.mixin', 'mail.thread', 'mail.activity.mixin']

    apu_tarea_id = fields.Many2one('apu.tareas', 'Parent Apus', index=True, ondelete='cascade', required=True)
    tarea_id = fields.Many2one('tareas', 'Tareas', required=True, check_company=True)
    tipo_actividad = fields.Selection([
        ('manoobra', 'Mano de Obra'),
        ('material', 'Materiales'),
        ('equipo', 'Equipos'),
        ('transporte', 'Transporte')], 'Tipo Actividad',
        default='material', required=True)
    unidad = fields.Many2one('uom.uom', 'Unidad')
    cantidad = fields.Float('Cantidad', default=1.0, digits=(16,4), required=True)
    tarifa = fields.Float('Tarifa', default=0.0, digits=(16,4), required=True)
    costo_hora_actividad = fields.Float(
        'Costo Hora Actividad',
        compute='_compute_costos',
        digits=(16,4)
    )
    rendimiento = fields.Float('Rendimiento', default=1.0, digits='Activity performance', required=True)
    costo_final_actividad = fields.Float(
        'Costo Final Actividad',
        compute='_compute_costos',
        digits=(16,4)
    )
    company_id = fields.Many2one(related='apu_tarea_id.company_id', store=True, index=True, readonly=True)

    @api.depends('cantidad', 'tarifa', 'rendimiento')
    def _compute_costos(self):
        for rec in self:
            rec.costo_hora_actividad = rec.cantidad * rec.tarifa
            rec.costo_final_actividad = rec.costo_hora_actividad * rec.rendimiento