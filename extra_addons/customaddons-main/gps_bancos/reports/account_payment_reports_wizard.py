# coding: utf-8
from odoo import api, fields, models, exceptions, tools, _
from ...message_dialog.tools import FileManager
from ...calendar_days.tools import DateManager
from ...calendar_days.tools import CalendarManager
dtFile=FileManager()        
dateO=DateManager()
calendarO=CalendarManager()
import openpyxl
import logging
_logger = logging.getLogger(__name__)

from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font, PatternFill
import openpyxl.utils as utils
from openpyxl.styles import NamedStyle

class report_pagos_report_xlsx(models.AbstractModel):
    _name = "report.gps_bancos.report_pagos_report_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Pagos/Cobros"
    
    def create_xlsx_report(self, docids, data):
        EXT="xlsx"
        dir_path=dtFile.get_path_file(__file__)
        new_filename=dtFile.create(EXT)  
        filename=dtFile.join(dir_path,"reporte_pagos.xlsx")
        dtFile.copyfile(filename,new_filename)
        wb=False
        try:
            brw_wizard=self.env["account.payment.reports.wizard"].sudo().browse(docids[-1])

            wb, ws, target = self.open_xlsx(new_filename, load_sheet=False)
            ws = wb["PAGOS"]
            self.create_report_pagos(brw_wizard,ws)
            ws = wb["DETALLE DE PAGOS"]
            self.create_report_report_pagos(brw_wizard, ws)
            wb = self.save_wb(wb, target)
            #########################################
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

    def create_report_pagos(self,brw_wizard,ws):
        date_from = brw_wizard.date_from
        date_to = brw_wizard.date_to
        companys = ', '.join(brw_wizard.mapped('company_ids').mapped('name'))
        company_ids = brw_wizard.company_ids.ids + [-1, -1]
        self._cr.execute(f""";with variables as (
        	SELECT ARRAY{company_ids} AS COMPany_ids,
          '{date_from}'::date as date_from,
          '{date_to}'::date as date_to
        )

        select rc.name as company_name,
        ap.id,
        am.name,
        rp.vat,
        rp.name as partner_name,
        ap.amount,
        ap.amount_residual,
        am.date,
        coalesce(am.ref,'') as ref,
        coalesce(ap.bank_reference,'') as bank_reference,
        case when(ap.is_prepayment) then 'SI' else 'NO' end es_anticipo,
        aj.name as journal_name,
        coalesce(aa.code,'') as codigo_cuenta,
        coalesce(aa.name,'') as nombre_cuenta
        from variables
        inner join account_payment ap on 1=1
        inner join account_move am on am.id=ap.move_id and am.company_id=any(variables.company_ids) 
        	and 
        	(
        		am.date>=variables.date_from	and 
        		am.date<=variables.date_to	
        	)
        inner join res_company rc on rc.id=am.company_id
        inner join res_partner rp on rp.id=ap.partner_id 
        inner join account_journal aj on aj.id=am.journal_id
        left join account_account aa on aa.id=ap.prepayment_account_id  
        where am.state='posted' and ap.payment_type='outbound'
		and coalesce(ap.is_internal_transfer,false)!=true  and ap.reversed_payment_id is null  
        order by rc.name asc,am.date asc,ap.id asc """, (tuple(company_ids), date_from, date_to))
        result = self._cr.dictfetchall()
        if result:
            i, INDEX_ROW = 0, 5
            for each_result in result:
                row = str(INDEX_ROW + i)
                ws['A' + row] = each_result["company_name"]
                ws['B' + row] = each_result["id"]
                ws['C' + row] = each_result["name"]
                ws['D' + row] = each_result["journal_name"]
                ws['E' + row] = each_result["date"]
                ws['F' + row] = each_result["vat"]
                ws['G' + row] = each_result["partner_name"]
                ws['H' + row] = each_result["amount"]
                ws['I' + row] = each_result["amount_residual"]

                ws['J' + row] = each_result["bank_reference"]
                ws['K' + row] = each_result["ref"]
                ws['L' + row] = each_result["es_anticipo"]
                ws['M' + row] = each_result["codigo_cuenta"]
                ws['N' + row] = each_result["nombre_cuenta"]
                i += 1
            last_row = INDEX_ROW + i
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':N' + str(last_row - 1), border)
            ws['A1'] = companys
            ws['B2'] = brw_wizard.date_from
            ws['B3'] = brw_wizard.date_to
            ws['E2'] = len(result)

    def create_report_report_pagos(self,brw_wizard,ws):
        date_from = brw_wizard.date_from
        date_to = brw_wizard.date_to
        companys = ', '.join(brw_wizard.mapped('company_ids').mapped('name'))
        company_ids = brw_wizard.company_ids.ids + [-1, -1]
        self._cr.execute(f"""with variables as (
	SELECT ARRAY{company_ids} AS COMPany_ids,
  '{date_from}'::date as date_from,
  '{date_to}'::date as date_to
),
pagos as (
	select rc.name as company_name,
		ap.id,
		ap.move_id,
		am.name,
		rp.vat,
		rp.name as partner_name,
		ap.amount,
		ap.amount_residual,
		am.date,
		coalesce(am.ref,'') as ref,
		coalesce(ap.bank_reference,'') as bank_reference,
		case when(ap.is_prepayment) then 'SI' else 'NO' end es_anticipo,
		aj.name as journal_name,
		coalesce(aa.code,'') as payment_account_code,
		coalesce(aa.name,'') as payment_account_name,
		ap.purchase_id
		from variables
		inner join account_payment ap on 1=1
		inner join account_move am on am.id=ap.move_id and am.company_id=any(variables.company_ids) 
			and 
			(
				am.date>=variables.date_from	and 
				am.date<=variables.date_to	
			)
		inner join res_company rc on rc.id=am.company_id
		inner join res_partner rp on rp.id=ap.partner_id 
		inner join account_journal aj on aj.id=am.journal_id
		left join account_account aa on aa.id=ap.prepayment_account_id  
		where am.state='posted' and ap.payment_type='outbound'
		and coalesce(ap.is_internal_transfer,false)!=true  and ap.reversed_payment_id is null 
),
aplicacion_documentos as (
	   SELECT
		'move,'||fact.id as doc_id,AM.id as move_id, 
			ROUND(SUM(APR.AMOUNT),2) AS aplicado 
		FROM ACCOUNT_PARTIAL_rECONCILE APR  
		INNER JOIN ACCOUNT_MOVE_LINE AMLD ON AMLD.ID=APR.DEBIT_MOVE_ID 
		INNER JOIN ACCOUNT_MOVE AM ON AM.ID=AMLD.MOVE_ID 
		INNER JOIN account_payment p ON p.move_id=AM.id  
		INNER JOIN ACCOUNT_MOVE_LINE factl ON factl.ID=APR.CREDIT_MOVE_ID 
		INNER JOIN ACCOUNT_MOVE fact ON fact.ID=factl.MOVE_ID and not coalesce(fact.prepayment_Assignment,false)
		GROUP BY 'move,'||fact.id,AM.id 
union
SELECT
		'move,'||fact.id as doc_id,AM.id as move_id, 
			ROUND(SUM(aprpg.AMOUNT),2) AS aplicado 
		FROM ACCOUNT_PARTIAL_rECONCILE APR  
		INNER JOIN ACCOUNT_MOVE_LINE AMLD ON AMLD.ID=APR.DEBIT_MOVE_ID 
		INNER JOIN ACCOUNT_MOVE AM ON AM.ID=AMLD.MOVE_ID 
		INNER JOIN account_payment p ON p.move_id=AM.id  
		INNER JOIN ACCOUNT_MOVE_LINE moveantl ON moveantl.ID=APR.CREDIT_MOVE_ID 
		INNER JOIN ACCOUNT_MOVE moveant ON moveant.ID=moveantl.MOVE_ID and 
			coalesce(moveant.prepayment_Assignment,false) 
		INNER JOIN ACCOUNT_MOVE_LINE movepgl ON movepgl.move_id=moveant.id and movepgl.account_id!= moveantl.account_id
		/*hasta aqui relacion con asiento de contrapartida y cyenta a cruzar*/
		inner join ACCOUNT_PARTIAL_rECONCILE aprpg on aprpg.debit_move_id=movepgl.id
		INNER JOIN ACCOUNT_MOVE_LINE facl ON facl.ID=aprpg.credit_move_id 
		INNER JOIN ACCOUNT_MOVE fact ON fact.ID=facl.MOVE_ID 
		GROUP BY 'move,'||fact.id,AM.id 
			union
		SELECT
			'purchase,'||PO.id as doc_id,AM.id as move_id, 
				ROUND(SUM(AP.AMOUNT),2) AS aplicado 
			FROM pagos AP
			inner join account_move am on am.id=ap.move_id 
			INNER JOIN PURCHASE_ORDER PO ON PO.ID=AP.PURCHASE_ID
			LEFT JOIN ACCOUNT_PAYMENT_TERM APT ON APT.ID=po.PAYMENT_TERM_ID	
			left join ( 
				SELECT AM.id as move_id 
                    FROM ACCOUNT_PARTIAL_rECONCILE APR  
                    INNER JOIN ACCOUNT_MOVE_LINE AMLD ON AMLD.ID=APR.DEBIT_MOVE_ID 
                    INNER JOIN ACCOUNT_MOVE AM ON AM.ID=AMLD.MOVE_ID 
                    INNER JOIN pagos p ON p.move_id=AM.id  
                    INNER JOIN ACCOUNT_MOVE_LINE factl ON factl.ID=APR.CREDIT_MOVE_ID 
                    INNER JOIN ACCOUNT_MOVE fact ON fact.ID=factl.MOVE_ID  
                group by AM.id 
			) x on x.move_id=am.id
			WHERE am.STATE='posted' AND PO.STATE in ('done','purchase')
			and x.move_id is null 
		GROUP BY 'purchase,'||PO.id,AM.id
),
facturas_oc as (
    select ad.doc_id, 
		string_agg(distinct po.id::text, ', ') AS order_ids,
		string_agg(distinct po.name, ', ') AS order_names , 
		string_agg(DISTINCT TO_CHAR(po.date_order, 'YYYY-MM-DD'), ', ') AS order_dates, 
		string_agg(distinct APT.name::JSON->>'es_EC'::varchar, ', ') AS order_payment_Terms
	from 
	aplicacion_documentos ad
	INNER JOIN account_move_line aml ON 'move,'||aml.move_id = ad.doc_id
	INNER JOIN purchase_order_line pol ON pol.id = aml.purchase_line_id
	INNER JOIN purchase_order po ON po.id = pol.order_id 
	LEFT JOIN ACCOUNT_PAYMENT_TERM APT ON APT.ID=po.PAYMENT_TERM_ID	
	where ad.doc_id ilike 'move,%%'
	group by ad.doc_id
),
documentos as (
	select af.doc_id,
		aj.name as doc_journal_name,
		am.name as doc_name,
		am.date as doc_date ,
	coalesce(apt.name::JSON->>'es_EC'::varchar,'') as doc_termino_pago   
	from aplicacion_documentos af
	inner join account_move am on 'move,'||am.id=af.doc_id
	inner join account_journal aj on aj.id=am.journal_id
	LEFT JOIN ACCOUNT_PAYMENT_TERM APT ON APT.ID=AM.INVOICE_PAYMENT_TERM_ID 
	group by af.doc_id,aj.name,am.name,am.date ,coalesce(apt.name::JSON->>'es_EC'::varchar,'')
	union 
	select af.doc_id,
		'' as doc_journal_name,
		PO.name as doc_name,
		PO.date_order::date as doc_date ,
	    coalesce(apt.name::JSON->>'es_EC'::varchar,'') as doc_termino_pago
	FROM aplicacion_documentos af	
	INNER JOIN PURCHASE_ORDER PO ON 'purchase,'||PO.ID=af.doc_id
	LEFT JOIN ACCOUNT_PAYMENT_TERM APT ON APT.ID=po.PAYMENT_TERM_ID	  
	group by af.doc_id ,po.name,po.date_order::date,
		coalesce(apt.name::JSON->>'es_EC'::varchar,'')
	
),CLASIFICACION_CTAS AS (                              
 	select agrl.name,aat.code,AAT.company_id,aat.id
		from account_Group_report agr
		inner join variables on 1=1
		inner join account_group_report_line agrl on agrl.report_id=agr.id
		inner join report_template_account_acc_rel rtpl on rtpl.report_line_id=agrl.id 
		inner join account_account  aat on aat.id=rtpl.account_id
		where agr.CODE='FLUJO' and aat.company_id = ANY (variables.company_ids )                                                                   
) ,
CTAS_GASTOS as (
        SELECT doc.doc_id, 
				AA.NAME AS CUENTA_NAME,
				AA.CODE AS CUENTA_CODE,
				C.name AS CLASIFICACION ,
		SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY doc.doc_id) AS TOTAL,
		case when(SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY doc.doc_id)>0) 
		then SUM(AML.PRICE_TOTAL)/SUM(SUM(AML.PRICE_TOTAL)) 
		OVER(PARTITION BY doc.doc_id) else 1.00 end  as percentage 
			
        	FROM documentos doc
        	INNER JOIN ACCOUNT_MOVE_LINE AML ON 'move,'||AML.MOVE_ID=doc.doc_id
        	INNER JOIN ACCOUNT_ACCOUNT AA ON AA.ID=AML.ACCOUNT_ID 	AND 
        		AA.ACCOUNT_TYPE != 'liability_payable'  
        		AND AML.DEBIT>0 
        	LEFT JOIN CLASIFICACION_CTAS C ON C.id=AA.ID 
        	GROUP BY doc.doc_id,AA.NAME,AA.CODE,C.name
        union
        SELECT ARQ.DOC_ID,COALESCE(AA.NAME,AAC.NAME) AS CUENTA_NAME,
                    COALESCE(AA.CODE,AAC.CODE) AS CUENTA_CODE,C.name AS CLASIFICACION ,
                    
                    SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.DOC_ID) AS TOTAL,
		case when(SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.DOC_ID)>0) 
		then SUM(AML.PRICE_TOTAL)/SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.DOC_ID)
			else 1.00 end  as percentage 

        	FROM documentos ARQ
        	INNER JOIN PURCHASE_ORDER_LINE AML ON 'purchase,'||AML.ORDER_ID=ARQ.DOC_ID AND AML.PRODUCT_ID IS NOT NULL 
        	INNER JOIN PURCHASE_ORDER PO ON PO.ID=AML.ORDER_ID 
			INNER JOIN PRODUCT_PRODUCT PP ON PP.ID=AML.PRODUCT_ID 
			INNER JOIN PRODUCT_TEMPLATE PT ON PT.ID=PP.PRODUCT_TMPL_ID 
			
			LEFT JOIN IR_PROPERTY IPPT ON IPPT.RES_ID=('product.template,'||PT.id )::VARCHAR
				AND IPPT.TYPE='many2one' AND IPPT.name='property_account_expense_id'
				AND IPPT.COMPANY_ID=PO.COMPANY_ID 
			
			LEFT JOIN ACCOUNT_ACCOUNT AA ON ('account.account,'||AA.ID)::VARCHAR=IPPT.VALUE_REFERENCE::VARCHAR AND IPPT.ID IS NOT NULL AND AA.COMPANY_ID=PO.COMPANY_ID 

	

			LEFT JOIN IR_PROPERTY IPPTC ON IPPTC.RES_ID=('product.category,'||PT.CATEG_ID )::VARCHAR
				AND IPPTC.TYPE='many2one' AND IPPTC.name='property_account_expense_categ_id'
				AND IPPTc.COMPANY_ID=PO.COMPANY_ID 
		LEFT JOIN ACCOUNT_ACCOUNT AAC ON ('account.account,'||AAC.ID)::VARCHAR=IPPTC.VALUE_REFERENCE::VARCHAR AND IPPTC.ID IS NOT NULL AND AAC.COMPANY_ID=PO.COMPANY_ID 

			LEFT JOIN CLASIFICACION_CTAS C ON C.id=COALESCE(AA.ID ,AAC.ID) 
			
				group by ARQ.DOC_ID,COALESCE(AA.NAME,AAC.NAME),COALESCE(AA.CODE,AAC.CODE),C.name
) 


select p.*, f.doc_journal_name,
	f.doc_name ,
	f.doc_date,
	f.doc_termino_pago,
	af.doc_id,
	af.aplicado,
	gtos.total,
	case when(af.move_id is null) then 1.00  else coalesce(gtos.percentage,1.00) end as percentage,
	case when(af.move_id is null) then p.amount 
	    else coalesce(gtos.percentage,1.00)*af.aplicado end as proporcional_aplicado,
	gtos.cuenta_name,gtos.cuenta_code,
	gtos.CLASIFICACION,
	foc.order_ids,
	foc.order_names , 
	foc.order_dates, 
	foc.order_payment_terms 		
from pagos p
left join APLICACION_documentos af on af.move_id=p.move_id
left join ctas_Gastos gtos on  gtos.doc_id=af.doc_id
left join documentos f on f.doc_id=af.doc_id
left join facturas_oc foc on foc.doc_id=af.doc_id and foc.doc_id=f.doc_id 
order by p.company_name asc,p.date asc,p.id asc,af.move_id asc  
""", (tuple(company_ids), date_from, date_to))
        result = self._cr.dictfetchall()
        if result:
            i, INDEX_ROW = 0, 5
            for each_result in result:
                row = str(INDEX_ROW + i)
                ws['A' + row] = each_result["company_name"]
                ws['B' + row] = each_result["id"]
                ws['C' + row] = each_result["name"]
                ws['D' + row] = each_result["journal_name"]
                ws['E' + row] = each_result["date"]

                ws['F' + row] = each_result["vat"]

                ws['G' + row] = each_result["partner_name"]
                ws['H' + row] = each_result["amount"]
                ws['I' + row] = each_result["amount_residual"]

                ws['J' + row] = each_result["bank_reference"]
                ws['K' + row] = each_result["ref"]
                ws['L' + row] = each_result["es_anticipo"]
                ws['M' + row] = each_result["payment_account_code"]
                ws['N' + row] = each_result["payment_account_name"]
                type_id=each_result["doc_id"] or ""
                if type_id.startswith("move,"):
                    ws['O' + row] = type_id.replace("move,","")
                    ws['P' + row] = each_result["doc_name"]
                    ws['Q' + row] = each_result["doc_date"]
                    ws['R' + row] = each_result["doc_journal_name"]
                    ws['S' + row] = each_result["doc_termino_pago"]
                    ws['T' + row] = each_result["order_ids"]
                    ws['U' + row] = each_result["order_names"]
                    ws['V' + row] = each_result["order_dates"]
                    ws['W' + row] = each_result["order_payment_terms"]
                if type_id.startswith("purchase,"):
                    ws['T' + row] = type_id.replace("purchase,","")
                    ws['U' + row] = each_result["doc_name"]
                    ws['V' + row] = each_result["doc_date"]
                    ws['W' + row] = each_result["doc_termino_pago"]

                ws['X' + row] = each_result["clasificacion"]
                ws['Y' + row] = each_result["cuenta_code"]
                ws['Z' + row] = each_result["cuenta_name"]
                ws['AA' + row] = each_result["aplicado"]
                ws['AB' + row] = each_result["percentage"]
                ws['AC' + row] =  each_result["proporcional_aplicado"]


                i += 1
            last_row = INDEX_ROW + i
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':AC' + str(last_row - 1), border)
            ws['A1'] = companys
            ws['B2'] = brw_wizard.date_from
            ws['B3'] = brw_wizard.date_to
            ws['E2'] = len(result)

class report_oc_pagos_report_xlsx(models.AbstractModel):
    _name = "report.gps_bancos.report_oc_pagos_report_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de OC con Pagos"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)
        filename = dtFile.join(dir_path, "reporte_oc_pagos.xlsx")
        dtFile.copyfile(filename, new_filename)
        wb = False
        try:
            brw_wizard = self.env["account.payment.reports.wizard"].sudo().browse(docids[-1])

            wb, ws, target = self.open_xlsx(new_filename, load_sheet=False)
            ws = wb["OC_PAGOS"]
            self.create_report_report_oc_pagos(brw_wizard, ws)
            wb = self.save_wb(wb, target)
            #########################################
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

    def create_report_report_oc_pagos(self, brw_wizard, ws):
        date_from = brw_wizard.date_from
        date_to = brw_wizard.date_to
        companys = ', '.join(brw_wizard.mapped('company_ids').mapped('name'))
        company_ids = brw_wizard.company_ids.ids + [-1, -1]
        self._cr.execute(f"""WITH variables AS (
    SELECT ARRAY{company_ids} AS company_ids,
           '{date_from}'::date AS date_from,
           '{date_to}'::date AS date_to
)

SELECT
    po.id AS order_id,
    po.name AS order_name,
    po.date_approve::date AS date_approve,
    rc.name AS company_name,
    rp.id AS partner_id,
    rp.vat,
    rp.name AS partner_name,
    po.amount_untaxed AS subtotal,
    po.amount_tax AS iva,
    po.amount_total AS total_oc,
    po.total_payments_advances,
    po.total_dif_payments_advances,

    ap.id AS id_pago,
    am.name AS pago,
    am.date::date AS fecha_pago,
    am.amount_total AS pago_individual,

    -- Pago acumulado tipo kardex por OC
    SUM(am.amount_total) OVER (PARTITION BY po.id ORDER BY am.date, am.id ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW) 
    AS pago_acumulado_oc,
    round(po.amount_untaxed-SUM(am.amount_total) OVER (PARTITION BY po.id ORDER BY am.date, am.id ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW) 
    ,2) AS dif_pago_acumulado_oc,
    case when(am.date is not null) then row_number() OVER (PARTITION BY po.id ORDER BY am.date, am.id  asc ) else null end as secuencia_pago

FROM variables
INNER JOIN purchase_order po 
    ON po.company_id = ANY(variables.company_ids)
    AND po.date_approve::date BETWEEN variables.date_from AND variables.date_to
INNER JOIN res_company rc ON rc.id = po.company_id
INNER JOIN res_partner rp ON rp.id = po.partner_id
LEFT JOIN purchase_order_payment_line popl ON popl.order_id = po.id
LEFT JOIN account_payment ap ON ap.id = popl.payment_id
LEFT JOIN account_move am ON am.id = ap.move_id
WHERE po.state IN ('done', 'purchase')
ORDER BY po.name, am.date;


  """, (company_ids, date_from, date_to))
        result = self._cr.dictfetchall()
        if result:
            i, INDEX_ROW = 0, 5
            for each_result in result:
                row = str(INDEX_ROW + i)
                ws['A' + row] = each_result["company_name"]
                ws['B' + row] = each_result["order_id"]
                ws['C' + row] = each_result["order_name"]
                ws['D' + row] = each_result["date_approve"]
                ws['E' + row] = each_result["partner_id"]
                ws['F' + row] = each_result["vat"]
                ws['G' + row] = each_result["partner_name"]
                ws['H' + row] = each_result["subtotal"]
                ws['I' + row] = each_result["iva"]

                ws['J' + row] = each_result["total_oc"]
                ws['K' + row] = each_result["total_payments_advances"]
                ws['L' + row] = each_result["total_dif_payments_advances"]

                ws['M' + row] = each_result["secuencia_pago"]

                ws['N' + row] = each_result["id_pago"]
                ws['O' + row] = each_result["pago"]
                ws['P' + row] = each_result["fecha_pago"]
                ws['Q' + row] = each_result["pago_individual"]

                ws['R' + row] = each_result["pago_acumulado_oc"]
                ws['S' + row] = each_result["dif_pago_acumulado_oc"]

                i += 1
            last_row = INDEX_ROW + i
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':S' + str(last_row - 1), border)
            ws['A1'] = companys
            ws['B2'] = brw_wizard.date_from
            ws['B3'] = brw_wizard.date_to
            ws['E2'] = len(result)

class report_soli_pagos_report_xlsx(models.AbstractModel):
    _name = "report.gps_bancos.report_soli_pagos_report_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Solicitudes de Pagos"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)
        filename = dtFile.join(dir_path, "reporte_solicitudes_pagos.xlsx")
        dtFile.copyfile(filename, new_filename)
        wb = False
        try:
            brw_wizard = self.env["account.payment.reports.wizard"].sudo().browse(docids[-1])

            wb, ws, target = self.open_xlsx(new_filename, load_sheet=False)
            ws = wb["SOLICITUDES_PAGOS"]
            self.create_report_report_solicitudes_pagos(brw_wizard, ws)
            wb = self.save_wb(wb, target)
            #########################################
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

    def create_report_report_solicitudes_pagos(self, brw_wizard, ws):
        date_from = brw_wizard.date_from
        date_to = brw_wizard.date_to
        companys = ', '.join(brw_wizard.mapped('company_ids').mapped('name'))
        company_ids = brw_wizard.company_ids.ids + [-1, -1]
        self._cr.execute(f"""WITH variables AS (
    SELECT ARRAY{company_ids} AS company_ids,
           '{date_from}'::date AS date_from,
           '{date_to}'::date AS date_to
)

SELECT
   rc.name as company_name,
 case when(apr.origin='automatic') then 'AUTOMATICO' else 'MANUAL' end as origen,
  case when( apr.type='account.move') then 'ASIENTO/FACTURA'
  	when( apr.type='purchase.order') then 'ORDEN DE COMPRA'
  else 'SOLICITUD' end as tipo,
   case when(apr.type_Document='document') then 'DOCUMENTO' else 'CUOTA' END as tipo_documento,
   rp.vat as identificacion_proveedor,
   rp.name as proveedor,
   RP.ID AS id_proveedor,
   apr.description_motive as motivo,
   apr.id as request_id,
   apr.quota as cuota,
   apr.date_maturity as vencimiento,
   apr.date as fecha_solicitud,
   apr.amount as valor,
   apr.paid as pagado,
   apr.pending as pendiente,

    CASE 
        WHEN apr.state = 'draft' THEN 'Preliminar'
        WHEN apr.state = 'confirmed' THEN 'Confirmado'
        WHEN apr.state = 'done' THEN 'Realizado'
        WHEN apr.state = 'cancelled' THEN 'Anulado'
        WHEN apr.state = 'locked' THEN 'Bloqueado'
        ELSE 'Desconocido'
    END AS estado
	
FROM variables
INNER JOIN account_payment_request apr 
    ON apr.company_id = ANY(variables.company_ids)
    AND apr.date::date BETWEEN variables.date_from AND variables.date_to
INNER JOIN res_company rc ON rc.id = apr.company_id
INNER JOIN res_partner rp ON rp.id = apr.partner_id
WHERE apr.state IN ('draft', 'confirmed','done','locked') 
  """, (company_ids, date_from, date_to))
        result = self._cr.dictfetchall()
        if result:
            i, INDEX_ROW = 0, 5
            for each_result in result:
                row = str(INDEX_ROW + i)
                ws['A' + row] = each_result["company_name"]
                ws['B' + row] = each_result["origen"]
                ws['C' + row] = each_result["tipo"]
                ws['D' + row] = each_result["tipo_documento"]
                ws['E' + row] = each_result["identificacion_proveedor"]

                ws['F' + row] = each_result["id_proveedor"]
                ws['G' + row] = each_result["identificacion_proveedor"]
                ws['H' + row] = each_result["proveedor"]

                ws['I' + row] = each_result["vencimiento"]
                ws['J' + row] = each_result["fecha_solicitud"]

                ws['K' + row] = each_result["valor"]
                ws['L' + row] = each_result["pagado"]
                ws['M' + row] = each_result["pendiente"]

                ws['N' + row] = each_result["estado"]

                i += 1
            last_row = INDEX_ROW + i
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':N' + str(last_row - 1), border)
            ws['A1'] = companys
            ws['B2'] = brw_wizard.date_from
            ws['B3'] = brw_wizard.date_to
            ws['E2'] = len(result)

class report_resumen_pagos_report_xlsx(models.AbstractModel):
    _name = "report.gps_bancos.report_resumen_pagos_report_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Resumen de OC con Pagos"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)
        filename = dtFile.join(dir_path, "reporte_resumen_oc_pagos.xlsx")
        dtFile.copyfile(filename, new_filename)
        wb = False
        try:
            brw_wizard = self.env["account.payment.reports.wizard"].sudo().browse(docids[-1])

            wb, ws, target = self.open_xlsx(new_filename, load_sheet=False)
            ws = wb["RESUMEN_OC_PAGOS"]
            self.create_report_report_oc_pagos(brw_wizard, ws)
            wb = self.save_wb(wb, target)
            #########################################
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

    def create_report_report_oc_pagos(self, brw_wizard, ws):
        date_from = brw_wizard.date_from
        date_to = brw_wizard.date_to
        companys = ', '.join(brw_wizard.mapped('company_ids').mapped('name'))
        company_ids = brw_wizard.company_ids.ids + [-1, -1]
        self._cr.execute(f"""WITH variables AS (
    SELECT ARRAY{company_ids} AS company_ids,
           '{date_from}'::date AS date_from,
           '{date_to}'::date AS date_to
) ,anticipos as (
	select pl.order_id,sum(pl.amount) as amount,count(1) as num_pagos
		from purchase_order_payment_line PL 
		inner join purchase_order po on po.id=pl.order_id
		inner join variables v on po.company_id = any(v.company_ids)
		where po.date_approve::date>=v.date_from and po.date_approve::date<=v.date_to
		group by pl.order_id
)


select 
rc.name as company_name,
po.id as order_id,
po.name as order_name,
rp.id as partner_id,
rp.vat as partner_vat,
rp.name as partner_name,
po.date_approve::date as date_approve,
coalesce(apt.name::JSON->>'es_EC'::varchar,'') as payment_term_name,
po.date_advance_payment as fecha_primer_anticipo,
count(pr.id) as conteo_solicitudes,
po.amount_untaxed as subtotal,
po.amount_tax  as iva,
po.amount_total as total_oc,
sum(pr.amount) as solicitudes,
round(po.amount_untaxed-sum(pr.amount),2) as dif_solicitudes,
a.amount as anticipado ,
a.num_pagos as num_pagos ,
round(po.amount_untaxed-a.amount,2) as dif_anticipos
from purchase_order po 
inner join variables v on po.company_id = any(v.company_ids)
inner join res_company rc on rc.id=po.company_id 
inner join res_partner rp on rp.id=po.partner_id
left join account_payment_request pr on pr.order_id=po.id
left join anticipos a on a.order_id=po.id  
left join account_payment_term apt on apt.id=po.payment_term_id 
where po.amount_total>0.00 and po.state in ('done','purchase') 
	and (po.date_advance_payment is not null)
	and
		( po.date_approve::date>=v.date_from and po.date_approve::date<=v.date_to)
	
group by rc.name,po.id,
po.name,
rp.id,
rp.vat,
rp.name,po.date_approve::date,apt.name,a.amount,a.num_pagos	 """, (company_ids, date_from, date_to))
        result = self._cr.dictfetchall()
        if result:
            i, INDEX_ROW = 0, 5
            for each_result in result:
                row = str(INDEX_ROW + i)
                ws['A' + row] = each_result["company_name"]
                ws['B' + row] = each_result["order_id"]
                ws['C' + row] = each_result["order_name"]
                ws['D' + row] = each_result["date_approve"]
                ws['E' + row] = each_result["partner_id"]
                ws['F' + row] = each_result["partner_vat"]
                ws['G' + row] = each_result["partner_name"]
                ws['H' + row] = each_result["payment_term_name"]
                ws['I' + row] = each_result["fecha_primer_anticipo"]
                ws['J' + row] = each_result["conteo_solicitudes"]

                ws['K' + row] = each_result["subtotal"]
                ws['L' + row] = each_result["iva"]

                ws['M' + row] = each_result["total_oc"]
                ws['N' + row] = each_result["solicitudes"]
                ws['O' + row] = each_result["dif_solicitudes"]

                ws['P' + row] = each_result["anticipado"]
                ws['Q' + row] = each_result["num_pagos"]
                ws['R' + row] = each_result["dif_anticipos"]



                i += 1
            last_row = INDEX_ROW + i
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':R' + str(last_row - 1), border)
            ws['A1'] = companys
            ws['B2'] = brw_wizard.date_from
            ws['B3'] = brw_wizard.date_to
            ws['E2'] = len(result)

class report_resumen_pagos_macros_report_xlsx(models.AbstractModel):
    _name = "report.gps_bancos.report_resumen_pagos_macros_report_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Pagos con Macros"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)
        filename = dtFile.join(dir_path, "reporte_macros_bancos.xlsx")
        dtFile.copyfile(filename, new_filename)
        wb = False
        try:
            brw_wizard = self.env["account.payment.reports.wizard"].sudo().browse(docids[-1])

            wb, ws, target = self.open_xlsx(new_filename, load_sheet=False)
            ws = wb["PAGOS_MACROS"]
            self.create_report_pagos_macros(brw_wizard, ws)
            ########
            ws = wb["DETALLE_PAGOS_MACROS"]
            self.create_report_detalle_pagos_macros(brw_wizard, ws)
            wb = self.save_wb(wb, target)
            #########################################
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

    def create_report_detalle_pagos_macros(self, brw_wizard, ws):
        date_from = brw_wizard.date_from
        date_to = brw_wizard.date_to
        companys = ', '.join(brw_wizard.mapped('company_ids').mapped('name'))
        company_ids = brw_wizard.company_ids.ids + [-1, -1]
        self._cr.execute(f""";with variables as (
    	select ARRAY{company_ids}:: int[] AS company_ids, 
    		{brw_wizard.weeks or 0}::int as weeks
    ),
   resumen_pagos as (
	select apbk.id,
		apbkl.id as bank_line_id,
		rc.name as company_name,
		aj.name as journal_name,
		apbk.name as pbk_name,
		apbk.date_request,
		apbk.date_payment,
		apbk.number_week,
		apbk.ref,
		apbkl.amount,
		apbkl.is_prepayment,
		apbkl.ref,
		apbkl.comments, 
		aa.code as account_code,
		aa.name as account_name,
		apr.invoice_id,
		apr.invoice_line_id,
		apr.order_id,
		coalesce(am.name,po.name,'SIN DOCUMENTO') as num_documento ,
		appm.name as payment_documento  ,
		rp.vat as partner_Vat,
		rp.name as partner_name 
		from variables
		inner join account_payment_bank_macro apbk on apbk.company_id = ANY(variables.company_ids)
			and (
				(variables.weeks=0) or 
				(variables.weeks!=0 and apbk.number_week=variables.weeks)
			)
		inner join res_company rc on rc.id=apbk.company_id
		inner join account_payment_bank_macro_line apbkl on apbk.id=apbkl.bank_macro_id
		inner join account_payment ap on ap.id=apbkl.payment_id and ap.reversed_payment_id is null 
		inner join res_partner rp on rp.id=ap.partner_id 
		inner join account_journal aj on aj.id=apbk.journal_id
		inner join account_payment_request apr on apr.id=apbkl.request_id 		 
		left join account_account aa on aa.id=apbkl.payment_account_id
		left join account_move am on am.id=apr.invoice_id
		left join purchase_order po on po.id=apr.ordeR_id 
		 
		left join account_move appm on appm.id=ap.move_id 
		
		where apbk.state='done'  
) ,

     CTAS_GASTOS as (
        	SELECT ARQ.invoice_id , 
			AA.NAME AS CUENTA_NAME,
			AA.CODE AS CUENTA_CODE,
			SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.invoice_id) AS TOTAL, 
			case when(SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.invoice_id)>0)  then SUM(AML.PRICE_TOTAL)/SUM(SUM(AML.PRICE_TOTAL)) OVER(PARTITION BY ARQ.invoice_id) else 1.00 end  as percentage 
        	FROM resumen_pagos ARQ
        	INNER JOIN ACCOUNT_MOVE_LINE AML ON AML.MOVE_ID=ARQ.invoice_id 
        	INNER JOIN ACCOUNT_ACCOUNT AA ON AA.ID=AML.ACCOUNT_ID 	AND 
        		AA.ACCOUNT_TYPE != 'liability_payable' AND AML.PRODUCT_ID IS NOT NULL 
        		AND AML.DEBIT>0  
        	GROUP BY ARQ.invoice_id,AA.NAME,AA.CODE

        )	 ,CTAS_GASTOS_COMPRAS as (
        
        SELECT ARQ.ORDER_ID,COALESCE(AA.NAME,AAC.NAME) AS CUENTA_NAME,
                    COALESCE(AA.CODE,AAC.CODE) AS CUENTA_CODE,
                    
                    SUM(SUM(POl.PRICE_TOTAL)) OVER(PARTITION BY ARQ.ORDER_ID) AS TOTAL,
		case when(SUM(SUM(POl.PRICE_TOTAL)) OVER(PARTITION BY ARQ.ORDER_ID)>0) 
		then SUM(POl.PRICE_TOTAL)/SUM(SUM(POl.PRICE_TOTAL)) OVER(PARTITION BY ARQ.ORDER_ID)
			else 1.00 end  as percentage 

        	FROM resumen_pagos ARQ 
        	INNER JOIN PURCHASE_ORDER PO ON PO.ID=ARQ.ORDER_ID 
			INNER JOIN PURCHASE_ORDER_line POl ON POl.ORDER_ID=PO.ID 
			INNER JOIN PRODUCT_PRODUCT PP ON PP.ID=POl.PRODUCT_ID 
			INNER JOIN PRODUCT_TEMPLATE PT ON PT.ID=PP.PRODUCT_TMPL_ID 
			
			LEFT JOIN IR_PROPERTY IPPT ON IPPT.RES_ID=('product.template,'||PT.id )::VARCHAR
				AND IPPT.TYPE='many2one' AND IPPT.name='property_account_expense_id'
				AND IPPT.COMPANY_ID=PO.COMPANY_ID 
			
			LEFT JOIN ACCOUNT_ACCOUNT AA ON ('account.account,'||AA.ID)::VARCHAR=IPPT.VALUE_REFERENCE::VARCHAR AND IPPT.ID IS NOT NULL AND AA.COMPANY_ID=PO.COMPANY_ID 

	

			LEFT JOIN IR_PROPERTY IPPTC ON IPPTC.RES_ID=('product.category,'||PT.CATEG_ID )::VARCHAR
				AND IPPTC.TYPE='many2one' AND IPPTC.name='property_account_expense_categ_id'
				AND IPPTc.COMPANY_ID=PO.COMPANY_ID 
		LEFT JOIN ACCOUNT_ACCOUNT AAC ON ('account.account,'||AAC.ID)::VARCHAR=IPPTC.VALUE_REFERENCE::VARCHAR AND IPPTC.ID IS NOT NULL AND AAC.COMPANY_ID=PO.COMPANY_ID 

			 		group by ARQ.ORDER_ID,COALESCE(AA.NAME,AAC.NAME),COALESCE(AA.CODE,AAC.CODE) 
        	
        
        )


select rp.*,
coalesce(cg.CUENTA_CODE,cgc.CUENTA_CODE,'') as CUENTA_CODE,
coalesce(cg.CUENTA_NAME,cgc.CUENTA_NAME,'') as CUENTA_NAME,
coalesce(cg.percentage,cgc.percentage,1.00) as percentage,
coalesce(cg.percentage,cgc.percentage,1.00)*rp.amount as proporcional_total
from resumen_pagos rp
 left join CTAS_GASTOS cg on cg.invoice_id=rp.invoice_id 
 left join CTAS_GASTOS_COMPRAS cgc on cgc.ORDER_ID=rp.order_id  """, (company_ids, date_from, date_to))
        result = self._cr.dictfetchall()
        if result:
            i, INDEX_ROW = 0, 4
            for each_result in result:
                row = str(INDEX_ROW + i)
                ws['A' + row] = each_result["company_name"]
                ws['B' + row] = each_result["id"]
                ws['C' + row] = each_result["bank_line_id"]
                ws['D' + row] = each_result["journal_name"]
                ws['E' + row] = each_result["pbk_name"]
                ws['F' + row] = each_result["partner_vat"]
                ws['G' + row] = each_result["partner_name"]
                ws['H' + row] = each_result["date_request"]
                ws['I' + row] = each_result["date_payment"]
                ws['J' + row] = each_result["number_week"]
                ws['K' + row] = each_result["amount"]
                ws['L' + row] = each_result["is_prepayment"] and "SI" or "NO"
                ws['M' + row] = each_result["account_code"]
                ws['N' + row] = each_result["account_name"]
                ws['O' + row] = each_result["comments"]
                ws['P' + row] = each_result["num_documento"]
                ws['Q' + row] = each_result["payment_documento"]

                ws['R' + row] = each_result["cuenta_code"]
                ws['S' + row] = each_result["cuenta_name"]
                ws['T' + row] = each_result["percentage"]
                ws['U' + row] = each_result["proporcional_total"]

                i += 1
            last_row = INDEX_ROW + i
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':U' + str(last_row - 1), border)
            ws['A1'] = companys
            ws['B2'] = brw_wizard.weeks or "TODAS"
            ws['D2'] = len(result)

    def create_report_pagos_macros(self, brw_wizard, ws):
        date_from = brw_wizard.date_from
        date_to = brw_wizard.date_to
        companys = ', '.join(brw_wizard.mapped('company_ids').mapped('name'))
        company_ids = brw_wizard.company_ids.ids + [-1, -1]
        self._cr.execute(f""";with variables as (
	select ARRAY{company_ids}::int[] AS company_ids, 
		{brw_wizard.weeks or 0}::int as weeks
),
resumen_pagos as (
	select apbk.id,
		apbkl.id as bank_line_id,
		rc.name as company_name,
		aj.name as journal_name,
		apbk.name as pbk_name,
		apbk.date_request,
		apbk.date_payment,
		apbk.number_week,
		apbk.ref,
		apbkl.amount,
		apbkl.is_prepayment,
		apbkl.ref,
		apbkl.comments, 
		aa.code as account_code,
		aa.name as account_name,
		apr.invoice_id,
		apr.invoice_line_id,
		apr.order_id,
		coalesce(am.name,po.name,'SIN DOCUMENTO') as num_documento , 
		appm.name as payment_documento ,
		rp.vat as partner_Vat,
		rp.name as partner_name 
		from variables
		inner join account_payment_bank_macro apbk on apbk.company_id = ANY(variables.company_ids)
			and (
				(variables.weeks=0) or 
				(variables.weeks!=0 and apbk.number_week=variables.weeks)
			)
		inner join res_company rc on rc.id=apbk.company_id 
		inner join account_payment_bank_macro_line apbkl on apbk.id=apbkl.bank_macro_id and apbkl.apply
		inner join account_payment ap on ap.id=apbkl.payment_id and ap.reversed_payment_id is null 
		inner join account_journal aj on aj.id=apbk.journal_id
		inner join account_payment_request apr on apr.id=apbkl.request_id 	and apr.state in ('confirmed','done','locked')
		
		inner join res_partner rp on rp.id=ap.partner_id 
		left join account_account aa on aa.id=apbkl.payment_account_id
		left join account_move am on am.id=apr.invoice_id
		left join purchase_order po on po.id=apr.ordeR_id  
		left join account_move appm on appm.id=ap.move_id 
		where apbk.state='done'   
) 

select * from resumen_pagos """, (company_ids, date_from, date_to))
        result = self._cr.dictfetchall()
        if result:
            i, INDEX_ROW = 0, 4
            for each_result in result:
                row = str(INDEX_ROW + i)
                ws['A' + row] = each_result["company_name"]
                ws['B' + row] = each_result["id"]
                ws['C' + row] = each_result["bank_line_id"]
                ws['D' + row] = each_result["journal_name"]
                ws['E' + row] = each_result["pbk_name"]
                ws['F' + row] = each_result["partner_vat"]
                ws['G' + row] = each_result["partner_name"]
                ws['H' + row] = each_result["date_request"]
                ws['I' + row] = each_result["date_payment"]
                ws['J' + row] = each_result["number_week"]
                ws['K' + row] = each_result["amount"]
                ws['L' + row] = each_result["is_prepayment"] and "SI" or "NO"
                ws['M' + row] = each_result["account_code"]
                ws['N' + row] = each_result["account_name"]
                ws['O' + row] = each_result["comments"]
                ws['P' + row] = each_result["num_documento"]
                ws['Q' + row] = each_result["payment_documento"]
                i += 1
            last_row = INDEX_ROW + i
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':Q' + str(last_row - 1), border)
            ws['A1'] = companys
            ws['B2'] = brw_wizard.weeks or "TODAS"
            ws['D2'] = len(result)

class report_resumen_contratos_report_xlsx(models.AbstractModel):
    _name = "report.gps_bancos.report_resumen_contratos_report_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Cartera por Cobrar Contratos"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)
        filename = dtFile.join(dir_path, "reporte_resumen_contratos.xlsx")
        dtFile.copyfile(filename, new_filename)
        wb = False
        try:
            brw_wizard = self.env["account.payment.reports.wizard"].sudo().browse(docids[-1])

            wb, ws, target = self.open_xlsx(new_filename, load_sheet=False)
            ws = wb["RESUMEN_CONTRATOS"]
            self.create_resumen_contratos(brw_wizard, ws)
            wb = self.save_wb(wb, target)
            #########################################
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

    def create_resumen_contratos(self, brw_wizard, ws):
        date_from = brw_wizard.date_from
        date_to = brw_wizard.date_to
        companys = ', '.join(brw_wizard.mapped('company_ids').mapped('name'))
        company_ids = brw_wizard.company_ids.ids + [-1, -1]
        self._cr.execute(f""";with variables as (
    	select ARRAY{company_ids}:: int[] AS company_ids, 
    		'{date_from}'::date as date_from,
			'{date_to}'::date as date_to
    )


	SELECT
     rc.name AS compania,
    df.id AS id_contrato,
    rp.name AS emisor,
    dfl.id AS linea_id,
    df.name AS documento,
    df.date_process AS fecha_emision,
    df.date_maturity AS fecha_vencimiento,
    aa.name AS proyecto,
    dfl.quota AS cuota,
    dfl.date_process AS fecha_vencimiento_cuota,
    dfl.amount AS valor,
    dfl.total_to_paid AS por_aplicar,
    dfl.total_paid AS aplicado,
    dfl.total_pending AS pendiente


	from variables
		inner join document_financial df on df.company_id = ANY(variables.company_ids)
		inner join document_financial_line dfl on dfl.document_id=df.id 
		AND dfl.date_process BETWEEN variables.date_from AND variables.date_to -- rango de fecha de vencimiento
inner JOIN
    res_company rc ON df.company_id = rc.id
inner JOIN
    res_partner rp ON df.partner_id = rp.id
	LEFT JOIN
    account_analytic_account aa ON df.account_analytic_id = aa.id
	
WHERE
    df.type = 'contrato' -- solo tipo contrato
    AND df.state not in ('draft', 'cancelled')  -- evitar documentos anulados (opcional)
    AND dfl.total_pending > 0 -- cuotas con saldo pendiente
    

ORDER BY
df.id asc,    dfl.date_process asc  """, (company_ids, date_from, date_to))
        result = self._cr.dictfetchall()
        if result:
            i, INDEX_ROW = 0, 4
            for each_result in result:
                row = str(INDEX_ROW + i)
                ws['A' + row] = each_result["compania"]
                ws['B' + row] = each_result["id_contrato"]
                ws['C' + row] = each_result["emisor"]
                ws['D' + row] = each_result["linea_id"]
                ws['E' + row] = each_result["documento"]
                ws['F' + row] = each_result["fecha_emision"]
                ws['G' + row] = each_result["fecha_vencimiento"]
                ws['H' + row] = each_result["proyecto"]
                ws['I' + row] = each_result["cuota"]
                ws['J' + row] = each_result["fecha_vencimiento_cuota"]
                ws['K' + row] = each_result["valor"]
                ws['L' + row] = each_result["por_aplicar"]
                ws['M' + row] = each_result["aplicado"]
                ws['N' + row] = each_result["pendiente"]
                i += 1
            last_row = INDEX_ROW + i
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':N' + str(last_row - 1), border)
            ws['A1'] = companys
            ws['B2'] = brw_wizard.date_from
            ws['D2'] = brw_wizard.date_to
            ws['F2'] = len(result)

class report_facturas_proveedores_report_xlsx(models.AbstractModel):
    _name = "report.gps_bancos.report_facturas_proveedores_report_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Facturas de Proveedores"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)
        filename = dtFile.join(dir_path, "reporte_facturas_proveedores.xlsx")
        dtFile.copyfile(filename, new_filename)
        wb = False
        try:
            brw_wizard = self.env["account.payment.reports.wizard"].sudo().browse(docids[-1])

            wb, ws, target = self.open_xlsx(new_filename, load_sheet=False)
            ws = wb["FACTURAS"]
            self.create_report_facturas(brw_wizard, ws)

            wb = self.save_wb(wb, target)
            #########################################
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

    def create_report_facturas(self, brw_wizard, ws):
        date_from = brw_wizard.date_from
        date_to = brw_wizard.date_to
        companys = ', '.join(brw_wizard.mapped('company_ids').mapped('name'))
        partnerS = ', '.join(brw_wizard.mapped('partner_ids').mapped('name'))

        company_ids = brw_wizard.company_ids.ids + [-1, -1]
        partner_ids = brw_wizard.partner_ids.ids

        rp_contratista_id = 0
        if brw_wizard.contratista:
            rp_contratista_id = self.env.ref('gps_bancos.rp_contratista').id
        ##################################################################
        self._cr.execute(f""";with variables as (
    	select ARRAY{company_ids}:: int[] AS company_ids, 
    		'{date_from}'::date as date_From,
			'{date_to}'::date as date_to ,
			array{partner_ids}::int [] as partner_ids ,
			{rp_contratista_id}::int as contratista_id 
    ), 
	partner_rows AS (
	    SELECT unnest(partner_ids) AS partner_id
		    FROM variables
		) ,
		facturas as (
select 


		rc.name as company_name,
		aj.name as journal_name,
		rp.vat as partner_Vat,
		rp.name as partner_name ,
		am.id as invoice_id,
		am.name as invoice_name,
		coalesce(am.invoice_date,am.date)::DATE as invoice_date ,
		am.ref as invoice_ref,
		am.amount_untaxed,
		am.amount_tax,
		am.amount_total,
		am.amount_residual

from variables
inner join account_move am on am.company_id = any(variables.company_ids)
		inner join res_company rc on rc.id=am.company_id  
		inner join res_partner rp on rp.id=am.partner_id and ((variables.contratista_id >0 and rp.id in (
		    select  rpcl.partner_id from res_partner_res_partner_category_rel rpcl where rpcl.category_id=variables.contratista_id  
		)) or (variables.contratista_id=0))
INNER join account_journal aj on aj.id=am.journal_id 
		left join partner_rows pr on pr.partner_id=rp.id 
WHERE am.state='posted' and am.move_type in ('in_invoice','in_receipt') AND 
			(
				am.date::date>=variables.date_from and 
				am.date::date<=variables.date_to
			) and 

			(
			    cardinality(variables.partner_ids) = 0
			    OR (cardinality(variables.partner_ids)!=0 and  pr.partner_id is not null  )
		)
	),
facturas_retenciones as (
	select rp.invoice_id,
	aml.move_id,
	am.name as withhold_name,
	sum(aml.l10n_ec_withhold_tax_amount) as monto_rte 
	from facturas rp 
	inner join account_move_line aml on aml.l10n_ec_withhold_invoice_id=rp.invoice_id
	inner join account_move am on am.id=aml.move_id
	where am.state='posted'
	group by rp.invoice_id,aml.move_id,am.name 
)	,
facturas_oc as (
    select ad.invoice_id, 
		string_agg(distinct po.id::text, ', ') AS order_ids,
		string_agg(distinct po.name, ', ') AS order_names , 
		string_agg(DISTINCT TO_CHAR(po.date_order, 'YYYY-MM-DD'), ', ') AS order_dates 
	from 
	facturas ad
	INNER JOIN account_move_line aml ON  aml.move_id = ad.invoice_id 
	INNER JOIN purchase_order_line pol ON pol.id = aml.purchase_line_id
	INNER JOIN purchase_order po ON po.id = pol.order_id  
	 group by ad.invoice_id
)

 

select 
f.*  ,fr.withhold_name,fr.monto_rte ,foc.order_names ,foc.order_dates,
coalesce((f.amount_total-f.amount_residual)-coalesce(fr.monto_rte,0),0.00) as total_payments,
coalesce((f.amount_total-f.amount_residual),0.00) as total_paid

from facturas f 
left join facturas_retenciones fr on fr.invoice_id=F.invoice_id
left join facturas_oc foc on foc.invoice_id=F.invoice_id 
  """, (company_ids, date_from, date_to,partner_ids))
        result = self._cr.dictfetchall()
        if result:
            i, INDEX_ROW = 0, 5
            for each_result in result:
                row = str(INDEX_ROW + i)
                ws['A' + row] = each_result["company_name"]
                ws['B' + row] = each_result["invoice_id"]
                ws['C' + row] = each_result["invoice_name"]
                ws['D' + row] = each_result["journal_name"]
                ws['E' + row] = each_result["invoice_date"]

                ws['F' + row] = each_result["invoice_ref"]

                ws['G' + row] = each_result["partner_vat"]
                ws['H' + row] = each_result["partner_name"]


                ws['I' + row] = each_result["amount_total"]
                ws['J' + row] = each_result["amount_residual"]



                ws['K' + row] = each_result["order_names"]
                ws['L' + row] = each_result["order_dates"]

                ws['M' + row] = each_result["amount_untaxed"]
                ws['N' + row] = each_result["amount_tax"]
                ws['O' + row] = each_result["amount_total"]

                ws['P' + row] = each_result["withhold_name"]

                ws['Q' + row] = each_result["monto_rte"]


                ws['R' + row] =each_result["total_payments"]
                ws['S' + row] =each_result["total_paid"]

                i += 1
            last_row = INDEX_ROW + i
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':S' + str(last_row - 1), border)
            ws['A1'] = companys
            ws['B2'] = date_from
            ws['B3'] =date_to
            ws['E2'] = partner_ids and partnerS or "TODOS"
            ws['E3'] = len(result)

class report_pagos_proveedores_report_xlsx(models.AbstractModel):
    _name = "report.gps_bancos.report_pagos_proveedores_report_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Pagos de Proveedores"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)
        filename = dtFile.join(dir_path, "reporte_pagos_proveedores.xlsx")
        dtFile.copyfile(filename, new_filename)
        wb = False
        try:
            brw_wizard = self.env["account.payment.reports.wizard"].sudo().browse(docids[-1])

            wb, ws, target = self.open_xlsx(new_filename, load_sheet=False)
            ws = wb["PAGOS"]
            self.create_report_pagos(brw_wizard, ws)

            wb = self.save_wb(wb, target)
            #########################################
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

    def create_report_pagos(self, brw_wizard, ws):
        date_from = brw_wizard.date_from
        date_to = brw_wizard.date_to
        companys = ', '.join(brw_wizard.mapped('company_ids').mapped('name'))
        partnerS = ', '.join(brw_wizard.mapped('partner_ids').mapped('name'))

        company_ids = brw_wizard.company_ids.ids + [-1, -1]
        partner_ids = brw_wizard.partner_ids.ids

        rp_contratista_id = 0
        if brw_wizard.contratista:
            rp_contratista_id = self.env.ref('gps_bancos.rp_contratista').id
        ##################################################################

        self._cr.execute(f""";with variables as (
    	select ARRAY{company_ids}:: int[] AS company_ids, 
    		'{date_from}'::date as date_From,
			'{date_to}'::date as date_to ,
			array{partner_ids}::int [] as partner_ids ,
			{rp_contratista_id}::int as contratista_id 
    ), 
	partner_rows AS (
	    SELECT unnest(partner_ids) AS partner_id
		    FROM variables
		),
resumen_pagos as (
	select apbk.id,
		apbkl.id as bank_line_id,
		rc.name as company_name,
		aj.name as journal_name,
		apbk.name as pbk_name,
		apbk.date_request,
		apbk.date_payment,
		apbk.number_week,
		apbk.ref,
		apbkl.amount,
		apbkl.is_prepayment,
		apbkl.ref,
		apbkl.comments, 
		aa.code as account_code,
		aa.name as account_name,
		apr.invoice_id,
		--apr.invoice_line_id,
		apr.order_id,
		coalesce(am.name,po.name,'SIN DOCUMENTO') as num_documento , 
		appm.name as payment_documento ,
		appm.date as payment_date,
		rp.vat as partner_Vat,
		rp.name as partner_name ,
		
		ap.id as payment_id,
		coalesce(am.amount_untaxed,po.amount_untaxed) as amount_untaxed,
		coalesce(am.amount_tax,po.amount_tax) as amount_tax,
		coalesce(am.amount_total,po.amount_total) as amount_total ,
		
		coalesce(am.ref,po.partner_ref) as documento_ref 
		
		
		from variables
		inner join account_payment_bank_macro apbk on apbk.company_id = ANY(variables.company_ids)
		
		inner join res_company rc on rc.id=apbk.company_id 
		inner join account_payment_bank_macro_line apbkl on apbk.id=apbkl.bank_macro_id and apbkl.apply
		inner join account_payment ap on ap.id=apbkl.payment_id and ap.reversed_payment_id is null 
		inner join account_journal aj on aj.id=apbk.journal_id
		inner join account_payment_request apr on apr.id=apbkl.request_id 	and apr.state in ('confirmed','done','locked')

		
		inner join res_partner rp on rp.id=ap.partner_id and ((variables.contratista_id >0 and rp.id in (
		    select  rpcl.partner_id from res_partner_res_partner_category_rel rpcl where rpcl.category_id=variables.contratista_id  
		)) or (variables.contratista_id=0))
		
		left join account_account aa on aa.id=apbkl.payment_account_id
		left join account_move am on am.id=apr.invoice_id
		left join purchase_order po on po.id=apr.ordeR_id  
		left join account_move appm on appm.id=ap.move_id 

		left join partner_rows pr on pr.partner_id=rp.id 
		
		where apbk.state='done'   

		and 

			(
			    cardinality(variables.partner_ids) = 0
			    OR (cardinality(variables.partner_ids)!=0 and  pr.partner_id is not null  )
		) and 
		
		    (appm.date::date>= variables.date_From and appm.date::Date<= variables.date_to )
		
) ,

facturas_retenciones as (
	select rp.invoice_id,
	aml.move_id,
	am.name as withhold_name,
	sum(aml.l10n_ec_withhold_tax_amount) as monto_rte 
	from resumen_pagos rp 
	inner join account_move_line aml on aml.l10n_ec_withhold_invoice_id=rp.invoice_id
	inner join account_move am on am.id=aml.move_id
	where am.state='posted'
	group by rp.invoice_id,aml.move_id,am.name 
),
facturas_oc as (
    select ad.invoice_id, 
		string_agg(distinct po.id::text, ', ') AS order_ids,
		string_agg(distinct po.name, ', ') AS order_names , 
		string_agg(DISTINCT TO_CHAR(po.date_order, 'YYYY-MM-DD'), ', ') AS order_dates 
	from 
	resumen_pagos ad
	INNER JOIN account_move_line aml ON  aml.move_id = ad.invoice_id 
	INNER JOIN purchase_order_line pol ON pol.id = aml.purchase_line_id
	INNER JOIN purchase_order po ON po.id = pol.order_id  
	 group by ad.invoice_id
)


select 
rp.* ,
fr.withhold_name,fr.monto_rte ,fr.monto_rte ,foc.order_names ,foc.order_dates
from resumen_pagos rp 
left join facturas_retenciones fr on fr.invoice_id=rp.invoice_id
left join facturas_oc foc on foc.invoice_id=rp.invoice_id 
  """, (company_ids, date_from, date_to, partner_ids))
        result = self._cr.dictfetchall()
        if result:
            i, INDEX_ROW = 0, 5
            for each_result in result:
                row = str(INDEX_ROW + i)

                ws['A' + row] = each_result["company_name"]
                ws['B' + row] = each_result["journal_name"]
                ws['C' + row] = each_result["partner_vat"]
                ws['D' + row] = each_result["partner_name"]
                ws['E' + row] = each_result["date_payment"]
                ws['F' + row] = each_result["number_week"]
                ws['G' + row] = each_result["amount"]
                ws['H' + row] = each_result["comments"]
                ws['I' + row] = each_result["num_documento"]


                ws['J' + row] = each_result["documento_ref"]
                ws['K' + row] = each_result["payment_documento"]

                ws['L' + row] = each_result["amount_untaxed"]
                ws['M' + row] = each_result["amount_tax"]
                ws['N' + row] = each_result["amount_total"]

                ws['O' + row] = each_result["withhold_name"]
                ws['P' + row] = each_result["monto_rte"]




                i += 1
            last_row = INDEX_ROW + i
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':P' + str(last_row - 1), border)
            ws['A1'] = companys
            ws['B2'] = date_from
            ws['B3'] = date_to
            ws['E2'] = partner_ids and partnerS or "TODOS"
            ws['E3'] = len(result)