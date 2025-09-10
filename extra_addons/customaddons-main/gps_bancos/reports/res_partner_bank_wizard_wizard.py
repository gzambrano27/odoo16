# coding: utf-8
from odoo import api, fields, models, exceptions, tools, _
from ...message_dialog.tools import FileManager
from ...calendar_days.tools import DateManager
from ...calendar_days.tools import CalendarManager
dtFile=FileManager()        
dateO=DateManager()
calendarO=CalendarManager()
import openpyxl
import logging
_logger = logging.getLogger(__name__)

from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font, PatternFill
import openpyxl.utils as utils
from openpyxl.styles import NamedStyle

class report_res_partner_bank_wizard_report_xlsx(models.AbstractModel):
    _name = "report.gps_bancos.report_res_partner_bank_wizard_report_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Pagos/Cobros"
    
    def create_xlsx_report(self, docids, data):
        EXT="xlsx"
        dir_path=dtFile.get_path_file(__file__)
        new_filename=dtFile.create(EXT)  
        filename=dtFile.join(dir_path,"reporte_cuentas.xlsx")
        dtFile.copyfile(filename,new_filename)
        wb=False
        try:
            brw_wizard=self.env["res.partner.bank.wizard"].sudo().browse(docids[-1])

            wb, ws, target = self.open_xlsx(new_filename, load_sheet=False)
            ws = wb["CUENTAS BANCARIAS"]
            self.create_report_cuentas(brw_wizard,ws)
            wb = self.save_wb(wb, target)
            #########################################
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

    def create_report_cuentas(self,brw_wizard,ws):
        companys=self.env["res.company"].sudo().search([]).mapped('name')
        companys=",".join(companys)
        self._cr.execute(f""";WITH banco_rel as (
	select imd.name, rb.bic,rbc.bank_id,rbc.code
	from ir_model_Data imd
	inner join res_bank rb on rb.id=imd.res_id and  imd.model='res.bank' and imd.module='l10n_ec' and imd.name in ('bank_12','bank_8')
	inner join res_bank_code rbc on rbc.bank_main_id=rb.id
) ,
usuarios as (
	select ru.id,rp.name 
	from res_users ru  
	inner join res_partner rp on ru.partner_id=rp.id
),
cuentas_registradas as (
    SELECT
    rp.id AS partner_id,
    -- Columnas booleanas por cada banco
    BOOL_OR(CASE WHEN imd.name = 'bank_12' THEN true ELSE false END) AS bank_12,
    BOOL_OR(CASE WHEN imd.name = 'bank_8' THEN true ELSE false END) AS bank_8
    FROM 
        res_partner_bank_for_pay_rel rel
    INNER JOIN res_partner rp ON rp.id = rel.partner_id
    INNER JOIN res_bank rb ON rb.id = rel.bank_id
    INNER JOIN ir_model_data imd ON imd.res_id = rb.id 
        AND imd.model = 'res.bank' 
        AND imd.module = 'l10n_ec'
        AND imd.name IN ('bank_12', 'bank_8')
    GROUP BY
        rp.id
    ORDER BY
        rp.id
)


select 
rc.name as company_name,
rp.id as partner_id,
idt.name->>'en_US' as tipo_identificacion,
rp.vat,
rp.name,
rp.email,
rpb.id as partner_bank_id,
rbk.BIC as codigo_banco,
rbk.name as banco,
rpb.acc_number,
rpb.tipo_cuenta,
rpb.partner_email,
case when(rpb.tercero) then 'SI' else 'NO' end as tercero,
case when(rpb.tercero) then idtt.name->>'en_US' else '' end  as tipo_identificacion_tercero,
case when(rpb.tercero) then rpb.identificacion_tercero else '' end  as identificacion_tercero,
case when(rpb.tercero) then rpb.nombre_tercero else '' end  as nombre_tercero ,
coalesce(rb1.code,'') as codigo_bolivariano,
coalesce(rb2.code,'') as codigo_internacional,
case when(rp.supplier_rank>1) then 'PROVEEDOR'
		when(rp.customer_rank>1) then 'CLIENTE' 
		else 
		case WHEN(he.id is null) then '' else 'EMPLEADO' end
		end as tipo_contacto , 
uc.name as usuario_ingreso_contacto,
ucc.name as usuario_ingreso_cuenta,
case when(crg.bank_12) then 'SI' else 'NO' end as registrado_bolivariano,
case when(crg.bank_8) then 'SI' else 'NO' end as registrado_internacional 
from res_partner rp
left join res_company rc on rc.id=rp.company_id  
left join l10n_latam_identification_type idt on idt.id=rp.l10n_latam_identification_type_id
inner join res_partner_bank rpb on rpb.partner_id=rp.id
inner join res_bank rbk on rbk.id=rpb.bank_id 
left join l10n_latam_identification_type idtt on idtt.id=rpb.l10n_latam_identification_tercero_id
left join banco_rel rb1 on rb1.name='bank_12' and rb1.bank_id=rbk.id
left join banco_rel rb2 on rb2.name='bank_8' and rb2.bank_id=rbk.id
inner join usuarios uc on uc.id=rp.create_uid 
inner join usuarios ucc on ucc.id=rpb.create_uid 
left join hr_employee he on he.work_contact_id=rp.id
left join cuentas_registradas crg on crg.partner_id=rp.id 
where rp.active 
order by rp.vat,rp.name asc """)
        result = self._cr.dictfetchall()
        if result:
            i, INDEX_ROW = 0, 4
            for each_result in result:
                row = str(INDEX_ROW + i)
                ws['A' + row] = each_result["company_name"]
                ws['B' + row] = each_result["tipo_contacto"]
                ws['C' + row] = each_result["partner_id"]
                ws['D' + row] = each_result["tipo_identificacion"]
                ws['E' + row] = each_result["vat"]
                ws['F' + row] = each_result["name"]
                ws['G' + row] = each_result["email"]
                ws['H' + row] = each_result["partner_bank_id"]
                ws['I' + row] = each_result["codigo_banco"]
                ws['J' + row] = each_result["banco"]
                ws['K' + row] = each_result["tipo_cuenta"]
                ws['L' + row] = each_result["acc_number"]
                ws['M' + row] = each_result["partner_email"]
                ws['N' + row] = each_result["tercero"]
                ws['O' + row] = each_result["tipo_identificacion_tercero"]
                ws['P' + row] = each_result["identificacion_tercero"]
                ws['Q' + row] = each_result["nombre_tercero"]
                ws['R' + row] =each_result["codigo_bolivariano"]
                ws['S' + row] =each_result["codigo_internacional"]
                ws['T' + row] = each_result["registrado_bolivariano"]
                ws['U' + row] = each_result["registrado_internacional"]
                ws['V' + row] = each_result["usuario_ingreso_contacto"]
                ws['W' + row] = each_result["usuario_ingreso_cuenta"]
                i += 1
            last_row = INDEX_ROW + i
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':W' + str(last_row - 1), border)
            ws['A1'] = companys
            ws['B2'] = len(result)
