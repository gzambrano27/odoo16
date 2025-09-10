# coding: utf-8
from odoo import api, fields, models, exceptions, tools, _
from ...message_dialog.tools import FileManager
from ...calendar_days.tools import DateManager
from ...calendar_days.tools import CalendarManager

dtFile = FileManager()
dateO = DateManager()
calendarO = CalendarManager()
import openpyxl
import logging

_logger = logging.getLogger(__name__)

from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font, PatternFill
import openpyxl.utils as utils
from openpyxl.styles import NamedStyle


class report_lote_sol_cxp_report_xlsx(models.AbstractModel):
    _name = "report.gps_bancos.report_lote_sol_cxp_report_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Lote de Sol. de Pagos"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)
        filename = dtFile.join(dir_path, "reporte_solicitudes_cxp.xlsx")
        dtFile.copyfile(filename, new_filename)
        wb = False
        try:
            brw_wizard = self.env["account.payment.analysis.request.wizard"].sudo().browse(docids[-1])

            wb, ws, target = self.open_xlsx(new_filename, load_sheet=False)
            ws = wb["SOLICITUDES"]
            self.create_report_cuentas(brw_wizard, ws)
            wb = self.save_wb(wb, target)
            #########################################
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

    def create_report_cuentas(self, brw_wizard, ws):

        if brw_wizard:
            i, INDEX_ROW = 0, 5
            for request_line in brw_wizard.request_line_ids:
                row = str(INDEX_ROW + i)
                ws['A' + row] =request_line.invoice_line_id.move_id.name
                ws['B' + row] = request_line.invoice_line_id.id
                ws['C' + row] = request_line.date_maturity
                ws['D' + row] =request_line.partner_id and request_line.partner_id.id or None
                ws['E' + row] =request_line.partner_id and request_line.partner_id.vat or None
                ws['F' + row] = request_line.partner_id and request_line.partner_id.name or None
                ws['G' + row] = request_line.account_id.code
                ws['H' + row] =  request_line.account_id.name
                ws['I' + row] = request_line.name
                ws['J' + row] = request_line.debit
                ws['K' + row] = request_line.credit
                ws['L' + row] =  request_line.amount
                ws['M' + row] = abs(request_line.amount_residual)
                i += 1
            last_row = INDEX_ROW + i
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':M' + str(last_row - 1), border)
            ws['A1'] = brw_wizard.company_id.name
            ws['B2'] = brw_wizard.date_request
            ws['F2'] =len( brw_wizard.request_line_ids)

            ws['B3'] = (brw_wizard.request_type_id.name)
            ws['F3'] = (brw_wizard.comments)
