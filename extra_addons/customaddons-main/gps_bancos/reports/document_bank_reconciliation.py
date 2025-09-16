# coding: utf-8
from odoo import api, fields, models, exceptions, tools, _
from ...message_dialog.tools import FileManager
from ...calendar_days.tools import DateManager
from ...calendar_days.tools import CalendarManager

dtFile = FileManager()
dateO = DateManager()
calendarO = CalendarManager()
import openpyxl
import logging

_logger = logging.getLogger(__name__)

from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font, PatternFill
import openpyxl.utils as utils
from openpyxl.styles import NamedStyle


class report_document_bank_reconciliation_xlsx(models.AbstractModel):
    _name = "report.gps_bancos.report_document_bank_recon_report_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Conciliacion Bancaria"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)
        filename = dtFile.join(dir_path, "reporte_conciliacion.xlsx")
        dtFile.copyfile(filename, new_filename)
        wb = False
        try:
            brw_document_bank = self.env["document.bank.reconciliation"].sudo().browse(docids[-1])
            #
            wb, ws, target = self.open_xlsx(new_filename, load_sheet=False)

            ws = wb["ASIENTOS"]
            self.create_report_cuentas(brw_document_bank, ws)

            ws = wb["BANCOS"]
            self.create_report_bancos(brw_document_bank, ws)

            ws = wb["RESUMEN"]
            self.create_report_resumen(brw_document_bank, ws)

            wb = self.save_wb(wb, target)
            #########################################
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

    def create_report_cuentas(self, brw_document, ws):
        print(brw_document.company_id,brw_document.journal_id,brw_document.account_id)
        self._cr.execute(f""";WITH VARIABLES AS (
    SELECT 
        {brw_document.company_id.id}::INT AS COMPANY_ID,
        {brw_document.journal_id.id}::INT AS JOURNAL_ID,
        {brw_document.account_id.id}::INT AS ACCOUNT_ID,
        '{brw_document.date_from}'::DATE AS DATE_FROM,
        '{brw_document.date_to}'::DATE AS DATE_TO,
		{brw_document.id}::int as DOCUMENT_ID   
) ,
DETALLE  AS (
    SELECT 
        AM.ID AS MOVE_ID,
        AML.ID AS LINE_ID,
        AJ.NAME AS JOURNAL_NAME,
        AM.DATE AS MOVE_DATE,
        AM.REF AS MOVE_REF,
        AM.NAME AS MOVE_NAME,
        AML.NAME AS MOVE_LINE_NAME,
        AML.REF AS MOVE_LINE_REF,
        RP.NAME AS PARTNER_NAME,
        AML.DEBIT,
        AML.CREDIT
    FROM VARIABLES
    INNER JOIN ACCOUNT_MOVE AM 
        ON AM.COMPANY_ID = VARIABLES.COMPANY_ID 
        --AND AM.JOURNAL_ID = VARIABLES.JOURNAL_ID
        AND AM.STATE = 'posted'
    INNER JOIN ACCOUNT_JOURNAL AJ 
        ON AJ.ID = AM.JOURNAL_ID 
    INNER JOIN ACCOUNT_MOVE_LINE AML 
        ON AML.MOVE_ID = AM.ID 
        AND AML.ACCOUNT_ID = VARIABLES.ACCOUNT_ID
        AND AML.DATE >= VARIABLES.DATE_FROM 
        AND AML.DATE <= VARIABLES.DATE_TO
    LEFT JOIN RES_PARTNER RP 
        ON RP.ID = COALESCE(AML.PARTNER_ID, AM.PARTNER_ID)
),
RESUMEN  AS (
    SELECT 
        0 AS MOVE_ID,
        0 AS LINE_ID,
        '' AS JOURNAL_NAME,
        (VARIABLES.DATE_FROM - INTERVAL '1 day')::DATE AS MOVE_DATE,
        '' AS MOVE_REF,
        '' AS MOVE_NAME,
        '' AS MOVE_LINE_NAME,
        '' AS MOVE_LINE_REF,
        '' AS PARTNER_NAME,
        SUM(AML.DEBIT) AS DEBIT,
        SUM(AML.CREDIT) AS CREDIT
    FROM VARIABLES
    INNER JOIN ACCOUNT_MOVE AM 
        ON AM.COMPANY_ID = VARIABLES.COMPANY_ID 
        --AND AM.JOURNAL_ID = VARIABLES.JOURNAL_ID
        AND AM.STATE = 'posted'
    INNER JOIN ACCOUNT_JOURNAL AJ 
        ON AJ.ID = AM.JOURNAL_ID 
    INNER JOIN ACCOUNT_MOVE_LINE AML 
        ON AML.MOVE_ID = AM.ID 
        AND AML.ACCOUNT_ID = VARIABLES.ACCOUNT_ID 
        AND AML.DATE < VARIABLES.DATE_FROM 
    GROUP BY 
         VARIABLES.DATE_FROM
),
UNION_FINAL AS (
    SELECT 0 AS TIPO,* FROM RESUMEN
    UNION ALL
    SELECT 1 AS TIPO,* FROM DETALLE
),
ASIENTOS_EMPAREJADOS AS 
(

SELECT DISTINCT ON (RGml.move_line_id)
        RGml.move_line_id,
        rg.id AS group_id,
        rg.reference,
        rg.description
    FROM VARIABLES
    INNER JOIN document_bank_reconciliation dr ON dr.id = VARIABLES.DOCUMENT_ID
    INNER JOIN document_bank_reconciliation_line_group rg ON rg.document_id = dr.id
    INNER JOIN rel_reconciliation_line_group_move_line RGml ON RGml.reconciliation_id = rg.id
)



SELECT 
    UF.*,
    AE.reference AS RECONCILIATION_REFERENCE,
    AE.description AS RECONCILIATION_DESCRIPTION,
    SUM(UF.DEBIT - UF.CREDIT) OVER (
        ORDER BY UF.TIPO ASC, UF.MOVE_DATE, UF.MOVE_ID, UF.LINE_ID
        ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
    ) AS SALDO_ACUMULADO
FROM UNION_FINAL UF
LEFT JOIN ASIENTOS_EMPAREJADOS AE ON AE.move_line_id = UF.LINE_ID
ORDER BY UF.TIPO ASC, UF.MOVE_DATE, UF.MOVE_ID, UF.LINE_ID;""" )
        result = self._cr.dictfetchall()
        if result:
            i, INDEX_ROW = 0, 5
            for each_result in result:
                row = str(INDEX_ROW + i)
                ws['A' + row] = each_result["journal_name"]  # Cuenta
                ws['B' + row] = each_result["move_id"]  # ID Asiento
                ws['C' + row] = each_result["line_id"]  # Línea
                ws['D' + row] = each_result["move_name"]  # Asiento
                ws['E' + row] = each_result["partner_name"]  # Contacto
                ws['F' + row] = each_result["move_date"]  # Fecha
                ws['G' + row] = each_result["move_ref"]  # Comunicación
                ws['H' + row] = each_result["debit"]  # Débito
                ws['I' + row] = each_result["credit"]  # Crédito
                ws['J' + row] = each_result["saldo_acumulado"]
                ws['K' + row] = each_result["reconciliation_reference"]
                ws['L' + row] = each_result["reconciliation_description"]
                i += 1
            last_row = INDEX_ROW + i
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':L' + str(last_row - 1), border)
            ws['A1'] = brw_document.company_id.name
            ws['B3'] = brw_document.date_from
            ws['F3'] = brw_document.date_to
            ws['F2'] = len(result)
            ws['B2'] = brw_document.journal_id.name

    def create_report_bancos(self, brw_document, ws):
        print(brw_document.company_id, brw_document.journal_id, brw_document.account_id)
        self.env.cr.execute("""
            WITH VARIABLES AS (
                SELECT %s::INT AS document_id
            )
            SELECT 
                l.sequence AS numero,
                l.date AS fecha,
                l.reference AS referencia,
                case when(l.transaction_type='debit') then '+' else '-' end AS tipo,
                l.amount AS valor,
                CASE l.transaction_type WHEN 'debit' THEN l.amount ELSE -l.amount END AS valor_con_signo,
                l.available_balance AS saldo_disponible,
                l.description AS descripcion,
                t.name AS tipo_nombre
            FROM document_bank_reconciliation_line l
            LEFT JOIN document_bank_reconciliation_type t ON l.type_id = t.id
            inner JOIN VARIABLES v ON l.document_id = v.document_id 
            ORDER BY l.sequence
        """, [brw_document.id])
        result = self._cr.dictfetchall()
        if result:
            i, INDEX_ROW = 0, 5
            for each_result in result:
                row = str(INDEX_ROW + i)
                ws[f'A{row}'] = brw_document.journal_id.name
                ws[f'B{row}'] = each_result["numero"]
                ws[f'C{row}'] = each_result["fecha"]
                ws[f'D{row}'] = each_result["referencia"]
                ws[f'E{row}'] = each_result["tipo"]
                ws[f'F{row}'] = each_result["valor"]
                ws[f'G{row}'] = each_result["valor_con_signo"]
                ws[f'H{row}'] = each_result["saldo_disponible"]
                ws[f'I{row}'] = each_result["descripcion"]
                ws[f'J{row}'] = each_result["tipo_nombre"]
                i += 1

            last_row = INDEX_ROW + i
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, f'A{INDEX_ROW}:J{last_row - 1}', border)

            # Información de cabecera
            ws['A1'] = brw_document.company_id.name
            ws['B2'] = brw_document.journal_id.name
            ws['B3'] = brw_document.date_from
            ws['F3'] = brw_document.date_to
            ws['F2'] = len(result)

    def create_report_resumen(self, brw_document, ws):
        DEC=2
        ws['C4'] = brw_document.company_id.name
        ws['F4'] = "FECHA %s" % (brw_document.date,)
        ws['B8'] = (brw_document.journal_id.name).upper()
        ws['B10'] = ("MES DE %s DEL %s " % (self.env["calendar.month"].sudo().get_month_name(brw_document.date_to.month),int(brw_document.date_to.year))).upper()

        ws['D13'] = "SALDO INICIAL LIBROS AL %s" % (brw_document.date_from,)
        ws['F13'] = brw_document.saldo_inicial_meses_anteriores

        values=brw_document.get_amounts_grouped_by_type()
        ws['F17'] =round(abs(values.get('cheque',0.00)),DEC)
        ws['F18'] =round(abs(values.get('realizadas',0.00)),DEC)
        ws['F19'] =round(abs(values.get('comision',0.00)),DEC)
        ws['F21'] =round(abs(values.get('recibidas',0.00)),DEC)

        ws['F24'] = brw_document.final_balance

        ws['D23'] = "SALDO FINAL LIBROS %s" % (brw_document.date_to,)
        ws['D24'] = "SALDO FINAL ESTADO DE CUENTA %s" % (brw_document.date_to,)

        ws['B28'] = brw_document.comments or ''



