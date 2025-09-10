# coding: utf-8
from odoo.exceptions import ValidationError
from odoo import api, fields, models, _, SUPERUSER_ID
from dateutil.relativedelta import relativedelta
import re
import base64
from datetime import datetime
from odoo import api,fields, models,_
from xlrd import open_workbook
from odoo.exceptions import ValidationError
from ...calendar_days.tools import CalendarManager,DateManager
from ...message_dialog.tools import FileManager
dtObj=DateManager()
clObj=CalendarManager()
flObj=FileManager()
from dateutil.relativedelta import relativedelta
from openpyxl.utils.datetime import from_excel

class DocumentFinancialVersionWizard(models.Model):
    _name = "document.financial.version.wizard"

    _description = "Asistente de Version"

    @api.model
    def _get_default_name(self):
        active_ids=self._context.get('active_ids',[])
        if active_ids:
            brw_active=self.env["document.financial"].browse(active_ids)
            if brw_active.state=='draft':
                return brw_active.version
            return brw_active.version+1
        return 1

    @api.model
    def _get_default_document_id(self):
        active_ids = self._context.get('active_ids', [])
        return active_ids and active_ids[0] or False

    @api.model
    def _get_default_line_ids(self):
        active_ids = self._context.get('active_ids', [])
        if active_ids:
            brw_active=self.env["document.financial"].browse(active_ids)
            line_ids=[]
            for brw_line in brw_active.line_ids:
                line_ids.append((0,0,{
                    "quota"  :  brw_line.quota,
                    "date_process":brw_line.date_process,
                    "percentage_amortize":brw_line.percentage_amortize,
                    "percentage_interest": brw_line.percentage_interest,
                    "payment_overdue_interest":brw_line.payment_overdue_interest,
                    "payment_capital": brw_line.payment_capital,
                    "payment_interest": brw_line.payment_interest,
                    "payment_other": brw_line.payment_other,
                    "amount": brw_line.amount,
                    "amount_interes": brw_line.amount_interes,
                    "parent_line_id":brw_line.id,
                    "total_paid":brw_line.total_paid,
                }))
            return line_ids
        return [(5,)]

    document_id = fields.Many2one(
        "document.financial",
        string="Documento Financiero", on_delete="cascade",default=_get_default_document_id
    )

    company_id = fields.Many2one(related="document_id.company_id", store=False, readonly=True)
    currency_id = fields.Many2one(related="company_id.currency_id", store=False, readonly=True)

    name=fields.Integer("# Version",required=True,default=_get_default_name)
    date=fields.Date("Fecha",required=True,default=fields.Date.context_today)

    line_ids=fields.One2many('document.financial.version.line.wizard','wizard_id','Detalle',default=_get_default_line_ids)

    total = fields.Monetary("Total", default=0.00, store=True, compute="_compute_total")
    total_to_paid = fields.Monetary("Por Aplicar", default=0.00, required=False, store=True, compute="_compute_total")
    total_paid = fields.Monetary("Aplicado", default=0.00, required=False, store=True, compute="_compute_total")
    total_pending = fields.Monetary("Pendiente", default=0.00, required=False, store=True, compute="_compute_total")

    parent_line_id = fields.Many2one('document.financial.line', "Linea Origen")

    comments = fields.Text("Comentarios")

    import_type=fields.Selection([('text','Texto'),
                                  ('file','Archivo')
                                  ],string="Tipo",default="file")
    quotas_values = fields.Text("Cuotas", default=None)

    file = fields.Binary("Archivo", required=False, filters='*.xlsx')
    file_name = fields.Char("Nombre de Archivo", required=False, size=255)

    @api.onchange('line_ids','line_ids.total', 'line_ids.total_to_paid', 'line_ids.total_paid', 'line_ids.total_pending' )
    @api.depends('line_ids','line_ids.total', 'line_ids.total_to_paid', 'line_ids.total_paid', 'line_ids.total_pending' )
    def _compute_total(self):
        DEC = 2
        for brw_each in self:
            total,total_to_paid,total_paid,total_pending=0.00,0.00,0.00,0.00
            for brw_line in brw_each.line_ids:
                total+=brw_line.total
                total_to_paid+=brw_line.total_to_paid
                total_paid += brw_line.total_paid
                total_pending += brw_line.total_pending
            brw_each.total = round(total, DEC)
            brw_each.total_to_paid = round(total_to_paid, DEC)
            brw_each.total_paid = round(total_paid, DEC)
            brw_each.total_pending = round(total_pending, DEC)

    def process(self):
        DEC=2
        OBJ_VERSION=self.env["document.financial.version"]
        def update_document(brw_each,srch_version):
            line_version_ids = [(5,)]
            all_lines=brw_each.document_id.line_ids
            updates_all_lines = self.env["document.financial.line"]
            for brw_line in all_lines:
                line_version_ids.append((0, 0, brw_line.copy_financial_line()))
            srch_version.write({"line_ids":line_version_ids})
            ###
            new_line_ids=[]
            for brw_new_line in brw_each.line_ids:
                if not brw_new_line.parent_line_id:
                    new_line_values = {
                            'company_id': brw_each.document_id.company_id.id,
                            'document_id': brw_each.document_id.id,
                            'quota': brw_new_line.quota,
                            'date_process': brw_new_line.date_process,
                            'date_maturity_payment': brw_new_line.date_process,
                            'percentage_amortize': brw_new_line.percentage_amortize,
                            'percentage_interest': brw_new_line.percentage_interest,
                            'payment_capital': brw_new_line.payment_capital,
                            'payment_interest': brw_new_line.payment_interest,
                            'payment_other': brw_new_line.payment_other,
                            'amount': brw_new_line.amount,
                            'amount_original': brw_new_line.amount,
                            'amount_interes': brw_new_line.amount_interes,

                            #'original_amount': brw_new_line.original_amount,
                            # 'total': self.total,
                            # 'total_to_paid': self.total_to_paid,
                            # 'total_paid': self.total_paid,
                            # 'total_pending': self.total_pending,
                            # 'total_invoiced': self.total_invoiced,
                            # 'total_to_invoice': self.total_to_invoice,
                            # 'overdue': self.overdue,
                            # 'name': self.name,
                            'last_version': True,
                            'is_copy': False,
                            'copy_payment_capital': brw_new_line.payment_capital,
                            'copy_payment_interest': brw_new_line.payment_interest,
                            'copy_payment_overdue_interest': brw_new_line.payment_overdue_interest,
                            'copy_payment_other': brw_new_line.payment_other,
                            'copy_paid': brw_new_line.total_paid,
                            'attachment_ids': [(6, 0, brw_new_line.parent_line_id and brw_new_line.parent_line_id.attachment_ids.ids or [] )],
                        }
                    new_line_ids.append((0,0,new_line_values))
                else:
                    if brw_new_line.parent_line_id:
                        if brw_new_line.parent_line_id.total_pending<=0.00:#mantener igual
                            pass#solo se actualiza lo que tenga un valor pendiente x eso se deja asi
                        else:#CON ALGUN valor pendiente o simplemente se actualiza los valores
                            new_line_values = {
                                'quota': brw_new_line.quota,
                                'date_process': brw_new_line.date_process,
                                'date_maturity_payment': brw_new_line.date_process,
                                'percentage_amortize': brw_new_line.percentage_amortize,
                                'percentage_interest': brw_new_line.percentage_interest,
                                'payment_capital': brw_new_line.payment_capital,
                                'payment_interest': brw_new_line.payment_interest,
                                'payment_other': brw_new_line.payment_other,
                                'amount': brw_new_line.amount,
                                'amount_original': brw_new_line.amount,
                                'amount_interes': brw_new_line.amount_interes,
                                'last_version': True,
                                'is_copy': False,
                            }
                            new_line_ids.append((1, brw_new_line.parent_line_id.id, new_line_values))
                        updates_all_lines+=brw_new_line.parent_line_id
            excluded_lines = all_lines - updates_all_lines
            if excluded_lines:#se borra lo que no este en el nuevo detalle
                excluded_lines=excluded_lines.with_context(validate_unlink=False)
                excluded_lines.unlink()
            brw_each.document_id.write({"last_version_id": srch_version.id,
                                        "version":brw_each.document_id.version+1,
                                        "line_ids":new_line_ids
                                        })
            brw_each.document_id.message_post(
                body=brw_each.comments
            )
        for brw_each in self:
            if brw_each.document_id.state!='draft':###si NOes preliminar

                for brw_new_line in brw_each.line_ids:
                    if round(brw_new_line.total_pending,DEC)<0.00:
                        raise ValidationError(_("El valor de la cuota %s queda en negativo con %s") % (brw_new_line.quota,round(brw_new_line.total_pending,DEC)))


                nounlink_doclines=brw_each.document_id.line_ids.filtered(lambda x:x.total_paid!=0.00)
                nounlink_lines=brw_each.line_ids.filtered(lambda x:x.parent_line_id and x.parent_line_id.total_paid!=0.00).mapped('parent_line_id')
                dif_lines= nounlink_doclines- nounlink_lines
                #print(dif_lines)
                if dif_lines:
                    dscr_lineas=",".join(dif_lines.mapped('name'))
                    raise ValidationError(_("No puedes borrar las cuotas que tiene un valor pagado mayor a 0.00.Validar lineas %s") % (dscr_lineas,))
                srch_version=OBJ_VERSION.search([('document_id','=',brw_each.document_id.id),
                                                 ('name','=',brw_each.document_id.version)
                                                 ])
                if not srch_version:
                    srch_version=OBJ_VERSION.create({
                        'document_id':brw_each.document_id.id,
                        'name':brw_each.document_id.version,
                        'date': brw_each.date,
                    })
                    update_document(brw_each,srch_version)

                else:
                    srch_version.write({
                        'name': brw_each.document_id.version,
                        'date': brw_each.date,
                    })
                    update_document(brw_each, srch_version)

                for brw_doc_line in brw_each.document_id.line_ids:
                    if round(brw_doc_line.total_pending,DEC)<0.00:
                        raise ValidationError(_("El valor de la cuota %s queda en negativo con %s") % (brw_doc_line.quota,round(brw_doc_line.total_pending,DEC)))
            else:
                new_line_ids=[(5,)]
                for brw_new_line in brw_each.line_ids:
                    if round(brw_new_line.total_pending,DEC)<0.00:
                        raise ValidationError(_("El valor de la cuota %s queda en negativo con %s") % (brw_new_line.quota,round(brw_new_line.total_pending,DEC)))

                    new_line_values = {
                            'company_id': brw_each.document_id.company_id.id,
                            'document_id': brw_each.document_id.id,
                            'quota': brw_new_line.quota,
                            'date_process': brw_new_line.date_process,
                            'date_maturity_payment': brw_new_line.date_process,
                            'percentage_amortize': brw_new_line.percentage_amortize,
                            'percentage_interest': brw_new_line.percentage_interest,
                            'payment_capital': brw_new_line.payment_capital,
                            'payment_interest': brw_new_line.payment_interest,
                            'payment_other': brw_new_line.payment_other,
                            'amount': brw_new_line.amount,
                            'amount_original': brw_new_line.amount,
                            'amount_interes':brw_new_line.amount_interes,
                            'last_version': True,
                            'is_copy': False,
                    }
                    new_line_ids.append((0, 0, new_line_values))
                ###
                brw_each.document_id.write({"line_ids": new_line_ids })
                brw_each.document_id.update_total()
                for brw_doc_line in brw_each.document_id.line_ids:
                    if round(brw_doc_line.total_pending,DEC)<0.00:
                        raise ValidationError(_("El valor de la cuota %s queda en negativo con %s") % (brw_doc_line.quota,round(brw_doc_line.total_pending,DEC)))
                brw_each.document_id.message_post(
                    body=brw_each.comments
                )
        return True

    def process_update(self):


        return True

    @api.model
    def parse_quotas_values(self, text):
        def parse_fecha(s):
            s = s.strip()
            # Intenta varios formatos comunes: 29/9/2023, 29/09/2023, 2023-10-29, etc.
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                try:
                    return datetime.strptime(s, fmt).date()  # devuelve date
                except ValueError:
                    continue
            # Si no matchea ninguno, lanza para que la línea se salte
            raise ValueError(f"Formato de fecha no reconocido: {s}")

        cuotas = []
        for raw in (text or '').strip().splitlines():
            line = raw.strip()
            if not line:
                continue
            # Separar por cualquier cantidad de espacios o tabs
            parts = re.split(r'\s+', line)

            # La fila válida debe tener al menos: nro, fecha, capital, interes, ..., otros
            if len(parts) < 5:
                continue

            try:
                nro = int(parts[0])
                fecha = parse_fecha(parts[1])

                # capital = parts[2], interes = parts[3], "otros" lo tomo como la última columna
                def to_float(x):
                    # elimina separador de miles en estilo "19,003.15" -> "19003.15"
                    return float(x.replace(',', ''))

                capital = to_float(parts[2])
                interes = to_float(parts[3])
                otros = to_float(parts[-1])

                cuotas.append({
                    'nro': nro,
                    'fecha': fecha,  # date; Odoo Date lo acepta o lo puedes formatear
                    # 'fecha' : fecha.strftime('%Y-%m-%d'),  # si prefieres string YYYY-MM-DD
                    'capital': capital,
                    'interes': interes,
                    'otros': otros,
                })
            except Exception:
                # Cabeceras o líneas no parseables → se ignoran
                continue

        return cuotas

    @api.model
    def parse_quotas_file_values(self, file,file_name):
        """
        Lee un archivo Excel con cuotas y retorna un arreglo de diccionarios:
        [{'nro': int, 'fecha': date, 'capital': float, 'interes': float, 'otros': float}, ...]
        """
        if not file:
            return []

        ext = flObj.get_ext(file_name)
        fileName = flObj.create(ext)
        flObj.write(fileName, flObj.decode64(file))
        book = open_workbook(fileName)
        sheet = book.sheet_by_index(0)
        NRO,FECHA,CAPITAL,INTERES,OTROS=0,1,2,3,4
        cuotas=[]
        for row_index in range(1, sheet.nrows):
            if row_index==1:
                continue
            nro = int(sheet.cell(row_index, NRO).value)

            detalle_fecha = sheet.cell(row_index, FECHA).value
            if isinstance(detalle_fecha, datetime):
                detalle_fecha = detalle_fecha.strftime('%Y-%m-%d')
            elif isinstance(detalle_fecha, (int, float)):
                detalle_fecha = from_excel(detalle_fecha).strftime('%Y-%m-%d')
            elif isinstance(detalle_fecha, str) and detalle_fecha.strip():
                for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
                    try:
                        detalle_fecha = datetime.strptime(detalle_fecha, fmt).strftime('%Y-%m-%d')
                        break
                    except ValueError:
                        continue
                else:
                    detalle_fecha = None  # o lanzar un error si es necesario
            else:
                detalle_fecha = None
            capital = float(str(sheet.cell(row_index, CAPITAL).value).replace(',', ''))
            interes = float(str(sheet.cell(row_index, INTERES).value).replace(',', ''))
            otros = float(str(sheet.cell(row_index, OTROS   ).value).replace(',', ''))
            cuota = {
                'nro': nro,
                'fecha': detalle_fecha,
                'capital': capital,
                'interes':interes,
                'otros': otros
            }
            cuotas.append(cuota)
        return cuotas

    @api.onchange('import_type','quotas_values','file')
    def onchange_quotas_values(self):
        def update_each_line(updated,each_resultado,self,all_lines):
            if updated:
                if  round(updated.total_pending) != 0.00:
                    if len(updated) > 1:
                        raise ValidationError(_("Se detecto mas de una linea en la cuota %s") % (each_resultado["nro"],))
                    else:
                        # updated.date_process=each_resultado["nro"]
                        updated.date_process = each_resultado["fecha"]
                        updated.payment_capital = each_resultado["capital"]
                        updated.payment_interest = each_resultado["interes"]
                        updated.payment_other = each_resultado["otros"]
                return False
            else:
                vals = {
                    'quota': each_resultado["nro"],
                    'date_process': each_resultado["fecha"],
                    'payment_capital': each_resultado["capital"],
                    'payment_interest': each_resultado["interes"],
                    'payment_other': each_resultado["otros"],
                    'company_id':self.company_id.id
                }
                new_line = [(0,0,vals)]#self.env['document.financial.version.line.wizard'].new(vals)  # TODO: nombre de modelo línea
                return new_line

        if self.import_type=='text':
            if self.quotas_values:
                resultado = self.parse_quotas_values(self.quotas_values)
                all_lines = self.line_ids
                new_lines = []
                for each_resultado in resultado:
                    updated = all_lines.filtered(
                        lambda x: x.parent_line_id and int(x.parent_line_id.quota) == int(each_resultado["nro"]))
                    new_line = update_each_line(updated, each_resultado, self, all_lines)
                    if new_line:
                        new_lines += new_line
                self.line_ids = new_lines
                self.file=None
        if self.import_type=='file':
            if self.file:
                resultado = self.parse_quotas_file_values(self.file,self.file_name)
                all_lines = self.line_ids
                new_lines=[]
                for each_resultado in resultado:
                    updated = all_lines.filtered(
                        lambda x: x.parent_line_id and int(x.parent_line_id.quota) == int(each_resultado["nro"]))
                    new_line = update_each_line(updated, each_resultado, self, all_lines)
                    if new_line:
                        new_lines+=new_line
                self.line_ids=new_lines
                self.quotas_values=None



class DocumentFinancialVersionLineWizard(models.Model):
    _name = "document.financial.version.line.wizard"

    wizard_id=fields.Many2one('document.financial.version.wizard','Asistente',ondelete="cascade")
    company_id = fields.Many2one(
        "res.company",
        string="Compañia",
        required=True,
        copy=False,
        default=lambda self: self.env.company,
    )
    currency_id = fields.Many2one(related="company_id.currency_id", store=False, readonly=True)
    quota = fields.Integer(string="Cuota", default=1, required=True)
    date_process = fields.Date("Fecha de Vencimiento", default=fields.Date.today(), required=True)

    percentage_amortize = fields.Float("% por Amortizar", default=0.00, digits=(16, 6))
    percentage_interest = fields.Float("% de Interes", default=0.00, digits=(16, 6))

    payment_capital = fields.Monetary("Capital", default=0.00, required=False)
    payment_interest = fields.Monetary("Interés", default=0.00, required=False)

    payment_overdue_interest = fields.Monetary("Interés Mora", default=0.00, required=False)
    payment_other = fields.Monetary("Otros", default=0.00, required=False)

    #amount = fields.Monetary("Valor", default=0.00)

    amount_interes = fields.Monetary("Interés Generado", default=0.00, tracking=True)
    amount = fields.Monetary("Valor", default=0.00, tracking=True)

    #original_amount = fields.Monetary("Valor Original", default=0.00, compute="_compute_amount", store=True,
    #                                  readonly=True)
    total = fields.Monetary("Total", default=0.00, store=True, compute="_compute_total")
    total_to_paid = fields.Monetary("Por Aplicar", default=0.00, required=False, store=True, compute="_compute_total")
    total_paid = fields.Monetary("Aplicado", default=0.00, required=False )
    total_pending = fields.Monetary("Pendiente", default=0.00, required=False, store=True, compute="_compute_total")

    parent_line_id = fields.Many2one('document.financial.line', "Linea Origen")

    inactive=fields.Boolean('Inactivar',default=False)

    @api.onchange('payment_capital', 'payment_interest', 'payment_overdue_interest', 'payment_other', 'amount','amount_interes')
    @api.depends('payment_capital','payment_interest','payment_overdue_interest','payment_other','amount','amount_interes' )
    def _compute_total(self):
        DEC = 2
        for brw_each in self:
            if brw_each.wizard_id.document_id.type!='contrato':
                total = round(
                        brw_each.payment_capital + brw_each.payment_interest + brw_each.payment_other,
                        DEC)
            else:
                total = round(
                    brw_each.amount+brw_each.amount_interes)
            total = total + brw_each.payment_overdue_interest  # se suma a los interese e valor por pagar
            total_to_paid = total  # siempre tomara el valor a pagar como reflejo de lo que debera pagar
            brw_each.total = total
            brw_each.total_to_paid = round(total_to_paid, DEC)
            #brw_each.total_paid = round(brw_each.total_paid, DEC)
            brw_each.total_pending = round(total_to_paid - brw_each.total_paid, DEC)

    def unlink(self):
        for brw_each in self:
            if self._context.get("validate_unlink", True):
                if brw_each.parent_line_id:
                    if brw_each.parent_line_id and brw_each.parent_line_id.total_paid>0.00:
                        raise ValidationError(_("No puedes borrar un registro con una valor pagado diferente a 0.00 %s ,id %s") % (brw_each.parent_line_id.quota,brw_each.parent_line_id.id))
        return super(DocumentFinancialVersionLineWizard, self).unlink()



