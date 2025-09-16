# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.
from odoo import fields, models, api, _
from odoo.exceptions import ValidationError,UserError
from datetime import timedelta
import re
import base64
import io
from openpyxl import load_workbook
from datetime import datetime
from collections import defaultdict

class DocumentBankReconciliation(models.Model):
    _name = 'document.bank.reconciliation'
    _description = "Documento de Conciliación Bancaria"

    _inherit = ['portal.mixin', 'mail.thread', 'mail.activity.mixin']

    @api.model
    def _get_default_date(self):
        return fields.Date.context_today(self)

    @api.model
    def _get_default_date_from(self):
        return self._get_default_date_to() - timedelta(days=7)

    @api.model
    def _get_default_date_to(self):
        return fields.Date.context_today(self) - timedelta(days=1)


    name=fields.Char("Nombre",required=True,size=255)
    company_id = fields.Many2one(
        "res.company",
        string="Compañia",
        required=True,
        copy=False,
        default=lambda self: self.env.company, tracking=True,
    )
    currency_id = fields.Many2one(related="company_id.currency_id", store=False, readonly=True)
    journal_id=fields.Many2one("account.journal","Diario",required=True)
    date= fields.Date(string="Fecha de Documento", default=_get_default_date,tracking=True)

    date_from = fields.Date(string="Fecha Inicio", default=_get_default_date_from, tracking=True)
    date_to = fields.Date(string="Fecha Fin", default=_get_default_date_to, tracking=True)

    comments = fields.Text("Comentarios",tracking=True)
    state = fields.Selection([('draft', 'Preliminar'),
                              ('generated', 'Generado'),
                              ('confirmed', 'Confirmado'),
                              ('done', 'Realizado'),
                              ('cancelled', 'Anulado'),
                              ], string="Estado", default="draft", tracking=True)


    file = fields.Binary(string="Archivo Banco")
    filename = fields.Char(stirng="Nombre Macro")
    line_ids=fields.One2many('document.bank.reconciliation.line','document_id','Detalle')
    group_ids = fields.One2many('document.bank.reconciliation.line.group', 'document_id', 'Referencias')

    initial_balance = fields.Monetary(string="Saldo Inicial", default=0.00)
    final_balance = fields.Monetary(string="Saldo Final", default=0.00)

    calculated_final_balance = fields.Monetary(
        string="Saldo Final Calculado",
        compute="_compute_balances",
        store=True
    )
    calculated_signed_amount = fields.Monetary(
        string="Saldo Movimientos",
        compute="_compute_balances",
        store=True
    )
    calculated_reconciled_signed_amount = fields.Monetary(
        string="Saldo Movimientos Conciliados",
        compute="_compute_balances",
        store=True
    )
    difference = fields.Monetary(
        string="Diferencia",
        compute="_compute_balances",
        store=True
    )
    account_id = fields.Many2one('account.account', 'Cuenta',required=True)


    move_line_ids=fields.Many2many('account.move.line',compute="_compute_move_lines",store=False,string="Detalle de Asientos")

    @api.onchange('journal_id','company_id')
    def _onchange_journal_id_set_previous_data(self):
        if self.journal_id:
            last_doc = self.env['document.bank.reconciliation'].search(
                [('company_id', '=', self.company_id.id),
                 ('journal_id', '=', self.journal_id.id),
                 ('state', '!=', 'cancelled')],
                order='date_to desc',
                limit=1
            )
            if last_doc:
                self.date_from = last_doc.date_to + timedelta(days=1)
                self.initial_balance = last_doc.final_balance
            else:
                self.date_from = self._get_default_date_from()
                self.initial_balance = 0.0

            self.date_to = self._get_default_date_to()

    @api.onchange('company_id', 'journal_id', 'account_id', 'date_from', 'date_to')
    @api.depends('company_id','journal_id','account_id','date_from','date_to')
    def _compute_move_lines(self):
        for brw_each in self:
            move_line_ids=self.env["account.move.line"].sudo().search(
                [
                    ('move_id.company_id', '=', brw_each.company_id.id),
                    #('move_id.journal_id','=',brw_each.journal_id.id),
                    ('account_id', '=', brw_each.account_id.id),
                    ('move_id.state','=','posted'),
                    ('date', '>=', brw_each.date_from),
                    ('date', '<=', brw_each.date_to),
                ]
            )
            brw_each.move_line_ids=move_line_ids

    def match_lines(self):
        #DEC=2
        OBJ_MOVE_LINE=self.env["account.move.line"].sudo()
        OBJ_PAYMENT=self.env["account.payment"].sudo()
        for brw_each in self:
            if brw_each.group_ids:
                for brw_group in brw_each.group_ids:
                    srch_move_line = self.env["account.move.line"]

                    #########################macros#############3
                    srch_bank_macro = self.env["account.payment.bank.macro"].search([('journal_id', '=', brw_each.journal_id.id),
                                                               ('company_id', '=', brw_each.company_id.id),
                                                               ('state', '=', 'done'),
                                                               ('ref', '=',brw_group.reference),
                                                               ('date_request','>=',brw_each.date_from)
                                                               ])
                    if srch_bank_macro and len(srch_bank_macro)==1:
                        brw_group.write({"bank_macro_id":  srch_bank_macro.id   })
                        brw_group.onchange_bank_macro_id()
                        brw_group.onchange_move_ids()
                    #########################macros#############3
                    if not srch_bank_macro:
                        srch_employee_payment = self.env["hr.employee.payment"].search(
                                [('payment_journal_id', '=', brw_each.journal_id.id),
                                 ('company_id', '=', brw_each.company_id.id),
                                 ('state', '=', 'approved'),
                                 ('ref', '=', brw_group.reference),
                                  ('date_process','>=',brw_each.date_from)
                                 ])
                        if srch_employee_payment and len(srch_employee_payment) == 1:
                            brw_group.write({"bank_employee_id": srch_employee_payment.id})
                            brw_group.onchange_bank_employee_id()
                            brw_group.onchange_move_ids()
                        #########################macros#############3
                        if not srch_employee_payment and not srch_bank_macro:
                            srch_payment=OBJ_PAYMENT.search([('move_id.company_id','=',brw_each.company_id.id),
                                                                 ('move_id.journal_id','=',brw_each.journal_id.id),
                                                                 ('move_id.state','=','posted'),
                                                                 ('bank_reference', '=', brw_group.reference),
                                                                 #('amount','<=',round(brw_group.amount,DEC)),
                                                                 ('date','>=',brw_each.date_from)
                                                                 ])
                            if srch_payment and len(srch_payment)==1:
                                #print("111111",srch_payment)
                                brw_group.write({"payment_ids":[(4, line.id) for line in srch_payment   ] })
                                brw_group.onchange_payment_ids()
                                brw_group.onchange_move_ids()
                            if not srch_employee_payment and not srch_bank_macro and not srch_payment:
                                #field=(brw_group.transaction_type=='debit') and 'debit' or 'credit'
                                srch_move_line=OBJ_MOVE_LINE.search([('move_id.company_id','=',brw_each.company_id.id),
                                                                         ('move_id.state','=','posted'),
                                                                       ('move_id','in',brw_group.parent_move_ids.ids),
                                                                       ('account_id','=',brw_each.account_id.id),
                                                                         ('date','>=',brw_each.date_from),
                                                                         '|',
                                                                         ('move_id.ref', 'ilike', brw_group.reference),
                                                                         ('ref', 'ilike', brw_group.reference),

                                                                         ])
                                if srch_move_line and len(srch_move_line[0])==1:
                                    brw_group.write({"move_line_ids":srch_move_line,
                                                         "move_ids":srch_move_line.move_id
                                        })
                    other_srch_move_line = OBJ_MOVE_LINE.search(
                                            [('move_id.company_id', '=', brw_each.company_id.id),
                                             ('move_id.state', '=', 'posted'),
                                             ('account_id', '!=', brw_each.account_id.id),
                                             ('name', '=', brw_group.reference),
                                             ('move_id', 'in', brw_group.parent_move_ids.ids),
                                             ('id','not in',srch_move_line.ids+[-1,-1])
                                             ])
                    if other_srch_move_line:
                         brw_group.write({'other_move_line_ids': [(4, line.id) for line in other_srch_move_line]  })
            brw_each.action_generate_summary()
        return True

    @api.onchange('journal_id')
    def onchange_journal_id(self):
        if not self.journal_id:
            self.account_id=False
        else:
            self.account_id = self.journal_id.default_account_id  and self.journal_id.default_account_id.id or False

    @api.onchange('journal_id', 'date_from', 'date_to')
    def _onchange_fill_name(self):
        for rec in self:
            if rec.journal_id and rec.date_from and rec.date_to:
                rec.name = f"Conciliación diario {rec.journal_id.name} del {rec.date_from.strftime('%Y-%m-%d')} al {rec.date_to.strftime('%Y-%m-%d')}".upper()


    @api.onchange('initial_balance', 'final_balance', 'line_ids', 'group_ids.total_signed_related_amount', 'group_ids.full_reconciled')
    @api.depends('initial_balance', 'final_balance','line_ids','group_ids.total_signed_related_amount','group_ids.full_reconciled')
    def _compute_balances(self):
        DEC=2
        for rec in self:
            calculated_signed_amount=sum(rec.group_ids.mapped('signed_amount'))
            calculated_reconciled_signed_amount=sum(rec.group_ids.filtered(lambda x:round(x.total_signed_related_amount,DEC)!=0.00).mapped('total_signed_related_amount'))
            rec.calculated_signed_amount = round(calculated_signed_amount,DEC)
            rec.calculated_reconciled_signed_amount = round(calculated_reconciled_signed_amount, DEC)
            rec.calculated_final_balance=round((rec.initial_balance-calculated_reconciled_signed_amount),DEC)
            rec.difference=round((rec.final_balance-rec.calculated_final_balance    ),DEC)


    @api.constrains('initial_balance', 'final_balance')
    def _check_balances_non_negative(self):
        for record in self:
            if record.initial_balance < 0:
                raise ValidationError("El campo 'Saldo Inicial' no puede ser menor que 0.")
            if record.final_balance < 0:
                raise ValidationError("El campo 'Saldo Final' no puede ser menor que 0.")


    _order="id desc"

    _rec_name = "name"
    _check_company_auto = True

    @api.returns('self', lambda value: value.id)
    def copy(self, default=None):
        raise ValidationError(_("No puedes duplicar este documento"))

    #@api.onchange('file')
    def update_file(self):

        def parse_excel_date(value):
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, str):
                for fmt in ('%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d'):
                    try:
                        return datetime.strptime(value.strip(), fmt).date()
                    except ValueError:
                        continue
            raise ValidationError(f"Formato de fecha no reconocido: {value}")

        def _match_type_from_description(self, texto):
            # Normalizar texto (sin acentos, lowercase)
            def normalize(txt):
                import unicodedata
                return unicodedata.normalize('NFKD', txt).encode('ASCII', 'ignore').decode().lower()

            normalized_input = normalize(texto)

            tipos = self.env['document.bank.reconciliation.type'].search([])
            for tipo in tipos:
                for desc in tipo.descriptions.split(','):
                    if normalize(desc.strip()) in normalized_input:
                        return tipo
            return False

        uncategorized_type = self.env['document.bank.reconciliation.type'].search([('code', '=', 'sin_tipo')], limit=1)


        for record in self:
            journal_name = record.journal_id.name.lower()
            parsed_lines=[(5,)]
            if not record.file:
                raise ValidationError("Debe adjuntar un archivo para procesar.")
            if 'bolivariano' in journal_name:
                try:
                    content = base64.b64decode(record.file)
                    excel_file = io.BytesIO(content)
                    wb = load_workbook(excel_file, data_only=True)
                    sheet = wb.active  # usa la primera hoja
                except Exception as e:
                    raise ValidationError(
                        "No se pudo leer el archivo Excel. Asegúrate de que sea un archivo válido (.xlsx).")

                # Buscar fila donde comienza la tabla (cabecera)
                start_index = None
                for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    if row and "FECHA" in row and "DESCRIPCION" in row:
                        start_index = idx + 1  # inicia después de la cabecera
                        break

                if start_index is None:
                    raise ValidationError("El archivo no contiene una tabla válida de transacciones.")

                #parsed_lines = []

                for row in sheet.iter_rows(min_row=start_index, values_only=True):
                    if not row or all(cell in [None, '', 0] for cell in row):
                        continue

                    try:

                        descripcion_texto = str(row[10]).strip() if row[10] else ""
                        type_obj = _match_type_from_description(self,descripcion_texto)
                        tipo_id = type_obj.id if type_obj else uncategorized_type.id

                        parsed_lines.append((0, 0, {
                            'sequence': int(row[0]),
                            'date': parse_excel_date(row[1]),
                            'reference': str(row[4]).strip().replace("'", "") if row[4] else "",
                            'location': str(row[3]).strip() if row[3] else "",
                            'detail': str(row[5]).strip() if row[5] else "",
                            'transaction_type': 'debit' if str(row[6]).strip() == '+' else 'credit',
                            'amount': float(str(row[7]).replace(',', '').strip()) if row[7] else 0.0,
                            'available_balance': float(str(row[8]).replace(',', '').strip()) if row[8] else 0.0,
                            'book_balance': float(str(row[9]).replace(',', '').strip()) if row[9] else 0.0,
                            'description': descripcion_texto,
                            'company_id': record.company_id.id,
                            'type_id':tipo_id
                        }))
                    except Exception as e:
                        raise ValidationError(f"Ocurrió un error al procesar la línea:\n{row}\n\nError: {str(e)}")
            elif 'internacional' in journal_name:

                try:
                    content = base64.b64decode(record.file)
                    excel_file = io.BytesIO(content)
                    wb = load_workbook(excel_file, data_only=True)
                    sheet = wb.active
                except Exception:
                    raise ValidationError(
                        "No se pudo leer el archivo Excel. Asegúrate de que esté en formato .xlsx válido.")
                #parsed_lines = []
                i = 1
                for row in sheet.iter_rows(min_row=5, values_only=True):
                    if not row or all(cell in [None, '', 0] for cell in row):
                        continue
                    try:
                        fecha = parse_excel_date(row[1])  # Fecha (columna desplazada)
                        descripcion = str(row[3]).strip() if row[3] else ""
                        numeros = re.findall(r'\d{5,}', descripcion)
                        reference = '-'.join(numeros) if numeros else ""

                        # Corregido: ya no son tuplas
                        debito = float(str(row[7]).replace(',', '').strip()) if row[7] else 0.0
                        credito = float(str(row[8]).replace(',', '').strip()) if row[8] else 0.0
                        saldo = float(str(row[9]).replace(',', '').strip()) if row[9] else 0.0
                        ciudad = str(row[10]).strip() if row[10] else ""

                        #descripcion_texto = str(row[10]).strip() if row[10] else ""
                        type_obj = _match_type_from_description(self,descripcion)
                        tipo_id = type_obj.id if type_obj else uncategorized_type.id

                        parsed_lines.append((0, 0, {
                            'date': fecha,
                            'reference': reference,
                            'description': descripcion,
                            'detail': descripcion,
                            'transaction_type': 'credit' if debito > 0 else 'debit',
                            'amount': debito if debito > 0 else credito,
                            'available_balance': saldo,
                            'book_balance': saldo,
                            'location': ciudad,
                            'company_id': record.company_id.id,
                            'sequence': i,
                            'type_id':tipo_id
                        }))
                        i += 1
                    except Exception as e:
                        raise ValidationError(f"Error al procesar la línea:\n{row}\n\n{str(e)}")
            elif 'produbanco' in journal_name:
                try:
                    decoded_file = base64.b64decode(record.file)
                    workbook = load_workbook(filename=io.BytesIO(decoded_file), data_only=True)
                    sheet = workbook.active

                    #parsed_lines = [(5,)]
                    headers = []
                    start_row = 0

                    # Buscar fila con encabezados
                    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                        if row and "Fecha" in row:
                            headers = [str(cell).strip() if cell else '' for cell in row]
                            start_row = row_idx + 1
                            break

                    if not headers:
                        raise ValidationError("No se encontraron encabezados válidos en el archivo.")

                    # Mapear columnas esperadas
                    def col(name):
                        return headers.index(name) if name in headers else -1

                    for i, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=1):
                        if not row or not row[col('Fecha')]:
                            continue

                        try:
                            fecha = row[col('Fecha')]
                            if isinstance(fecha, str):
                                fecha = datetime.strptime(fecha.strip(), '%Y-%m-%d %H:%M:%S')
                            elif isinstance(fecha, datetime):
                                pass
                            else:
                                continue  # No es una fecha válida

                            referencia = str(row[col('Referencia')]).strip() if col('Referencia') != -1 else ''
                            transaccion = str(row[col('Transacción')]).strip() if col('Transacción') != -1 else ''
                            signo = str(row[col('Signo')]).strip() if col('Signo') != -1 else ''
                            valor = float(str(row[col('Valor')]).replace('$', '').replace(',', '').strip()) if col(
                                'Valor') != -1 and row[col('Valor')] else 0.0
                            saldo_contable = float(
                                str(row[col('Saldo Contable')]).replace('$', '').replace(',', '').strip()) if col(
                                'Saldo Contable') != -1 and row[col('Saldo Contable')] else 0.0
                            disponible = float(
                                str(row[col('Disponible')]).replace('$', '').replace(',', '').strip()) if col(
                                'Disponible') != -1 and row[col('Disponible')] else 0.0
                            oficina = str(row[col('Oficina')]).strip() if col('Oficina') != -1 else ''
                            referencia1 = str(row[col('Referencia 1')]).strip() if col('Referencia 1') != -1 and row[
                                col('Referencia 1')] else ''
                            referencia2 = str(row[col('Referencia 2')]).strip() if col('Referencia 2') != -1 and row[
                                col('Referencia 2')] else ''

                            tipo = 'credit' if signo == '(-)' else 'debit'
                            ref_numeros = '-'.join(filter(None, [referencia1, referencia2])) or referencia

                            type_obj = _match_type_from_description(self,transaccion)
                            tipo_id = type_obj.id if type_obj else uncategorized_type.id

                            parsed_lines.append((0, 0, {
                                'date': fecha.date(),
                                'reference': referencia1,
                                'description': transaccion,
                                'transaction_type': tipo,
                                'amount': valor,
                                'available_balance': disponible,
                                'book_balance': saldo_contable,
                                'location': oficina,
                                'company_id': record.company_id.id,
                                'sequence': i,
                                'type_id':tipo_id
                            }))

                        except Exception as e:
                            raise ValidationError(f"Error al procesar fila {i + start_row}:\n{e}")
                except Exception as e:
                        raise ValidationError("Error al leer el archivo Excel: " + str(e))
            else:
                raise ValidationError(
                    "El nombre del diario no coincide con un banco soportado (bolivariano o internacional).")
            record.line_ids = parsed_lines
            record._onchange_generate_groups()
            record.match_lines()

    def action_draft(self):
        for rec in self:
            rec.write({'state': 'draft'})
        return True

    def action_generate(self):
        for rec in self:
            rec.update_file()
            rec.write({'state': 'generated'})
        return True

    def action_confirm(self):
        for rec in self:
            rec.write({'state': 'confirmed'})
        return True

    def action_done(self):
        DEC = 2
        for rec in self:
            total_summary = 0.00
            for summary in rec.summary_line_ids:
                total_summary += summary.total_amount
            total_summary = round(total_summary, DEC)
            total_lines = 0.00
            for line in rec.line_ids:
                total_lines += line.signed_amount
            total_lines = round(total_lines, DEC)
            if total_summary != total_lines:
                raise ValidationError(
                    _("La suma del detalle %s debe ser igual a la suma del resumen %s") % (total_lines, total_summary))
            rec.write({'state': 'done'})
        return True

    def action_cancel(self):
        for rec in self:
            rec.write({'state': 'cancelled'})
        return True

    def unlink(self):
        for rec in self:
            if rec.state != 'draft':
                raise UserError("Solo se puede eliminar si el estado es 'Preliminar'.")
        return super().unlink()

    @api.onchange('line_ids')
    def _onchange_generate_groups(self):
        # Limpiar grupos actuales
        self.group_ids = [(5, 0, 0)]

        # Agrupar líneas por (date, reference, transaction_type)
        groups_dict = defaultdict(list)
        initial_balance = 0.00
        final_balance = 0.00
        i=0
        max=len(self.line_ids)
        for line in self.line_ids:
            i+=1
            if i==1:
                initial_balance=line.available_balance+line.amount
            if i==max:
                final_balance = line.available_balance
            key = (line.date, line.reference, line.transaction_type)
            groups_dict[key].append(line)
        self.initial_balance=initial_balance
        self.final_balance = final_balance
        new_groups = []
        for (date, reference, transaction_type), lines in groups_dict.items():
            total_amount = sum(l.amount for l in lines)
            group_vals = {
                'company_id': self.company_id.id,
                'date': date,
                'reference': reference,
                'transaction_type': transaction_type,
                'amount': total_amount,
                'description': lines[0].description,
                'line_ids': [(6, 0, [l.id for l in lines])],
            }
            new_groups.append((0, 0, group_vals))

        self.group_ids = new_groups

    def print_report(self):
        # Obtenemos la acción de reporte
        action = self.env.ref('gps_bancos.report_document_bank_recon_xlsx_act')

        if not action:
            raise  ValidationError("No se pudo encontrar el reporte.")

        # Retornamos la acción para imprimir el reporte
        return action.report_action(self)

    cheques_no_cobrados = fields.Monetary(
        string='(-) Cheques Girados No Cobrados',
        compute='_compute_saldos',
        store=False,readonly=True
    )
    nd_no_reg_bancos = fields.Monetary(
        string='(-) N/D No Registrados en Bancos',
        compute='_compute_saldos',
        store=False,readonly=True
    )
    depositos_no_banco = fields.Monetary(
        string='(+) Depósitos No en el Banco',
        compute='_compute_saldos',
        store=False,readonly=True
    )
    depositos_nc_no_libros = fields.Monetary(
        string='(-) Depósitos y N/C No Registrados en Libros',
        compute='_compute_saldos',
        store=False,readonly=True
    )
    nd_no_reg_libros = fields.Monetary(
        string='(+) N/D No Registrado en Libros',
        compute='_compute_saldos',
        store=False,readonly=True
    )
    saldo_inicial_meses_anteriores = fields.Monetary(
        string='Saldo Inicial Meses Anteriores',
        compute='_compute_saldos',
        store=False, readonly=True
    )
    saldo_conciliacion = fields.Monetary(
        string='Saldo Conciliación',
        compute='_compute_saldos',
        store=False, readonly=True
    )

    calculated_difference = fields.Monetary(
        string='Diferencia con Saldo Conciliación',
        compute='_compute_saldos',
        store=False,readonly=True
    )

    saldo_fecha_limite = fields.Monetary(
        string='Saldo Fecha Limite',
        compute='_compute_saldos',
        store=False, readonly=True
    )


    @api.depends('company_id','journal_id','account_id','date_from','date_to','initial_balance','final_balance','line_ids')
    def _compute_saldos(self):
        DEC=2
        for rec in self:

            ################################3################3################3################3
            srch_depositos_nc_no_libros=self.env["document.bank.reconciliation.line"].sudo().search([
                ('document_id','=',rec.id),
                ('transaction_type', '=','credit'),
                ('group_id.full_reconciled', '!=', True),

            ])

            srch_nd_no_reg_libros = self.env["document.bank.reconciliation.line"].sudo().search([
                ('document_id', '=', rec.id),
                ('transaction_type', '=', 'debit'),
                ('group_id.full_reconciled', '!=', True),

            ])

            depositos_nc_no_libros =sum(srch_depositos_nc_no_libros.mapped('amount'))
            nd_no_reg_libros = sum(srch_nd_no_reg_libros.mapped('amount'))
            ################################3################3################3################3
            cheques_no_cobrados = 0.0
            #nd_no_reg_bancos = 0.0
            #depositos_no_banco = 0.0

            lines=rec.move_line_ids
            matched_line_ids=rec.group_ids.mapped('move_line_ids')
            unmatched_lines = lines-matched_line_ids
            nd_no_reg_bancos=sum(unmatched_lines.filtered(lambda x: x.debit>0.00).mapped('debit'))
            depositos_no_banco = sum(unmatched_lines.filtered(lambda x: x.credit>0.00).mapped('credit'))
            ####################################################################
            self._cr.execute(''';WITH VARIABLES AS (
    SELECT 
        %s::INT AS COMPANY_ID,
        %s::INT AS JOURNAL_ID,
        %s::INT AS ACCOUNT_ID ,
         %s::DATE AS DATE_FROM,
        %s::DATE AS DATE_TO
),
SALDO_INICIAL AS (
    SELECT 
        COALESCE(SUM(AML.DEBIT), 0) AS debit_inicial,
        COALESCE(SUM(AML.CREDIT), 0) AS credit_inicial
    FROM VARIABLES
    INNER JOIN ACCOUNT_MOVE AM 
        ON AM.COMPANY_ID = VARIABLES.COMPANY_ID 
        AND AM.STATE = 'posted'
    INNER JOIN ACCOUNT_MOVE_LINE AML 
        ON AML.MOVE_ID = AM.ID 
        AND AML.ACCOUNT_ID = VARIABLES.ACCOUNT_ID 
        AND AML.DATE < VARIABLES.DATE_FROM
),
SALDO_FINAL AS (
    SELECT 
        COALESCE(SUM(AML.DEBIT), 0) AS debit_final,
        COALESCE(SUM(AML.CREDIT), 0) AS credit_final
    FROM VARIABLES
    INNER JOIN ACCOUNT_MOVE AM 
        ON AM.COMPANY_ID = VARIABLES.COMPANY_ID 
        AND AM.STATE = 'posted'
    INNER JOIN ACCOUNT_MOVE_LINE AML 
        ON AML.MOVE_ID = AM.ID 
        AND AML.ACCOUNT_ID = VARIABLES.ACCOUNT_ID 
        AND AML.DATE <= VARIABLES.DATE_TO
)

SELECT 
    si.debit_inicial,
    si.credit_inicial,
    si.debit_inicial - si.credit_inicial AS saldo_inicial,
    sf.debit_final,
    sf.credit_final,
    sf.debit_final - sf.credit_final AS saldo_final
FROM SALDO_INICIAL si, SALDO_FINAL sf;  ''',(rec.company_id.id,rec.journal_id.id,rec.account_id.id,rec.date_from,rec.date_to))
            print(rec.company_id.id,rec.journal_id.id,rec.account_id.id,rec.date_to)
            result_saldo_fecha_limite=self._cr.dictfetchone()
            print(result_saldo_fecha_limite)
            saldo_fecha_limite=0.00
            saldo_inicial_meses_anteriores = 0.0
            if result_saldo_fecha_limite:
                saldo_fecha_limite=result_saldo_fecha_limite['saldo_final']
                saldo_inicial_meses_anteriores = result_saldo_fecha_limite['saldo_inicial']
            ####################################################################
            rec.cheques_no_cobrados = cheques_no_cobrados
            rec.nd_no_reg_bancos = nd_no_reg_bancos
            rec.depositos_no_banco = depositos_no_banco
            rec.depositos_nc_no_libros = depositos_nc_no_libros
            rec.nd_no_reg_libros = nd_no_reg_libros
            rec.saldo_inicial_meses_anteriores = saldo_inicial_meses_anteriores
            rec.saldo_fecha_limite=saldo_fecha_limite
            rec.saldo_conciliacion = rec.final_balance
            calculated_difference = round(saldo_fecha_limite-rec.final_balance,DEC)
            rec.calculated_difference=calculated_difference

    def action_open_reconciliation_lines(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': 'Líneas de Conciliacion Bancaria',
            'res_model': 'document.bank.reconciliation.line',
            'view_mode': 'tree,form',
            'views': [(self.env.ref('gps_bancos.view_document_bank_reconciliation_line_tree').id, 'tree')],
            'domain': [('document_id', '=', self.id)],  # ajusta el dominio si aplica
            'context': {'default_document_id': self.id,'search_default_full_reconciled_no':1},
            'target': 'current',
        }

    summary_line_ids = fields.One2many(
        'document.bank.reconciliation.summary',
        'document_id',
        string="Resumen por Tipo"
    )

    def action_generate_summary(self):
        uncategorized_type = self.env['document.bank.reconciliation.type'].search([('code', '=', 'sin_tipo')], limit=1)
        if not uncategorized_type:
            raise ValidationError("Debe existir un tipo con código 'sin_tipo' para registrar líneas sin clasificación.")

        for rec in self:
            rec.summary_line_ids.unlink()  # Borra resumen previo

            com_type_id=self.env.ref('gps_bancos.reconciliation_comisiones').id
            key_com = (com_type_id, rec.company_id.id, rec.company_id.currency_id.id)

            data = {key_com:0.00}
            for line in rec.line_ids:
                type_id = line.type_id.id if line.type_id else uncategorized_type.id
                sign = 1 if line.transaction_type == 'debit' else -1
                key = (type_id, line.company_id.id, line.company_id.currency_id.id)
                data.setdefault(key, 0.0)
                data[key] += sign * line.amount

            for (type_id, company_id, currency_id), total in data.items():
                brw_type=self.env['document.bank.reconciliation.type'].sudo().browse(type_id)
                self.env['document.bank.reconciliation.summary'].create({
                    'document_id': rec.id,
                    'type_id': type_id,
                    'company_id': company_id,
                    'currency_id': currency_id,
                    'total_amount': total,
                    'can_edit':brw_type.can_edit
                })


    def get_amounts_grouped_by_type(self):
        self.ensure_one()
        rec=self
        lines = rec.summary_line_ids
        result = {}
        for line in lines:
            code = line.type_id.code or 'sin_tipo'
            result[code] = result.get(code, 0.0) + line.total_amount
        return result