# coding: utf-8
from odoo import api, fields, models, exceptions, tools, _
from ...message_dialog.tools import FileManager
from ...calendar_days.tools import DateManager
from ...calendar_days.tools import CalendarManager
dtFile=FileManager()        
dateO=DateManager()
calendarO=CalendarManager()
import openpyxl
import logging
_logger = logging.getLogger(__name__)

from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font, PatternFill
import openpyxl.utils as utils
from openpyxl.styles import NamedStyle

class report_acct_analytic_acct_xlsx(models.AbstractModel):
    _name = "report.gps_cuentas_analiticas.report_acct_analytic_acct_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Cuentas Analiticas"
    
    def create_xlsx_report(self, docids, data):
        EXT="xlsx"
        dir_path=dtFile.get_path_file(__file__)
        new_filename=dtFile.create(EXT)  
        filename=dtFile.join(dir_path,"reporte_cuentas_analiticas.xlsx")
        dtFile.copyfile(filename,new_filename)
        wb=False
        try:
            brw_wizard=self.env["account.analytic.account.wizard"].sudo().browse(docids[-1])
            wb, ws, target = self.open_xlsx(new_filename, load_sheet=False)
            ws = wb["REPORTE DE CUENTAS ANALITICAS"]
            self.create_report_cuentas_analiticas(brw_wizard,ws)
            wb = self.save_wb(wb, target)
            #########################################
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

    def clean_text(self,value):
        import re
        """ Elimina caracteres ilegales que pueden causar errores en Excel """
        if isinstance(value, str):
            # Reemplazar caracteres no imprimibles o ilegales con una cadena vacía
            return re.sub(r'[^\x09\x0A\x0D\x20-\x7E]', '', value)
        return value

    def create_report_cuentas_analiticas(self,brw_wizard,ws):
        companys = ', '.join(brw_wizard.mapped('company_ids').mapped('name'))
        analitic_accounts=', '.join(brw_wizard.mapped('analytic_ids').mapped('name'))
        company_ids = brw_wizard.company_ids.ids
        analytic_ids = brw_wizard.analytic_ids.ids
        journal_ids = brw_wizard.journal_ids.ids
        self._cr.execute(f""";WITH VARIABLES AS (
	SELECT
		array{company_ids}::INT[] AS COMPANY_IDS,
		array{analytic_ids}::INT[] AS ANALYTIC_IDS,
		array{journal_ids}::INT[] AS JOURNAL_IDS
),
cuentas_analiticas AS (
    SELECT aaa.id
    FROM variables 
    INNER JOIN account_analytic_account aaa ON aaa.company_id = ANY(variables.company_ids)
    WHERE (
           cardinality(variables.analytic_ids) = 0 OR
		   	(cardinality(variables.analytic_ids) >0 and aaa.id = ANY(variables.analytic_ids))
		   )
),diarios_filtrados AS (
    SELECT aj.id
    FROM variables 
    INNER JOIN account_journal aj ON aj.company_id = ANY(variables.company_ids)
    WHERE (
           cardinality(variables.journal_ids) = 0 OR
           (cardinality(variables.journal_ids) > 0 AND aj.id = ANY(variables.journal_ids))
    )
),
resumen_asientos as (
	select  rc.name as company_name,
			aaa.id as cuenta_analitica_id,
			coalesce(aaa.code,'') as cuenta_analitica_code,
			aaa.name as cuenta_analitica_name,
			coalesce(rp.vat,'') as cliente_identificacion,
			rp.name as cliente_nombre,
			coalesce(aap.name,'') as plan,
			cctp.name as proyecto_name,
			coalesce(aaa.nro_contrato,'') as nro_contrato,
			aal.date as fecha,
			aal.amount as importe,
			aal.name as descr,
			aal.move_line_id  as linea_asiento_id,
			aml.move_id as asiento_id ,
			aml.debit as linea_debito,
			aml.credit as linea_credito,
			aml.credit-aml.debit as balance,
			am.name as numero_asiento,
			am.date as fecha_asiento,
			aj.name as diario_asiento,
			coalesce(am.ref,'') as ref_asiento,
			case when(am.state='posted') then 'PUBLICADO'
			when(am.state='draft') then 'PRELIMINAR' 
			when(am.state='cancel') then 'ANULADO'
			ELSE '' end as estado_asiento,
			CASE 
			        WHEN am.move_type = 'entry' THEN 'Asiento Contable'
			        WHEN am.move_type = 'out_invoice' THEN 'Factura de Cliente'
			        WHEN am.move_type = 'in_invoice' THEN 'Factura de Proveedor'
			        WHEN am.move_type = 'out_refund' THEN 'Nota de Crédito de Cliente'
			        WHEN am.move_type = 'in_refund' THEN 'Nota de Crédito de Proveedor'
			        WHEN am.move_type = 'out_receipt' THEN 'Recibo de Cliente'
			        WHEN am.move_type = 'in_receipt' THEN 'Recibo de Proveedor'
        ELSE 'Tipo Desconocido' END as tipo_asiento,
	    aa.code as account_code,
	    aa.name as account_name,
	    coalesce(lldt.name,'') as tipo_documento_name ,
		case when(aaa.state='active' ) then 'ACTIVO' else 'INACTIVO' end as estado,
		coalesce(replace((pt.NAME::JSON->'es_EC')::varchar,'"',''),'')    as product_name,
		pt.default_Code as product_default_code ,
		aml.purchase_line_id
	from 
	account_analytic_account aaa
	inner join cuentas_analiticas  can on can.id=aaa.id 
	inner join account_analytic_line aal on aal.account_id=aaa.id
	inner join account_move_line aml on aml.id=aal.move_line_id 
	left join product_product pp on pp.id=aml.product_id 
	left join product_template pt on pt.id=pp.product_tmpl_id  
	inner join account_move am on am.id=aml.move_id 
	inner join diarios_filtrados  dft on dft.id=am.journal_id 
	left join account_journal aj on aj.id=am.journal_id 
	left join account_analytic_plan aap on aap.id=aaa.plan_id 
	left join centro_costo_tipo_proyecto cctp on cctp.id=aaa.tipo_proy_id
	inner join res_company rc on rc.id=aaa.company_id 
	left join res_partner rp on rp.id=aaa.partner_id 
	left join account_account aa on aa.id=aml.account_id
	left join l10n_latam_document_type lldt on lldt.id=am.l10n_latam_document_type_id  
	
	order by  rc.name asc,aaa.name asc,am.date asc
),
ordenes_compra as (
	select ra.linea_asiento_id,   
		STRING_AGG(pol.id::TEXT, ',') as PURCHASE_order_line_ids,
		STRING_AGG(po.id::TEXT, ',')  AS purchase_order_ids,
		STRING_AGG(po.name::TEXT, ',')  as purchase_order_names, 
		STRING_AGG((po.date_approve::date)::TEXT, ',') as purchase_order_dates
	from resumen_asientos ra
	inner join purchase_order_line pol on pol.id=ra.purchase_line_id
	inner join purchase_order po on po.id=pol.order_id 
	GROUP BY ra.linea_asiento_id
),
ordenes_venta as (
	SELECT 
    aml.id AS linea_asiento_id,
	STRING_AGG(sol.id::TEXT, ',') AS sale_order_line_ids,
    STRING_AGG(so.id::TEXT, ',') AS sale_order_ids,
    STRING_AGG(so.name, ',') AS sale_order_names,
    STRING_AGG((so.date_order::date)::TEXT, ',') AS sale_order_dates 
	FROM account_move_line aml
	inner JOIN sale_order_line_invoice_rel rel ON rel.invoice_line_id = aml.id
	inner JOIN sale_order_line sol ON rel.order_line_id = sol.id
	inner JOIN sale_order so ON sol.order_id = so.id
	inner JOIN account_move am ON aml.move_id = am.id
	GROUP BY aml.id
),
rel_asientos_picking as (
	SELECT 
    aml.id AS account_move_line_id,
    svl.id AS stock_valuation_layer_id,
    sm.id AS stock_move_id,
   	sp.id AS stock_picking_id 
	FROM account_move_line aml
	JOIN account_move am ON aml.move_id = am.id
	JOIN stock_valuation_layer svl ON svl.account_move_id = am.id
	JOIN stock_move sm ON svl.stock_move_id = sm.id
	JOIN stock_picking sp ON sm.picking_id = sp.id 
	GROUP BY 
	aml.id,svl.id ,sm.id,sp.id 
),
pickings as (
	select rap.account_move_line_id as linea_asiento_id,
			rap.stock_picking_id,
			spk.date_done as stock_picking_date,
			spk.name as stock_picking_name,
		  	null as PURCHASE_order_line_ids,
			STRING_AGG(po.id::TEXT, ',')  AS purchase_order_ids,
			STRING_AGG(po.name::TEXT, ',')  as purchase_order_names, 
			STRING_AGG((po.date_approve::date)::TEXT, ',') as purchase_order_dates,
			null AS sale_order_line_ids,
	    	STRING_AGG(so.id::TEXT, ',') AS sale_order_ids,
	    	STRING_AGG(so.name, ',') AS sale_order_names,
	    	STRING_AGG((so.date_order::date)::TEXT, ',') AS sale_order_dates 
			
		from rel_asientos_picking rap	
		inner join stock_picking spk on spk.id=rap.stock_picking_id
		left join purchase_order po on po.name=spk.origin and po.name ilike 'OC%'
		left join sale_order so on so.name=spk.origin and so.name ilike 'SO%'
		group by rap.account_move_line_id,rap.stock_picking_id,spk.date_done,spk.name
)

select ra.*,

coalesce(oc.PURCHASE_order_line_ids,pk.PURCHASE_order_line_ids) as purchase_order_line_ids,
coalesce(oc.purchase_order_ids,pk.purchase_order_ids) as purchase_order_ids,
coalesce(oc.purchase_order_names,pk.purchase_order_names) as purchase_order_names,
coalesce(oc.purchase_order_dates,pk.purchase_order_dates) as purchase_order_dates,

coalesce(ov.sale_order_line_ids,pk.sale_order_line_ids) as sale_order_line_ids,
coalesce(ov.sale_order_ids,pk.sale_order_ids) as sale_order_ids,
coalesce(ov.sale_order_names,pk.sale_order_names) as sale_order_names,
coalesce(ov.sale_order_dates,pk.sale_order_dates) as sale_order_dates,

pk.stock_picking_id,
pk.stock_picking_name,
pk.stock_picking_date

from resumen_asientos ra
left join ordenes_compra oc on oc.linea_asiento_id=ra.linea_asiento_id
left join ordenes_venta ov on ov.linea_asiento_id=ra.linea_asiento_id 
left join pickings pk on pk.linea_asiento_id=ra.linea_asiento_id  """)
        result = self._cr.dictfetchall()
        if result:
            i, INDEX_ROW = 0, 5
            for each_result in result:
                row = str(INDEX_ROW + i)
                ws['A' + row] = each_result["company_name"]
                ws['B' + row] = each_result["cuenta_analitica_id"]
                ws['C' + row] = self.clean_text(each_result["cuenta_analitica_code"])
                ws['D' + row] = self.clean_text(each_result["cuenta_analitica_name"])
                ws['E' + row] = self.clean_text(each_result["plan"])
                ws['F' + row] = each_result["cliente_identificacion"]
                ws['G' + row] = self.clean_text(each_result["cliente_nombre"])
                ws['H' + row] = each_result["nro_contrato"]
                ws['I' + row] = self.clean_text(each_result["proyecto_name"])
                ws['J' + row] = each_result["cuenta_analitica_id"]
                ws['K' + row] = each_result["asiento_id"]
                ws['L' + row] = each_result["linea_asiento_id"]
                ws['M' + row] = self.clean_text(each_result["descr"])
                ws['N' + row] = each_result["fecha"]
                ws['O' + row] = each_result["importe"]
                ws['P' + row] = each_result["account_code"]
                ws['Q' + row] = self.clean_text(each_result["account_name"])
                ws['R' + row] = each_result["linea_debito"]
                ws['S' + row] = each_result["linea_credito"]
                ws['T' + row] = each_result["balance"]
                ws['U' + row] = each_result["tipo_asiento"]
                ws['V' + row] = each_result["tipo_documento_name"]
                ws['W' + row] = each_result["fecha_asiento"]
                ws['X' + row] = each_result["numero_asiento"]
                ws['Y' + row] = each_result["diario_asiento"]
                ws['Z' + row] = each_result["ref_asiento"]
                ws['AA' + row] = each_result["estado_asiento"]
                ws['AB' + row] = each_result["estado"]

                ws['AC' + row] = each_result["product_default_code"]
                ws['AD' + row] = each_result["product_name"]

                ws['AE' + row] = each_result["purchase_order_line_ids"]
                ws['AF' + row] = each_result["purchase_order_ids"]
                ws['AG' + row] = each_result["purchase_order_names"]
                ws['AH' + row] = each_result["sale_order_dates"]

                ws['AI' + row] = each_result["sale_order_line_ids"]
                ws['AJ' + row] = each_result["sale_order_ids"]
                ws['AK' + row] = each_result["sale_order_names"]
                ws['AL' + row] = each_result["sale_order_dates"]

                ws['AM' + row] = each_result["stock_picking_id"]
                ws['AN' + row] = each_result["stock_picking_name"]
                ws['AO' + row] = each_result["stock_picking_date"]

                i += 1
            last_row = INDEX_ROW + i
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':AO' + str(last_row - 1), border)
            ws['A1'] = companys
            ws['B2'] = analytic_ids and analitic_accounts or 'TODAS'
            ws['B3'] = len(result)

class account_anlyt_acct_move_xlsx(models.AbstractModel):
    _name = "report.gps_cuentas_analiticas.account_anlyt_acct_move_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Asientos por Cuentas Analiticas"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)
        filename = dtFile.join(dir_path, "reporte_asientos_cuentas_analiticas.xlsx")
        dtFile.copyfile(filename, new_filename)
        wb = False
        try:
            brw_wizard = self.env["account.analytic.account.wizard"].sudo().browse(docids[-1])
            wb, ws, target = self.open_xlsx(new_filename, load_sheet=False)
            ws = wb["REPORTE DE ASIENTOS DE CUENTAS"]
            self.create_report_cuentas_analiticas(brw_wizard, ws)
            wb = self.save_wb(wb, target)
            #########################################
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

    def clean_text(self, value):
        import re
        """ Elimina caracteres ilegales que pueden causar errores en Excel """
        if isinstance(value, str):
            # Reemplazar caracteres no imprimibles o ilegales con una cadena vacía
            return re.sub(r'[^\x09\x0A\x0D\x20-\x7E]', '', value)
        return value

    def create_report_cuentas_analiticas(self, brw_wizard, ws):
        companys = ', '.join(brw_wizard.mapped('company_ids').mapped('name'))
        analitic_accounts = ', '.join(brw_wizard.mapped('analytic_ids').mapped('name'))
        company_ids = brw_wizard.company_ids.ids
        analytic_ids = brw_wizard.analytic_ids.ids
        journal_ids = brw_wizard.journal_ids.ids
        date_from=brw_wizard.date_from
        date_to = brw_wizard.date_to
        self._cr.execute(f""";WITH VARIABLES AS (
	SELECT
		array{company_ids}::INT[] AS COMPANY_IDS,
		array{analytic_ids}::INT[] AS ANALYTIC_IDS,
		'{date_from}'::date as date_from,
		'{date_to}'::date as date_to,
		array{journal_ids}::INT[] AS JOURNAL_IDS
),
cuentas_analiticas AS (
    SELECT aaa.id
    FROM variables 
    INNER JOIN account_analytic_account aaa ON aaa.company_id = ANY(variables.company_ids)
    WHERE (
           cardinality(variables.analytic_ids) = 0 OR
		   	(cardinality(variables.analytic_ids) >0 and aaa.id = ANY(variables.analytic_ids))
		   )
),diarios_filtrados AS (
    SELECT aj.id
    FROM variables 
    INNER JOIN account_journal aj ON aj.company_id = ANY(variables.company_ids)
    WHERE (
           cardinality(variables.journal_ids) = 0 OR
           (cardinality(variables.journal_ids) > 0 AND aj.id = ANY(variables.journal_ids))
    )
),
resumen_asientos as (
	select  rc.name as company_name,
			aaa.id as cuenta_analitica_id,
			coalesce(aaa.code,'') as cuenta_analitica_code,
			aaa.name as cuenta_analitica_name,
			coalesce(rp.vat,'') as cliente_identificacion,
			rp.name as cliente_nombre,
			rp.vat as cliente_identificacion,
			coalesce(aap.name,'') as plan,
			cctp.name as proyecto_name,
			coalesce(aaa.nro_contrato,'') as nro_contrato,
			aal.date as fecha,
			case when(aaa.id  is not null ) then	aal.amount else  aml.credit-aml.debit end as importe,
			aal.name as descr,  
			aml.id  as linea_asiento_id,
			aml.move_id as asiento_id ,
			aml.debit as linea_debito,
			aml.credit as linea_credito,
			aml.credit-aml.debit as balance,
			
			case when(aaa.id  is not null ) then
		        ROUND(case when(abs(aml.credit-aml.debit)>0) then aal.amount/(aml.credit-aml.debit) else 1.00 end,4)
		else 1 end  as porcentaje,
		
			am.name as numero_asiento,
			am.date as fecha_asiento,
			aj.name as diario_asiento,
			coalesce(am.ref,'') as ref_asiento,
			case when(am.state='posted') then 'PUBLICADO'
			when(am.state='draft') then 'PRELIMINAR' 
			when(am.state='cancel') then 'ANULADO'
			ELSE '' end as estado_asiento,
			CASE 
			        WHEN am.move_type = 'entry' THEN 'Asiento Contable'
			        WHEN am.move_type = 'out_invoice' THEN 'Factura de Cliente'
			        WHEN am.move_type = 'in_invoice' THEN 'Factura de Proveedor'
			        WHEN am.move_type = 'out_refund' THEN 'Nota de Crédito de Cliente'
			        WHEN am.move_type = 'in_refund' THEN 'Nota de Crédito de Proveedor'
			        WHEN am.move_type = 'out_receipt' THEN 'Recibo de Cliente'
			        WHEN am.move_type = 'in_receipt' THEN 'Recibo de Proveedor'
        ELSE 'Tipo Desconocido' END as tipo_asiento,
	    aa.code as account_code,
	    aa.name as account_name,
		 case when(aaa.id  is not null ) then
		        (case when(aaa.state='active' ) then 'ACTIVO' else 'INACTIVO' end )
		else '' end as estado_cta,
		coalesce(replace((pt.NAME::JSON->'es_EC')::varchar,'"',''),'')    as product_name,
		pt.default_Code as product_default_code 
		
	from VARIABLES 
	INNER JOIN res_company rc ON rc.id = ANY(variables.company_ids)
	inner join account_move am   on rc.id=am.company_id and am.date>=variables.date_from  and am.date<=variables.date_to 
	inner join diarios_filtrados  dft on dft.id=am.journal_id  
	inner join account_journal aj on aj.id=am.journal_id 
	inner join account_move_line aml  on am.id=aml.move_id 	
	inner join account_account aa on aa.id=aml.account_id 
	left join res_partner rp on rp.id=coalesce(aml.partner_id ,am.partner_id)
	left join l10n_latam_document_type lldt on lldt.id=am.l10n_latam_document_type_id  
	left join product_product pp on pp.id=aml.product_id 
	left join product_template pt on pt.id=pp.product_tmpl_id  
	left join account_analytic_line aal on aal.move_line_id=aml.id
	left join account_analytic_account aaa on aaa.id=aal.account_id	
	left join account_analytic_plan aap on aap.id=aaa.plan_id 
	left join centro_costo_tipo_proyecto cctp on cctp.id=aaa.tipo_proy_id
	WHERE (
           cardinality(variables.analytic_ids) = 0 OR
		   	(cardinality(variables.analytic_ids) >0 and aaa.id = ANY(variables.analytic_ids))
		   )
)

select ra.* 
from resumen_asientos ra """)
        result = self._cr.dictfetchall()
        if result:
            i, INDEX_ROW = 0, 5
            for each_result in result:
                row = str(INDEX_ROW + i)
                ws['A' + row] = each_result["company_name"]
                ws['B' + row] = each_result["asiento_id"]

                ws['C' + row] = each_result["cliente_identificacion"]
                ws['D' + row] = each_result["cliente_nombre"]

                ws['E' + row] = each_result["tipo_asiento"]
                ws['F' + row] = each_result["fecha_asiento"]
                ws['G' + row] = each_result["numero_asiento"]
                ws['H' + row] = each_result["diario_asiento"]
                ws['I' + row] = each_result["ref_asiento"]
                ws['J' + row] = each_result["estado_asiento"]

                ws['K' + row] = each_result["linea_asiento_id"]
                ws['L' + row] = each_result["product_default_code"]
                ws['M' + row] = each_result["product_name"]
                ws['N' + row] = each_result["account_code"]
                ws['O' + row] = self.clean_text(each_result["account_name"])
                ws['P' + row] = self.clean_text(each_result["cuenta_analitica_code"])
                ws['Q' + row] = self.clean_text(each_result["cuenta_analitica_name"])
                ws['R' + row] = self.clean_text(each_result["plan"])
                ws['S' + row] = each_result["linea_debito"]
                ws['T' + row] = each_result["linea_credito"]
                ws['U' + row] = each_result["balance"]

                ws['V' + row] = each_result["importe"]
                ws['W' + row] = each_result["porcentaje"]


                ws['X' + row] = each_result["estado_cta"]

                i += 1
            last_row = INDEX_ROW + i
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':X' + str(last_row - 1), border)
            ws['A1'] = companys
            ws['B3'] = analytic_ids and analitic_accounts or 'TODAS'
            ws['B2'] = date_from
            ws['D2'] = date_to
            ws['D3'] = len(result)