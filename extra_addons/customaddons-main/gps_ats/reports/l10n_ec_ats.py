# coding: utf-8
from odoo import api, fields, models, exceptions, tools, _
from ...message_dialog.tools import FileManager
from ...calendar_days.tools import DateManager
from ...calendar_days.tools import CalendarManager
dtFile=FileManager()        
dateO=DateManager()
calendarO=CalendarManager()

import logging
_logger = logging.getLogger(__name__)

from openpyxl.styles.borders import Border, Side

class report_ventas_report_xlsx(models.AbstractModel):
    _name = "report.gps_ats.report_ventas_xlsx"    
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Ventas"
    
    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        new_filename = dtFile.create(EXT)
        def get_ws_objects(template_file_name):
            dir_path = dtFile.get_path_file(__file__)
            filename = dtFile.join(dir_path, template_file_name)
            dtFile.copyfile(filename, new_filename)
            wb, ws, target = self.open_xlsx(new_filename, load_sheet=True)
            return wb, ws, target
        wb=False
        try:
            brw_wizard=self.env["l10n.ec.ats"].sudo().browse(docids[-1])
            start_date,end_date=brw_wizard.date_start,brw_wizard.date_end##lectura de parametros

            if brw_wizard.type_info=="grouped":
                wb, ws, target = get_ws_objects ("reporte_ventas.xlsx")
                self._cr.execute(""";with variables as (
        select %s::date as fecha_inicial,
        %s::date as fecha_final ,
        %s::int as company_id   
    
    ),
    retenciones_detalle as (
                    select account_move.id,
                        sum(aml.l10n_ec_withhold_tax_amount) as total,
                        atxg.l10n_ec_type
    
                    from res_company 
                    inner join variables on 1=1
                    inner join res_partner rp  on rp.id=res_company.partner_id
                    inner join account_move on account_move.company_id = res_company.id and account_move.move_type in ('out_invoice') and account_move.state in ('posted','paid')  and 
                    account_move.invoice_date>=variables.fecha_inicial and account_move.invoice_date<=variables.fecha_final and account_move.company_id=variables.company_id 
                    inner join account_journal aj on  aj.id=account_move.journal_id and coalesce(aj.l10n_latam_use_documents,false)=true 
                    inner join account_move_line aml on aml.l10n_ec_withhold_invoice_id=account_move.id 
                    inner join account_move_line_account_tax_rel atxrl on atxrl.account_move_line_id=aml.id 
                    inner join account_tax atx on atx.id=  atxrl.account_tax_id 
                    inner join account_tax_group atxg on atxg.id=atx.tax_group_id
                    inner join account_move awt on awt.id=aml.move_id and awt.state in ('posted','paid')  
                    where aml.l10n_ec_withhold_tax_amount>0.00 
                    group by     account_move.id        ,atxg.l10n_ec_type
    )
    
    
    select account_move.id,
    doc_document_type.name as  tipo_documento,
    replace(account_move.nAME,'-','') as numero_documento,
    account_move.invoice_date as fecha,
    case when(account_move.state='draft') then 'preliminar' 
    when(account_move.state='posted') then 'publicado'
    when(account_move.state='cancel') then 'anulado' end as estado,
    res_partner.name as cliente,
    res_partner.vat as id_cliente,
    account_move.authorization_type as tipo_autorizacion,
    coalesce(account_move.l10n_ec_authorization_number,'') as autorizacion,
    (case when(account_move.move_type='out_refund') then -1 else 1 end )*coalesce(account_move.amount_base0,0.00)+ coalesce(account_move.amount_baseno0,0.00) as subtotal,
    (case when(account_move.move_type='out_refund') then -1 else 1 end )*coalesce(account_move.amount_base0,0.00) as   base0,
    (case when(account_move.move_type='out_refund') then -1 else 1 end )*coalesce(account_move.amount_baseno0,0.00) as   baseno0,
    (case when(account_move.move_type='out_refund') then -1 else 1 end )*coalesce(account_move.amount_taxno0,0.00) as  iva,    
    (case when(account_move.move_type='out_refund') then -1 else 1 end )*coalesce(account_move.amount_total,0.00) as  total,   
    coalesce(rdfte.total,0.00) as  rte_fte,
    coalesce(rdiva.total,0.00) as  rte_iva,
    account_move.l10n_ec_authorization_date as fecha_autorizacion,
    case  when (coalesce(account_move.anulado_sri,false)=true) then 'SI' else 'NO' end as anulado_sri,
    coalesce(account_move.motivo_anulacion_sri,'') as motivo_anulacion 
                    from res_company 
                    inner join variables on 1=1
                    inner join res_partner rp  on rp.id=res_company.partner_id
                    inner join account_move on account_move.company_id = res_company.id and account_move.move_type in ('out_invoice','out_refund') and account_move.state in ('posted','paid')  and 
                    account_move.invoice_date>=variables.fecha_inicial and account_move.invoice_date<=variables.fecha_final  and account_move.company_id=variables.company_id 
                    inner join account_journal aj on  aj.id=account_move.journal_id and coalesce(aj.l10n_latam_use_documents,false)=true 
                    left join res_partner on account_move.partner_id = res_partner.id
                    left join l10n_latam_identification_type on res_partner.l10n_latam_identification_type_id = l10n_latam_identification_type.id
                    left join l10n_ec_sri_payment doc_payment_type on doc_payment_type.id = account_move.l10n_ec_sri_payment_id 
                    left join l10n_latam_document_type doc_document_type on doc_document_type.id = account_move.l10n_latam_document_type_id and doc_document_type.code!='00' 
                    left join retenciones_detalle rdfte on rdfte.id=account_move.id and rdfte.l10n_ec_type='withhold_vat_sale'
                    left join retenciones_detalle rdiva on rdiva.id=account_move.id and rdiva.l10n_ec_type='withhold_income_sale'
        """,(start_date,end_date,brw_wizard.company_id.id))
                result=self._cr.dictfetchall()
                if result:
                    i,INDEX_ROW=0,5
                    last_row=INDEX_ROW
                    for each_result in result:
                        row=str(INDEX_ROW+i)
                        ws['A'+row]= each_result["id"]
                        ws['B'+row]= each_result["tipo_documento"]
                        ws['C'+row]= each_result["numero_documento"]
                        ws['D'+row]= each_result["fecha"]
                        ws['E'+row]= each_result["estado"]
                        ws['F'+row]= each_result["cliente"]
                        ws['G'+row]= each_result["id_cliente"]
                        ws['H'+row]= each_result["tipo_autorizacion"]
                        ws['I'+row]= each_result["autorizacion"]

                        ws['J'+row]= each_result["fecha_autorizacion"]
                        ws['K'+row]= each_result["anulado_sri"]
                        ws['L'+row]= each_result["motivo_anulacion"]

                        ws['M'+row]= each_result["subtotal"]
                        ws['N'+row]= each_result["base0"]
                        ws['O'+row]= each_result["baseno0"]
                        ws['P'+row]= each_result["iva"]
                        ws['Q'+row]= each_result["total"]
                        ws['R'+row]= each_result["rte_fte"]
                        ws['S'+row]= each_result["rte_iva"]
                        i+=1
                        last_row=INDEX_ROW+i
                    if last_row>=INDEX_ROW:
                        thin = Side(border_style="thin", color="000000")
                        border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        self.set_border(ws,'A'+str(INDEX_ROW)+':S'+str(last_row-1),border)
                ws['B3']= len(result)
                ws['A1']= brw_wizard.company_id.name
                ws['B2']=start_date
                ws['D2']=end_date
                wb = self.save_wb(wb, target)
            else:
                wb, ws, target = get_ws_objects("reporte_ventas_detallado.xlsx")
                self._cr.execute(""";with variables as (
                        select %s::date as fecha_inicial,
                       %s::date as fecha_final ,
                       %s::int as company_id   

                    ),
                    retenciones_detalle as (
                                    select account_move.id,
                                        sum(aml.l10n_ec_withhold_tax_amount) as total,
                                        atxg.l10n_ec_type

                                    from res_company 
                                    inner join variables on 1=1
                                    inner join res_partner rp  on rp.id=res_company.partner_id
                                    inner join account_move on account_move.company_id = res_company.id and account_move.move_type in ('out_invoice') and account_move.state in ('posted','paid')  and 
                                    account_move.invoice_date>=variables.fecha_inicial and account_move.invoice_date<=variables.fecha_final and account_move.company_id=variables.company_id 
                                    inner join account_journal aj on  aj.id=account_move.journal_id and coalesce(aj.l10n_latam_use_documents,false)=true 
                                    inner join account_move_line aml on aml.l10n_ec_withhold_invoice_id=account_move.id 
                                    inner join account_move_line_account_tax_rel atxrl on atxrl.account_move_line_id=aml.id 
                                    inner join account_tax atx on atx.id=  atxrl.account_tax_id 
                                    inner join account_tax_group atxg on atxg.id=atx.tax_group_id
                                    inner join account_move awt on awt.id=aml.move_id and awt.state in ('posted','paid')  
                                    where aml.l10n_ec_withhold_tax_amount>0.00 
                                    group by     account_move.id        ,atxg.l10n_ec_type
                    )


                    select account_move.id,
                    doc_document_type.name as  tipo_documento,
                    replace(account_move.nAME,'-','') as numero_documento,
                    account_move.invoice_date as fecha,
                    case when(account_move.state='draft') then 'preliminar' 
                    when(account_move.state='posted') then 'publicado'
                    when(account_move.state='cancel') then 'anulado' end as estado,
                    res_partner.name as cliente,
                    res_partner.vat as id_cliente,
                    account_move.authorization_type as tipo_autorizacion,
                    coalesce(account_move.l10n_ec_authorization_number,'') as autorizacion,
                    (case when(account_move.move_type='out_refund') then -1 else 1 end )*(coalesce(account_move.amount_base0,0.00)+ coalesce(account_move.amount_baseno0,0.00)) as subtotal,
                    (case when(account_move.move_type='out_refund') then -1 else 1 end )*coalesce(account_move.amount_base0,0.00) as   base0,
                    (case when(account_move.move_type='out_refund') then -1 else 1 end )*coalesce(account_move.amount_baseno0,0.00) as   baseno0,
                    (case when(account_move.move_type='out_refund') then -1 else 1 end )*coalesce(account_move.amount_taxno0,0.00) as  iva,    
                    (case when(account_move.move_type='out_refund') then -1 else 1 end )*coalesce(account_move.amount_total,0.00) as  total,   
                    coalesce(rdfte.total,0.00) as  rte_fte,
                    coalesce(rdiva.total,0.00) as  rte_iva,
                    account_move.l10n_ec_authorization_date as fecha_autorizacion,
                    case  when (coalesce(account_move.anulado_sri,false)=true) then 'SI' else 'NO' end as anulado_sri,
                    coalesce(account_move.motivo_anulacion_sri,'') as motivo_anulacion ,
					pp.id AS product_id,
					pp.default_code,
					coalesce(pt.name::json->>'es_EC',pt.name::json->>'en_US') as product_name ,
					(case when(account_move.move_type='out_refund') then -1 else 1 end )*coalesce(aml.quantity,1) as cantidad,
					(case when(account_move.move_type='out_refund') then -1 else 1 end )*coalesce(aml.price_unit,0.00) as precio_unit,
					(case when(account_move.move_type='out_refund') then -1 else 1 end )*coalesce(aml.price_subtotal,0.00) as price_subtotal,
					(case when(account_move.move_type='out_refund') then -1 else 1 end )*(coalesce(aml.price_total,0.00)-coalesce(aml.price_subtotal,0.00)) as price_Tax,
					(case when(account_move.move_type='out_refund') then -1 else 1 end )*coalesce(aml.price_total,0.00) as price_total,
					case when (pt.detailed_type='product') then 'PRODUCTO INVENTARIABLE' 
					when (pt.detailed_type='consu') then 'CONSUMIBLE'
					when (pt.detailed_type='service') then 'SERVICIO'
					else '' end as tipo_producto 
                                    from res_company 
                                    inner join variables on 1=1
                                    inner join res_partner rp  on rp.id=res_company.partner_id
                                    inner join account_move on account_move.company_id = res_company.id and account_move.move_type in ('out_invoice','out_refund') and account_move.state in ('posted','paid')  and 
                                    account_move.invoice_date>=variables.fecha_inicial and account_move.invoice_date<=variables.fecha_final  and account_move.company_id=variables.company_id 
									inner join account_move_line aml on aml.move_id=account_move.id 
									inner join product_product pp on pp.id=aml.product_id 
									inner join product_template pt on pt.id=pp.product_Tmpl_id
                                    inner join account_journal aj on  aj.id=account_move.journal_id and coalesce(aj.l10n_latam_use_documents,false)=true 
                                    left join res_partner on account_move.partner_id = res_partner.id
                                    left join l10n_latam_identification_type on res_partner.l10n_latam_identification_type_id = l10n_latam_identification_type.id
                                    left join l10n_ec_sri_payment doc_payment_type on doc_payment_type.id = account_move.l10n_ec_sri_payment_id 
                                    left join l10n_latam_document_type doc_document_type on doc_document_type.id = account_move.l10n_latam_document_type_id and doc_document_type.code!='00' 
                                    left join retenciones_detalle rdfte on rdfte.id=account_move.id and rdfte.l10n_ec_type='withhold_vat_sale'
                                    left join retenciones_detalle rdiva on rdiva.id=account_move.id and rdiva.l10n_ec_type='withhold_income_sale'

  """, (start_date, end_date, brw_wizard.company_id.id))
                result = self._cr.dictfetchall()
                if result:
                    i, INDEX_ROW = 0, 5
                    last_row = INDEX_ROW
                    for each_result in result:
                        row = str(INDEX_ROW + i)
                        ws['A' + row] = each_result["id"]
                        ws['B' + row] = each_result["tipo_documento"]
                        ws['C' + row] = each_result["numero_documento"]
                        ws['D' + row] = each_result["fecha"]
                        ws['E' + row] = each_result["estado"]
                        ws['F' + row] = each_result["cliente"]
                        ws['G' + row] = each_result["id_cliente"]
                        ws['H' + row] = each_result["tipo_autorizacion"]
                        ws['I' + row] = each_result["autorizacion"]

                        ws['J' + row] = each_result["fecha_autorizacion"]
                        ws['K' + row] = each_result["anulado_sri"]
                        ws['L' + row] = each_result["motivo_anulacion"]

                        ws['M' + row] = each_result["subtotal"]
                        ws['N' + row] = each_result["base0"]
                        ws['O' + row] = each_result["baseno0"]
                        ws['P' + row] = each_result["iva"]
                        ws['Q' + row] = each_result["total"]
                        ws['R' + row] = each_result["rte_fte"]
                        ws['S' + row] = each_result["rte_iva"]

                        ws['T' + row] = each_result["product_id"]
                        ws['U' + row] = each_result["default_code"]
                        ws['V' + row] = each_result["product_name"]
                        ws['W' + row] = each_result["cantidad"]
                        ws['X' + row] = each_result["precio_unit"]
                        ws['Y' + row] = each_result["price_subtotal"]
                        ws['Z' + row] = each_result["price_tax"]
                        ws['AA' + row] = each_result["price_total"]
                        ws['AB' + row] = each_result["tipo_producto"]
                        i += 1
                        last_row = INDEX_ROW + i
                    if last_row >= INDEX_ROW:
                        thin = Side(border_style="thin", color="000000")
                        border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        self.set_border(ws, 'A' + str(INDEX_ROW) + ':AB' + str(last_row - 1), border)
                ws['B3'] = len(result)
                ws['A1'] = brw_wizard.company_id.name
                ws['B2'] = start_date
                ws['D2'] = end_date
                wb = self.save_wb(wb, target)

        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb=self.close_wb(wb)
        filecontent=dtFile.get_binary(new_filename)        
        return filecontent, EXT

class report_compras_report_xlsx(models.AbstractModel):
    _name = "report.gps_ats.report_compras_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Compras"
    
    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        new_filename = dtFile.create(EXT)
        def get_ws_objects(template_file_name):
            dir_path = dtFile.get_path_file(__file__)
            filename = dtFile.join(dir_path, template_file_name)
            dtFile.copyfile(filename, new_filename)
            wb, ws, target = self.open_xlsx(new_filename, load_sheet=True)
            return wb, ws, target
        wb=False
        try:
            brw_wizard=self.env["l10n.ec.ats"].sudo().browse(docids[-1])
            start_date,end_date=brw_wizard.date_start,brw_wizard.date_end##lectura de parametros

            if brw_wizard.type_info == "grouped":
                wb, ws, target = get_ws_objects("reporte_compras.xlsx")

                self._cr.execute(""";with variables as (
        select %s::date as fecha_inicial,
        %s::date as fecha_final ,
        %s::int as company_id     
    
    ),
    retenciones_detalle as (
                    select account_move.id,
                        sum(aml.l10n_ec_withhold_tax_amount) as total,
                        atxg.l10n_ec_type
                    from res_company 
                    inner join variables on 1=1
                    inner join res_partner rp  on rp.id=res_company.partner_id
                    inner join account_move on account_move.company_id = res_company.id and account_move.move_type in ('in_invoice') and account_move.state in ('posted','paid')  and 
                    account_move.invoice_date>=variables.fecha_inicial and account_move.invoice_date<=variables.fecha_final  and account_move.company_id=variables.company_id 
                    inner join account_journal aj on  aj.id=account_move.journal_id and coalesce(aj.l10n_latam_use_documents,false)=true  
                    inner join account_move_line aml on aml.l10n_ec_withhold_invoice_id=account_move.id 
                    inner join account_move_line_account_tax_rel atxrl on atxrl.account_move_line_id=aml.id 
                    inner join account_tax atx on atx.id=  atxrl.account_tax_id 
                    inner join account_tax_group atxg on atxg.id=atx.tax_group_id
                    inner join account_move awt on awt.id=aml.move_id and awt.state in ('posted','paid')  
                    where aml.l10n_ec_withhold_tax_amount>0.00
                    group by     account_move.id        ,atxg.l10n_ec_type,round(abs(atx.amount),2)
    )
    
    
    select account_move.id,
    doc_document_type.name as  tipo_documento,
    replace(account_move.nAME,'-','') as numero_documento,
    account_move.invoice_date as fecha,
    case when(account_move.state='draft') then 'preliminar' 
    when(account_move.state='posted') then 'publicado'
    when(account_move.state='cancel') then 'anulado' end as estado,
    res_partner.name as proveedor,
    l10n_latam_identification_type.name::json->'en_US'::varchar as tipo_identificacion,
    res_partner.vat as id_proveedor,
    account_move.authorization_type as tipo_autorizacion,
    coalesce(account_move.l10n_ec_authorization_number,'') as autorizacion,
    (case when(account_move.move_type='in_refund') then -1 else 1 end )*(coalesce(account_move.amount_base0,0.00)+ coalesce(account_move.amount_baseno0,0.00)) as subtotal,
    (case when(account_move.move_type='in_refund') then -1 else 1 end )*coalesce(account_move.amount_base0,0.00) as   base0,
     (case when(account_move.move_type='in_refund') then -1 else 1 end )*coalesce(account_move.amount_baseno0,0.00) as   baseno0,
     (case when(account_move.move_type='in_refund') then -1 else 1 end )*coalesce(account_move.amount_taxno0,0.00) as  iva,    
      (case when(account_move.move_type='in_refund') then -1 else 1 end )*coalesce(account_move.amount_total,0.00) as  total,   
    coalesce(rdfte.total,0.00) as  rte_fte,
    coalesce(rdiva.total,0.00) as  rte_iva,
    coalesce(account_move.l10n_ec_code_taxsupport,'') as l10n_ec_code_taxsupport
                    from res_company 
                    inner join variables on 1=1
                    inner join res_partner rp  on rp.id=res_company.partner_id
                    inner join account_move on account_move.company_id = res_company.id and account_move.move_type in ('in_invoice','in_refund') and account_move.state in ('posted','paid')  and 
                    account_move.invoice_date>=variables.fecha_inicial and account_move.invoice_date<=variables.fecha_final  and account_move.company_id=variables.company_id 
                    inner join account_journal aj on  aj.id=account_move.journal_id and coalesce(aj.l10n_latam_use_documents,false)=true 
                    
                    inner join res_partner on account_move.partner_id = res_partner.id
                    inner join l10n_latam_identification_type on res_partner.l10n_latam_identification_type_id = l10n_latam_identification_type.id
                    inner join l10n_ec_sri_payment doc_payment_type on doc_payment_type.id = account_move.l10n_ec_sri_payment_id 
                    inner join l10n_latam_document_type doc_document_type on doc_document_type.id = account_move.l10n_latam_document_type_id and doc_document_type.code!='00' 
                    left join retenciones_detalle rdfte on rdfte.id=account_move.id and rdfte.l10n_ec_type='withhold_income_purchase'
                    left join retenciones_detalle rdiva on rdiva.id=account_move.id and rdiva.l10n_ec_type!='withhold_income_purchase' 
        """,(start_date,end_date,brw_wizard.company_id.id))
                result=self._cr.dictfetchall()
                if result:
                    i,INDEX_ROW=0,5
                    last_row=INDEX_ROW
                    for each_result in result:
                        row=str(INDEX_ROW+i)
                        ws['A'+row]= each_result["id"]
                        ws['B'+row]= each_result["tipo_documento"]
                        ws['C'+row]= each_result["numero_documento"]
                        ws['D'+row]= each_result["fecha"]
                        ws['E'+row]= each_result["estado"]
                        ws['F'+row]= each_result["proveedor"]
                        ws['G' + row] = each_result["tipo_identificacion"]
                        ws['H'+row]= each_result["id_proveedor"]
                        ws['I'+row]= each_result["tipo_autorizacion"]
                        ws['J'+row]= each_result["autorizacion"]
                        ws['K'+row]= each_result["subtotal"]
                        ws['L'+row]= each_result["base0"]
                        ws['M'+row]= each_result["baseno0"]
                        ws['N'+row]= each_result["iva"]
                        ws['O'+row]= each_result["total"]
                        ws['P'+row]= each_result["rte_fte"]
                        ws['Q'+row]= each_result["rte_iva"]
                        ws['R' + row] = each_result["l10n_ec_code_taxsupport"]
                        i+=1
                        last_row=INDEX_ROW+i
                    if last_row>=INDEX_ROW:
                        thin = Side(border_style="thin", color="000000")
                        border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        self.set_border(ws,'A'+str(INDEX_ROW)+':R'+str(last_row-1),border)
                ws['B3']= len(result)
                ws['A1']= brw_wizard.company_id.name
                ws['B2']=start_date
                ws['D2']=end_date
                wb = self.save_wb(wb, target)
            else:
                wb, ws, target = get_ws_objects("reporte_compras_detallado.xlsx")
                self._cr.execute(""";with variables as (
                        select %s::date as fecha_inicial,
                        %s::date as fecha_final ,
                        %s::int as company_id     

                    ),
                    retenciones_detalle as (
                                    select account_move.id,
                                        sum(aml.l10n_ec_withhold_tax_amount) as total,
                                        atxg.l10n_ec_type
                                    from res_company 
                                    inner join variables on 1=1
                                    inner join res_partner rp  on rp.id=res_company.partner_id
                                    inner join account_move on account_move.company_id = res_company.id and account_move.move_type in ('in_invoice') and account_move.state in ('posted','paid')  and 
                                    account_move.invoice_date>=variables.fecha_inicial and account_move.invoice_date<=variables.fecha_final  and account_move.company_id=variables.company_id 
                                    inner join account_journal aj on  aj.id=account_move.journal_id and coalesce(aj.l10n_latam_use_documents,false)=true  
                                    inner join account_move_line aml on aml.l10n_ec_withhold_invoice_id=account_move.id 
                                    inner join account_move_line_account_tax_rel atxrl on atxrl.account_move_line_id=aml.id 
                                    inner join account_tax atx on atx.id=  atxrl.account_tax_id 
                                    inner join account_tax_group atxg on atxg.id=atx.tax_group_id
                                    inner join account_move awt on awt.id=aml.move_id and awt.state in ('posted','paid')  
                                    where aml.l10n_ec_withhold_tax_amount>0.00
                                    group by     account_move.id        ,atxg.l10n_ec_type,round(abs(atx.amount),2)
                    )


                    select distinct account_move.id,
                    doc_document_type.name as  tipo_documento,
                    replace(account_move.nAME,'-','') as numero_documento,
                    account_move.invoice_date as fecha,
                    case when(account_move.state='draft') then 'preliminar' 
                    when(account_move.state='posted') then 'publicado'
                    when(account_move.state='cancel') then 'anulado' end as estado,
                    res_partner.name as proveedor,
                    l10n_latam_identification_type.name->'en_US'::varchar as tipo_identificacion,
                    res_partner.vat as id_proveedor,
                    account_move.authorization_type as tipo_autorizacion,
                    coalesce(account_move.l10n_ec_authorization_number,'') as autorizacion,
                    (case when(account_move.move_type='in_refund') then -1 else 1 end )*(coalesce(account_move.amount_base0,0.00)+ coalesce(account_move.amount_baseno0,0.00)) as subtotal,
                    (case when(account_move.move_type='in_refund') then -1 else 1 end )*coalesce(account_move.amount_base0,0.00) as   base0,
                    (case when(account_move.move_type='in_refund') then -1 else 1 end )*coalesce(account_move.amount_baseno0,0.00) as   baseno0,
                    (case when(account_move.move_type='in_refund') then -1 else 1 end )*coalesce(account_move.amount_taxno0,0.00) as  iva,    
                    (case when(account_move.move_type='in_refund') then -1 else 1 end )*coalesce(account_move.amount_total,0.00) as  total,   
                    coalesce(rdfte.total,0.00) as  rte_fte,
                    coalesce(rdiva.total,0.00) as  rte_iva,
                    coalesce(account_move.l10n_ec_code_taxsupport,'') as l10n_ec_code_taxsupport,
					pp.id as product_id,
					pp.default_code,
					coalesce(pt.name::json->>'es_EC',pt.name::json->>'en_US') as product_name ,
					 (case when(account_move.move_type='in_refund') then -1 else 1 end )*coalesce(aml.quantity,1) as cantidad,
					 (case when(account_move.move_type='in_refund') then -1 else 1 end )*coalesce(aml.price_unit,0.00) as precio_unit,
					 (case when(account_move.move_type='in_refund') then -1 else 1 end )*coalesce(aml.price_subtotal,0.00) as price_subtotal,
					 (case when(account_move.move_type='in_refund') then -1 else 1 end )*(coalesce(aml.price_total,0.00)-coalesce(aml.price_subtotal,0.00)) as price_Tax,
					 (case when(account_move.move_type='in_refund') then -1 else 1 end )*coalesce(aml.price_total,0.00) as price_total,
					case when (pt.detailed_type='product') then 'PRODUCTO INVENTARIABLE' 
					when (pt.detailed_type='consu') then 'CONSUMIBLE'
					when (pt.detailed_type='service') then 'SERVICIO'
					else '' end as tipo_producto 
                    
                                    from res_company 
                                    inner join variables on 1=1
                                    inner join res_partner rp  on rp.id=res_company.partner_id
                                    inner join account_move on account_move.company_id = res_company.id and account_move.move_type in ('in_invoice','in_refund') and account_move.state in ('posted','paid')  and 
                                    account_move.invoice_date>=variables.fecha_inicial and account_move.invoice_date<=variables.fecha_final  and account_move.company_id=variables.company_id 
                                    
                                    inner join account_move_line aml on aml.move_id=account_move.id and aml.display_type='product'
									inner join account_account aa on aa.id=aml.account_id 
									left join product_product pp on pp.id=aml.product_id   and aml.display_type='product'
									left join product_template pt on pt.id=pp.product_Tmpl_id and aml.display_type='product'
                                    inner join account_journal aj on  aj.id=account_move.journal_id and coalesce(aj.l10n_latam_use_documents,false)=true 
                                   
                                 
                                    inner join res_partner on account_move.partner_id = res_partner.id
                                    inner join l10n_latam_identification_type on res_partner.l10n_latam_identification_type_id = l10n_latam_identification_type.id
                                    inner join l10n_ec_sri_payment doc_payment_type on doc_payment_type.id = account_move.l10n_ec_sri_payment_id 
                                    inner join l10n_latam_document_type doc_document_type on doc_document_type.id = account_move.l10n_latam_document_type_id and doc_document_type.code!='00' 
                                    left join retenciones_detalle rdfte on rdfte.id=account_move.id and rdfte.l10n_ec_type='withhold_income_purchase'
                                    left join retenciones_detalle rdiva on rdiva.id=account_move.id and rdiva.l10n_ec_type!='withhold_income_purchase' 
                        """, (start_date, end_date, brw_wizard.company_id.id))
                result = self._cr.dictfetchall()
                if result:
                    i, INDEX_ROW = 0, 5
                    last_row = INDEX_ROW
                    for each_result in result:
                        row = str(INDEX_ROW + i)
                        ws['A' + row] = each_result["id"]
                        ws['B' + row] = each_result["tipo_documento"]
                        ws['C' + row] = each_result["numero_documento"]
                        ws['D' + row] = each_result["fecha"]
                        ws['E' + row] = each_result["estado"]
                        ws['F' + row] = each_result["proveedor"]
                        ws['G' + row] = each_result["tipo_identificacion"]
                        ws['H' + row] = each_result["id_proveedor"]
                        ws['I' + row] = each_result["tipo_autorizacion"]
                        ws['J' + row] = each_result["autorizacion"]
                        ws['K' + row] = each_result["subtotal"]
                        ws['L' + row] = each_result["base0"]
                        ws['M' + row] = each_result["baseno0"]
                        ws['N' + row] = each_result["iva"]
                        ws['O' + row] = each_result["total"]
                        ws['P' + row] = each_result["rte_fte"]
                        ws['Q' + row] = each_result["rte_iva"]
                        ws['R' + row] = each_result["l10n_ec_code_taxsupport"]

                        ws['S' + row] = each_result["product_id"]
                        ws['T' + row] = each_result["default_code"]
                        ws['U' + row] = each_result["product_name"]
                        ws['V' + row] = each_result["cantidad"]
                        ws['W' + row] = each_result["precio_unit"]
                        ws['X' + row] = each_result["price_subtotal"]
                        ws['Y' + row] = each_result["price_tax"]
                        ws['Z' + row] = each_result["price_total"]
                        ws['AA' + row] = each_result["tipo_producto"]

                        i += 1
                        last_row = INDEX_ROW + i
                    if last_row >= INDEX_ROW:
                        thin = Side(border_style="thin", color="000000")
                        border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        self.set_border(ws, 'A' + str(INDEX_ROW) + ':AA' + str(last_row - 1), border)
                ws['B3'] = len(result)
                ws['A1'] = brw_wizard.company_id.name
                ws['B2'] = start_date
                ws['D2'] = end_date
                wb=self.save_wb(wb,target)
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb=self.close_wb(wb)
        filecontent=dtFile.get_binary(new_filename)        
        return filecontent, EXT

class report_impuestos_report_xlsx(models.AbstractModel):
    _name = "report.gps_ats.report_impuestos_xlsx"    
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Impuestos"
    
    def create_xlsx_report(self, docids, data):
        EXT="xlsx"
        dir_path=dtFile.get_path_file(__file__)
        new_filename=dtFile.create(EXT)  
        filename=dtFile.join(dir_path,"reporte_impuestos.xlsx")
        dtFile.copyfile(filename,new_filename)
        wb=False
        try:
            brw_wizard=self.env["l10n.ec.ats"].sudo().browse(docids[-1])
            start_date,end_date=brw_wizard.date_start,brw_wizard.date_end##lectura de parametros
            wb,ws,target=self.open_xlsx(new_filename,load_sheet=True)
            ws_compras = wb["compras"]
            ws_ventas = wb["ventas"]
            self.escribir_compras(brw_wizard, ws_compras, start_date, end_date)
            self.escribir_ventas(brw_wizard, ws_ventas, start_date, end_date)
            wb=self.save_wb(wb,target)
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb=self.close_wb(wb)
        filecontent=dtFile.get_binary(new_filename)        
        return filecontent, EXT
    
    def escribir_compras(self,brw_wizard,ws,start_date,end_date):
        self._cr.execute(""";with variables as (
    select %s::date as fecha_inicial,
    %s::date as fecha_final   ,
    %s::int as company_id    

),
retenciones_detalle as (
                select distinct account_move.id as invoice_id,
                   abs( aml.l10n_ec_withhold_tax_amount) as retencion_total,
                    atx.name as retencion_impuesto_nombre,
    
                    atx.l10n_ec_code_ats as retencion_codigo,    
               case when(atxg.l10n_ec_type='withhold_income_purchase') then 'FUENTE'  else 'IVA' END AS retencion_tipo_impuesto        ,
    awt.NAME as retencion,
    abs(aml.balance) as retencion_base,
    coalesce(aml.l10n_ec_withhold_tax_percn,0.00) as retencion_porcentaje,
    awtj.name as retencion_diario ,
   case when(awt.state='draft') then 'preliminar' 
when(awt.state='posted') then 'publicado'
when(awt.state='cancel') then 'anulado' end as retencion_estado
                from res_company 
                inner join variables on 1=1
                inner join res_partner rp  on rp.id=res_company.partner_id
                inner join account_move on account_move.company_id = res_company.id and account_move.move_type in ('in_invoice') and account_move.state in ('posted','paid')  and 
                account_move.invoice_date>=variables.fecha_inicial and account_move.invoice_date<=variables.fecha_final  and account_move.company_id=variables.company_id 
                
                inner join account_journal aj on  aj.id=account_move.journal_id and coalesce(aj.l10n_latam_use_documents,false)=true 
                
                inner join account_move_line aml on aml.l10n_ec_withhold_invoice_id=account_move.id  
                inner join account_move_line_account_tax_rel atxrl on atxrl.account_move_line_id=aml.id 
                inner join account_tax atx on atx.id=  atxrl.account_tax_id 
                inner join account_tax_group atxg on atxg.id=atx.tax_group_id
                inner join account_move awt on awt.id=aml.move_id and awt.state in ('posted','paid')  and awt.move_type='entry' 
                   inner join account_journal awtj on awtj.id=awt.journal_id  
)    

select account_move.id,
doc_document_type.name as  tipo_documento,
replace(account_move.nAME,'-','') as numero_documento,
account_move.invoice_date as fecha,
case when(account_move.state='draft') then 'preliminar' 
when(account_move.state='posted') then 'publicado'
when(account_move.state='cancel') then 'anulado' end as estado,
res_partner.name as proveedor,
res_partner.vat as id_proveedor,
account_move.authorization_type as tipo_autorizacion,
coalesce(account_move.l10n_ec_authorization_number,'') as autorizacion,
coalesce(account_move.amount_base0,0.00)+ coalesce(account_move.amount_baseno0,0.00) as subtotal,
coalesce(account_move.amount_base0,0.00) as   base0,
                coalesce(account_move.amount_baseno0,0.00) as   baseno0,
                coalesce(account_move.amount_taxno0,0.00) as  iva,    
                coalesce(account_move.amount_total,0.00) as  total,   
rdt.*
                from res_company 
                inner join variables on 1=1
                inner join res_partner rp  on rp.id=res_company.partner_id
                inner join account_move on account_move.company_id = res_company.id and account_move.move_type in ('in_invoice') and account_move.state in ('posted','paid')  and 
                account_move.invoice_date>=variables.fecha_inicial and account_move.invoice_date<=variables.fecha_final and account_move.company_id=variables.company_id 
                
                inner join account_journal aj on  aj.id=account_move.journal_id and coalesce(aj.l10n_latam_use_documents,false)=true 
                
                inner join res_partner on account_move.partner_id = res_partner.id
                inner join l10n_latam_identification_type on res_partner.l10n_latam_identification_type_id = l10n_latam_identification_type.id
                inner join l10n_ec_sri_payment doc_payment_type on doc_payment_type.id = account_move.l10n_ec_sri_payment_id 
                inner join l10n_latam_document_type doc_document_type on doc_document_type.id = account_move.l10n_latam_document_type_id and doc_document_type.code!='00' 
                left join retenciones_detalle rdt on rdt.invoice_id=account_move.id     """,(start_date,end_date,brw_wizard.company_id.id))
        result=self._cr.dictfetchall()
        if result:
            i,INDEX_ROW=0,5
            last_row=INDEX_ROW
            for each_result in result:    
                row=str(INDEX_ROW+i)
                ws['A'+row]= each_result["id"]
                ws['B'+row]= each_result["tipo_documento"]
                ws['C'+row]= each_result["numero_documento"]
                ws['D'+row]= each_result["fecha"]
                ws['E'+row]= each_result["estado"]
                ws['F'+row]= each_result["proveedor"]
                ws['G'+row]= each_result["id_proveedor"]
                ws['H'+row]= each_result["tipo_autorizacion"]
                ws['I'+row]= each_result["autorizacion"]
                ws['J'+row]= each_result["subtotal"]
                ws['K'+row]= each_result["base0"]
                ws['L'+row]= each_result["baseno0"]
                ws['M'+row]= each_result["iva"]
                ws['N'+row]= each_result["total"]
                ws['O'+row]= each_result["retencion"]
                ws['P'+row]= each_result["retencion_base"]
                ws['Q'+row]= each_result["retencion_porcentaje"]                
                ws['R'+row]= each_result["retencion_total"]
                ws['S'+row]= each_result["retencion_tipo_impuesto"]               
                ws['T'+row]= each_result["retencion_codigo"]
                ws['U'+row]= each_result["retencion_impuesto_nombre"]                
                ws['V'+row]= each_result["retencion_diario"]
                ws['W'+row]= each_result["retencion_estado"]
                i+=1
                last_row=INDEX_ROW+i
            if last_row>=INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws,'A'+str(INDEX_ROW)+':W'+str(last_row-1),border)
        ws['B3']= len(result)
        ws['A1']= brw_wizard.company_id.name
        ws['B2']=start_date
        ws['D2']=end_date
        return ws
      
    def escribir_ventas(self,brw_wizard,ws,start_date,end_date):
        self._cr.execute(""";with variables as (
    select %s::date as fecha_inicial,
    %s::date as fecha_final,
    %s::int as company_id        

),
retenciones_detalle as (
                select distinct account_move.id as invoice_id,
                   abs( aml.l10n_ec_withhold_tax_amount) as retencion_total,
                    atx.name as retencion_impuesto_nombre,
    
                    atx.l10n_ec_code_ats as retencion_codigo,    
               case when(atxg.l10n_ec_type='withhold_income_sale') then 'FUENTE'  else 'IVA' END AS retencion_tipo_impuesto        ,
    awt.NAME as retencion,
    abs(aml.balance) as retencion_base,
    coalesce(abs(aml.l10n_ec_withhold_tax_percn),0.00) as retencion_porcentaje,
    awtj.name as retencion_diario  ,
   case when(awt.state='draft') then 'preliminar' 
when(awt.state='posted') then 'publicado'
when(awt.state='cancel') then 'anulado' end as retencion_estado
                from res_company 
                inner join variables on 1=1
                inner join res_partner rp  on rp.id=res_company.partner_id
                inner join account_move on account_move.company_id = res_company.id and account_move.move_type in ('out_invoice') and account_move.state in ('posted','paid')  and 
                account_move.invoice_date>=variables.fecha_inicial and account_move.invoice_date<=variables.fecha_final  and account_move.company_id=variables.company_id 
                inner join account_journal aj on  aj.id=account_move.journal_id and coalesce(aj.l10n_latam_use_documents,false)=true 
                
                inner join account_move_line aml on aml.l10n_ec_withhold_invoice_id=account_move.id  
                inner join account_move_line_account_tax_rel atxrl on atxrl.account_move_line_id=aml.id 
                inner join account_tax atx on atx.id=  atxrl.account_tax_id 
                inner join account_tax_group atxg on atxg.id=atx.tax_group_id
                inner join account_move awt on awt.id=aml.move_id and awt.state in ('posted','paid')  and awt.move_type='entry'
                inner join account_journal awtj on awtj.id=awt.journal_id 
)    

select account_move.id,
doc_document_type.name as  tipo_documento,
replace(account_move.nAME,'-','') as numero_documento,
account_move.invoice_date as fecha,
case when(account_move.state='draft') then 'preliminar' 
when(account_move.state='posted') then 'publicado'
when(account_move.state='cancel') then 'anulado' end as estado,
res_partner.name as proveedor,
res_partner.vat as id_proveedor,
account_move.authorization_type as tipo_autorizacion,
coalesce(account_move.l10n_ec_authorization_number,'') as autorizacion,
coalesce(account_move.amount_base0,0.00)+ coalesce(account_move.amount_baseno0,0.00) as subtotal,
coalesce(account_move.amount_base0,0.00) as   base0,
                coalesce(account_move.amount_baseno0,0.00) as   baseno0,
                coalesce(account_move.amount_taxno0,0.00) as  iva,    
                coalesce(account_move.amount_total,0.00) as  total,   
rdt.*
                from res_company 
                inner join variables on 1=1
                inner join res_partner rp  on rp.id=res_company.partner_id
                inner join account_move on account_move.company_id = res_company.id and account_move.move_type in ('out_invoice') and account_move.state in ('posted','paid')  and 
                account_move.invoice_date>=variables.fecha_inicial and account_move.invoice_date<=variables.fecha_final  and account_move.company_id=variables.company_id 
                inner join account_journal aj on  aj.id=account_move.journal_id and coalesce(aj.l10n_latam_use_documents,false)=true  
                inner join res_partner on account_move.partner_id = res_partner.id
                inner join l10n_latam_identification_type on res_partner.l10n_latam_identification_type_id = l10n_latam_identification_type.id
                inner join l10n_ec_sri_payment doc_payment_type on doc_payment_type.id = account_move.l10n_ec_sri_payment_id 
                inner join l10n_latam_document_type doc_document_type on doc_document_type.id = account_move.l10n_latam_document_type_id and doc_document_type.code!='00'
                left join retenciones_detalle rdt on rdt.invoice_id=account_move.id     """,(start_date,end_date,brw_wizard.company_id.id))
        result=self._cr.dictfetchall()
        if result:
            i,INDEX_ROW=0,5
            last_row=INDEX_ROW
            for each_result in result:    
                row=str(INDEX_ROW+i)
                ws['A'+row]= each_result["id"]
                ws['B'+row]= each_result["tipo_documento"]
                ws['C'+row]= each_result["numero_documento"]
                ws['D'+row]= each_result["fecha"]
                ws['E'+row]= each_result["estado"]
                ws['F'+row]= each_result["proveedor"]
                ws['G'+row]= each_result["id_proveedor"]
                ws['H'+row]= each_result["tipo_autorizacion"]
                ws['I'+row]= each_result["autorizacion"]
                ws['J'+row]= each_result["subtotal"]
                ws['K'+row]= each_result["base0"]
                ws['L'+row]= each_result["baseno0"]
                ws['M'+row]= each_result["iva"]
                ws['N'+row]= each_result["total"]
                ws['O'+row]= each_result["retencion"]
                ws['P'+row]= each_result["retencion_base"]
                ws['Q'+row]= each_result["retencion_porcentaje"]                
                ws['R'+row]= each_result["retencion_total"]
                ws['S'+row]= each_result["retencion_tipo_impuesto"]               
                ws['T'+row]= each_result["retencion_codigo"]
                ws['U'+row]= each_result["retencion_impuesto_nombre"]                
                ws['V'+row]= each_result["retencion_diario"]
                ws['W'+row]= each_result["retencion_estado"]
                i+=1
                last_row=INDEX_ROW+i
            if last_row>=INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws,'A'+str(INDEX_ROW)+':W'+str(last_row-1),border)
        ws['B3']= len(result)
        ws['A1']= brw_wizard.company_id.name
        ws['B2']=start_date
        ws['D2']=end_date
        return ws

class report_compras_acct_report_xlsx_act(models.AbstractModel):
    _name = "report.gps_ats.report_compras_acct_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Compras por Cuentas"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        new_filename = dtFile.create(EXT)

        def get_ws_objects(template_file_name):
            dir_path = dtFile.get_path_file(__file__)
            filename = dtFile.join(dir_path, template_file_name)
            dtFile.copyfile(filename, new_filename)
            wb, ws, target = self.open_xlsx(new_filename, load_sheet=True)
            return wb, ws, target

        wb = False
        try:
            brw_wizard = self.env["l10n.ec.ats"].sudo().browse(docids[-1])
            start_date, end_date = brw_wizard.date_start, brw_wizard.date_end  ##lectura de parametros
            wb, ws, target = get_ws_objects("reporte_compras_cuentas_detallado.xlsx")

            self._cr.execute(""";with variables as (
            select %s::date as fecha_inicial,
            %s::date as fecha_final ,
            %s::int as company_id     
    
        )
        
            select account_move.id,
    doc_document_type.name as  tipo_documento,
    replace(account_move.nAME,'-','') as numero_documento,
    account_move.invoice_date as fecha,
    case when(account_move.state='draft') then 'preliminar' 
    when(account_move.state='posted') then 'publicado'
    when(account_move.state='cancel') then 'anulado' end as estado,
    res_partner.name as proveedor,
    l10n_latam_identification_type.name::json->'en_US'::varchar as tipo_identificacion,
    res_partner.vat as id_proveedor,
    account_move.authorization_type as tipo_autorizacion,
    coalesce(account_move.l10n_ec_authorization_number,'') as autorizacion,
	aa.code as account_code,
	aa.name as account_name ,
	coalesce(pt.default_code,'') as codigo_producto,
	coalesce(pt.name->>'es_EC','') as nombre_producto,	
	(case when(account_move.move_type='in_refund') then -1 else 1 end )*aml.quantity as cantidad,
	aml.price_unit as precio_unitario,
	(case when(account_move.move_type='in_refund') then -1 else 1 end )*coalesce((coalesce(aml.quantity,1.00)*aml.price_unit)*(coalesce(aml.discount,0.00)/100.00) ,0.00) as descuento,
	(case when(account_move.move_type='in_refund') then -1 else 1 end )*aml.price_subtotal as subtotal,
	(case when(account_move.move_type='in_refund') then -1 else 1 end )*aml.price_total as total,
	(case when(account_move.move_type='in_refund') then -1 else 1 end )*(aml.price_total-aml.price_subtotal)  as iva,
	aml.debit as debito,
	aml.credit  as credito,
	aml.name as descripcion
	
	
                    from res_company 
                    inner join variables on 1=1
                    inner join res_partner rp  on rp.id=res_company.partner_id
                    inner join account_move on account_move.company_id = res_company.id and account_move.move_type in ('in_invoice','in_refund') and account_move.state in ('posted','paid')  and 
                    account_move.invoice_date>=variables.fecha_inicial and account_move.invoice_date<=variables.fecha_final  and account_move.company_id=variables.company_id 
                    inner join account_journal aj on  aj.id=account_move.journal_id and coalesce(aj.l10n_latam_use_documents,false)=true 

                    inner join res_partner on account_move.partner_id = res_partner.id
                    inner join l10n_latam_identification_type on res_partner.l10n_latam_identification_type_id = l10n_latam_identification_type.id
                    inner join l10n_ec_sri_payment doc_payment_type on doc_payment_type.id = account_move.l10n_ec_sri_payment_id 
                    inner join l10n_latam_document_type doc_document_type on doc_document_type.id = account_move.l10n_latam_document_type_id and doc_document_type.code!='00' 
					inner join account_move_line aml on aml.move_id=account_move.id 	and aml.display_type='product'			
					inner join account_account aa on aa.id=aml.account_id 
 					LEFT JOIN PROduct_product pp on pp.id=aml.product_id and aml.display_type='product'
					left join product_template pt on pt.id=pp.product_tmpl_id  and aml.display_type='product' 

					
        """, (start_date, end_date, brw_wizard.company_id.id))
            result = self._cr.dictfetchall()
            if result:
                    i, INDEX_ROW = 0, 5
                    last_row = INDEX_ROW
                    for each_result in result:
                        row = str(INDEX_ROW + i)
                        ws['A' + row] = each_result["id"]
                        ws['B' + row] = each_result["tipo_documento"]
                        ws['C' + row] = each_result["numero_documento"]
                        ws['D' + row] = each_result["fecha"]
                        ws['E' + row] = each_result["estado"]
                        ws['F' + row] = each_result["proveedor"]
                        ws['G' + row] = each_result["tipo_identificacion"]
                        ws['H' + row] = each_result["id_proveedor"]
                        ws['I' + row] = each_result["tipo_autorizacion"]
                        ws['J' + row] = each_result["autorizacion"]
                        ###
                        ws['K' + row] = each_result["codigo_producto"]
                        ws['L' + row] = each_result["nombre_producto"]
                        ws['M' + row] = each_result["descripcion"]
                        ws['N' + row] = each_result["account_code"]
                        ws['O' + row] = each_result["account_name"]
                        ###
                        ws['P' + row] = each_result["cantidad"]
                        ws['Q' + row] = each_result["precio_unitario"]
                        ws['R' + row] = each_result["descuento"]
                        ws['S' + row] = each_result["subtotal"]
                        ws['T' + row] = each_result["iva"]
                        ws['U' + row] = each_result["total"]
                        ws['V' + row] = each_result["debito"]
                        ws['W' + row] = each_result["credito"]
                        i += 1
                        last_row = INDEX_ROW + i
                    if last_row >= INDEX_ROW:
                        thin = Side(border_style="thin", color="000000")
                        border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        self.set_border(ws, 'A' + str(INDEX_ROW) + ':W' + str(last_row - 1), border)
            ws['B3'] = len(result)
            ws['A1'] = brw_wizard.company_id.name
            ws['B2'] = start_date
            ws['D2'] = end_date
            wb = self.save_wb(wb, target)
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT
