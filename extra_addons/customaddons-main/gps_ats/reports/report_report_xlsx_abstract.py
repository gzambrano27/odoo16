# -*- coding: utf-8 -*-
# -*- encoding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.
import logging
from odoo import models,_,api
from odoo.exceptions import ValidationError
_logger = logging.getLogger(__name__)
from openpyxl import load_workbook
from openpyxl.styles import  Alignment, Font, PatternFill

class ReportReportXlsxAbstract(models.AbstractModel):
    _inherit = "report.report_xlsx.abstract"
    
    @api.model
    def open_xlsx(self,target,load_sheet=True):
        wb = load_workbook(target)
        if not load_sheet:
            return wb,False,target
        sheet_name=self._get_sheet_names(wb)
        ws = wb[sheet_name]
        return wb,ws,target
    
    @api.model
    def xls_font(self,bold=False,size=8,name="Tahoma"):
        return Font(bold=bold, size=size, name=name)
    
    @api.model
    def xls_alignment(self,wrapText=True ,horizontal='left',vertical='bottom'):
        return Alignment(wrapText=wrapText, horizontal=horizontal, vertical=vertical)
    
    @api.model
    def xls_pattern_fill(self,bgColor="F2F2F2", fill_type = "solid"):
        return PatternFill(start_color=bgColor, end_color=bgColor, fill_type = fill_type)
    
    @api.model
    def set_border(self,ws,range,border):
        pass
        # for row in ws[range]:
        #     for cell in row:
        #         cell.border = border
    
    @api.model              
    def write_cell_by_index(self,ws,letter,index,value,alignment=False,font=False,number_format=False,default='',fill=False):
        pos="%s%s" % (letter,index)
        return self.write_cell(ws, pos, value, alignment=alignment, font=font, number_format=number_format, default=default,fill=fill)
    
    @api.model
    def merge_cells(self,ws,range_text):
        ws.merge_cells(range_text)  
        
    @api.model
    def write_cell(self,ws,pos,value,alignment=False,font=False,number_format=False,default='',fill=False):
        if not default or default is None:
            default=''
        if alignment:
            ws[pos].alignment = alignment
        if font:
            ws[pos].font = font
        if number_format:
            ws[pos].number_format = number_format     
        if fill:
            ws[pos].fill =fill       
        ws[pos]= (value or default)
    
    @api.model           
    def txt_format(self,txt,size=255):
        if len(txt) > size:
            ret = txt[:(size-1)]
            return ret
        return txt
    
    @api.model
    def _get_sheet_names(self,wb):
        for sheet_name in wb.worksheets:
            return sheet_name.title 
        return "Hoja 1"
    
    @api.model
    def save_wb(self,wb,target):
        wb.save(target)  
        return self.close_wb(wb,)
    
    @api.model
    def close_wb(self,wb):
        if wb:
            wb.close()
            wb=False   
        return wb 
