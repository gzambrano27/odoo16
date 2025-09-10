# coding: utf-8
from odoo import api, fields, models, exceptions, tools, _
from ...message_dialog.tools import FileManager
from ...calendar_days.tools import DateManager
from ...calendar_days.tools import CalendarManager
dtFile=FileManager()        
dateO=DateManager()
calendarO=CalendarManager()

import logging
_logger = logging.getLogger(__name__)

from openpyxl.styles.borders import Border, Side

class report_movements_report_xlsx(models.AbstractModel):
    _name = "report.gps_hr.report_movements_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Movimientos"
    
    def create_xlsx_report(self, docids, data):
        EXT="xlsx"
        dir_path=dtFile.get_path_file(__file__)
        new_filename=dtFile.create(EXT)
        OBJ_MOVEMENT=self.env["hr.employee.movement"].sudo()
        brw_movement = OBJ_MOVEMENT.browse(docids[-1])
        COLUMNS = OBJ_MOVEMENT._fields.copy()
        STATES=dict(COLUMNS["state"].selection).copy()
        file_path="reporte_movimientos.xlsx"
        is_discount=brw_movement.type=="discount"
        if is_discount:
            file_path = "reporte_movimientos_diferidos.xlsx"
        filename=dtFile.join(dir_path,file_path)
        dtFile.copyfile(filename,new_filename)
        wb=False
        try:
            wb,ws,target=self.open_xlsx(new_filename,load_sheet=True)
            if brw_movement.line_ids:
                i,INDEX_ROW=0,6
                last_row=INDEX_ROW
                for brw_line in brw_movement.line_ids:
                    row=str(INDEX_ROW+i)
                    ws['A' + row] = brw_line.id
                    ws['B' + row] = brw_line.employee_id.name
                    ws['C' + row] = brw_line.employee_id.identification_id or ''
                    ws['D' + row] = brw_line.quota
                    ws['E' + row] = brw_line.date_process
                    ws['F' + row] = brw_line.total
                    ws['G' + row] = brw_line.name
                    ws['H' + row] = brw_line.comments or ''
                    i+=1
                    last_row=INDEX_ROW+i
                if last_row>=INDEX_ROW:
                    thin = Side(border_style="thin", color="000000")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    self.set_border(ws, 'A' + str(INDEX_ROW) + ':H' + str(last_row - 1), border)
            ws['A1']=brw_movement.company_id.name
            ws['B2']=brw_movement.id
            ws['B3']= brw_movement.rule_id.name
            ws['B4']=brw_movement.name
            ws['E2']=brw_movement.date_process
            ws['E3'] = brw_movement.total
            ws['E4'] = len(brw_movement.line_ids)
            ws['H2'] = STATES[brw_movement.state].upper()
            if is_discount:
                ws['H3'] = brw_movement.total_pending
                ws['H4'] = brw_movement.employee_id.name
            wb=self.save_wb(wb,target)  
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb=self.close_wb(wb)
        filecontent=dtFile.get_binary(new_filename)        
        return filecontent, EXT