# coding: utf-8
from odoo import api, fields, models, tools, _
from ...message_dialog.tools import FileManager
from ...calendar_days.tools import DateManager
from ...calendar_days.tools import CalendarManager
dtFile=FileManager()
dateO=DateManager()
calendarO=CalendarManager()
from odoo.exceptions import ValidationError

import logging
_logger = logging.getLogger(__name__)

from openpyxl.styles.borders import Border, Side

class hr_employee_payslip_reports_movements_wizard_xlsx(models.AbstractModel):
    _name = "report.gps_hr.hr_employee_payslip_reports_movements_wizard_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Movimientos"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)
        brw_wizard = self.env["hr.employee.payslip.reports.wizard"].sudo().browse(docids[0])

        OBJ_REPORT_MOVEMENT_LINE = self.env["hr.employee.movement.line"].sudo()
        domain = [('process_id.month_id', '=', brw_wizard.month_id.id),
                  ('process_id.year', '=', brw_wizard.year),
                  ('process_id.state', '!=', 'cancelled')
                  ]
        if brw_wizard.company_id:
            domain += [('process_id.company_id', '=', brw_wizard.company_id.id)]
        if brw_wizard.rule_ids:
            domain += [('rule_id', 'in', brw_wizard.rule_ids.ids)]
        if brw_wizard.employee_ids:
            domain += [('employee_id', 'in', brw_wizard.employee_ids.ids)]

        srch_movement_line = OBJ_REPORT_MOVEMENT_LINE.search(domain)
        if not srch_movement_line:
            raise ValidationError(_("Sin resultados"))

        COLUMNS = self.env["hr.employee.movement"].sudo()._fields.copy()
        STATES = dict(COLUMNS["state"].selection).copy()
        file_path = "reporte_resumen_movimientos.xlsx"
        filename = dtFile.join(dir_path, file_path)
        dtFile.copyfile(filename, new_filename)
        wb = False
        try:
            wb, ws, target = self.open_xlsx(new_filename, load_sheet=True)
            i, INDEX_ROW = 0, 6
            last_row = INDEX_ROW
            for brw_line in srch_movement_line:
                row = str(INDEX_ROW + i)
                ws['A' + row] = brw_line.company_id.name
                ws['B' + row] = brw_line.rule_id.name
                ws['C' + row] = brw_line.process_id.id
                ws['D' + row] = brw_line.process_id.name
                ws['E' + row] = brw_line.id
                ws['F' + row] = brw_line.employee_id.name or ''
                ws['G' + row] = brw_line.employee_id.identification_id or ''
                ws['H' + row] = brw_line.quota
                ws['I' + row] = brw_line.date_process
                ws['J' + row] = brw_line.total
                ws['K' + row] = brw_line.comments or ''
                i += 1
                last_row = INDEX_ROW + i
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':K' + str(last_row - 1), border)
            ws['A1'] = ",".join(brw_wizard.mapped('company_id').mapped('name'))
            ws['B2'] = brw_wizard.date_start
            ws['E2'] = brw_wizard.date_end
            ws['B3'] = brw_wizard.rule_ids and ",".join(brw_wizard.mapped('rule_ids').mapped('name')) or "TODOS"
            ws['B4'] = brw_wizard.employee_ids and ",".join(brw_wizard.mapped('employee_ids').mapped('name')) or "TODOS"
            ws['E3'] = len(srch_movement_line)
            wb = self.save_wb(wb, target)
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT
