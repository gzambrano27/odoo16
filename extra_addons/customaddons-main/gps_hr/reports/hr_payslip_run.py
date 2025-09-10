# coding: utf-8
from odoo import api, fields, models, exceptions, tools, _
from openpyxl.styles import PatternFill

from ...message_dialog.tools import FileManager
from ...calendar_days.tools import DateManager
from ...calendar_days.tools import CalendarManager

dtFile = FileManager()
dateO = DateManager()
calendarO = CalendarManager()

import logging

_logger = logging.getLogger(__name__)

from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font, PatternFill
import openpyxl.utils as utils
from openpyxl.styles import NamedStyle

class report_payslip_runs_xlsx(models.AbstractModel):
    _name = "report.gps_hr.report_payslip_runs_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Roles de Empleados"

    def create_xlsx_report(self, docids, data):
        DEC=2
        finiquito_ids =data.get('finiquito_ids',[])
        rol_ids=data.get('rol_ids',[])

        if finiquito_ids and rol_ids:
            tipo = "ROLES Y FINIQUITOS"
        elif rol_ids:
            tipo = "ROLES"
        elif finiquito_ids:
            tipo = "FINIQUITOS"
        else:
            tipo = ""

        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)
        OBJ_PAYSLIP_RUN = self.env["hr.payslip.run"].sudo()
        #brw_payslip_run = OBJ_PAYSLIP_RUN.browse(rol_ids)
        #run_ids =[str(id) for id in brw_payslip_run.ids]
        dscr_periods = ""
        companys=""#"', '.join(brw_payslip_run.mapped('company_id').mapped('name'))
        dscr_roles=""
        #for brw_each_run in brw_payslip_run:
        #    #companys=brw_each_run.company_id.name
        #    dscr_roles=brw_each_run.type_struct_id.name
        #    dscr_periods="%s/%s" % (brw_each_run.month_id.name,str(brw_each_run.year))
        file_path = "reporte_roles.xlsx"
        filename = dtFile.join(dir_path, file_path)
        dtFile.copyfile(filename, new_filename)
        wb = False
        try:
            wb, ws, target = self.open_xlsx(new_filename, load_sheet=True)
            LIST_IDS=tuple(rol_ids+[-1,-1])
            FINIQUITO_IDS = tuple(finiquito_ids + [-1, -1])
            COLUMN_DATA_TYPE=""
            COLUMN_BASE_QUERY = """ SELECT (CASE WHEN(HSRC.CODE='IN') THEN 1 
		WHEN(HSRC.CODE='OUT') THEN 3 
			WHEN(HSRC.CODE='PRO') THEN 5 ELSE 999999999999999 END ) AS  SEQUENCE,
			HSR.CODE AS RULE_CODE,HSR.NAME AS RULE_NAME ,HSR.SEQUENCE RULE_ORDER
                FROM HR_PAYSLIP_RUN HPR 
                INNER JOIN HR_PAYSLIP HP ON HP.PAYSLIP_RUN_ID=HPR.ID
                INNER JOIN HR_PAYSLIP_LINE HPL ON HPL.SLIP_ID=HP.ID
                INNER JOIN (SELECT CATEGORY_ID,SEQUENCE,ID,CODE,(NAME::JSON->'en_US')::TEXT AS NAME FROM HR_SALARY_RULE WHERE ACTIVE=TRUE ) HSR ON HSR.ID=HPL.SALARY_RULE_ID 
                INNER JOIN HR_SALARY_RULE_CATEGORY HSRC ON HSRC.ID= HSR.CATEGORY_ID
                WHERE HPR.ID IN %s  
                GROUP BY HSRC.CODE,HSR.CODE ,HSR.NAME  ,HSR.SEQUENCE 
                UNION
SELECT (CASE WHEN(HSRC.CODE='IN') THEN 1 
		WHEN(HSRC.CODE='OUT') THEN 3 
			WHEN(HSRC.CODE='PRO') THEN 5 ELSE 999999999999999 END ) AS  SEQUENCE,
			HSR.CODE AS RULE_CODE,HSR.NAME AS RULE_NAME ,HSR.SEQUENCE RULE_ORDER             
			FROM  HR_EMPLOYEE_LIQUIDATION HP 
                INNER JOIN HR_EMPLOYEE_LIQUIDATION_LINE HPL ON HPL.LIQUIDATION_ID=HP.ID 				
                INNER JOIN (SELECT CATEGORY_ID,SEQUENCE,ID,CODE,(NAME::JSON->'en_US')::TEXT AS NAME FROM HR_SALARY_RULE WHERE ACTIVE=TRUE ) HSR ON HSR.ID=HPL.RULE_ID  				
                INNER JOIN HR_SALARY_RULE_CATEGORY HSRC ON HSRC.ID= HSR.CATEGORY_ID
                WHERE HP.ID IN %s AND HPL.type='liquidation' 
                GROUP BY HSRC.CODE,HSR.CODE ,HSR.NAME  ,HSR.SEQUENCE	
                UNION 
                SELECT 2 AS SEQUENCE,'TOTAL_INGRESOS' AS RULE_CODE ,'Total Ingresos' AS RULE_NAME ,99999 AS RULE_ORDER
                UNION 
                SELECT 4 AS SEQUENCE,'TOTAL_EGRESOS' AS RULE_CODE,'Total Egresos' AS RULE_NAME   ,99999 AS RULE_ORDER
                UNION 
                SELECT 6 AS SEQUENCE,'TOTAL_PROVISION' AS RULE_CODE ,'Total Provision' AS RULE_NAME  ,99999 AS RULE_ORDER
                UNION 
                SELECT 7 AS SEQUENCE,'TOTAL' AS RULE_CODE,'Total a Recibir' AS RULE_NAME   ,99999 AS RULE_ORDER
            """ % (LIST_IDS,FINIQUITO_IDS)
            self._cr.execute(
                "SELECT X.* FROM (" + COLUMN_BASE_QUERY + " ) X   ORDER BY X.SEQUENCE ASC,X.RULE_ORDER ASC")
            result_rules = self._cr.fetchall()
            columns_dscr={}
            if result_rules:
                for sequence, rule_code, rule_name, rule_order in result_rules:
                    role_code_pk=rule_code.lower()
                    if role_code_pk not in columns_dscr:
                        columns_dscr[role_code_pk]=str(rule_name).replace('"',"")
                columns_dscr["company_name"]="Empresa"
                columns_dscr["identification_id"]="# Identificación"
                columns_dscr["name"] = "Empleado"
                columns_dscr["wage"] = "Salario"
                columns_dscr["worked_days"] = "Dias Trabajados"
                columns_dscr["payslip_id"] = "ID # Nómina"
                columns_dscr["departamento"] = "Departamento"
                columns_dscr["cargo"] = "Cargo"
                columns_dscr["dscr_payslip"] = "Descripción"
                columns_dscr["fecha_contrato"] = "Fecha de Contrato"
            COLUMN_QUERY = """SELECT   X.RULE_CODE FROM 
                (
                    %s
                ) X 
                ORDER BY X.SEQUENCE ASC,X.RULE_ORDER ASC """ % (COLUMN_BASE_QUERY,)
            #print("0000000000000000000000000000")
            self._cr.execute(COLUMN_QUERY)
            #print("0.111111111111111111111111")
            result_columns = self._cr.fetchall()
            if result_columns:
                COLUMN_DATA_TYPE = ""
                for rule_code, in result_columns:
                    COLUMN_DATA_TYPE += "," + rule_code + " float"
            ####################################
            PAYSLIP_RUN_REPORT_QUERY = """
            ;WITH TOTAL_PAYSLIP AS (
             SELECT * FROM CROSSTAB(
              $$ 
                
                SELECT 'payslip_'||HP.ID AS PAYSLIP_ID,
       HSR.CODE AS RULE_CODE,
       SUM(ABS(HPL.TOTAL)) AS TOTAL
FROM HR_PAYSLIP_RUN HPR
INNER JOIN HR_PAYSLIP HP ON HP.PAYSLIP_RUN_ID = HPR.ID
INNER JOIN HR_PAYSLIP_LINE HPL ON HPL.SLIP_ID = HP.ID
INNER JOIN HR_SALARY_RULE HSR ON HSR.ID = HPL.SALARY_RULE_ID
WHERE HPR.ID IN {0}
GROUP BY HP.ID, HSR.CODE

UNION
SELECT 'payslip_'||HP.ID, 'TOTAL_INGRESOS', COALESCE(HP.TOTAL_IN,0.00) AS TOTAL 
FROM HR_PAYSLIP_RUN HPR
INNER JOIN HR_PAYSLIP HP ON HP.PAYSLIP_RUN_ID = HPR.ID
WHERE HPR.ID IN {0}

UNION
SELECT 'payslip_'||HP.ID, 'TOTAL_EGRESOS', ABS(COALESCE(HP.TOTAL_OUT,0.00)) AS TOTAL 
FROM HR_PAYSLIP_RUN HPR
INNER JOIN HR_PAYSLIP HP ON HP.PAYSLIP_RUN_ID = HPR.ID
WHERE HPR.ID IN {0}

UNION
SELECT 'payslip_'||HP.ID, 'TOTAL_PROVISION', COALESCE(HP.TOTAL_PROVISION,0.00) AS TOTAL 
FROM HR_PAYSLIP_RUN HPR
INNER JOIN HR_PAYSLIP HP ON HP.PAYSLIP_RUN_ID = HPR.ID
WHERE HPR.ID IN {0}

UNION
SELECT 'payslip_'||HP.ID, 'TOTAL', COALESCE(HP.total_payslip,0.00) AS TOTAL  
FROM HR_PAYSLIP_RUN HPR
INNER JOIN HR_PAYSLIP HP ON HP.PAYSLIP_RUN_ID = HPR.ID
WHERE HPR.ID IN {0}


UNION
SELECT 'finiquito_'||HP.ID,
       HSR.CODE,
       SUM(ABS(HPL.AMOUNT)) AS TOTAL
FROM HR_EMPLOYEE_LIQUIDATION HP
INNER JOIN HR_EMPLOYEE_LIQUIDATION_LINE HPL ON HPL.LIQUIDATION_ID = HP.ID
INNER JOIN HR_SALARY_RULE HSR ON HSR.ID = HPL.RULE_ID
WHERE HP.ID IN {3} AND HPL.type = 'liquidation'
GROUP BY HP.ID, HSR.CODE

UNION
SELECT 'finiquito_'||HP.ID, 'TOTAL_INGRESOS', COALESCE(HP.TOTAL_IN_LIQUIDATION,0.00)  AS TOTAL
FROM HR_EMPLOYEE_LIQUIDATION HP
WHERE HP.ID IN {3} 

UNION
SELECT 'finiquito_'||HP.ID, 'TOTAL_EGRESOS', ABS(COALESCE(HP.TOTAL_OUT_LIQUIDATION,0.00))  AS TOTAL
FROM HR_EMPLOYEE_LIQUIDATION HP
WHERE HP.ID IN {3} 

UNION
SELECT 'finiquito_'||HP.ID, 'TOTAL_PROVISION', COALESCE(HP.TOTAL_PROVISION_LIQUIDATION,0.00)  AS TOTAL
FROM HR_EMPLOYEE_LIQUIDATION HP
WHERE HP.ID IN {3} 

UNION
SELECT 'finiquito_'||HP.ID, 'TOTAL', COALESCE(HP.TOTAL_LIQUIDATION,0.00)  AS TOTAL 
FROM HR_EMPLOYEE_LIQUIDATION HP
WHERE HP.ID IN {3} 


-- ==== ACUMULADOS GLOBALES (payslip + finiquito) ====
UNION
SELECT 'global_'||0 AS PAYSLIP_ID,
       HSR.CODE,
       SUM(ABS(COALESCE(P.TOTAL,0))) AS TOTAL
FROM (
    -- Totales detalle payslip
    SELECT HSR.CODE, SUM(ABS(HPL.TOTAL)) AS TOTAL
    FROM HR_PAYSLIP_RUN HPR
    INNER JOIN HR_PAYSLIP HP ON HP.PAYSLIP_RUN_ID = HPR.ID
    INNER JOIN HR_PAYSLIP_LINE HPL ON HPL.SLIP_ID = HP.ID
    INNER JOIN HR_SALARY_RULE HSR ON HSR.ID = HPL.SALARY_RULE_ID
    WHERE HPR.ID IN {0}
    GROUP BY HSR.CODE

    UNION ALL
    -- Totales detalle finiquito
    SELECT HSR.CODE, SUM(ABS(HPL.AMOUNT))  AS TOTAL 
    FROM HR_EMPLOYEE_LIQUIDATION HP
    INNER JOIN HR_EMPLOYEE_LIQUIDATION_LINE HPL ON HPL.LIQUIDATION_ID = HP.ID
    INNER JOIN HR_SALARY_RULE HSR ON HSR.ID = HPL.RULE_ID
    WHERE HP.ID IN {3} AND HPL.type='liquidation'
    GROUP BY HSR.CODE
) P
INNER JOIN HR_SALARY_RULE HSR ON HSR.CODE = P.CODE
GROUP BY HSR.CODE

UNION
SELECT 'global_'||0 AS PAYSLIP_ID,
       'TOTAL_INGRESOS' AS RULE_CODE,
       COALESCE((
           SELECT SUM(COALESCE(HP.TOTAL_IN, 0.00))
           FROM HR_PAYSLIP_RUN HPR
           JOIN HR_PAYSLIP HP ON HP.PAYSLIP_RUN_ID = HPR.ID
           WHERE HPR.ID IN {0}
       ), 0.00)
       +
       COALESCE((
           SELECT SUM(COALESCE(HP.TOTAL_IN_LIQUIDATION, 0.00))
           FROM HR_EMPLOYEE_LIQUIDATION HP
           WHERE HP.ID IN {3}
       ), 0.00) AS TOTAL

UNION
SELECT 'global_'||0 AS PAYSLIP_ID,
       'TOTAL_EGRESOS' AS RULE_CODE,
       COALESCE((
           SELECT SUM(ABS(COALESCE(HP.TOTAL_OUT, 0.00)))
           FROM HR_PAYSLIP_RUN HPR
           JOIN HR_PAYSLIP HP ON HP.PAYSLIP_RUN_ID = HPR.ID
           WHERE HPR.ID IN {0}
       ), 0.00)
       +
       COALESCE((
           SELECT SUM(ABS(COALESCE(HP.TOTAL_OUT_LIQUIDATION, 0.00)))
           FROM HR_EMPLOYEE_LIQUIDATION HP
           WHERE HP.ID IN {3}
       ), 0.00) AS TOTAL

UNION
SELECT 'global_'||0 AS PAYSLIP_ID,
       'TOTAL_PROVISION' AS RULE_CODE,
       COALESCE((
           SELECT SUM(COALESCE(HP.TOTAL_PROVISION, 0.00))
           FROM HR_PAYSLIP_RUN HPR
           JOIN HR_PAYSLIP HP ON HP.PAYSLIP_RUN_ID = HPR.ID
           WHERE HPR.ID IN {0}
       ), 0.00)
       +
       COALESCE((
           SELECT SUM(COALESCE(HP.TOTAL_PROVISION_LIQUIDATION, 0.00))
           FROM HR_EMPLOYEE_LIQUIDATION HP
           WHERE HP.ID IN {3}
       ), 0.00) AS TOTAL

UNION
SELECT 'global_'||0 AS PAYSLIP_ID,
       'TOTAL' AS RULE_CODE,
       COALESCE((
           SELECT SUM(COALESCE(HP.TOTAL_PAYSLIP, 0.00))
           FROM HR_PAYSLIP_RUN HPR
           JOIN HR_PAYSLIP HP ON HP.PAYSLIP_RUN_ID = HPR.ID
           WHERE HPR.ID IN {0}
       ), 0.00)
       +
       COALESCE((
           SELECT SUM(COALESCE(HP.TOTAL_LIQUIDATION, 0.00))
           FROM HR_EMPLOYEE_LIQUIDATION HP
           WHERE HP.ID IN {3}
       ), 0.00) AS TOTAL
       
               ORDER BY 1,2 ASC 
            $$,
            $$ 
                {1}
            $$
                ) AS CT(
            PAYSLIP_ID VARCHAR {2}  
                ) 
            )

            SELECT 1 SEQUENCE_FORMAT,RC.NAME AS COMPANY_NAME,HE.IDENTIFICATION_ID,
            HE.NAME, hd.name AS DEPARTAMENTO,(Hj.NAME::JSON->'es_EC')::varchar AS CARGO,
            HC.DATE_START AS FECHA_CONTRATO,'ROL '||CASE WHEN(HPR.INACTIVE) THEN 'INACTIVOS' ELSE 'ACTIVOS' END  AS DSCR_PAYSLIP,hct.name AS TIPO_CONTRATO,
            HP.WAGE AS WAGE,
            HP.total_worked_days AS WORKED_DAYS,
            T.*  
            FROM TOTAL_PAYSLIP T
            INNER JOIN HR_PAYSLIP HP ON 'payslip_'||HP.ID=T.PAYSLIP_ID
            INNER JOIN HR_PAYSLIP_RUN HPR ON HPR.ID=HP.PAYSLIP_RUN_ID 
            INNER JOIN RES_COMPANY RC ON Rc.ID=HPR.COMPANY_id 
            INNER JOIN HR_EMPLOYEE HE ON HE.ID=HP.EMPLOYEE_ID
            INNER JOIN HR_DEPARTMENT HD ON HD.ID=HP.DEPARTMENT_ID
            INNER JOIN HR_JOB HJ ON HJ.ID=HP.JOB_ID
            INNER JOIN HR_CONTRACT HC ON HC.ID=HP.CONTRACT_ID
            inner join hr_contract_type hct on hct.id=hc.type_id
            union 
            SELECT 1 SEQUENCE_FORMAT,RC.NAME AS COMPANY_NAME,HE.IDENTIFICATION_ID,
            HE.NAME, hd.name AS DEPARTAMENTO,(Hj.NAME::JSON->'es_EC')::varchar AS CARGO,
            HC.DATE_START AS FECHA_CONTRATO,'FINIQUITO'  AS DSCR_PAYSLIP,hct.name AS TIPO_CONTRATO,
            HP.WAGE AS WAGE,
            HP.nopaid_Days AS WORKED_DAYS,
            T.*  
            FROM TOTAL_PAYSLIP T
            INNER JOIN HR_EMPLOYEE_LIQUIDATION HP ON 'finiquito_'||HP.ID=T.PAYSLIP_ID
            INNER JOIN RES_COMPANY RC ON Rc.ID=hp.COMPANY_id 
            INNER JOIN HR_EMPLOYEE HE ON HE.ID=HP.EMPLOYEE_ID
            INNER JOIN HR_DEPARTMENT HD ON HD.ID=HP.DEPARTMENT_ID
            INNER JOIN HR_JOB HJ ON HJ.ID=HP.JOB_ID
            INNER JOIN HR_CONTRACT HC ON HC.ID=HP.CONTRACT_ID
            inner join hr_contract_type hct on hct.id=hc.type_id 
            UNION ALL

-- 👇 Totales globales
SELECT 9999 AS SEQUENCE_FORMAT,
       '*** TOTAL GENERAL ***' AS COMPANY_NAME,
       NULL AS IDENTIFICATION_ID,
       NULL AS NAME,
       NULL AS DEPARTAMENTO,
       NULL AS CARGO,
       NULL AS FECHA_CONTRATO,
       'TOTAL GLOBAL' AS DSCR_PAYSLIP,
       NULL AS TIPO_CONTRATO,
       NULL AS WAGE,
       NULL AS WORKED_DAYS,
       T.*
FROM TOTAL_PAYSLIP T
WHERE T.PAYSLIP_ID = 'global_0'
            
            ORDER BY 1,3 ASC """.replace("{0}", str(tuple(LIST_IDS))).replace("{1}", COLUMN_QUERY).replace("{2}", COLUMN_DATA_TYPE).replace("{3}", str(tuple(FINIQUITO_IDS)))
            #print("11111111111111111")
            self._cr.execute(PAYSLIP_RUN_REPORT_QUERY)
            headers = [d.name for d in self._cr.description]
            #print(headers)
            result_payslip_run = self._cr.dictfetchall()
            i, INDEX_ROW = 0, 4
            last_row = INDEX_ROW
            ultima_columna_registros = 1
            COUNT = len(result_payslip_run)
            total=0.00
            if result_payslip_run:
                RANGE_COLUMNS = range(1, len(headers) )
                PATTERN_FILL=PatternFill(start_color='D9D9D9', end_color='D9D9D9',
                                            fill_type='solid')
                FONT_HEADER=Font(  bold=True, color='000000')
                ALIGNMENT = Alignment(horizontal='center', vertical='center')
                ALIGNMENT_RIGHT = Alignment(horizontal='right', vertical='center')
                ALIGNMENT_LEFT = Alignment(horizontal='left', vertical='center')

                DECIMAL_STYLE = NamedStyle(name="decimal_style", number_format="0.00")

                BLACK_LIST=[]
                FONT_BOLD=Font( bold=True)
                FONT_BOLD_LIST=["total_ingresos","total_egresos","total_provision","total"]
                ultima_columna=1
                #print("fffffffffff")
                for j in RANGE_COLUMNS:
                    column_name=headers[j].lower()
                    if column_name not in BLACK_LIST:
                        EACH_CELL = ws.cell(row=last_row, column=ultima_columna)
                        EACH_CELL.value = columns_dscr.get(column_name,column_name).capitalize()
                        EACH_CELL.fill = PATTERN_FILL # Yellow background
                        EACH_CELL.font =FONT_HEADER
                        ALIGNMENT.alignment=ALIGNMENT
                        ultima_columna+=1
                last_row += 1
                for each_payslip in result_payslip_run:
                    ultima_columna_registros = 1
                    for j in RANGE_COLUMNS:
                        column_name = headers[j].lower()
                        if column_name not in BLACK_LIST:
                            ROW_EACH_CELL = ws.cell(row=last_row, column=ultima_columna_registros)
                            value_write=each_payslip[column_name ]
                            if type(value_write)==str:
                                value_write = each_payslip[column_name].replace('"','').upper()
                            ROW_EACH_CELL.value=value_write
                            if type(value_write) == str:
                                ROW_EACH_CELL.alignment=ALIGNMENT_LEFT
                            if type(value_write) in (float,int):
                                ROW_EACH_CELL.alignment=ALIGNMENT_RIGHT
                                if type(value_write)==float:
                                    ROW_EACH_CELL.style=DECIMAL_STYLE
                            if column_name in FONT_BOLD_LIST or each_payslip["sequence_format"]==2:
                                ROW_EACH_CELL.font = FONT_BOLD
                            ultima_columna_registros += 1
                    last_row+=1
                    if each_payslip["sequence_format"]==1:
                        total+=each_payslip.get("total",0.00)

            detalles = []
            if rol_ids:
                detalles.append("ROLES: " + ",".join(map(str, rol_ids)))
            if finiquito_ids:
                detalles.append("FINIQUITOS: " + ",".join(map(str, finiquito_ids)))

            ids_texto = " | ".join(detalles)

            ws['A1'] = companys
            ws['B2'] =ids_texto
            ws['B3']= tipo
            ws['E2'] = dscr_periods
            ws['E3'] = round(total,DEC)
            ws['H2'] = COUNT-1
            ultima_letra = "A"
            if ultima_columna_registros>1:
                ultima_letra=utils.get_column_letter(ultima_columna_registros-1)
            if last_row >= INDEX_ROW:
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                self.set_border(ws, 'A' + str(INDEX_ROW) + ':'+ultima_letra + str(last_row - 1), border)
            wb = self.save_wb(wb, target)
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT