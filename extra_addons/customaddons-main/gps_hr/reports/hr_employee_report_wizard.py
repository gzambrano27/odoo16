# coding: utf-8
from odoo import api, fields, models, exceptions, tools, _
from ...message_dialog.tools import FileManager
from ...calendar_days.tools import DateManager
from ...calendar_days.tools import CalendarManager
dtFile=FileManager()
dateO=DateManager()
calendarO=CalendarManager()

import logging
_logger = logging.getLogger(__name__)

from openpyxl.styles.borders import Border, Side

class hr_employee_report_wizard_cumpleanios_xlsx(models.AbstractModel):
    _name = "report.gps_hr.hr_employee_report_wizard_cumpleanios_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Cumpleanios"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)
        OBJ_MOVEMENT = self.env["hr.employee.report.wizard"].sudo()
        brw_movement = OBJ_MOVEMENT.browse(docids[-1])
        file_path = "reporte_cumpleanios.xlsx"
        filename = dtFile.join(dir_path, file_path)
        dtFile.copyfile(filename, new_filename)
        wb = False
        try:
            wb, ws, target = self.open_xlsx(new_filename, load_sheet=True)
            self._cr.execute("""SELECT he.id,
he.name,
he.identification_id,
he.birthday,
EXTRACT(YEAR FROM AGE(he.birthday)) AS age
from hr_employee he
inner join hr_contract hc on hc.employee_id=he.id
inner join res_Company rc on rc.id=hc.company_id
where hc.state='open' and rc.id in %s order by he.name asc """,(tuple(brw_movement.company_ids.ids+[-1,-1]),))
            result=self._cr.dictfetchall()
            if result:
                i, INDEX_ROW = 0, 4
                last_row = INDEX_ROW
                for each_result in result:
                    row = str(INDEX_ROW + i)
                    ws['A' + row] = each_result["id"]
                    ws['B' + row] = each_result["name"]
                    ws['C' + row] = each_result["identification_id"]
                    ws['D' + row] = each_result["birthday"]
                    ws['E' + row] = each_result["age"]
                    i += 1
                    last_row = INDEX_ROW + i
                if last_row >= INDEX_ROW:
                    thin = Side(border_style="thin", color="000000")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    self.set_border(ws, 'A' + str(INDEX_ROW) + ':E' + str(last_row - 1), border)
            ws['A1'] = ','.join(brw_movement.company_ids.mapped('name'))
            ws['B2'] = len(result)
            wb = self.save_wb(wb, target)
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT


class hr_employee_report_wizard_cargas_familiares_xlsx(models.AbstractModel):
    _name = "report.gps_hr.hr_employee_report_wizard_cargas_familiares_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Cargas Familiares"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)
        OBJ_MOVEMENT = self.env["hr.employee.report.wizard"].sudo()
        brw_movement = OBJ_MOVEMENT.browse(docids[-1])
        file_path = "cargas_familiares.xlsx"
        filename = dtFile.join(dir_path, file_path)
        dtFile.copyfile(filename, new_filename)
        wb = False
        try:
            wb, ws, target = self.open_xlsx(new_filename, load_sheet=True)
            self._cr.execute("""SELECT distinct he.id,
he.name,
he.identification_id,
thf.name AS carga,
thf.birth_date,
case when(thf.relationship='wife_husband') then 'CONYUGUE' else 'HIJO' END as parentesco,
 EXTRACT(YEAR FROM AGE(thf.birth_date)) AS age,
 case when(thf.relationship='wife_husband') then 'SI' ELSE 
        (case when(EXTRACT(YEAR FROM AGE(thf.birth_date))<=18) then 'SI' ELSe 'NO' END  ) end as activo 
from hr_employee he
inner join hr_contract hc on hc.employee_id=he.id
inner join res_Company rc on rc.id=hc.company_id
inner join th_family_burden thf on thf.employee_id=he.id 
WHERE rc.id in %s  order by he.name asc """, (tuple(brw_movement.company_ids.ids + [-1, -1]),))
            result = self._cr.dictfetchall()
            if result:
                i, INDEX_ROW = 0, 4
                last_row = INDEX_ROW
                for each_result in result:
                    row = str(INDEX_ROW + i)
                    ws['A' + row] = each_result["id"]
                    ws['B' + row] = each_result["name"]
                    ws['C' + row] = each_result["identification_id"]
                    ws['D' + row] = each_result["carga"]
                    ws['E' + row] = each_result["birth_date"]
                    ws['F' + row] = each_result["age"]
                    ws['G' + row] = each_result["parentesco"]
                    ws['H' + row] = each_result["activo"]
                    i += 1
                    last_row = INDEX_ROW + i
                if last_row >= INDEX_ROW:
                    thin = Side(border_style="thin", color="000000")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    self.set_border(ws, 'A' + str(INDEX_ROW) + ':H' + str(last_row - 1), border)
            ws['A1'] = ','.join(brw_movement.company_ids.mapped('name'))
            ws['B2'] = len(result)
            wb = self.save_wb(wb, target)
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT