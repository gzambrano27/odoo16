# -*- coding: utf-8 -*-
# -*- encoding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.
from odoo import api,fields, models,_
from xlrd import open_workbook
import base64
import tempfile
import xlrd
from odoo.exceptions import ValidationError,UserError
from ...calendar_days.tools import CalendarManager,DateManager
from ...message_dialog.tools import FileManager
dtObj=DateManager()
clObj=CalendarManager()
flObj=FileManager()
from openpyxl import load_workbook
import json

class PurchaseOrderWizard(models.TransientModel):
    _name="purchase.order.wizard"
    _description="Asistente de Orden de Compra"
    
    @api.model
    def get_default_movement_id(self):
        return self._context.get("active_ids",False) and self._context["active_ids"][0] or False
    
    @api.model
    def get_default_company_id(self):
        if self._context.get("active_ids",False):
            for brw_each in self.env["purchase.order"].sudo().browse(self._context["active_ids"]):
                return brw_each.company_id.id
        return False


    movement_id=fields.Many2one("purchase.order", "Requisicion",required=False,default=get_default_movement_id)
    company_id=fields.Many2one("res.company",string="Compañia",required=False,default=get_default_company_id)
    currency_id = fields.Many2one("res.currency", "Moneda", related="company_id.currency_id", store=False,
                                  readonly=True)
    origin=fields.Selection([('file','Archivo'),],string="Origen",default='file')
    file=fields.Binary("Archivo",required=False,filters='*.xlsx')
    file_name=fields.Char("Nombre de Archivo",required=False,size=255)
    
    

    def process(self):
        for brw_each in self:
            if brw_each.origin=="file":
                brw_each.process_file()
        return True
    
    def process_file(self):
        OBJ_PRODUCT = self.env["product.product"].sudo()
        OBJ_ANALITICA = self.env["account.analytic.account"].sudo()
        ACTIVE_LINE_ID, Producto, Descripcion, Rubro, Tipo,CtaAnalitica, Unidad, Cantidad, Precio, Pvp = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9
        
        for brw_each in self:
            line_ids = []
            ext = flObj.get_ext(brw_each.file_name)
            fileName = flObj.create(ext)
            flObj.write(fileName, flObj.decode64(brw_each.file))

            book = load_workbook(fileName, data_only=True)
            sheet = book.active
            
            for row_index, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
                line_value = row[ACTIVE_LINE_ID]
                line_id = int(line_value) if isinstance(line_value, (int, float)) else 0

                default_code = str(row[Producto] or '').strip()
                descripcion = str(row[Descripcion] or '').strip()
                rubro = str(row[Rubro] or '').strip()
                ctaanalitica = str(row[CtaAnalitica] or '').strip()
                tipo_costo = str(row[Tipo] or '').strip()
                qty_value = row[Cantidad] or 0
                precio_unit = row[Precio] or 0
                precio_vta = row[Pvp] or 0
                json_dict = json.loads(ctaanalitica)
                #voy a buscar con el nombre el tipo de costo
                tipo_c = self.env['tipo.costo'].search([('name','=',tipo_costo)])
                # Obtener la primera clave
                ctaanalitica = list(json_dict.keys())[0]

                if not default_code:
                    continue
                
                if not isinstance(qty_value, (int, float)):
                    raise ValidationError(_("La cantidad debe ser numérica. Revisa la fila %s") % (row_index,))

                srch_product = OBJ_PRODUCT.search([('default_code', '=', default_code)])
                if not srch_product:
                    raise ValidationError(_("No hay producto con código %s. Revisa la fila %s") % (default_code, row_index))
                
                srch_cta = OBJ_ANALITICA.search([('id', '=', int(ctaanalitica)), ('company_id', '=', self.company_id.id)]) if ctaanalitica else None
                if ctaanalitica and not srch_cta:
                    raise ValidationError(_("No hay cuenta analitica con nombre %s. Revisa la fila %s") % (ctaanalitica, row_index))
                
                analytic_distribution = {str(srch_cta[0].id): 100} if srch_cta else {}

                values = {
                    "product_qty": qty_value,
                    "price_unit": precio_unit,
                    "precio_venta": precio_vta,
                    "product_id": srch_product[0].id,
                    "name": descripcion,
                    "rubro": rubro,
                    "tipo_costo_id": tipo_c[0].id,
                    'product_uom': srch_product[0].uom_po_id.id,
                    'analytic_distribution': analytic_distribution,
                }
                line_ids.append((1 if line_id > 0 else 0, line_id, values))
            
            brw_each.movement_id.write({"order_line": line_ids})
        return True
    
    def process_fileXX(self):
        OBJ_PRODUCT = self.env["product.product"].sudo()
        OBJ_ANALITICA = self.env["account.analytic.account"].sudo()
        OBJ_LOCATION = self.env["stock.location"].sudo()
        OBJ_LINE = self.env["purchase.order.line"].sudo()
        ACTIVE_LINE_ID, Producto, Descripcion, Rubro, CtaAnalitica, Tipo, Unidad, Cantidad, Precio, Pvp = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9
        for brw_each in self:
            line_ids = []
            ext = flObj.get_ext(brw_each.file_name)
            fileName = flObj.create(ext)
            flObj.write(fileName, flObj.decode64(brw_each.file))
    
            # Cargar el archivo con openpyxl
            book = load_workbook(fileName, data_only=True)
            sheet = book.active  # Selecciona la hoja activa
        
            for row_index, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):  # Comienza desde la fila 4
                line_value = row[ACTIVE_LINE_ID]
                line_id = 0
                if line_value:
                    if isinstance(line_value, (int, float)):
                        line_id = int(line_value)
                    else:
                        if line_value != "":
                            raise ValidationError(_("El # ID debe ser numérico o vacío. Revisa la fila %s") % (row_index,))
                # Este código se mantiene igual que el original, adaptando cómo se acceden a los valores de las celdas.
                default_code = str(row[Producto] or '').strip()
                qty_value = row[Cantidad]
                precio_unit = row[Precio]
                tipo_costo = row[Tipo]
                precio_vta = row[Pvp]
                descripcion = row[Descripcion]
                rubro = row[Rubro]
                ctaanalitica = row[CtaAnalitica]
                quantity = int(qty_value)
                if not isinstance(qty_value, (int, float)):
                    raise ValidationError(_("La cantidad debe ser numérica. Revisa la fila %s") % (row_index,))
                if not default_code:
                    continue
                srch_product = OBJ_PRODUCT.search([('default_code', '=', default_code)])
                if not srch_product:
                    raise ValidationError(_("No hay producto con código %s. Revisa la fila %s") % (default_code, row_index))
                if len(srch_product) > 1:
                    raise ValidationError(_("Existe más de un producto con código %s en la base de datos. Revisa la fila %s") % (default_code, row_index))    
                srch_cta = ''
                if ctaanalitica:
                    srch_cta = OBJ_ANALITICA.search([('name', '=', ctaanalitica),('company_id', '=', self.company_id.id)])
                    if not srch_cta:
                        raise ValidationError(_("No hay cuenta analitica con nombre %s. Revisa la fila %s") % (ctaanalitica, row_index))
                values = {
                    "product_qty": quantity,
                    "price_unit": precio_unit,
                    "precio_venta": precio_vta,
                    "product_id": srch_product[0].id,
                    "name": descripcion,
                    "rubro": rubro,
                    "tipo_costo": tipo_costo,
                    'product_uom': srch_product[0].uom_po_id.id,
                    'analytic_distribution': {str(srch_cta[0].id): 100}
                }
                line_ids.append((1 if line_id > 0 else 0, line_id, values))
            brw_each.movement_id.write({"order_line": line_ids})
        return True