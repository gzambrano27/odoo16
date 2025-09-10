import base64
import io
import xlsxwriter
from openpyxl import load_workbook
from odoo import models, fields, api, _
from odoo.exceptions import ValidationError
from odoo.exceptions import UserError
class Planilla(models.Model):
    _name = 'planillas.planilla'
    _inherit = ['portal.mixin', 'mail.thread', 'mail.activity.mixin']
    _description = 'Planilla'

    name = fields.Char(string='Nombre', readonly=True, copy=False, default='PLANILLA - New')
    usuario_id = fields.Many2one(
        'res.users',
        string='Usuario',
        tracking=True,
        default=lambda self: self.env.user,
        readonly=True
    )
    cuenta_analitica_id = fields.Many2one('account.analytic.account', string='Cuenta Analítica', tracking=True)
    notebook_ids = fields.One2many('planillas.notebook', 'planilla_id', string='Notebook', tracking=True)
    descuento = fields.Float(string='Descuento (%)', tracking=True, help="Descuento aplicado a la planilla", default=0.0, digits=(3,2))

    @api.constrains('descuento')
    def _check_descuento_range(self):
        for record in self:
            if not (0 <= record.descuento <= 100):
                raise ValidationError("El descuento debe estar entre 0 y 100.")

    total_subtotal = fields.Monetary(string='Total ', compute='_compute_totales', store=True, readonly=True)
    total_subtotal_descuento = fields.Monetary(string='Total Descuento', compute='_compute_totales', store=True, readonly=True)
    total_diferencia_subtotal = fields.Monetary(string='Total Diferencia', compute='_compute_totales', store=True, readonly=True)
    total_subtotal_planilla = fields.Monetary( string='Total Planilla', compute='_compute_totales', store=True, readonly=True)
    total_subtotal_planilla_descuento = fields.Monetary(string='Total Planilla Descuento', compute='_compute_totales', store=True, readonly=True)
    total_diferencia_subtotal_descuento = fields.Monetary(string='Total Diferencia Planilla', compute='_compute_totales', store=True, readonly=True)

    company_id = fields.Many2one("res.company", string="Compañía", copy=False, default=lambda self: self.env.company, _check_company_auto=True)

    currency_id = fields.Many2one(related="company_id.currency_id", string="Moneda", store=False, readonly=True)

    @api.depends('notebook_ids.subtotal', 'notebook_ids.subtotal_descuento',
                 'notebook_ids.subtotal_planilla', 'notebook_ids.subtotal_planilla_descuento')
    def _compute_totales(self):
        for record in self:
            record.total_subtotal = sum(notebook.subtotal for notebook in record.notebook_ids)
            record.total_subtotal_descuento = sum(notebook.subtotal_descuento for notebook in record.notebook_ids)
            record.total_diferencia_subtotal = sum(notebook.diferencia_subtotal for notebook in record.notebook_ids)
            record.total_subtotal_planilla = sum(notebook.subtotal_planilla for notebook in record.notebook_ids)
            record.total_subtotal_planilla_descuento = sum(notebook.subtotal_planilla_descuento for notebook in record.notebook_ids)
            record.total_diferencia_subtotal_descuento = sum(notebook.diferencia_subtotal_descuento for notebook in record.notebook_ids)

    state = fields.Selection(
        [('draft', 'Borrador'),
         ('to_approve', 'Por Aprobar'),
         ('approved', 'Aprobado'),
         ('cancel', 'Cancelado')],
        string='Estado', default='draft', tracking=True, readonly=True, copy=False
    )

    @api.model
    def create(self, vals):
        if vals.get('name', 'PLANILLA - New') == 'PLANILLA - New':
            vals['name'] = self.env['ir.sequence'].next_by_code('planillas.planilla') or 'PLANILLA - New'
        vals['state'] = 'draft'
        vals['usuario_id'] = self.env.user.id
        return super(Planilla, self).create(vals)

    def action_to_approve(self):
        """Cambiar el estado a 'Por Aprobar'."""
        self.write({'state': 'to_approve'})

    def action_approve(self):
        """Cambiar el estado a 'Aprobado'."""
        self.write({'state': 'approved'})

    def action_cancel(self):
        """Cambiar el estado a 'Cancelado'."""
        self.write({'state': 'cancel'})

    def action_reset_to_draft(self):
        """Reestablecer el estado a 'Borrador'."""
        self.write({'state': 'draft'})

    def export_to_excel(self):
        """Exporta los datos de la planilla en un archivo Excel con formato detallado."""
        for planilla in self:
            if not planilla.notebook_ids:
                raise UserError("La planilla no tiene Notebooks para exportar.")

            # Crear un buffer en memoria para el archivo Excel
            buffer = io.BytesIO()

            # Crear el archivo Excel
            workbook = xlsxwriter.Workbook(buffer, {'in_memory': True})
            worksheet = workbook.add_worksheet('Planilla')

            # Estilos básicos
            header_format = workbook.add_format({'bold': True, 'bg_color': '#D3D3D3', 'border': 1})
            text_format = workbook.add_format({'border': 1})
            number_format = workbook.add_format({'border': 1, 'num_format': '#,##0.00'})

            # Encabezados básicos
            headers = [
                'Planilla Nombre',
                'Notebook Rubro',
                'Notebook Descripción',
                'Notebook Precio Unitario',
                'Notebook Descuento',
                'Notebook Cantidad',
                'Notebook Subtotal',
                'Notebook Subtotal Descuento',
                'Notebook Consumo',
                'Notebook Diferencia Cantidad',
                'Notebook Subtotal Planilla',
                'Notebook Subtotal Planilla Descuento',
                'Notebook Diferencia Subtotal',
                'Notebook Diferencia Subtotal Descuento',
                'Notebook Observación'
            ]

            # Agregar encabezados dinámicos para los detalles
            max_detalles = max(len(notebook.detalle_ids) for notebook in planilla.notebook_ids)
            for i in range(1, max_detalles + 1):
                headers.extend([f'Detalle {i} - Tipo', f'Detalle {i} - Valor'])

            # Escribir encabezados en Excel
            for col_num, header in enumerate(headers):
                worksheet.write(0, col_num, header, header_format)

            # Escribir datos
            row = 1  # La fila inicial para los datos
            for notebook in planilla.notebook_ids:
                # Datos básicos del notebook
                notebook_data = [
                    planilla.name,  # Nombre de la planilla
                    notebook.rubro_id.name if notebook.rubro_id else '',
                    notebook.descripcion or '',
                    notebook.precio_unitario or 0.0,
                    notebook.descuento or 0.0,
                    notebook.cantidad or 0.0,
                    notebook.subtotal or 0.0,
                    notebook.subtotal_descuento or 0.0,
                    notebook.consumo or 0.0,
                    notebook.diferencia_cantidad or 0.0,
                    notebook.subtotal_planilla or 0.0,
                    notebook.subtotal_planilla_descuento or 0.0,
                    notebook.diferencia_subtotal or 0.0,
                    notebook.diferencia_subtotal_descuento or 0.0,
                    notebook.observacion or ''
                ]

                # Agregar los detalles del notebook
                detalles = []
                for detalle in notebook.detalle_ids:
                    detalles.extend([detalle.tipo or '', detalle.valor or 0.0])

                # Rellenar con celdas vacías si faltan detalles
                while len(detalles) < max_detalles * 2:
                    detalles.extend(['', ''])

                # Combinar datos del notebook con los detalles
                row_data = notebook_data + detalles

                # Escribir datos en Excel
                for col_num, cell_data in enumerate(row_data):
                    if isinstance(cell_data, (int, float)):
                        worksheet.write(row, col_num, cell_data, number_format)
                    else:
                        worksheet.write(row, col_num, cell_data, text_format)

                row += 1

            workbook.close()
            buffer.seek(0)

            # Crear archivo adjunto
            file_data = buffer.read()
            buffer.close()

            attachment = self.env['ir.attachment'].create({
                'name': f'{planilla.name}.xlsx',
                'type': 'binary',
                'datas': base64.b64encode(file_data).decode('utf-8'),
                'res_model': self._name,
                'res_id': planilla.id,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            })

            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?download=true',
                'target': 'new',
            }

    import_file = fields.Binary(string="Archivo Excel")
    import_filename = fields.Char(string="Nombre del Archivo")

    def import_from_excel(self):
        """Importa datos desde un archivo Excel al modelo planillas.planilla."""
        if not self.import_file or not self.import_filename.endswith('.xlsx'):
            raise UserError(_("Por favor, cargue un archivo Excel válido."))

        try:
            # Decodificar y leer el archivo Excel
            file_data = base64.b64decode(self.import_file)
            file_content = io.BytesIO(file_data)

            # Cargar el archivo Excel usando openpyxl
            workbook = load_workbook(file_content)
            sheet = workbook.active  # Seleccionar la hoja activa

            # Obtener las filas del archivo Excel
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise UserError(_("El archivo Excel está vacío."))

            headers = rows[0]  # Encabezados
            data_rows = rows[1:]  # Filas de datos

            # Verificar que estamos trabajando con el registro activo
            planilla = self  # Se trabaja sobre el registro actual de planilla

            for row in data_rows:
                if len(row) < len(headers):
                    raise UserError(_("El archivo Excel tiene filas con menos columnas que los encabezados."))

                # Asignar el nombre de la planilla si se encuentra en la fila
                if row[0]:
                    planilla.write({'name': row[0]})  # Nombre de la planilla

                # Crear notebooks asociados a la planilla
                notebook_data = {
                    'planilla_id': planilla.id,  # Relacionar con la planilla existente
                    'rubro_id': self._get_rubro_id(row[1]),  # Rubro
                    'descripcion': row[2],  # Descripción
                    'precio_unitario': float(row[3]) if row[3] else 0.0,  # Precio Unitario
                    'descuento': float(row[4]) if row[4] else 0.0,  # Descuento
                    'cantidad': float(row[5]) if row[5] else 0.0,  # Cantidad
                    'subtotal': float(row[6]) if row[6] else 0.0,  # Subtotal
                    'subtotal_descuento': float(row[7]) if row[7] else 0.0,  # Subtotal Descuento
                    'consumo': float(row[8]) if row[8] else 0.0,  # Consumo
                    'diferencia_cantidad': float(row[8]) if row[8] else 0.0,  # Consumo
                    'subtotal_planilla': float(row[10]) if row[10] else 0.0,  # Subtotal Planilla
                    'subtotal_planilla_descuento': float(row[11]) if row[11] else 0.0,  # Subtotal Planilla Descuento
                    'diferencia_subtotal': float(row[12]) if row[12] else 0.0,  # Diferencia Subtotal
                    'diferencia_subtotal_descuento': float(row[13]) if row[13] else 0.0,
                    # Diferencia Subtotal Descuento
                    'observacion': row[14],  # Observación
                }
                notebook = planilla.notebook_ids.create(notebook_data)

                # Procesar detalles asociados al notebook
                detalle_index = 15  # La columna 14 y siguientes son detalles
                while detalle_index < len(headers):
                    if "Detalle" not in headers[detalle_index]:
                        break
                    tipo = row[detalle_index]  # Tipo de detalle
                    valor = row[detalle_index + 1]  # Valor del detalle

                    # Validar que 'tipo' no sea nulo
                    if not tipo:
                        detalle_index += 2  # Saltar al siguiente detalle
                        continue

                    detalle_data = {
                        'notebook_id': notebook.id,  # Relacionar el detalle con el notebook
                        'tipo': tipo,  # Tipo de detalle
                        'valor': float(valor) if valor else 0.0,  # Valor del detalle (0.0 por defecto)
                    }
                    notebook.detalle_ids.create(detalle_data)
                    detalle_index += 2  # Cada detalle ocupa 2 columnas (tipo y valor)

            file_content.close()
            return True

        except Exception as e:
            raise UserError(_("Error al procesar el archivo Excel: %s") % str(e))

    # Métodos auxiliares
    def _get_rubro_id(self, rubro_name):
        """Obtiene el ID del rubro a partir de su nombre."""
        if not rubro_name:
            return False
        rubro = self.env['planillas.rubro'].search([('name', '=', rubro_name)], limit=1)
        if not rubro:
            raise UserError(f"Rubro no encontrado: {rubro_name}")
        return rubro.id

class Notebook(models.Model):
    _name = 'planillas.notebook'
    _description = 'Notebook'

    planilla_id = fields.Many2one('planillas.planilla', string='Planilla', required=True, ondelete='cascade')
    rubro_id = fields.Many2one('planillas.rubro', string='Rubro', tracking=True, domain=[('estado', '=', 'activo')])
    descripcion = fields.Text(string='Descripción', tracking=True)
    unidad_medida = fields.Many2one(string='Unidad', related='rubro_id.unidad_medida', readonly=True, store=True)
    precio_unitario = fields.Monetary(string='PVP', tracking=True)
    descuento = fields.Float(string='Desc (%)', tracking=True)

    @api.constrains('descuento')
    def _check_descuento_range(self):
        for record in self:
            if not (0 <= record.descuento <= 100):
                raise ValidationError("El descuento debe estar entre 0 y 100 en el Notebook.")

    cantidad = fields.Float(string='Cantidad', tracking=True)
    subtotal = fields.Monetary(string='Subtotal', compute='_compute_subtotal', store=True, readonly=True, tracking=True)
    subtotal_descuento = fields.Monetary(string='Subtotal Descuento', compute='_compute_subtotal_descuento', store=True, readonly=True, tracking=True)
    diferencia_subtotal = fields.Monetary(string='Diferencia subtotal', compute='_compute_diferencia_subtotal', store=True, readonly=True, tracking=True)
    consumo = fields.Float(string='Consumo', compute='_compute_consumo', store=True, readonly=True)
    diferencia_cantidad = fields.Float(string='Diferencia cantidad', compute='_compute_diferencia_cantidad', store=True, readonly=True, tracking=True)
    subtotal_planilla = fields.Monetary(string='Sub. Planilla', compute='_compute_subtotal_planilla', store=True, readonly=True, tracking=True)
    subtotal_planilla_descuento = fields.Monetary(string='Sub. Planilla Descuento', compute='_compute_subtotal_planilla_descuento', store=True, readonly=True, tracking=True)
    diferencia_subtotal_descuento = fields.Monetary(string='Dif. Sub. Descuento', compute='_compute_diferencia_subtotal_descuento', store=True, readonly=True, tracking=True)
    detalle_ids = fields.One2many('planillas.detalle', 'notebook_id', string='Detalles', tracking=True)
    observacion = fields.Text(string='Observación', tracking=True)

    currency_id = fields.Many2one(related="planilla_id.currency_id", string="Moneda", store=False, readonly=True)

    @api.onchange('rubro_id')
    def _onchange_rubro(self):
        """Actualizar el precio al seleccionar un rubro."""
        if self.rubro_id:
            self.precio_unitario = self.rubro_id.precio_unitario

    @api.depends('detalle_ids.valor')
    def _compute_consumo(self):
        for record in self:
            record.consumo = sum(detalle.valor for detalle in record.detalle_ids)

    @api.depends('precio_unitario', 'cantidad')
    def _compute_subtotal(self):
        for record in self:
            record.subtotal = record.precio_unitario * record.cantidad if record.cantidad > 0 else 0

    @api.depends('precio_unitario', 'cantidad', 'descuento')
    def _compute_subtotal_descuento(self):
        for record in self:
            if record.cantidad > 0:
                total = record.precio_unitario * record.cantidad
                descuento = total * (record.descuento / 100)
                record.subtotal_descuento = total - descuento
            else:
                record.subtotal_descuento = 0

    @api.depends('subtotal', 'subtotal_planilla')
    def _compute_diferencia_subtotal(self):
        for record in self:
            record.diferencia_subtotal = record.subtotal - record.subtotal_planilla

    @api.depends('precio_unitario', 'consumo')
    def _compute_subtotal_planilla(self):
        for record in self:
            record.subtotal_planilla = record.precio_unitario * record.consumo if record.consumo > 0 else 0

    @api.depends('precio_unitario', 'consumo', 'descuento')
    def _compute_subtotal_planilla_descuento(self):
        for record in self:
            if record.consumo > 0:
                total = record.precio_unitario * record.consumo
                descuento = total * (record.descuento / 100)
                record.subtotal_planilla_descuento = total - descuento
            else:
                record.subtotal_planilla_descuento = 0

    @api.depends('subtotal_descuento', 'subtotal_planilla_descuento')
    def _compute_diferencia_subtotal_descuento(self):
        for record in self:
            record.diferencia_subtotal_descuento = record.subtotal_descuento - record.subtotal_planilla_descuento

    @api.depends('cantidad', 'consumo')
    def _compute_diferencia_cantidad(self):
        for record in self:
            record.diferencia_cantidad = record.cantidad - record.consumo

class Detalle(models.Model):
    _name = 'planillas.detalle'
    _description = 'Detalle del Notebook'

    notebook_id = fields.Many2one('planillas.notebook', string='Notebook', required=True, ondelete='cascade')
    tipo = fields.Char(string='Tipo', required=True, help="Tipo de detalle asociado al Notebook")
    valor = fields.Float(string='Valor', required=True, help="Valor asociado al detalle")
