# coding: utf-8
from odoo import api, fields, models, exceptions, tools, _
from ...message_dialog.tools import FileManager
from ...calendar_days.tools import DateManager
from ...calendar_days.tools import CalendarManager
dtFile = FileManager()
dateO = DateManager()
calendarO = CalendarManager()

import logging

_logger = logging.getLogger(__name__)

from openpyxl.styles.borders import Border, Side

class report_inventory_document_template_xlsx(models.AbstractModel):
    _name = "report.gps_inventario.report_inventory_document_template_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Plantilla de Inventario"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)

        wb = False
        try:
            filename = "reporte_ajuste.xlsx"
            filename=dtFile.join(dir_path, filename)
            dtFile.copyfile(filename, new_filename)

            brw_document= self.env["inventory.document.adjust"].sudo().browse(docids[-1])

            wb, ws, target = self.open_xlsx(new_filename, load_sheet=True)
            i, INDEX_ROW = 0, 6
            last_row = INDEX_ROW
            for brw_line in brw_document.line_ids:
                row = str(INDEX_ROW + i)
                if brw_line.origin == 'original':
                    ws['A' + row] = brw_line.id
                ws['B' + row] = brw_line.stock_location_id.name
                ws['C' + row] = brw_line.referencia_anterior or ''
                ws['D' + row] = brw_line.product_id.default_code or ''
                ws['E' + row] = brw_line.product_id.name
                ws['F' + row] = brw_line.quantity
                ws['G' + row] = brw_line.comments or ''
                ws['H' + row] = brw_line.stock
                i += 1
                last_row = INDEX_ROW + i
                if last_row >= INDEX_ROW:
                    thin = Side(border_style="thin", color="000000")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    LETRA_FINAL="G"
                    self.set_border(ws, 'A' + str(INDEX_ROW) + ':' + LETRA_FINAL+str(last_row - 1), border)
            ws['G2'] = len(brw_document.line_ids)
            ws['A1'] = brw_document.company_id.name
            ws['B2'] = brw_document.date_from
            product_tmpl_ids=brw_document.line_ids.mapped('product_id').mapped('product_tmpl_id')
            product_tmpl_ids=product_tmpl_ids.filtered(lambda x: x.default_code )
            ws['B3'] = ",".join(product_tmpl_ids.mapped('default_code'))
            ws['C3'] = ",".join(product_tmpl_ids.mapped('name'))
            ws['B4'] = brw_document.stock_location_id.warehouse_id.name+" "+brw_document.stock_location_id.name
            ws['G3'] = ",".join(brw_document.mapped('user_ids').mapped('name'))

            wb = self.save_wb(wb, target)
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT
