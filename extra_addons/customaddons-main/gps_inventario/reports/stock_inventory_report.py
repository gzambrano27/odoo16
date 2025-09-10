# coding: utf-8
from odoo import api, fields, models, exceptions, tools, _
from ...message_dialog.tools import FileManager
from ...calendar_days.tools import DateManager
from ...calendar_days.tools import CalendarManager
import datetime
dtFile = FileManager()
dateO = DateManager()
calendarO = CalendarManager()
import base64
import xlsxwriter
import io
import logging

_logger = logging.getLogger(__name__)

from openpyxl.styles.borders import Border, Side


def get_list_names(location_ids):
    locations_list = []
    # Iterar sobre cada almacén
    for location in location_ids:
        # Crear el formato 'warehouse_id/location_id'
        location_format = f"{location.warehouse_id.name}/{location.name}"
        locations_list.append(location_format)
    return locations_list

class report_all_movement_xlsx(models.AbstractModel):
    _name = "report.gps_inventario.report_all_movement_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Movimientos de Inventario"

    def create_xlsx_report(self, docids, data):


        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)

        wb = False
        try:
            acceso_costo = self.user_has_groups('account_payment_purchase.group_costo_usuario')
            filename = "reporte_movimiento_inventario.xlsx"
            if acceso_costo:
                filename = "reporte_movimiento_inventario_costo.xlsx"
            filename = dtFile.join(dir_path, filename)
            dtFile.copyfile(filename, new_filename)

            brw_wizard = self.env["stock.inventory.report"].sudo().browse(docids[-1])
            company_ids = brw_wizard.mapped('company_ids').ids
            product_ids = brw_wizard.mapped('product_ids').ids

            warehouse_ids = brw_wizard.mapped('warehouse_ids').ids

            location_ids = brw_wizard.mapped('location_ids').ids

            date_from = brw_wizard.date_from if brw_wizard.date_from is not None else datetime.date(2023, 1, 1)

            date_to = brw_wizard.date_to if brw_wizard.date_to is not None else fields.Date.context_today(self)


            wb, ws, target = self.open_xlsx(new_filename, load_sheet=True)
            user_tz = self.env.user.tz or 'UTC'
            self._cr.execute(""";WITH variables AS (
                SELECT %s::date as date_from,%s::date as date_to, %s::INTEGER[] AS company_ids,
                       %s::INTeger[] AS product_ids, 
            		  %s::INTeger[] AS WAREHOUSE_IDs,
            		  %s::INTeger[] AS location_ids ,
            		  %s::bool as grouped
            ), 
            warehouse_locations as ( 
            	select sl.id as location_id ,
            	case when(slv.usage='view') then sl.id else  sl.location_id end as parent_location_id,
            	case when(slv.usage='view' or not variables.grouped) then sl.id else  sl.location_id end as final_location_id,
            	sl.name,
            	sl.usage ,sl.warehouse_id
            	FROM variables
            	inner join stock_location sl on 1=1 
            	inner join stock_location slv on sl.location_id=slv.id
            	where sl.company_id=any(variables.company_ids) and 
            		sl.warehouse_id=any(variables.warehouse_ids)  
            		and sl.usage='internal' 
            		and (not variables.grouped and  
            		sl.id=any(variables.location_ids) ) or 
            		(variables.grouped and  
            			(
            				(sl.location_id=any(variables.location_ids)) or 
            				(sl.id=any(variables.location_ids))
            			)
            		)
            ),
            not_categories as (
                select pc.id,pcp.name as parent_name,pc.name from product_category pc
                inner join product_category pcp on pcp.id=pc.parent_id
                where pcp.name ilike 'ACTIVO%%'
                union
                select pcp.id,'' as parent_name,pcp.name from 
                product_category  pcp
                where pcp.name ilike 'ACTIVO%%' and pcp.parent_id is null
                order by 1 asc,2 asc,3 asc  
            ),
            all_stock_movements as ( 
            	SELECT variables.grouped,
            	   sm.company_id,
            	   pp.id as product_id,
                   pp.product_tmpl_id,
                   pp.default_code,
                   pp.barcode,
                   COALESCE(pt.name::json->>'es_EC', pt.name::json->>'en_US') AS product_template_name,
            	   sml.id as stock_move_line_id,
                   sm.id AS stock_move_id,
                   sm.picking_id,
                   sm.name,
                   sm.reference,
                   sml.qty_done as quantity_done,
            	   sm.location_id,
            	    sm.location_dest_id,
            		coalesce(slo.final_location_id,sm.location_id) as parent_location_id,
            	    coalesce(sld.final_location_id,sm.location_dest_id) as parent_location_dest_id,
            		slo.warehouse_id as warehouse_id,
            		sld.warehouse_id as warehouse_dest_id,
                   (sml.date AT TIME ZONE 'UTC' AT TIME ZONE 'America/Guayaquil' ) as adjusted_date,
            	   case when sld.usage='internal' then sml.qty_done else  0 end as qty_in,
            	   case when slo.usage='internal' then -sml.qty_done else  0 end as qty_out,
            	   case when sld.usage='internal' then sml.qty_done else  0 end+case when slo.usage='internal' then -sml.qty_done else  0 end as qty ,
            	   pt.categ_id ,
            	   pt.detailed_type,
            	   (sm.date AT TIME ZONE 'UTC' AT TIME ZONE 'America/Guayaquil' ) as fecha_programada_date  
            	FROM variables
            	INNER JOIN stock_move sm 
            	    ON sm.company_id = ANY(variables.company_ids) 
            	     AND (
                     (
                        array_length(variables.product_ids, 1) > 0 AND sm.product_id = ANY(variables.product_ids))
                        OR (array_length(variables.product_ids, 1) IS NULL OR array_length(variables.product_ids, 1) = 0
                      )
                    )
            	    AND sm.state = 'done' 
            	inner join stock_move_line  sml on sml.move_id=sm.id 
            	inner join product_product pp on pp.id=sm.product_id
            	INNER JOIN product_template pt  ON pt.id = pp.product_tmpl_id    
            	left join  not_categories on  not_categories.id=   pt.categ_id 	
            	left join warehouse_locations slo on slo.location_id=sm.location_id  
            	left join warehouse_locations sld on sld.location_id=sm.location_dest_id  
            	where (slo.location_id is not null or sld.location_id is not null)
            	and not_categories.id is null 
            ) , 
valuation_layers as (
	select svm.stock_move_id,svm.product_id,
	sum(svm.value) as cost_value,
	sum(svm.quantity) as cost_qty,
	(case when(sum(svm.quantity)!=0.00) then sum(svm.value)/sum(svm.quantity) else null end ) as cost_unit,
	ARRAY_AGG(DISTINCT svm.account_move_id) AS account_move_ids
	from all_stock_movements asm
	inner join stock_Valuation_layer svm on asm.stock_move_id=svm.stock_move_id
	group by svm.stock_move_id ,svm.product_id
) ,
product_config_account as (
		select AAC.company_id, pp.id as product_id,AAC.id as account_id
		from
		variables 
		inner join PRODUCT_PRODUCT PP on 1=1 
		INNER JOIN PRODUCT_TEMPLATE PT ON PT.ID=PP.PRODUCT_TMPL_ID 
		LEFT JOIN IR_PROPERTY IPPTC ON IPPTC.RES_ID=('product.category,'||PT.CATEG_ID )::VARCHAR
				AND IPPTC.TYPE='many2one' AND IPPTC.name='property_stock_valuation_account_id'
				AND IPPTc.COMPANY_ID=any(variables.COMPANY_IDs) 
		LEFT JOIN ACCOUNT_ACCOUNT AAC ON ('account.account,'||AAC.ID)::VARCHAR=IPPTC.VALUE_REFERENCE::VARCHAR AND IPPTC.ID IS NOT NULL 
		group by AAC.company_id,pp.id,AAC.id
),
valuation_layers_account as (
	select vl.*,
	        max(am.date) as move_date,
			sum(aml.debit) as debit,
			sum(aml.credit) as credit,
			aml.account_id,
			am.state as move_state, 
			aa.code as move_code,
			aa.name as move_name
	from valuation_layers vl
	inner join account_move am on am.id= any(vl.account_move_ids)
	inner join account_move_line aml on aml.move_id=am.id
	inner join account_account aa on aa.id=aml.account_id and aa.company_id=am.company_id 
	inner join product_config_account pca on pca.company_id=aa.company_id  and pca.account_id=aa.id  and pca.product_id=vl.product_id 
	group by vl.stock_move_id,vl.product_id,vl.cost_value,vl.cost_qty,vl.cost_unit,vl.account_move_ids,
	aml.account_id,am.state,aa.code,aa.name 
),
valoracion_ini as (
     SELECT 
         svl.company_id,
        svl.product_id, 
		svl.quantity,
		sm.date::date as date,
        sm.date AS fecha_hasta, 
        svl.unit_cost,
		(case when(svl.quantity!=0.00) then svl.value/svl.quantity else 0.00 end ) as cost_unit_compute,
        (LAG(DATE(sm.date), 1, '2019-12-31') OVER (PARTITION BY svl.product_id ORDER BY DATE(sm.date))+ INTERVAL '1 DAY')::date AS fecha_Desde
    FROM stock_valuation_layer svl
	inner join variables on svl.company_id=ANY(variables.company_ids)  
				AND (
         (
            array_length(variables.product_ids, 1) > 0 AND svl.product_id = ANY(variables.product_ids))
            OR (array_length(variables.product_ids, 1) IS NULL OR array_length(variables.product_ids, 1) = 0
          )
        )
    INNER JOIN stock_move sm ON sm.id = svl.stock_move_id
    WHERE sm.date = (
            SELECT MAX(sm2.date)
            FROM stock_move sm2
            INNER JOIN stock_valuation_layer svl2 ON sm2.id = svl2.stock_move_id
			inner join variables on svl2.company_id=ANY(variables.company_ids) 
					AND (
         (
            array_length(variables.product_ids, 1) > 0 AND svl2.product_id = ANY(variables.product_ids))
            OR (array_length(variables.product_ids, 1) IS NULL OR array_length(variables.product_ids, 1) = 0
          )
        )
            WHERE svl2.product_id = svl.product_id
                AND svl2.company_id = svl.company_id
                AND DATE(sm2.date) = DATE(sm.date)
        )
),costeos as (
	SELECT 
		  company_id,
		  product_id,
		  fecha_Desde,
		  fecha_hasta,
		  -- Movimiento: la cantidad (positivo para entrada, negativo para salida)
		  quantity AS movimiento,
		  cost_unit_compute,
		  -- Valor del movimiento: cantidad x costo unitario de la transacción
		  quantity * cost_unit_compute AS valor_movimiento,
		  -- Saldo acumulado de cantidad
		  SUM(quantity) OVER (PARTITION BY product_id ORDER BY fecha_hasta, fecha_Desde) AS saldo_cantidad,
		  -- Saldo acumulado de valor
		  SUM(quantity * cost_unit_compute) OVER (PARTITION BY product_id ORDER BY fecha_hasta, fecha_Desde) AS saldo_valor,
		  -- Costo promedio del saldo (si el saldo de cantidad es 0, se evita la división)
		  CASE 
		    WHEN SUM(quantity) OVER (PARTITION BY product_id ORDER BY fecha_hasta, fecha_Desde) <> 0
		    THEN SUM(quantity * cost_unit_compute) OVER (PARTITION BY product_id ORDER BY fecha_hasta, fecha_Desde) /
		         SUM(quantity) OVER (PARTITION BY product_id ORDER BY fecha_hasta, fecha_Desde)
		    ELSE 0
		  END AS costo_promedio_saldo
		FROM valoracion_ini
		ORDER BY date, fecha_Desde--fecha_hasta
),
valoracion as (
    SELECT company_id,product_id, fecha_desde::date AS fecha_desde, 
		fecha_hasta::date AS fecha_hasta,  costo_promedio_saldo as cost_unit
    FROM costeos
    UNION     
    SELECT 
        v.company_id,
        v.product_id, 
        (MAX(v.fecha_hasta) OVER (PARTITION BY v.company_id, v.product_id) + INTERVAL '1 DAY')::date AS fecha_desde, 
        (CURRENT_DATE + INTERVAL '1 DAY')::date AS fecha_hasta,
        LAST_VALUE(v.costo_promedio_saldo) OVER (PARTITION BY v.company_id, v.product_id ORDER BY v.fecha_hasta ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING) AS cost_unit
    FROM costeos v
),
preliminar_kardex as (
	select s.*,sp.name as picking_name, 
	case when(sw.id  is not null) then sw.NAME||'/'||sl.NAME else NULL end as origen,
	case when(swd.id  is not null) then swd.NAME||'/'||sld.NAME else NULL end as destino ,
	/*vly.cost_counter,
	vly.cost_value,
	vly.cost_qty,*/
	coalesce(vly.cost_unit,null,/*,v.cost_unit necesito sacar el ultimo costo es decir el ultimo vly.cost_unit diferente a null*/
		0.00 --FIRST_VALUE(vly.cost_unit) OVER (PARTITION BY s.id ORDER BY s.adjusted_date DESC, s.stock_move_line_id DESC)
	) as cost_unit,
	sum(s.qty) over (partition by s.product_id order by s.adjusted_date asc, s.stock_move_line_id asc) as qty_acum,
	sl.usage as usage_origin,
	sld.usage as usage_destino ,
	(sp.CREATE_DATE AT TIME ZONE 'UTC' AT TIME ZONE 'America/Guayaquil') as picking_CREATE_DATE,
	(sp.SCHEDULED_DATE AT TIME ZONE 'UTC' AT TIME ZONE 'America/Guayaquil')  as picking_SCHEDULED_DATE   ,
	sp.force_date::date as picking_force_DATE,
	(sp.date_done AT TIME ZONE 'UTC' AT TIME ZONE 'America/Guayaquil')    as picking_date_done ,
			vla.debit,
			vla.credit,
			vla.account_id,
			vla.move_state, 
			vla.move_code,
			vla.move_name, 	
			vla.move_date 
	from all_stock_movements s  
	inner join stock_location sl on sl.id= s.parent_location_id 
	inner join stock_location sld on sld.id= s.parent_location_dest_id 
	left join valuation_layers  vly on vly.stock_move_id=s.stock_move_id 
	/*left join valoracion v on (vly.stock_move_id is null and (v.company_id=s.company_id and v.product_id=s.product_id and 
	(
	    s.adjusted_date::date>=v.fecha_desde  and s.adjusted_date::date<=v.fecha_hasta
	)))*/ 	
	left join valuation_layers_account vla on vla.stock_move_id=vly.stock_move_id and vla.stock_move_id=vly.stock_move_id
	left join stock_warehouse sw on sw.id=s.warehouse_id
	left join stock_warehouse swd on swd.id=s.warehouse_dest_id
	left join stock_picking sp on sp.id=s.picking_id 
	ORDER BY  s.adjusted_date ASC
),
            kardex as (
            	select k.product_id,k.default_code,k.product_template_name,k.stock_move_line_id,k.name,k.reference,
            			k.adjusted_date,
            			k.origen,k.destino,
            			k.qty_in,k.qty_out,k.qty,k.qty_acum,
            			k.cost_unit*k.qty as cost_value,
            			k.cost_unit,k.location_id,k.location_dest_id,
            			k.qty*k.cost_unit as cost_acum ,
            			k.usage_origin,
            			k.usage_destino ,k.categ_id,
            			case when(k.detailed_type='consu') then 'CONSUMIBLE'
            			    when(k.detailed_type='product') then 'PRODUCTO'
            			    when(k.detailed_type='service') then 'SERVICIO'
            			    else 'SIN DEFINIR' end as tipo_producto,
						k.picking_CREATE_DATE,
						k.picking_SCHEDULED_DATE,
						k.picking_force_DATE,
						k.picking_date_done ,
						K.fecha_programada_date ,
						k.debit,
			k.credit,
			k.account_id,
			k.move_state, 
			k.move_code,
			k.move_name, 	
			k.move_date  				
            			from preliminar_kardex k 
            			order by k.adjusted_date asc
            )


            select k.product_id,k.default_code,k.product_template_name,k.stock_move_line_id,k.name,k.reference,
            			k.adjusted_date,
            			coalesce(k.origen, case when(sw.id  is not null) then sw.NAME||'/'||sl.NAME else sl.NAMe end ) as origen,
            			coalesce(k.destino,case when(swd.id  is not null) then swd.NAME||'/'||sld.NAME else sld.NAMe end) as destino,
            			k.qty_in,k.qty_out,k.qty,k.qty_acum,
            			k.cost_unit,
            			k.cost_value,
            			k.cost_acum ,
            			k.usage_origin||'-'||k.usage_destino as tipo_movimiento ,
            			pc.name as categ_name,
            			k.tipo_producto,
						k.picking_CREATE_DATE,
						k.picking_SCHEDULED_DATE,
						k.picking_force_DATE,
						k.picking_date_done ,
						K.fecha_programada_date  ,
						k.debit,
						k.credit,
						coalesce(k.debit,0)-coalesce(k.credit,0) as balance,
						k.account_id,
						k.move_state, 
						k.move_code,
						k.move_name ,
						k.move_date ,
						k.cost_value - (coalesce(k.debit,0)-coalesce(k.credit,0)) as dif_value_contabilidad 
            			from 
            variables inner join 		
            kardex k on k.adjusted_date::date>=variables.date_from and k.adjusted_date::date<=variables.date_to 
            inner join product_category pc on pc.id=k.categ_id
            left join stock_location sl on sl.id=k.location_id and k.origen is null
            left join stock_location sld on sld.id=k.location_dest_id  and k.destino is null
            left join stock_warehouse sw on sw.id=sl.warehouse_id and k.origen is null
            left join stock_warehouse swd on swd.id=sld.warehouse_id  and k.destino is null
			order by 7 asc,5 asc 
            """, (brw_wizard.date_from,brw_wizard.date_to, company_ids, product_ids, warehouse_ids, location_ids,
                  brw_wizard.enable_parent_location))
            print(brw_wizard.date_from, company_ids)
            print(product_ids)
            print(warehouse_ids)
            print(location_ids)
            result = self._cr.dictfetchall()
            if result:
                i, INDEX_ROW = 0, 7
                last_row = INDEX_ROW
                for each_result in result:
                    row = str(INDEX_ROW + i)
                    ws['A' + row] = each_result["product_id"]
                    ws['B' + row] = each_result["default_code"]
                    ws['C' + row] = each_result["product_template_name"]

                    ws['D' + row] = each_result["stock_move_line_id"]
                    ws['E' + row] = each_result["name"]
                    ws['F' + row] = each_result["reference"]

                    datetime_user_tz = each_result["adjusted_date"]#fields.Datetime.context_timestamp(self, each_result["adjusted_date"])
                    #print(datetime_user_tz)
                    ws['G' + row] = datetime_user_tz.strftime("%Y-%m-%d %H:%M:%S")
                    ws['H' + row] = each_result["origen"]
                    ws['I' + row] = each_result["destino"]
                    ws['J' + row] = each_result["qty_in"]
                    ws['K' + row] = each_result["qty_out"]
                    ws['L' + row] = each_result["qty"]

                    if acceso_costo:
                        ws['M' + row] = each_result["cost_unit"]
                        ws['N' + row] = each_result["cost_value"]
                        ws['O' + row] = each_result["tipo_movimiento"]
                        ws['P' + row] = each_result["categ_name"]
                        ws['Q' + row] = each_result["tipo_producto"]
                        ws['R' + row] = each_result["picking_create_date"] and each_result[
                            "picking_create_date"].strftime("%Y-%m-%d %H:%M:%S") or None
                        ws['S' + row] = each_result["picking_scheduled_date"] and each_result[
                            "picking_scheduled_date"].strftime("%Y-%m-%d %H:%M:%S") or None
                        ws['T' + row] = each_result["picking_force_date"] and each_result[
                            "picking_force_date"].strftime("%Y-%m-%d") or None
                        ws['U' + row] = each_result["picking_date_done"] and each_result["picking_date_done"].strftime(
                            "%Y-%m-%d %H:%M:%S") or None
                        ws['V' + row] = each_result["fecha_programada_date"] and each_result[
                            "fecha_programada_date"].strftime(
                            "%Y-%m-%d %H:%M:%S") or None
                        ws['W' + row] = datetime_user_tz.strftime("%Y-%m-%d %H:%M:%S")

                        ws['X' + row] = each_result["move_date"]
                        ws['Y' + row] =each_result["move_code"]
                        ws['Z' + row] = each_result["move_name"]
                        ws['AA' + row] = each_result["debit"]
                        ws['AB' + row] = each_result["credit"]
                        ws['AC' + row] = each_result["balance"]
                        ws['AD' + row] = each_result["dif_value_contabilidad"]
                    else:
                        ws['M' + row]=each_result["tipo_movimiento"]
                        ws['N' + row] = each_result["categ_name"]
                        ws['O' + row] = each_result["tipo_producto"]
                        ws['P' + row] = each_result["picking_create_date"] and each_result["picking_create_date"].strftime("%Y-%m-%d %H:%M:%S") or None
                        ws['Q' + row] = each_result["picking_scheduled_date"] and each_result["picking_scheduled_date"].strftime("%Y-%m-%d %H:%M:%S") or None
                        ws['R' + row] = each_result["picking_force_date"] and each_result["picking_force_date"].strftime("%Y-%m-%d") or None
                        ws['S' + row] = each_result["picking_date_done"] and each_result["picking_date_done"].strftime("%Y-%m-%d %H:%M:%S") or None
                        ws['T' + row] = each_result["fecha_programada_date"] and each_result[
                            "fecha_programada_date"].strftime(
                            "%Y-%m-%d %H:%M:%S") or None
                        ws['U' + row] = datetime_user_tz.strftime("%Y-%m-%d %H:%M:%S")

                    i += 1
                    last_row = INDEX_ROW + i
                if last_row >= INDEX_ROW:
                    thin = Side(border_style="thin", color="000000")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    LETRA_FINAL = acceso_costo and "AD" or 'U'
                    self.set_border(ws, 'A' + str(INDEX_ROW) + ':' + LETRA_FINAL + str(last_row - 1), border)
            ws['G2'] = len(result)
            ws['A1'] = ",".join(brw_wizard.mapped('company_id').mapped('name'))
            ws['B2'] = date_from
            ws['D2'] = date_to
            product_tmpl_ids = brw_wizard.mapped('product_ids').mapped('product_tmpl_id')
            ws['B3'] = product_tmpl_ids and ",".join(product_tmpl_ids.mapped('default_code')) or "TODOS"
            ws['C3'] = product_tmpl_ids and  ",".join(product_tmpl_ids.mapped('name'))  or "TODOS"
            ws['B4'] = ",".join(brw_wizard.mapped('warehouse_ids').mapped('name'))
            ws['B5'] = ",".join(get_list_names(brw_wizard.location_ids))
            wb = self.save_wb(wb, target)
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

class report_kardex_xlsx(models.AbstractModel):
    _name = "report.gps_inventario.report_kardex_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Kardex"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)

        wb = False
        try:#
            acceso_costo=False#self.user_has_groups('account_payment_purchase.group_costo_usuario')
            filename = "reporte_kardex.xlsx"
            # if acceso_costo:
            #     filename="reporte_kardex_costo.xlsx"
            filename=dtFile.join(dir_path, filename)
            dtFile.copyfile(filename, new_filename)

            brw_wizard = self.env["stock.inventory.report"].sudo().browse(docids[-1])
            company_ids=brw_wizard.mapped('company_ids').ids
            product_id=brw_wizard.product_id.id

            warehouse_ids = brw_wizard.mapped('warehouse_ids').ids

            location_ids = brw_wizard.mapped('location_ids').ids

            wb, ws, target = self.open_xlsx(new_filename, load_sheet=True)
            user_tz = self.env.user.tz or 'UTC'
            self._cr.execute(""";WITH variables AS (
    SELECT %s::date as date_from, %s::INTEGER[] AS company_ids,
           ARRAY[%s]::INTeger[] AS product_ids, 
		  %s::INTeger[] AS WAREHOUSE_IDs,
		  %s::INTeger[] AS location_ids ,
		  %s::bool as grouped
		  ),
warehouse_locations as ( 
	select sl.id as location_id ,
	case when(slv.usage='view') then sl.id else  sl.location_id end as parent_location_id,
	case when(slv.usage='view' or not variables.grouped) then sl.id else  sl.location_id end as final_location_id,
	sl.name,
	sl.usage ,sl.warehouse_id
	FROM variables
	inner join stock_location sl on 1=1 
	inner join stock_location slv on sl.location_id=slv.id
	where sl.company_id=any(variables.company_ids) and 
		sl.warehouse_id=any(variables.warehouse_ids)  
		and sl.usage='internal' 
		and (not variables.grouped and  
		sl.id=any(variables.location_ids) ) or 
		(variables.grouped and  
			(
				(sl.location_id=any(variables.location_ids)) or 
				(sl.id=any(variables.location_ids))
			)
		)
),
all_stock_movements as ( 
	SELECT variables.grouped,
	   sm.company_id,
	   pp.id as product_id,
       pp.product_tmpl_id,
       pp.default_code,
       pp.barcode,
       COALESCE(pt.name::json->>'es_EC', pt.name::json->>'en_US') AS product_template_name,
	   sml.id as stock_move_line_id,
       sm.id AS stock_move_id,
       sm.picking_id,
       sm.name,
       sm.reference,
       sml.qty_done as quantity_done,
	   sm.location_id,
	    sm.location_dest_id,
		coalesce(slo.final_location_id,sm.location_id) as parent_location_id,
	    coalesce(sld.final_location_id,sm.location_dest_id) as parent_location_dest_id,
		slo.warehouse_id as warehouse_id,
		sld.warehouse_id as warehouse_dest_id,
       sml.date as adjusted_date,
	   case when sld.usage='internal' then sml.qty_done else  0 end as qty_in,
	   case when slo.usage='internal' then -sml.qty_done else  0 end as qty_out,
	   case when sld.usage='internal' then sml.qty_done else  0 end+case when slo.usage='internal' then -sml.qty_done else  0 end as qty 
	   
	FROM variables
	INNER JOIN stock_move sm 
	    ON sm.company_id = ANY(variables.company_ids) 
	     AND (
         (
            array_length(variables.product_ids, 1) > 0 AND sm.product_id = ANY(variables.product_ids))
            OR (array_length(variables.product_ids, 1) IS NULL OR array_length(variables.product_ids, 1) = 0
          )
        )
	    AND sm.state = 'done' 
	inner join stock_move_line  sml on sml.move_id=sm.id 
	inner join product_product pp on pp.id=sm.product_id
	INNER JOIN product_template pt  ON pt.id = pp.product_tmpl_id 
	
	left join warehouse_locations slo on slo.location_id=sm.location_id  
	left join warehouse_locations sld on sld.location_id=sm.location_dest_id  
	where (slo.location_id is not null or sld.location_id is not null)
) ,
valuation_layers as (
	select svm.stock_move_id,
	sum(svm.value) as cost_value,
	sum(svm.quantity) as cost_qty,
	(case when(sum(svm.quantity)!=0.00) then sum(svm.value)/sum(svm.quantity) else null end ) as cost_unit
	from all_stock_movements asm
	inner join stock_Valuation_layer svm on asm.stock_move_id=svm.stock_move_id
	group by svm.stock_move_id 
),
valoracion_ini as (
     SELECT 
         svl.company_id,
        svl.product_id, 
		svl.quantity,
		sm.date::date as date,
        sm.date AS fecha_hasta, 
        svl.unit_cost,
		(case when(svl.quantity!=0.00) then svl.value/svl.quantity else 0.00 end ) as cost_unit_compute,
        (LAG(DATE(sm.date), 1, '2019-12-31') OVER (PARTITION BY svl.product_id ORDER BY DATE(sm.date))+ INTERVAL '1 DAY')::date AS fecha_Desde
    FROM stock_valuation_layer svl
	inner join variables on svl.company_id=ANY(variables.company_ids)  
				AND (
         (
            array_length(variables.product_ids, 1) > 0 AND svl.product_id = ANY(variables.product_ids))
            OR (array_length(variables.product_ids, 1) IS NULL OR array_length(variables.product_ids, 1) = 0
          )
        )
    INNER JOIN stock_move sm ON sm.id = svl.stock_move_id
    WHERE sm.date = (
            SELECT MAX(sm2.date)
            FROM stock_move sm2
            INNER JOIN stock_valuation_layer svl2 ON sm2.id = svl2.stock_move_id
			inner join variables on svl2.company_id=ANY(variables.company_ids) 
					AND (
         (
            array_length(variables.product_ids, 1) > 0 AND svl2.product_id = ANY(variables.product_ids))
            OR (array_length(variables.product_ids, 1) IS NULL OR array_length(variables.product_ids, 1) = 0
          )
        )
            WHERE svl2.product_id = svl.product_id
                AND svl2.company_id = svl.company_id
                AND DATE(sm2.date) = DATE(sm.date)
        )
),costeos as (
	SELECT 
		  company_id,
		  product_id,
		  fecha_Desde,
		  fecha_hasta,
		  -- Movimiento: la cantidad (positivo para entrada, negativo para salida)
		  quantity AS movimiento,
		  cost_unit_compute,
		  -- Valor del movimiento: cantidad x costo unitario de la transacción
		  quantity * cost_unit_compute AS valor_movimiento,
		  -- Saldo acumulado de cantidad
		  SUM(quantity) OVER (PARTITION BY product_id ORDER BY fecha_hasta, fecha_Desde) AS saldo_cantidad,
		  -- Saldo acumulado de valor
		  SUM(quantity * cost_unit_compute) OVER (PARTITION BY product_id ORDER BY fecha_hasta, fecha_Desde) AS saldo_valor,
		  -- Costo promedio del saldo (si el saldo de cantidad es 0, se evita la división)
		  CASE 
		    WHEN SUM(quantity) OVER (PARTITION BY product_id ORDER BY fecha_hasta, fecha_Desde) <> 0
		    THEN SUM(quantity * cost_unit_compute) OVER (PARTITION BY product_id ORDER BY fecha_hasta, fecha_Desde) /
		         SUM(quantity) OVER (PARTITION BY product_id ORDER BY fecha_hasta, fecha_Desde)
		    ELSE 0
		  END AS costo_promedio_saldo
		FROM valoracion_ini
		ORDER BY date, fecha_Desde--fecha_hasta
),
valoracion as (
    SELECT company_id,product_id, fecha_desde::date AS fecha_desde, 
		fecha_hasta::date AS fecha_hasta,  costo_promedio_saldo as cost_unit
    FROM costeos
    UNION     
    SELECT 
        v.company_id,
        v.product_id, 
        (MAX(v.fecha_hasta) OVER (PARTITION BY v.company_id, v.product_id) + INTERVAL '1 DAY')::date AS fecha_desde, 
        (CURRENT_DATE + INTERVAL '1 DAY')::date AS fecha_hasta,
        LAST_VALUE(v.costo_promedio_saldo) OVER (PARTITION BY v.company_id, v.product_id ORDER BY v.fecha_hasta ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING) AS cost_unit
    FROM costeos v
),
preliminar_kardex as ( 
	select s.*,sp.name as picking_name, 
	case when(sw.id  is not null) then sw.NAME||'/'||sl.NAME else NULL end as origen,
	case when(swd.id  is not null) then swd.NAME||'/'||sld.NAME else NULL end as destino ,
	/*vly.cost_counter,
	vly.cost_value,
	vly.cost_qty,*/
	coalesce(vly.cost_unit,v.cost_unit,/*necesito sacar el ultimo costo es decir el ultimo vly.cost_unit diferente a null*/
		0.00 --FIRST_VALUE(vly.cost_unit) OVER (PARTITION BY s.id ORDER BY s.adjusted_date DESC, s.stock_move_line_id DESC)
	) as cost_unit,
	sum(s.qty) over (partition by s.product_id order by s.adjusted_date asc, s.stock_move_line_id asc) as qty_acum
	from all_stock_movements s  
	inner join stock_location sl on sl.id= s.parent_location_id 
	inner join stock_location sld on sld.id= s.parent_location_dest_id 
	left join valuation_layers  vly on vly.stock_move_id=s.stock_move_id
	left join valoracion v on (vly.stock_move_id is null and (v.company_id=s.company_id and v.product_id=s.product_id and 
	(
	    s.adjusted_date::date>=v.fecha_desde  and s.adjusted_date::date<=v.fecha_hasta
	))) 	
	left join stock_warehouse sw on sw.id=s.warehouse_id
	left join stock_warehouse swd on swd.id=s.warehouse_dest_id
	left join stock_picking sp on sp.id=s.picking_id 
	ORDER BY  s.adjusted_date ASC
),
kardex as (
	select k.product_id,k.default_code,k.product_template_name,k.stock_move_line_id,k.name,k.reference,
			k.adjusted_date,
			k.origen,k.destino,
			k.qty_in,k.qty_out,k.qty,k.qty_acum,
			k.cost_unit*k.qty as cost_value,
			k.cost_unit,k.location_id,k.location_dest_id,
			k.qty*k.cost_unit as cost_acum 
			from preliminar_kardex k 
			order by k.adjusted_date asc
)


/*
select k.product_id,
	k.default_code,
	k.product_template_name,0 as stock_move_line_id,
    'INVENTARIO INICIAL' as name,
    'INVENTARIO INICIAL' as reference,
    variables.date_from as adjusted_date,
    '' as origen,
    '' as destino,
    sum(k.qty_in) as qty_in,
    sum(k.qty_out) as qty_out,
    sum(k.qty) as qty,
    sum(k.qty) as qty_acum,
    case when(sum(k.qty)!=0.00) then sum(k.cost_acum )/sum(k.qty) else 0.00 end as cost_unit,
    sum(k.cost_acum) as cost_value,
    sum(k.cost_acum) as cost_acum
from variables
inner join kardex k on 1=1
where  k.adjusted_date::date<=variables.date_from
group by k.product_id,k.default_code,k.product_template_name,variables.date_from 
union */
select k.product_id,k.default_code,k.product_template_name,k.stock_move_line_id,k.name,k.reference,
			k.adjusted_date,
			coalesce(k.origen, case when(sw.id  is not null) then sw.NAME||'/'||sl.NAME else sl.NAMe end ) as origen,
			coalesce(k.destino,case when(swd.id  is not null) then swd.NAME||'/'||sld.NAME else sld.NAMe end) as destino,
			k.qty_in,k.qty_out,k.qty,k.qty_acum,
			k.cost_unit,
			k.cost_value,
			k.cost_acum 
			from 
variables inner join 		
kardex k on k.adjusted_date::date>variables.date_from
left join stock_location sl on sl.id=k.location_id and k.origen is null
left join stock_location sld on sld.id=k.location_dest_id  and k.destino is null
left join stock_warehouse sw on sw.id=sl.warehouse_id and k.origen is null
left join stock_warehouse swd on swd.id=sld.warehouse_id  and k.destino is null
order by 7 asc,5 asc
""", (brw_wizard.date_from, company_ids,product_id,warehouse_ids,location_ids,brw_wizard.enable_parent_location))
            print(brw_wizard.date_from,company_ids)
            print(product_id)
            print(warehouse_ids)
            print(location_ids)
            result = self._cr.dictfetchall()
            print(result)
            warehouse_ids = brw_wizard.mapped('warehouse_ids').ids
            location_ids = brw_wizard.mapped('location_ids').ids
            i, INDEX_ROW = 0, 7
            last_row = INDEX_ROW

            from datetime import datetime, time
            # fecha_obj = datetime.strptime(date_to, "%Y-%m-%d")
            fecha_completa = datetime.combine(brw_wizard.date_from, time(23, 59, 59))
            fecha_corte = fecha_completa.strftime("%Y-%m-%d %H:%M:%S")

            product_srch = brw_wizard.product_id.with_context(to_date=fecha_corte)
            qty=0.00
            for each_product in product_srch:
                for each_location_id in location_ids:
                    brw_product = each_product.with_context(location=each_location_id)
                    qty+=brw_product.qty_available
            ######################################################
            row = str(INDEX_ROW + i)
            ws['A' + row] = product_srch.id
            ws['B' + row] =product_srch.default_code
            ws['C' + row] = product_srch.name
            ws['D' + row] = None
            ws['E' + row] = 'INVENTARIO INICIAL'
            ws['F' + row] =None
            #datetime_user_tz = fields.Datetime.context_timestamp(self,brw_wizard.date_from)
            ws['G' + row] = brw_wizard.date_from.strftime("%Y-%m-%d")
            ws['H' + row] = None
            ws['I' + row] = None
            ws['J' + row] = None
            ws['K' + row] = None
            ws['L' + row] = qty
            ws['M' + row] = qty
            # cost_acum+= each_result["cost_acum"] or 0.00
            # if acceso_costo:
            #     ws['N' + row] = each_result["cost_unit"]
            #     ws['O' + row] = each_result["cost_value"]
            #     ws['P' + row] =cost_acum
            i += 1
            last_row = INDEX_ROW + i
            #######################################################
            qty_acum=qty
            if result:

                for each_result in result:
                    row = str(INDEX_ROW + i)
                    ws['A' + row] = each_result["product_id"]
                    ws['B' + row] = each_result["default_code"]
                    ws['C' + row] = each_result["product_template_name"]
                    ws['D' + row] = each_result["stock_move_line_id"]
                    ws['E' + row] = each_result["name"]
                    ws['F' + row] = each_result["reference"]
                    datetime_user_tz = fields.Datetime.context_timestamp(self, each_result["adjusted_date"])
                    ws['G' + row] = datetime_user_tz.strftime("%Y-%m-%d %H:%M:%S")
                    ws['H' + row] = each_result["origen"]
                    ws['I' + row] = each_result["destino"]
                    ws['J' + row] = each_result["qty_in"]
                    ws['K' + row] = each_result["qty_out"]
                    ws['L' + row] = each_result["qty"]
                    qty_acum+=each_result["qty"]
                    ws['M' + row] =qty_acum
                    # cost_acum+= each_result["cost_acum"] or 0.00
                    # if acceso_costo:
                    #     ws['N' + row] = each_result["cost_unit"]
                    #     ws['O' + row] = each_result["cost_value"]
                    #     ws['P' + row] =cost_acum
                    i += 1
                    last_row = INDEX_ROW + i
                if last_row >= INDEX_ROW:
                    thin = Side(border_style="thin", color="000000")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    LETRA_FINAL=acceso_costo and "P" or 'M'
                    self.set_border(ws, 'A' + str(INDEX_ROW) + ':' + LETRA_FINAL+str(last_row - 1), border)
            ws['G2'] = len(result)
            ws['A1'] = ",".join( brw_wizard.mapped('company_id').mapped('name'))
            ws['B2'] = brw_wizard.date_from
            ws['B3'] = brw_wizard.product_id.default_code
            ws['C3'] =  brw_wizard.product_id.name
            ws['B4'] = ",".join(brw_wizard.mapped('warehouse_ids').mapped('name'))
            ws['B5'] = ",".join(get_list_names(brw_wizard.location_ids))
            ws['G3'] = brw_wizard.product_id.standard_price
            wb = self.save_wb(wb, target)
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

class report_date_inventory_xlsx(models.AbstractModel):
    _name = "report.gps_inventario.report_date_inventory_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Inventario a la fecha"

    def create_xlsx_report(self, docids, data):

        DEC=2
        EXT = "xlsx"
        dir_path = dtFile.get_path_file(__file__)
        new_filename = dtFile.create(EXT)

        wb = False
        try:
            acceso_costo = self.user_has_groups('account_payment_purchase.group_costo_usuario')
            filename = "reporte_inventario_fecha.xlsx"
            if acceso_costo:
                filename = "reporte_inventario_fecha_costo.xlsx"
            filename = dtFile.join(dir_path, filename)
            dtFile.copyfile(filename, new_filename)

            brw_wizard = self.env["stock.inventory.report"].sudo().browse(docids[-1])
            company_ids = brw_wizard.mapped('company_ids').ids
            product_ids = brw_wizard.mapped('product_ids').ids

            date_to = brw_wizard.date_to if brw_wizard.date_to is not None else fields.Date.context_today(self)
            wb, ws, target = self.open_xlsx(new_filename, load_sheet=True)

            domain = [('type', '=', 'product')]
            if product_ids:
                domain += [('id', 'in', product_ids)]
            product_srch = self.env["product.product"].sudo().search(domain)
            from datetime import datetime, time
            # fecha_obj = datetime.strptime(date_to, "%Y-%m-%d")
            fecha_completa = datetime.combine(date_to, time(23, 59, 59))
            fecha_corte = fecha_completa.strftime("%Y-%m-%d %H:%M:%S")
            if not brw_wizard.by_company:
                warehouse_ids = brw_wizard.mapped('warehouse_ids').ids
                location_ids = brw_wizard.mapped('location_ids').ids

    #             user_tz = self.env.user.tz or 'UTC'
    #             self._cr.execute(""";WITH variables AS (
    #     SELECT  %s::date as date_to,
    #             %s::INTEGER[] AS company_ids,
    #            %s::INTeger[] AS product_ids,
    #           %s::INTeger[] AS WAREHOUSE_IDs,
    #           %s::INTeger[] AS location_ids ,
    #           %s::bool as grouped
    # ),
    # warehouse_locations as (
    #     select sl.id as location_id ,
    #     case when(slv.usage='view') then sl.id else  sl.location_id end as parent_location_id,
    #     case when(slv.usage='view' or not variables.grouped) then sl.id else  sl.location_id end as final_location_id,
    #     sl.name,
    #     sl.usage ,sl.warehouse_id
    #     FROM variables
    #     inner join stock_location sl on 1=1
    #     inner join stock_location slv on sl.location_id=slv.id
    #     where sl.company_id=any(variables.company_ids) and
    #         sl.warehouse_id=any(variables.warehouse_ids)
    #         and sl.usage='internal'
    #         and (not variables.grouped and
    #         sl.id=any(variables.location_ids) ) or
    #         (variables.grouped and
    #             (
    #                 (sl.location_id=any(variables.location_ids)) or
    #                 (sl.id=any(variables.location_ids))
    #             )
    #         )
    # ),
    # all_stock_movements as (
    #     SELECT variables.grouped,
    #        sm.company_id,
    #        pp.id as product_id,
    #        pp.product_tmpl_id,
    #        pp.default_code,
    #        pp.barcode,
    #        COALESCE(pt.name::json->>'es_EC', pt.name::json->>'en_US') AS product_template_name,
    #        sml.id as stock_move_line_id,
    #        sm.id AS stock_move_id,
    #        sm.picking_id,
    #        sm.name,
    #        sm.reference,
    #        sml.qty_done as quantity_done,
    #        sm.location_id,
    #         sm.location_dest_id,
    #         coalesce(slo.final_location_id,sm.location_id) as parent_location_id,
    #         coalesce(sld.final_location_id,sm.location_dest_id) as parent_location_dest_id,
    #         slo.warehouse_id as warehouse_id,
    #         sld.warehouse_id as warehouse_dest_id,
    #        sml.date as adjusted_date,
    #        case when sld.usage='internal' then sml.qty_done else  0 end as qty_in,
    #        case when slo.usage='internal' then -sml.qty_done else  0 end as qty_out,
    #        case when sld.usage='internal' then sml.qty_done else  0 end+case when slo.usage='internal' then -sml.qty_done else  0 end as qty
    #
    #     FROM variables
    #     INNER JOIN stock_move sm
    #         ON sm.company_id = ANY(variables.company_ids)
    #          AND (
    #          (
    #             array_length(variables.product_ids, 1) > 0 AND sm.product_id = ANY(variables.product_ids))
    #             OR (array_length(variables.product_ids, 1) IS NULL OR array_length(variables.product_ids, 1) = 0
    #           )
    #         )
    #         AND sm.state in ('done'  )
    #     inner join stock_move_line  sml on sml.move_id=sm.id
    #     inner join product_product pp on pp.id=sm.product_id
    #     INNER JOIN product_template pt  ON pt.id = pp.product_tmpl_id
    #     left join warehouse_locations slo on slo.location_id=sm.location_id
    #     left join warehouse_locations sld on sld.location_id=sm.location_dest_id
    #     where (slo.location_id is not null or sld.location_id is not null)
    # ) ,
    # valuation_layers as (
    #     select svm.stock_move_id,
    #     sum(svm.value) as cost_value,
    #     sum(svm.quantity) as cost_qty,
    #     (case when(sum(svm.quantity)!=0.00) then sum(svm.value)/sum(svm.quantity) else null end ) as cost_unit
    #     from all_stock_movements asm
    #     inner join stock_Valuation_layer svm on asm.stock_move_id=svm.stock_move_id
    #     group by svm.stock_move_id
    # ),
    # valoracion_ini as (
    #      SELECT
    #          svl.company_id,
    #         svl.product_id,
    #         svl.quantity,
    #         sm.date::date as date,
    #         sm.date AS fecha_hasta,
    #         svl.unit_cost,
    #         (case when(svl.quantity!=0.00) then svl.value/svl.quantity else 0.00 end ) as cost_unit_compute,
    #         (LAG(DATE(sm.date), 1, '2019-12-31') OVER (PARTITION BY svl.product_id ORDER BY DATE(sm.date))+ INTERVAL '1 DAY')::date AS fecha_Desde
    #     FROM stock_valuation_layer svl
    #     inner join variables on svl.company_id=ANY(variables.company_ids)
    #                 AND (
    #          (
    #             array_length(variables.product_ids, 1) > 0 AND svl.product_id = ANY(variables.product_ids))
    #             OR (array_length(variables.product_ids, 1) IS NULL OR array_length(variables.product_ids, 1) = 0
    #           )
    #         )
    #     INNER JOIN stock_move sm ON sm.id = svl.stock_move_id
    #     WHERE sm.date = (
    #             SELECT MAX(sm2.date)
    #             FROM stock_move sm2
    #             INNER JOIN stock_valuation_layer svl2 ON sm2.id = svl2.stock_move_id
    #             inner join variables on svl2.company_id=ANY(variables.company_ids)
    #                     AND (
    #          (
    #             array_length(variables.product_ids, 1) > 0 AND svl2.product_id = ANY(variables.product_ids))
    #             OR (array_length(variables.product_ids, 1) IS NULL OR array_length(variables.product_ids, 1) = 0
    #           )
    #         )
    #             WHERE svl2.product_id = svl.product_id
    #                 AND svl2.company_id = svl.company_id
    #                 AND DATE(sm2.date) = DATE(sm.date)
    #         )
    # ),costeos as (
    #     SELECT
    #           company_id,
    #           product_id,
    #           fecha_Desde,
    #           fecha_hasta,
    #           -- Movimiento: la cantidad (positivo para entrada, negativo para salida)
    #           quantity AS movimiento,
    #           cost_unit_compute,
    #           -- Valor del movimiento: cantidad x costo unitario de la transacción
    #           quantity * cost_unit_compute AS valor_movimiento,
    #           -- Saldo acumulado de cantidad
    #           SUM(quantity) OVER (PARTITION BY product_id ORDER BY fecha_hasta, fecha_Desde) AS saldo_cantidad,
    #           -- Saldo acumulado de valor
    #           SUM(quantity * cost_unit_compute) OVER (PARTITION BY product_id ORDER BY fecha_hasta, fecha_Desde) AS saldo_valor,
    #           -- Costo promedio del saldo (si el saldo de cantidad es 0, se evita la división)
    #           CASE
    #             WHEN SUM(quantity) OVER (PARTITION BY product_id ORDER BY fecha_hasta, fecha_Desde) <> 0
    #             THEN SUM(quantity * cost_unit_compute) OVER (PARTITION BY product_id ORDER BY fecha_hasta, fecha_Desde) /
    #                  SUM(quantity) OVER (PARTITION BY product_id ORDER BY fecha_hasta, fecha_Desde)
    #             ELSE 0
    #           END AS costo_promedio_saldo
    #         FROM valoracion_ini
    #         ORDER BY date, fecha_Desde--fecha_hasta
    # ),
    # valoracion as (
    #     SELECT company_id,product_id, fecha_desde::date AS fecha_desde,
    #         fecha_hasta::date AS fecha_hasta,  costo_promedio_saldo as cost_unit
    #     FROM costeos
    #     UNION
    #     SELECT
    #         v.company_id,
    #         v.product_id,
    #         (MAX(v.fecha_hasta) OVER (PARTITION BY v.company_id, v.product_id) + INTERVAL '1 DAY')::date AS fecha_desde,
    #         (CURRENT_DATE + INTERVAL '1 DAY')::date AS fecha_hasta,
    #         LAST_VALUE(v.costo_promedio_saldo) OVER (PARTITION BY v.company_id, v.product_id ORDER BY v.fecha_hasta ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING) AS cost_unit
    #     FROM costeos v
    # ),
    # preliminar_kardex as (
    #     select s.*,sp.name as picking_name,
    #     case when(sw.id  is not null) then sw.NAME||'/'||sl.NAME else NULL end as origen,
    #     case when(swd.id  is not null) then swd.NAME||'/'||sld.NAME else NULL end as destino ,
    #     /*vly.cost_counter,
    #     vly.cost_value,
    #     vly.cost_qty,*/
    #     coalesce(vly.cost_unit,v.cost_unit,/*necesito sacar el ultimo costo es decir el ultimo vly.cost_unit diferente a null*/
    #         0.00 --FIRST_VALUE(vly.cost_unit) OVER (PARTITION BY s.id ORDER BY s.adjusted_date DESC, s.stock_move_line_id DESC)
    #     ) as cost_unit,
    #     sum(s.qty) over (partition by s.product_id order by s.adjusted_date asc, s.stock_move_line_id asc) as qty_acum
    #     from all_stock_movements s
    #     inner join stock_location sl on sl.id= s.parent_location_id
    #     inner join stock_location sld on sld.id= s.parent_location_dest_id
    #     left join valuation_layers  vly on vly.stock_move_id=s.stock_move_id
    #     left join valoracion v on (vly.stock_move_id is null and (v.company_id=s.company_id and v.product_id=s.product_id and
    #     (
    #         s.adjusted_date::date>=v.fecha_desde  and s.adjusted_date::date<=v.fecha_hasta
    #     )))
    #     left join stock_warehouse sw on sw.id=s.warehouse_id
    #     left join stock_warehouse swd on swd.id=s.warehouse_dest_id
    #     left join stock_picking sp on sp.id=s.picking_id
    #     ORDER BY  s.adjusted_date ASC
    # ),
    # kardex as (
    #     select k.product_id,k.default_code,k.product_template_name,k.stock_move_line_id,k.name,k.reference,
    #             k.adjusted_date,
    #             k.origen,k.destino,
    #             k.qty_in,k.qty_out,k.qty,k.qty_acum,
    #             k.cost_unit*k.qty as cost_value,
    #             k.cost_unit,k.location_id,k.location_dest_id,
    #             k.qty*k.cost_unit as cost_acum
    #             from preliminar_kardex k
    #             order by k.adjusted_date asc
    # )
    #
    # --
    # select k.product_id,k.default_code,k.product_template_name,
    #             coalesce(k.origen,k.destino) as ubicacion,
    #             sum(k.qty) as qty,
    #             sum(k.cost_acum) as cost_acum,
    #             cast(COALESCE(MAX(k.cost_unit) FILTER (WHERE k.cost_unit IS NOT NULL),
    #                 case when(sum(k.qty)!=0.00) then sum(k.cost_acum)/sum(k.qty) else 0.00 end,
    #                 0 -- Valor por defecto en caso de que todos sean NULL
    #             )AS DECIMAL(16,8)) AS cost_unit
    #             from
    # variables inner join
    # kardex k on k.adjusted_date::date<=variables.date_to
    # GROUP BY
    #     k.product_id,
    #     k.default_code,
    #     k.product_template_name,
    #     COALESCE(k.origen, k.destino)
    # ORDER BY
    #     k.default_code ASC,
    #     COALESCE(k.origen, k.destino) ASC;""", (date_to, company_ids, product_ids, warehouse_ids, location_ids,brw_wizard.enable_parent_location))
    #             print(brw_wizard.date_from, company_ids)
    #             print(product_ids)
    #             print(warehouse_ids)
    #             print(location_ids)
    #             result = self._cr. dictfetchall()
                product_srch = product_srch.with_context(to_date=fecha_corte)
                result = []
                for each_product in product_srch:
                    for each_location_id in location_ids:
                        brw_location=self.env["stock.location"].sudo().browse(each_location_id)
                        brw_product=each_product.with_context(location=each_location_id)
                        result.append({
                            "product_id": brw_product.id,
                            "default_code": brw_product.default_code,
                            "product_template_name": brw_product.name,
                            "ubicacion": (brw_location.warehouse_id and brw_location.warehouse_id.name+'/' or '')+brw_location.name,
                            "qty": brw_product.qty_available,
                            "cost_unit": brw_product.standard_price,
                            "cost_acum": brw_product.qty_available * brw_product.standard_price
                        })
            else:
                product_srch=product_srch.with_context(to_date=fecha_corte)
                result=[]
                for each_product in product_srch:
                    result.append({
                        "product_id":each_product.id,
                        "default_code": each_product.default_code,
                        "product_template_name": each_product.name,
                        "ubicacion": "",
                        "qty": each_product.qty_available,
                        "cost_unit":each_product.standard_price,
                        "cost_acum": each_product.qty_available*each_product.standard_price
                    })

            if result:
                i, INDEX_ROW = 0, 7
                last_row = INDEX_ROW
                for each_result in result:
                    row = str(INDEX_ROW + i)
                    ws['A' + row] = each_result["product_id"]
                    ws['B' + row] = each_result["default_code"]
                    ws['C' + row] = each_result["product_template_name"]

                    ws['D' + row] = each_result["ubicacion"]
                    ws['E' + row] = each_result["qty"]

                    if acceso_costo:
                        ws['F' + row] = round(each_result["cost_acum"],DEC)
                        ws['G' + row] = round(each_result["cost_unit"]*each_result["qty"],DEC)
                        ws['H' + row] = each_result["cost_unit"]
                    i += 1
                    last_row = INDEX_ROW + i
                if last_row >= INDEX_ROW:
                    thin = Side(border_style="thin", color="000000")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    LETRA_FINAL = acceso_costo and "H" or 'E'
                    self.set_border(ws, 'A' + str(INDEX_ROW) + ':' + LETRA_FINAL + str(last_row - 1), border)
            ws['F2'] = len(result)
            ws['A1'] = ",".join(brw_wizard.mapped('company_id').mapped('name'))
            ws['B2'] =date_to
            product_tmpl_ids = brw_wizard.mapped('product_ids').mapped('product_tmpl_id')
            warehouse_ids=brw_wizard.mapped('warehouse_ids')
            ws['B3'] = product_tmpl_ids and ",".join(product_tmpl_ids.mapped('default_code')) or "TODOS"
            ws['C3'] = product_tmpl_ids and ",".join(product_tmpl_ids.mapped('name')) or "TODOS"
            ws['B4'] = warehouse_ids and ",".join(warehouse_ids.mapped('name')) or "TODOS"
            ws['B5'] = brw_wizard.location_ids and ",".join(get_list_names(brw_wizard.location_ids)) or "TODOS"
            wb = self.save_wb(wb, target)
        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            wb = self.close_wb(wb)
        filecontent = dtFile.get_binary(new_filename)
        return filecontent, EXT

class report_adjust_inventory_xlsx(models.AbstractModel):
    _name = "report.gps_inventario.report_adjust_inventory_xlsx"
    _inherit = "report.report_xlsx.abstract"
    _description = "Reportes de Ajuste de Inventario"

    def create_xlsx_report(self, docids, data):
        EXT = "xlsx"
        filecontent=False
        OBJ_AJUST=self.env["inventory.document.adjust"].sudo()
        DSCR_STATES=dict(OBJ_AJUST._get_selection_state())
        try:
            acceso_costo = self.env.user.has_group('gps_inventario.group_ajuste_inventario_manager') or self.env.user.has_group(
                    'gps_inventario.group_ajuste_inventario_costo_manager')

            brw_wizard = self.env["stock.inventory.report"].sudo().browse(docids[-1])
            company_id = brw_wizard.company_id.id
            warehouse_ids = brw_wizard.mapped('warehouse_ids').ids

            OBJ_LOCATION = self.env["stock.location"].sudo()
            locations = OBJ_LOCATION.search([
                ('usage', '=', 'internal'),  # Solo ubicaciones internas
                #('id', 'not in', OBJ_LOCATION.search([('location_id', '!=', False)]).ids),  # Ubicaciones sin hijas
                ('warehouse_id', 'in', warehouse_ids and warehouse_ids or [-1, -1])
            ])

            date_from=brw_wizard.date_from
            date_to= brw_wizard.date_to

            record=OBJ_AJUST.search([('company_id','=',company_id),
                                                                      ('state','not in',('draft','annulled')),
                                                                      ('date_from','>=',date_from),
                                                                      ('date_from', '<=', date_to),
                                                                        ('stock_location_id','in',locations.ids )
                                                                      ])


            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output)
            sheet = workbook.add_worksheet('Adjustments')

            title_format = workbook.add_format({'bold': True, 'font_size': 14, 'align': 'center'})
            header_format = workbook.add_format({'bold': True, 'align': 'left'})
            value_format = workbook.add_format({'align': 'left'})

            row = 4
            # Encabezados de la tabla
            if acceso_costo:
                headers = ['# AJUSTE','ESTADO AJUSTE','#ID', 'Ubicación', 'Ref. Anterior', 'Ref. Interna', 'Nombre', 'Cantidad', 'Comentario',
                           'Unidad Medida', 'Stock', 'Ajuste','Costo', 'Aplicar']
            else:
                headers = ['# AJUSTE','ESTADO AJUSTE','#ID', 'Ubicación', 'Ref. Anterior', 'Ref. Interna', 'Nombre', 'Cantidad', 'Comentario',
                           'Unidad Medida', 'Aplicar']

            for col, header in enumerate(headers):
                sheet.write(row, col, header, header_format)
            row = row + 1
            i=1
            for each_record in record:
                for line in each_record.line_ids:
                    sheet.write(row, 0, each_record.id)
                    sheet.write(row, 1, DSCR_STATES[each_record.state])
                    sheet.write(row, 2, line.id)
                    sheet.write(row, 3, line.stock_location_id.display_name or '')
                    sheet.write(row, 4, line.product_id.referencia_anterior or '')
                    sheet.write(row, 5, line.product_id.default_code or '')
                    sheet.write(row, 6, line.product_id.name or '')
                    sheet.write(row, 7, line.quantity)
                    sheet.write(row, 8, line.comments or '')
                    sheet.write(row, 9, line.product_id.uom_id.name or '')
                    column=10
                    if acceso_costo:
                        sheet.write(row, column, line.stock or 0)
                        column += 1
                        sheet.write(row, column, line.adjust or 0)
                        column += 1
                        sheet.write(row, column, line.standard_price or 0)
                        column+=1
                    sheet.write(row, column, line.apply and 'SI' or 'NO')
                    row += 1
                    i+=1

            sheet.write('A1',brw_wizard.company_id.name, title_format)
            # sheet.merge_range('E1:H1', 'AJUSTE DE INVENTARIO', title_format)
            sheet.write('G1', 'REPORTE DE AJUSTES DE INVENTARIOS', title_format)

            # Detalles de la cabecera
            sheet.write('A2', 'RANGO DE FECHA:', header_format)
            sheet.write('B2',brw_wizard.date_from and  str(brw_wizard.date_from) or '', value_format)
            sheet.write('C3',brw_wizard.date_to and str(brw_wizard.date_to) or '', value_format)

            sheet.write('A3', 'BODEGAS:', header_format)
            sheet.write('B3',",".join(brw_wizard.mapped('warehouse_ids').mapped('name')), value_format)

            sheet.write('E2', '# LINEAS:', header_format)
            sheet.write('F2', i, value_format)

            sheet.write('E3', 'AJUSTES:', header_format)
            sheet.write('F3', ",".join([str(each_record.id) for each_record in record]), value_format)

            # Guardar el archivo en memoria
            workbook.close()
            output.seek(0)
            #print(sheet,workbook.filename)
            # Convertir el archivo a base64 y guardarlo en el campo
            filecontent = output.read()
            output.close()

        except Exception as e:
            _logger.warning("error create_xlsx_report %s" % (str(e),))
        finally:
            pass
        return filecontent, EXT
